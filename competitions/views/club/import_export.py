"""
Module pour l'import et l'export des données du club.
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.utils.translation import gettext_lazy as _
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.utils import timezone

from competitions.models import Practitioner, Club
from competitions.utils.decorators import club_required
from competitions.utils.permission_helpers import manual_permission_check

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime
import os


import logging
logger = logging.getLogger(__name__)

def get_user_club(request):
    """Récupère le club de l'utilisateur connecté."""
    club = None
    if hasattr(request.user, 'club') and request.user.club:
        club = request.user.club
    else:
        club = Club.objects.filter(owner=request.user).first()
    return club

@login_required
@manual_permission_check('club.manage_members')
def import_export_data(request):
    """Vue pour l'import/export de données en Excel."""
    # Récupérer le club de l'utilisateur
    club = get_user_club(request)

    if not club:
        messages.error(request, _("Vous devez être responsable de club pour accéder à cette page."))
        return redirect('competitions:dashboard')
    
    if request.method == 'POST':
        if 'export' in request.POST:
            # Logique d'export en Excel
            try:
                response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename="{club.name}_pratiquants_{datetime.now().strftime("%Y%m%d")}.xlsx"'
                
                # Créer un nouveau classeur Excel
                workbook = Workbook()
                worksheet = workbook.active
                worksheet.title = "Pratiquants"
                
                # Définir les en-têtes avec style
                headers = [
                    'Nom', 'Prénom', 'Date de naissance', 'Grade', 'Email', 'Licence', 
                    'Genre', 'Téléphone', 'Adresse', 'Ville', 'Code postal', 'Poids (kg)', 
                    'Taille (cm)', 'Certificat médical', 'Nationalité', 'Statut'
                ]
                
                for col_num, header in enumerate(headers, 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.value = header
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                    
                    # Ajouter une bordure
                    thin_border = Border(
                        left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin')
                    )
                    cell.border = thin_border
                
                # Vérifier si le club a une organisation associée
                club_organization = club.organization or getattr(club, 'as_organization', None)
                
                if not club_organization:
                    messages.warning(request, _("Aucune organisation associée trouvée pour ce club."))
                    practitioners = Practitioner.objects.none()
                else:
                    # Ajouter les données des pratiquants
                    practitioners = Practitioner.objects.filter(organization=club_organization).order_by('last_name', 'first_name')
                
                if not practitioners.exists():
                    # Message si aucun pratiquant n'est trouvé
                    worksheet.append(['Aucun pratiquant n\'a été trouvé dans votre club.'])
                    worksheet.merge_cells(f'A2:P2')
                    cell = worksheet.cell(row=2, column=1)
                    cell.alignment = Alignment(horizontal='center')
                else:
                    # Ajouter les données
                    for row_num, practitioner in enumerate(practitioners, 2):
                        # Nom
                        cell = worksheet.cell(row=row_num, column=1, value=practitioner.last_name)
                        cell.alignment = Alignment(horizontal='left')
                        
                        # Prénom
                        cell = worksheet.cell(row=row_num, column=2, value=practitioner.first_name)
                        cell.alignment = Alignment(horizontal='left')
                        
                        # Date de naissance
                        if practitioner.birth_date:
                            cell = worksheet.cell(row=row_num, column=3, value=practitioner.birth_date)
                            cell.number_format = 'yyyy-mm-dd'
                        else:
                            worksheet.cell(row=row_num, column=3, value="")
                        
                        # Grade
                        worksheet.cell(row=row_num, column=4, value=practitioner.grade or "")
                        
                        # Email
                        worksheet.cell(row=row_num, column=5, value=practitioner.email or "")
                        
                        # Licence
                        worksheet.cell(row=row_num, column=6, value=practitioner.license_number or "")
                        
                        # Genre
                        gender_map = {'male': 'Homme', 'female': 'Femme', 'other': 'Autre'}
                        worksheet.cell(row=row_num, column=7, value=gender_map.get(practitioner.gender, ""))
                        
                        # Téléphone
                        worksheet.cell(row=row_num, column=8, value=practitioner.phone or "")
                        
                        # Adresse
                        worksheet.cell(row=row_num, column=9, value=practitioner.address or "")
                        
                        # Ville
                        worksheet.cell(row=row_num, column=10, value=practitioner.city or "")
                        
                        # Code postal
                        worksheet.cell(row=row_num, column=11, value=practitioner.postal_code or "")
                        
                        # Poids
                        if hasattr(practitioner, 'weight') and practitioner.weight:
                            worksheet.cell(row=row_num, column=12, value=float(practitioner.weight))
                        else:
                            worksheet.cell(row=row_num, column=12, value="")
                        
                        # Taille
                        if hasattr(practitioner, 'height') and practitioner.height:
                            worksheet.cell(row=row_num, column=13, value=float(practitioner.height))
                        else:
                            worksheet.cell(row=row_num, column=13, value="")
                        
                        # Certificat médical
                        if hasattr(practitioner, 'medical_certificate_date') and practitioner.medical_certificate_date:
                            cell = worksheet.cell(row=row_num, column=14, value=practitioner.medical_certificate_date)
                            cell.number_format = 'yyyy-mm-dd'
                        else:
                            worksheet.cell(row=row_num, column=14, value="")
                        
                        # Nationalité
                        if hasattr(practitioner, 'nationality'):
                            worksheet.cell(row=row_num, column=15, value=practitioner.nationality or "")
                        else:
                            worksheet.cell(row=row_num, column=15, value="")
                        
                        # Statut
                        if hasattr(practitioner, 'status') and hasattr(Practitioner, 'STATUS_CHOICES'):
                            status_map = dict(Practitioner.STATUS_CHOICES)
                            worksheet.cell(row=row_num, column=16, value=status_map.get(practitioner.status, ""))
                        else:
                            worksheet.cell(row=row_num, column=16, value="Actif")
                
                # Ajuster la largeur des colonnes
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Sauvegarder le fichier Excel dans la réponse HTTP
                workbook.save(response)
                return response
                
            except Exception as e:
                logger.error(f"Erreur lors de l'export Excel: {str(e)}", exc_info=True)
                messages.error(request, _("Une erreur est survenue lors de l'export: {}").format(str(e)))
            
        elif 'import' in request.POST and 'excel_file' in request.FILES:
            # Logique d'import depuis Excel
            excel_file = request.FILES['excel_file']
            
            try:
                # Vérifier l'extension du fichier
                file_ext = os.path.splitext(excel_file.name)[1].lower()
                if file_ext not in ['.xlsx', '.xls']:
                    messages.error(request, _("Le fichier doit être au format Excel (.xlsx, .xls)"))
                    return redirect('competitions:club:import_export')
                
                # Ouvrir le classeur Excel
                workbook = openpyxl.load_workbook(excel_file)
                worksheet = workbook.active
                
                # Compteurs pour le résumé
                created_count = 0
                updated_count = 0
                errors_count = 0
                
                # Vérifier la structure du fichier (en-têtes)
                headers = [cell.value for cell in worksheet[1]]
                required_headers = ['Nom', 'Prénom', 'Date de naissance']
                
                # Vérifier que les en-têtes requis sont présents
                missing_headers = [header for header in required_headers if header not in headers]
                if missing_headers:
                    messages.error(request, _("Le fichier ne contient pas les colonnes requises: {}").format(", ".join(missing_headers)))
                    return redirect('competitions:club:import_export')
                
                # Cartographie des colonnes
                header_map = {}
                for idx, header in enumerate(headers, 1):
                    if header in ['Nom', 'Prénom', 'Date de naissance', 'Grade', 'Email', 'Licence', 
                            'Genre', 'Téléphone', 'Adresse', 'Ville', 'Code postal', 'Poids (kg)', 
                            'Taille (cm)', 'Certificat médical', 'Nationalité', 'Statut']:
                        header_map[header] = idx
                
                # Parcourir les lignes (en sautant l'en-tête)
                for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), 2):
                    try:
                        # Vérifier si la ligne contient des données
                        if not row or not any(row[:3]):  # Les 3 premiers champs sont obligatoires
                            continue
                        
                        # Extraire les données selon le mapping
                        last_name = row[header_map.get('Nom', 0) - 1] if 'Nom' in header_map else None
                        first_name = row[header_map.get('Prénom', 1) - 1] if 'Prénom' in header_map else None
                        birth_date = row[header_map.get('Date de naissance', 2) - 1] if 'Date de naissance' in header_map else None
                        
                        # Validation des champs obligatoires
                        if not (last_name and first_name and birth_date):
                            errors_count += 1
                            logger.warning(f"Ligne {row_idx}: Données obligatoires manquantes")
                            continue
                        
                        # Extraire les autres champs
                        grade = row[header_map.get('Grade', 3) - 1] if 'Grade' in header_map and len(row) > header_map.get('Grade', 3) - 1 else None
                        email = row[header_map.get('Email', 4) - 1] if 'Email' in header_map and len(row) > header_map.get('Email', 4) - 1 else None
                        license_number = row[header_map.get('Licence', 5) - 1] if 'Licence' in header_map and len(row) > header_map.get('Licence', 5) - 1 else None
                        
                        # Validation de l'email
                        if email and '@' not in str(email):
                            logger.warning(f"Ligne {row_idx}: Email invalide: {email}")
                            email = None  # Ignorer l'email invalide
                        
                        # Genre
                        gender_value = row[header_map.get('Genre', 6) - 1] if 'Genre' in header_map and len(row) > header_map.get('Genre', 6) - 1 else None
                        gender_map = {'Homme': 'male', 'Femme': 'female', 'Autre': 'other'}
                        gender = gender_map.get(gender_value, 'male')  # Valeur par défaut si non reconnu
                        
                        # Autres champs
                        phone = row[header_map.get('Téléphone', 7) - 1] if 'Téléphone' in header_map and len(row) > header_map.get('Téléphone', 7) - 1 else None
                        address = row[header_map.get('Adresse', 8) - 1] if 'Adresse' in header_map and len(row) > header_map.get('Adresse', 8) - 1 else None
                        city = row[header_map.get('Ville', 9) - 1] if 'Ville' in header_map and len(row) > header_map.get('Ville', 9) - 1 else None
                        postal_code = row[header_map.get('Code postal', 10) - 1] if 'Code postal' in header_map and len(row) > header_map.get('Code postal', 10) - 1 else None
                        
                        # Données numériques
                        weight_value = row[header_map.get('Poids (kg)', 11) - 1] if 'Poids (kg)' in header_map and len(row) > header_map.get('Poids (kg)', 11) - 1 else None
                        height_value = row[header_map.get('Taille (cm)', 12) - 1] if 'Taille (cm)' in header_map and len(row) > header_map.get('Taille (cm)', 12) - 1 else None
                        
                        # Conversion et validation des données numériques
                        weight = float(weight_value) if weight_value is not None and str(weight_value).replace('.', '', 1).isdigit() else None
                        height = float(height_value) if height_value is not None and str(height_value).replace('.', '', 1).isdigit() else None
                        
                        # Date du certificat médical
                        medical_certificate_date = row[header_map.get('Certificat médical', 13) - 1] if 'Certificat médical' in header_map and len(row) > header_map.get('Certificat médical', 13) - 1 else None
                        
                        # Autres informations
                        nationality = row[header_map.get('Nationalité', 14) - 1] if 'Nationalité' in header_map and len(row) > header_map.get('Nationalité', 14) - 1 else None
                        
                        # Statut
                        status_value = row[header_map.get('Statut', 15) - 1] if 'Statut' in header_map and len(row) > header_map.get('Statut', 15) - 1 else None
                        
                        status = 'active'  # Valeur par défaut si non reconnu
                        if hasattr(Practitioner, 'STATUS_CHOICES'):
                            status_reverse_map = {v: k for k, v in dict(Practitioner.STATUS_CHOICES).items()}
                            status = status_reverse_map.get(status_value, 'active')
                        
                        # Vérifier si le pratiquant existe déjà
                        if birth_date:
                            # Si la date de naissance est fournie, chercher par nom, prénom et date de naissance
                            practitioner = Practitioner.objects.filter(
                                club=club,
                                last_name=last_name,
                                first_name=first_name,
                                birth_date=birth_date
                            ).first()
                        else:
                            # Sinon, chercher seulement par nom et prénom
                            practitioner = Practitioner.objects.filter(
                                club=club,
                                last_name=last_name,
                                first_name=first_name
                            ).first()
                        
                        if practitioner:
                            # Mise à jour du pratiquant existant
                            if grade:
                                practitioner.grade = grade
                            if email:
                                practitioner.email = email
                            if license_number:
                                practitioner.license_number = license_number
                            if gender:
                                practitioner.gender = gender
                            if phone and hasattr(practitioner, 'phone'):
                                practitioner.phone = phone
                            if address and hasattr(practitioner, 'address'):
                                practitioner.address = address
                            if city and hasattr(practitioner, 'city'):
                                practitioner.city = city
                            if postal_code and hasattr(practitioner, 'postal_code'):
                                practitioner.postal_code = postal_code
                            if weight and hasattr(practitioner, 'weight'):
                                practitioner.weight = weight
                            if height and hasattr(practitioner, 'height'):
                                practitioner.height = height
                            if medical_certificate_date and hasattr(practitioner, 'medical_certificate_date'):
                                practitioner.medical_certificate_date = medical_certificate_date
                            if nationality and hasattr(practitioner, 'nationality'):
                                practitioner.nationality = nationality
                            if status and hasattr(practitioner, 'status'):
                                practitioner.status = status
                            
                            practitioner.save()
                            updated_count += 1
                        else:
                            # Création d'un nouveau pratiquant avec les champs de base
                            practitioner_data = {
                                'club': club,
                                'last_name': last_name,
                                'first_name': first_name,
                                'birth_date': birth_date,
                                'grade': grade or '',
                                'email': email,
                                'license_number': license_number or '',
                                'gender': gender,
                            }
                            
                            # Ajouter les champs optionnels s'ils existent dans le modèle
                            optional_fields = {
                                'phone': phone,
                                'address': address,
                                'city': city,
                                'postal_code': postal_code,
                                'weight': weight,
                                'height': height,
                                'medical_certificate_date': medical_certificate_date,
                                'nationality': nationality,
                                'status': status
                            }
                            
                            # Vérifier quels champs sont disponibles dans le modèle
                            model_fields = [f.name for f in Practitioner._meta.get_fields()]
                            for field_name, field_value in optional_fields.items():
                                if field_name in model_fields and field_value is not None:
                                    practitioner_data[field_name] = field_value
                            
                            # Créer le pratiquant
                            Practitioner.objects.create(**practitioner_data)
                            created_count += 1
                            
                    except Exception as e:
                        errors_count += 1
                        logger.error(f"Erreur lors de l'import d'un pratiquant (ligne {row_idx}): {str(e)}", exc_info=True)
                
                # Message de résumé
                if created_count > 0 or updated_count > 0:
                    messages.success(
                        request, 
                        _("{} pratiquant(s) créé(s), {} mis à jour, {} erreur(s)").format(
                            created_count, updated_count, errors_count
                        )
                    )
                elif errors_count > 0:
                    messages.warning(
                        request,
                        _("Aucun pratiquant créé ou mis à jour. {} erreur(s) rencontrée(s).").format(errors_count)
                    )
                else:
                    messages.info(request, _("Aucune modification n'a été effectuée."))
                    
            except Exception as e:
                messages.error(request, _("Une erreur est survenue lors de l'import: {}").format(str(e)))
                logger.error(f"Erreur lors de l'import Excel: {str(e)}", exc_info=True)
            
        elif 'template' in request.POST:
            # Générer un modèle d'import
            return generate_template_response()
    
    # Si c'est une requête GET avec le paramètre template, générer un modèle d'import
    if request.method == 'GET' and 'template' in request.GET:
        return generate_template_response()
    
    # Rendre le template avec le contexte
    return render(request, 'competitions/club/import_export.html', {
        'club': club,
        'page_title': _("Import/Export de données")
    })

@login_required
@manual_permission_check('club.manage_members')
def generate_template_response():
    """Génère un fichier Excel template pour l'import."""
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="modele_import_pratiquants.xlsx"'
    
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Modèle"
    
    # En-têtes
    headers = [
        'Nom', 'Prénom', 'Date de naissance', 'Grade', 'Email', 'Licence', 
        'Genre', 'Téléphone', 'Adresse', 'Ville', 'Code postal', 'Poids (kg)', 
        'Taille (cm)', 'Certificat médical', 'Nationalité', 'Statut'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Exemple de données
    example_data = [
        ['Dupont', 'Jean', '1990-01-15', 'Ceinture noire 1er dan', 'jean.dupont@example.com', 'LIC12345', 
        'Homme', '0612345678', '123 rue Example', 'Paris', '75001', '75.5', '180', '2023-06-01', 'Française', 'Actif'],
        ['Martin', 'Sophie', '1995-03-22', 'Ceinture marron', 'sophie.martin@example.com', 'LIC54321',
        'Femme', '0698765432', '456 avenue Example', 'Lyon', '69002', '60.2', '165', '2023-07-15', 'Française', 'Actif']
    ]
    
    for row_num, row_data in enumerate(example_data, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            
            # Formatage spécial pour les dates
            if col_num in [3, 14]:  # Date de naissance et certificat médical
                cell.number_format = 'yyyy-mm-dd'
    
    # Ajuster la largeur des colonnes
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    workbook.save(response)
    return response