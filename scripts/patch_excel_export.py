#!/usr/bin/env python3
"""Replace CSV export with formatted Excel export using openpyxl."""

filepath = "/var/www/vhosts/martialcomp.com/httpdocs/apps/competitions/views/management/participants.py"
with open(filepath, "r") as f:
    content = f.read()

old_func = '''def export_participants(request, competition_id):
    """
    Exporte la liste des participants au format CSV.
    """
    import csv
    from django.http import HttpResponse

    competition = get_object_or_404(Competition, pk=competition_id)

    # Filtres (comme dans la liste)
    status = request.GET.get('status')
    club_id = request.GET.get('club')
    category_id = request.GET.get('category')

    registrations = CompetitionRegistration.objects.filter(
        competition=competition,
        is_competitor=True
    ).select_related('practitioner', 'practitioner__organization')

    # Appliquer les filtres
    if status:
        registrations = registrations.filter(status=status)

    if club_id:
        # Filtrer par organisation via le club
        club = Club.objects.filter(id=club_id).first()
        if club:
            registrations = registrations.filter(practitioner__organization=club.organization)

    if category_id:
        registrations = registrations.filter(categories__id=category_id)

    # Cr\xc3\xa9er la r\xc3\xa9ponse HTTP avec l'en-t\xc3\x83\xc2\xaate CSV
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{competition.title}_participants.csv"'

    # \xc3\x83\xe2\x80\xb0crire le fichier CSV
    writer = csv.writer(response)
    writer.writerow([
        _('Nom'), _('Pr\xc3\xa9nom'), _('Club'), _('Grade'),
        _('Cat\xc3\xa9gories'), _('Statut'), _('Notes')
    ])

    for reg in registrations:
        p = reg.practitioner
        categories = ', '.join([c.name for c in reg.categories.all()])

        writer.writerow([
            p.last_name,
            p.first_name,
            p.organization.name if p.organization else "",
            p.grade,
            categories,
            reg.get_status_display(),
            reg.notes
        ])

    return response'''

# Since the old function may have encoding issues, let's use a simpler match
# Find the function start and end
import re

start_marker = "def export_participants(request, competition_id):"
end_marker = "\n\n@login_required\n@competition_management_permission_required\n@require_POST\ndef approve_participant"

start_idx = content.find(start_marker)
end_idx = content.find(end_marker)

if start_idx == -1 or end_idx == -1:
    print(f"ERROR: start={start_idx}, end={end_idx}")
    # Try alternate end marker
    end_marker = "\n\n\n@login_required"
    end_idx = content.find(end_marker, start_idx + 100)
    if end_idx == -1:
        print("ERROR: Could not find function boundaries")
        exit(1)

new_func = '''def export_participants(request, competition_id):
    """
    Exporte la liste des participants au format Excel (.xlsx) avec mise en forme.
    """
    import io
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    competition = get_object_or_404(Competition, pk=competition_id)

    # Filtres
    status = request.GET.get('status')
    club_id = request.GET.get('club')
    category_id = request.GET.get('category')

    registrations = CompetitionRegistration.objects.filter(
        competition=competition,
        is_competitor=True
    ).select_related(
        'practitioner', 'practitioner__organization', 'practitioner__grade'
    ).prefetch_related('categories', 'competition_types')

    if status:
        registrations = registrations.filter(status=status)
    if club_id:
        club = Club.objects.filter(id=club_id).first()
        if club:
            registrations = registrations.filter(practitioner__organization=club.organization)
    if category_id:
        registrations = registrations.filter(categories__id=category_id)

    registrations = registrations.order_by(
        'practitioner__organization__name', 'practitioner__last_name', 'practitioner__first_name'
    )

    wb = Workbook()

    # ==========================================
    # STYLES
    # ==========================================
    dark_bg = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    gold_font = Font(name="Calibri", bold=True, size=18, color="C9A227")
    white_font = Font(name="Calibri", size=10, color="FFFFFF")
    white_font_sm = Font(name="Calibri", size=9, color="FFFFFF")

    header_fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
    header_font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")

    male_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    female_fill = PatternFill(start_color="FCE7F3", end_color="FCE7F3", fill_type="solid")

    approved_font = Font(name="Calibri", size=9, color="16A34A", bold=True)
    pending_font = Font(name="Calibri", size=9, color="D97706", bold=True)
    rejected_font = Font(name="Calibri", size=9, color="DC2626", bold=True)

    cat_font = Font(name="Calibri", size=8, color="475569")
    type_font = Font(name="Calibri", size=8, color="0369A1", bold=True)

    thin_border = Border(
        bottom=Side(style="thin", color="E2E8F0")
    )
    wrap = Alignment(wrap_text=True, vertical="center")
    center = Alignment(horizontal="center", vertical="center")
    center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ==========================================
    # SHEET 1: All participants
    # ==========================================
    ws = wb.active
    ws.title = "Tous les inscrits"

    # Title row (merged)
    ws.merge_cells("A1:J1")
    ws["A1"] = competition.title
    ws["A1"].font = gold_font
    ws["A1"].fill = dark_bg
    ws["A1"].alignment = Alignment(vertical="center")
    for col in range(1, 11):
        ws.cell(row=1, column=col).fill = dark_bg
    ws.row_dimensions[1].height = 40

    # Info row
    ws.merge_cells("A2:J2")
    info = []
    if competition.start_date:
        info.append(competition.start_date.strftime("%d/%m/%Y"))
    if competition.venue_name:
        info.append(competition.venue_name)
    info.append(f"{registrations.count()} inscrits")
    ws["A2"] = " | ".join(info)
    ws["A2"].font = white_font_sm
    ws["A2"].fill = dark_bg
    for col in range(1, 11):
        ws.cell(row=2, column=col).fill = dark_bg
    ws.row_dimensions[2].height = 22

    # Empty row
    ws.row_dimensions[3].height = 8

    # Headers
    headers = [
        ("#", 5),
        ("Nom", 18),
        ("Prenom", 15),
        ("Genre", 7),
        ("Age", 6),
        ("Club", 18),
        ("Grade", 16),
        ("Types de competition", 25),
        ("Categories", 40),
        ("Statut", 12),
    ]

    for col_idx, (header, width) in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[4].height = 28

    # Data rows
    row_num = 5
    for idx, reg in enumerate(registrations, 1):
        p = reg.practitioner
        categories = "\\n".join([c.name for c in reg.categories.all()])
        comp_types = "\\n".join([t.name for t in reg.competition_types.all()])
        grade_name = p.grade.name if p.grade else "-"
        age = p.age if hasattr(p, "age") and p.age else "-"
        gender = ""
        if p.gender in ("M", "male"):
            gender = "H"
        elif p.gender in ("F", "female"):
            gender = "F"

        status_text = reg.get_status_display() if hasattr(reg, "get_status_display") else str(reg.status)

        ws.cell(row=row_num, column=1, value=idx).alignment = center
        ws.cell(row=row_num, column=2, value=p.last_name).font = Font(name="Calibri", bold=True, size=10)
        ws.cell(row=row_num, column=3, value=p.first_name)
        gender_cell = ws.cell(row=row_num, column=4, value=gender)
        gender_cell.alignment = center
        if gender == "H":
            gender_cell.font = Font(name="Calibri", bold=True, size=10, color="2563EB")
        elif gender == "F":
            gender_cell.font = Font(name="Calibri", bold=True, size=10, color="DB2777")
        ws.cell(row=row_num, column=5, value=age).alignment = center
        ws.cell(row=row_num, column=6, value=p.organization.name if p.organization else "-")
        ws.cell(row=row_num, column=7, value=grade_name)
        ws.cell(row=row_num, column=8, value=comp_types).font = type_font
        ws.cell(row=row_num, column=8).alignment = wrap
        ws.cell(row=row_num, column=9, value=categories).font = cat_font
        ws.cell(row=row_num, column=9).alignment = wrap

        status_cell = ws.cell(row=row_num, column=10, value=status_text)
        status_cell.alignment = center
        if "appro" in status_text.lower():
            status_cell.font = approved_font
        elif "attente" in status_text.lower() or "pending" in status_text.lower():
            status_cell.font = pending_font
        elif "rejet" in status_text.lower() or "reject" in status_text.lower():
            status_cell.font = rejected_font

        # Alternate row coloring and gender tint
        if gender == "H":
            for c in range(1, 11):
                ws.cell(row=row_num, column=c).fill = male_fill
        elif gender == "F":
            for c in range(1, 11):
                ws.cell(row=row_num, column=c).fill = female_fill
        elif idx % 2 == 0:
            alt_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
            for c in range(1, 11):
                ws.cell(row=row_num, column=c).fill = alt_fill

        for c in range(1, 11):
            ws.cell(row=row_num, column=c).border = thin_border

        # Auto row height based on categories count
        cat_count = reg.categories.count()
        ws.row_dimensions[row_num].height = max(20, cat_count * 14)

        row_num += 1

    # ==========================================
    # SHEET PER COMPETITION TYPE
    # ==========================================
    for comp_type in competition.competition_types.all().order_by("name"):
        from apps.competitions.models import CompetitionCategory
        categories = CompetitionCategory.objects.filter(
            competition=competition, competition_type=comp_type
        ).order_by("name")

        if not categories.exists():
            continue

        # Sheet name (max 31 chars)
        sheet_name = comp_type.name[:31]
        wst = wb.create_sheet(title=sheet_name)

        # Title
        wst.merge_cells("A1:H1")
        wst["A1"] = comp_type.name
        wst["A1"].font = gold_font
        wst["A1"].fill = dark_bg
        for c in range(1, 9):
            wst.cell(row=1, column=c).fill = dark_bg
        wst.row_dimensions[1].height = 35

        row = 3
        for cat in categories:
            regs = cat.registrations.all().select_related(
                "practitioner", "practitioner__organization", "practitioner__grade"
            ).order_by("practitioner__last_name")

            # Category header
            wst.merge_cells(f"A{row}:F{row}")
            cat_cell = wst.cell(row=row, column=1, value=cat.name)
            cat_cell.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
            cat_cell.fill = PatternFill(start_color="475569", end_color="475569", fill_type="solid")
            for c in range(1, 9):
                wst.cell(row=row, column=c).fill = PatternFill(start_color="475569", end_color="475569", fill_type="solid")

            # Badges
            count_cell = wst.cell(row=row, column=7, value=f"{regs.count()} inscrits")
            count_cell.font = Font(name="Calibri", bold=True, size=9, color="C9A227")
            count_cell.fill = PatternFill(start_color="475569", end_color="475569", fill_type="solid")
            count_cell.alignment = Alignment(horizontal="right", vertical="center")

            info_parts = []
            if cat.gender:
                info_parts.append("Hommes" if cat.gender in ("M", "male") else "Femmes" if cat.gender in ("F", "female") else "Mixte")
            if cat.min_age or cat.max_age:
                info_parts.append(f"{cat.min_age or 0}-{cat.max_age or 99} ans")
            info_cell = wst.cell(row=row, column=8, value=" | ".join(info_parts))
            info_cell.font = Font(name="Calibri", size=8, color="94A3B8")
            info_cell.fill = PatternFill(start_color="475569", end_color="475569", fill_type="solid")

            wst.row_dimensions[row].height = 28
            row += 1

            # Sub-headers
            sub_headers = ["#", "Nom", "Prenom", "Genre", "Age", "Club", "Grade", "Controle"]
            sub_widths = [5, 20, 15, 7, 6, 18, 16, 10]
            for ci, (sh, sw) in enumerate(zip(sub_headers, sub_widths), 1):
                cell = wst.cell(row=row, column=ci, value=sh)
                cell.font = Font(name="Calibri", bold=True, size=8, color="64748B")
                cell.alignment = center
                wst.column_dimensions[get_column_letter(ci)].width = max(
                    wst.column_dimensions[get_column_letter(ci)].width or 0, sw
                )
            row += 1

            # Practitioners
            for pidx, reg in enumerate(regs, 1):
                p = reg.practitioner
                gender = "H" if p.gender in ("M", "male") else "F" if p.gender in ("F", "female") else ""

                wst.cell(row=row, column=1, value=pidx).alignment = center
                wst.cell(row=row, column=2, value=p.last_name).font = Font(name="Calibri", bold=True, size=10)
                wst.cell(row=row, column=3, value=p.first_name)
                g_cell = wst.cell(row=row, column=4, value=gender)
                g_cell.alignment = center
                if gender == "H":
                    g_cell.font = Font(color="2563EB", bold=True)
                elif gender == "F":
                    g_cell.font = Font(color="DB2777", bold=True)
                wst.cell(row=row, column=5, value=p.age if hasattr(p, "age") and p.age else "-").alignment = center
                wst.cell(row=row, column=6, value=p.organization.name if p.organization else "-")
                wst.cell(row=row, column=7, value=p.grade.name if p.grade else "-")
                # Empty checkbox column for offline control
                ctrl = wst.cell(row=row, column=8, value="[ ]")
                ctrl.alignment = center
                ctrl.font = Font(size=12)

                # Row color
                if gender == "H":
                    for c in range(1, 9):
                        wst.cell(row=row, column=c).fill = male_fill
                elif gender == "F":
                    for c in range(1, 9):
                        wst.cell(row=row, column=c).fill = female_fill

                for c in range(1, 9):
                    wst.cell(row=row, column=c).border = thin_border

                row += 1

            if not regs.exists():
                wst.cell(row=row, column=2, value="Aucun inscrit").font = Font(italic=True, color="94A3B8")
                row += 1

            row += 1  # Space between categories

    # ==========================================
    # SHEET: Stats by club
    # ==========================================
    ws_stats = wb.create_sheet(title="Stats par club")
    ws_stats.merge_cells("A1:D1")
    ws_stats["A1"] = "Statistiques par club"
    ws_stats["A1"].font = gold_font
    ws_stats["A1"].fill = dark_bg
    for c in range(1, 5):
        ws_stats.cell(row=1, column=c).fill = dark_bg
    ws_stats.row_dimensions[1].height = 35

    stats_headers = ["Club", "Nb inscrits", "Hommes", "Femmes"]
    for ci, h in enumerate(stats_headers, 1):
        cell = ws_stats.cell(row=3, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
    ws_stats.column_dimensions["A"].width = 25
    ws_stats.column_dimensions["B"].width = 12
    ws_stats.column_dimensions["C"].width = 10
    ws_stats.column_dimensions["D"].width = 10

    from django.db.models import Count, Q
    club_stats = registrations.values(
        "practitioner__organization__name"
    ).annotate(
        total=Count("id"),
        males=Count("id", filter=Q(practitioner__gender__in=["M", "male"])),
        females=Count("id", filter=Q(practitioner__gender__in=["F", "female"])),
    ).order_by("-total")

    srow = 4
    for cs in club_stats:
        club_name = cs["practitioner__organization__name"] or "Sans club"
        ws_stats.cell(row=srow, column=1, value=club_name).font = Font(bold=True)
        ws_stats.cell(row=srow, column=2, value=cs["total"]).alignment = center
        m_cell = ws_stats.cell(row=srow, column=3, value=cs["males"])
        m_cell.alignment = center
        m_cell.font = Font(color="2563EB")
        f_cell = ws_stats.cell(row=srow, column=4, value=cs["females"])
        f_cell.alignment = center
        f_cell.font = Font(color="DB2777")
        for c in range(1, 5):
            ws_stats.cell(row=srow, column=c).border = thin_border
        srow += 1

    # Total row
    ws_stats.cell(row=srow, column=1, value="TOTAL").font = Font(bold=True, size=11)
    ws_stats.cell(row=srow, column=2, value=registrations.count()).font = Font(bold=True, size=11)
    ws_stats.cell(row=srow, column=2).alignment = center

    # ==========================================
    # Save and return
    # ==========================================
    # Freeze panes on main sheet
    ws.freeze_panes = "A5"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    safe_title = competition.title.replace(" ", "_").replace("/", "-")[:50]
    filename = f"{safe_title}_inscrits.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response'''

content = content[:start_idx] + new_func + content[end_idx:]

with open(filepath, "w") as f:
    f.write(content)
print("Excel export function replaced")
print("DONE")
