from django.core.exceptions import PermissionDenied
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.utils.translation import gettext_lazy as _
from django.contrib.auth.decorators import login_required
from django.db.models import Q, Count, Sum, Avg, Max, F
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from django.utils import timezone
from django.db import transaction

from apps.competitions.models import (
    Competition, CompetitionCategory, Match, CompetitionRegistration,
    Club
)
# Import unified scoring models
from apps.competitions.models.scoring_results import (
    CategoryRanking, RankingEntry, TechnicalPerformanceResult,
    PerformanceResult, CompetitionResult
)
# Import technical scoring models
from apps.competitions.models.technical_scoring import (
    TechnicalPerformance, CompetitionRanking
)
# Import combat models
from apps.competitions.models.combat import Combat, Poule
from apps.competitions.utils.decorators import competition_management_permission_required


@login_required
@competition_management_permission_required
def results_dashboard(request, competition_id):
    """
    Tableau de bord des résultats pour une compétition.
    Supporte les catégories techniques (kata, quyền) et combat.
    """
    # Récupérer la compétition
    competition = get_object_or_404(Competition, pk=competition_id)

    # Récupérer les catégories de la compétition avec les relations nécessaires
    categories = CompetitionCategory.objects.filter(
        competition=competition
    ).select_related('competition_type').prefetch_related('registrations')

    # Ajouter des informations sur l'état des résultats
    categories_list = []
    for category in categories:
        # Détecter le type de catégorie (combat ou technique)
        is_combat = _is_combat_category(category)
        category.is_combat = is_combat

        if is_combat:
            # Pour les catégories combat
            # Vérifier si des résultats existent (CategoryRanking ou combats terminés avec vainqueur)
            has_ranking = CategoryRanking.objects.filter(category=category).exists()
            has_completed_combats = Combat.objects.filter(
                poule__category=category,
                status='termine'
            ).exists()
            category.has_results = has_ranking or has_completed_combats

            # Nombre de combats terminés
            category.completed_performances_count = Combat.objects.filter(
                poule__category=category,
                status='termine'
            ).count()

            # Nombre total de combats
            category.total_performances_count = Combat.objects.filter(
                poule__category=category
            ).count()
        else:
            # Pour les catégories techniques
            # Vérifier si des résultats existent dans CompetitionRanking OU CategoryRanking
            has_competition_ranking = CompetitionRanking.objects.filter(
                competition=competition,
                category=category
            ).exists()
            has_category_ranking = CategoryRanking.objects.filter(category=category).exists()
            category.has_results = has_competition_ranking or has_category_ranking

            # Nombre de performances terminées (essayer les deux modèles)
            completed_tech = TechnicalPerformance.objects.filter(
                category=category,
                status='completed'
            ).count()
            completed_result = TechnicalPerformanceResult.objects.filter(
                category=category,
                status='completed'
            ).count() if TechnicalPerformanceResult else 0
            category.completed_performances_count = max(completed_tech, completed_result)

            # Nombre total de performances
            total_tech = TechnicalPerformance.objects.filter(category=category).count()
            total_result = TechnicalPerformanceResult.objects.filter(
                category=category
            ).count() if TechnicalPerformanceResult else 0
            category.total_performances_count = max(total_tech, total_result)

        # Vérifier si la catégorie est terminée
        category.is_completed = getattr(category, 'status', None) == 'completed'

        categories_list.append(category)

    # Extraire les types de compétition distincts pour le filtre
    competition_types = []
    seen_type_ids = set()
    for cat in categories_list:
        if cat.competition_type and cat.competition_type.pk not in seen_type_ids:
            seen_type_ids.add(cat.competition_type.pk)
            competition_types.append(cat.competition_type)
    competition_types.sort(key=lambda ct: ct.name)

    # Calculer les statistiques globales
    total_categories = len(categories_list)
    categories_with_results = sum(1 for cat in categories_list if getattr(cat, 'has_results', False))
    categories_completed = sum(1 for cat in categories_list if getattr(cat, 'is_completed', False))
    categories_in_progress = sum(1 for cat in categories_list if getattr(cat, 'status', None) == 'ongoing')

    context = {
        'competition': competition,
        'categories': categories_list,
        'competition_types': competition_types,
        'total_categories': total_categories,
        'categories_with_results': categories_with_results,
        'categories_completed': categories_completed,
        'categories_in_progress': categories_in_progress,
    }

    return render(request, 'competitions/management/results_dashboard.html', context)


def _is_combat_category(category):
    """
    Détermine si une catégorie est de type combat.
    """
    # Vérifier par le nom
    combat_keywords = ['combat', 'kumite', 'sparring', 'fighting', 'doi khang']
    name_lower = category.name.lower()
    if any(kw in name_lower for kw in combat_keywords):
        return True

    # Vérifier si des poules existent pour cette catégorie
    try:
        has_poules = Poule.objects.filter(category=category).exists()
        if has_poules:
            return True
    except Exception:
        pass

    return False


@login_required
@competition_management_permission_required
def category_results(request, competition_id, category_id):
    """
    Affiche les résultats d'une catégorie spécifique.
    """
    # Récupérer la compétition et la catégorie
    competition = get_object_or_404(Competition, pk=competition_id)
    category = get_object_or_404(CompetitionCategory, pk=category_id, competition=competition)
    
    # Vérifier si un classement existe
    category_ranking = CategoryRanking.objects.filter(
        competition=competition,
        category=category
    ).first()
    
    # Récupérer les résultats
    if category_ranking:
        rankings = RankingEntry.objects.filter(
            ranking=category_ranking
        ).select_related('practitioner').order_by('rank')
        has_results = rankings.exists()
    else:
        rankings = []
        has_results = False
    
    # Récupérer les configurations de scoring
    from apps.competitions.models.technical_scoring import ScoringConfiguration
    try:
        scoring_config = ScoringConfiguration.objects.get(category=category)
    except ScoringConfiguration.DoesNotExist:
        scoring_config = None
    
    context = {
        'competition': competition,
        'category': category,
        'rankings': rankings,
        'has_results': has_results,
        'scoring_config': scoring_config,
        'category_ranking': category_ranking
    }
    
    return render(request, 'competitions/management/category_results.html', context)


@login_required
@competition_management_permission_required
def calculate_category_results(request, competition_id, category_id):
    """
    Calcule les résultats pour une catégorie (redirects vers la vue de scoring).
    """
    # Importer la vue de scoring pour le calcul des résultats
    from apps.competitions.views.management.scoring import calculate_results
    
    # Utiliser la vue de scoring
    return calculate_results(request, competition_id, category_id)


@login_required
@competition_management_permission_required
def edit_ranking(request, competition_id, ranking_entry_id):
    """
    Modifie manuellement un classement.
    """
    # Récupérer la compétition et l'entrée de classement
    competition = get_object_or_404(Competition, pk=competition_id)
    ranking_entry = get_object_or_404(
        RankingEntry, 
        pk=ranking_entry_id, 
        ranking__competition=competition
    )
    
    if request.method == 'POST':
        # Récupérer les nouvelles valeurs
        new_rank = request.POST.get('rank')
        new_score = request.POST.get('score')
        is_tie = request.POST.get('is_tie') == 'on'
        
        try:
            # Valider les valeurs
            if new_rank:
                ranking_entry.rank = int(new_rank)
            
            if new_score:
                ranking_entry.score = float(new_score)
            
            ranking_entry.is_tie = is_tie
            ranking_entry.save()
            
            # Mettre Ã  jour le résultat de performance associé si existant
            if ranking_entry.performance and hasattr(ranking_entry.performance, 'result'):
                result = ranking_entry.performance.result
                result.rank = ranking_entry.rank
                result.is_tie = ranking_entry.is_tie
                result.save(update_fields=['rank', 'is_tie'])
            
            messages.success(request, _("Le classement a été mis Ã  jour avec succès."))
        except ValueError:
            messages.error(request, _("Valeurs invalides. Veuillez vérifier les données saisies."))
        
        return redirect('competitions:management:category_results', 
                      competition_id=competition_id, 
                      category_id=ranking_entry.ranking.category.id)
    
    context = {
        'competition': competition,
        'ranking_entry': ranking_entry,
    }
    
    return render(request, 'competitions/management/edit_ranking.html', context)


@login_required
@competition_management_permission_required
@require_POST
def delete_ranking(request, competition_id, ranking_entry_id):
    """
    Supprime une entrée de classement.
    """
    # Récupérer la compétition et l'entrée de classement
    competition = get_object_or_404(Competition, pk=competition_id)
    ranking_entry = get_object_or_404(
        RankingEntry, 
        pk=ranking_entry_id, 
        ranking__competition=competition
    )
    
    category_id = ranking_entry.ranking.category.id
    
    # Supprimer l'entrée de classement
    ranking_entry.delete()
    
    messages.success(request, _("Le classement a été supprimé."))
    return redirect('competitions:management:category_results', 
                  competition_id=competition_id, 
                  category_id=category_id)


@login_required
@competition_management_permission_required
def all_competition_results(request, competition_id):
    """
    Affiche tous les résultats de la compétition.
    Supporte les deux systèmes : CompetitionRanking (technique) et CategoryRanking (unifié).
    """
    # Récupérer la compétition
    competition = get_object_or_404(Competition, pk=competition_id)

    # Récupérer toutes les catégories de la compétition
    all_categories = CompetitionCategory.objects.filter(
        competition=competition
    ).select_related('competition_type')

    # Extraire les types de compétition distincts pour le filtre
    competition_types = []
    seen_type_ids = set()
    for cat in all_categories:
        if cat.competition_type and cat.competition_type.pk not in seen_type_ids:
            seen_type_ids.add(cat.competition_type.pk)
            competition_types.append(cat.competition_type)
    competition_types.sort(key=lambda ct: ct.name)

    # Filtrer les catégories qui ont des résultats (dans l'un ou l'autre système)
    categories_with_results = []
    results_by_category = {}

    for category in all_categories:
        rankings_list = []

        # 1. Vérifier CompetitionRanking (système technique)
        tech_rankings = CompetitionRanking.objects.filter(
            competition=competition,
            category=category
        ).select_related('practitioner', 'practitioner__organization').order_by('rank')

        if tech_rankings.exists():
            rankings_list.extend(list(tech_rankings))

        # 2. Vérifier CategoryRanking (système unifié)
        try:
            category_ranking = CategoryRanking.objects.get(
                competition=competition,
                category=category
            )
            unified_rankings = RankingEntry.objects.filter(
                ranking=category_ranking
            ).select_related('practitioner').order_by('rank')
            rankings_list.extend(list(unified_rankings))
        except CategoryRanking.DoesNotExist:
            pass

        # Si des résultats existent, ajouter la catégorie
        if rankings_list:
            categories_with_results.append(category)
            results_by_category[category.id] = rankings_list

    context = {
        'competition': competition,
        'categories': categories_with_results,
        'competition_types': competition_types,
        'results_by_category': results_by_category,
    }

    return render(request, 'competitions/management/all_results.html', context)


@login_required
@competition_management_permission_required
def export_all_results(request, competition_id):
    """
    Exporte tous les resultats de la competition (toutes categories).
    Supporte CSV, Excel et PDF via parametre ?format=
    """
    from apps.competitions.models.technical_scoring import CompetitionRanking
    import csv

    competition = get_object_or_404(Competition, pk=competition_id)
    export_format = request.GET.get('format', 'csv')

    # Recuperer toutes les categories avec des resultats
    category_ids = CompetitionRanking.objects.filter(
        competition=competition
    ).values_list('category_id', flat=True).distinct()
    categories = CompetitionCategory.objects.filter(id__in=category_ids).order_by('name')

    headers = ['Categorie', 'Rang', 'Nom', 'Prenom', 'Club', 'Score', 'Ex-aequo', 'Medaille']
    rows = []

    for category in categories:
        rankings = CompetitionRanking.objects.filter(
            competition=competition,
            category=category
        ).select_related('practitioner').order_by('rank')

        for ranking in rankings:
            p = ranking.practitioner
            club_name = ''
            try:
                if p.organization:
                    club_name = p.organization.name
            except Exception:
                pass
            if not club_name:
                try:
                    if p.club:
                        club_name = p.club.name
                except Exception:
                    pass

            medal = ''
            if ranking.rank == 1:
                medal = 'Or'
            elif ranking.rank == 2:
                medal = 'Argent'
            elif ranking.rank == 3:
                medal = 'Bronze'

            rows.append([
                category.name,
                ranking.rank,
                p.last_name,
                p.first_name,
                club_name,
                float(ranking.final_score),
                'Oui' if ranking.is_tie else '',
                medal,
            ])

    if export_format == 'pdf':
        try:
            from xhtml2pdf import pisa
            from io import BytesIO
            from apps.competitions.models.technical_scoring import TechnicalPerformance, TechnicalScore
            from django.conf import settings
            import os

            date_str = competition.start_date.strftime('%d/%m/%Y') if competition.start_date else ''
            media_root = getattr(settings, 'MEDIA_ROOT', '')

            # Construire le contenu par categorie (style page web)
            categories_html = ''
            for category in categories:
                cat_rankings = CompetitionRanking.objects.filter(
                    competition=competition, category=category
                ).select_related('practitioner').order_by('rank')

                if not cat_rankings.exists():
                    continue

                # Type badge
                type_name = category.competition_type.name if category.competition_type else ''

                rows_cat = ''
                for ranking in cat_rankings:
                    p = ranking.practitioner

                    # Photo
                    photo_html = ''
                    if p.photo:
                        photo_path = os.path.join(media_root, str(p.photo))
                        if os.path.exists(photo_path):
                            photo_html = f'<td width="30" align="center"><img src="{photo_path}" width="24" height="24"></td>'
                        else:
                            photo_html = '<td width="30" align="center"></td>'
                    else:
                        photo_html = '<td width="30" align="center"></td>'

                    # Club
                    club_name = ''
                    try:
                        if p.organization:
                            club_name = p.organization.name
                    except Exception:
                        pass
                    if not club_name:
                        try:
                            if p.club:
                                club_name = p.club.name
                        except Exception:
                            pass

                    # Medaille
                    medal_html = ''
                    if ranking.rank == 1:
                        medal_html = '<span style="background-color:#C9A227; color:white; padding:2px 8px; font-size:8px; border-radius:10px;">Or</span>'
                    elif ranking.rank == 2:
                        medal_html = '<span style="background-color:#A0AEC0; color:white; padding:2px 8px; font-size:8px; border-radius:10px;">Argent</span>'
                    elif ranking.rank == 3:
                        medal_html = '<span style="background-color:#CD7F32; color:white; padding:2px 8px; font-size:8px; border-radius:10px;">Bronze</span>'

                    # Ligne
                    rows_cat += f"""<tr style="border-bottom: 1px solid #e2e8f0;">
                        <td width="40" align="center" style="padding:6px; font-weight:bold; font-size:14px;">{ranking.rank}</td>
                        {photo_html}
                        <td style="padding:6px; font-size:10px;"><b>{p.last_name}</b> {p.first_name}</td>
                        <td style="padding:6px; font-size:10px; color:#666;">{club_name}</td>
                        <td width="60" align="center" style="padding:6px; font-weight:bold; font-size:11px; color:#C9A227;">{ranking.final_score}</td>
                        <td width="70" align="center" style="padding:6px;">{medal_html}</td>
                    </tr>"""

                categories_html += f"""
                <div style="margin-bottom: 15px;">
                    <table width="100%"><tr>
                        <td style="background-color:#2D3748; color:#C9A227; padding:8px 12px; font-size:12px; font-weight:bold; border-radius:4px;">
                            {category.name}
                            <span style="float:right; background-color:#C9A227; color:#000; padding:1px 8px; font-size:8px; border-radius:10px;">{type_name}</span>
                        </td>
                    </tr></table>
                    <table width="100%" style="margin-top:2px;">
                        <tr style="background-color:#f7fafc;">
                            <th width="40" style="padding:4px; font-size:8px; color:#C9A227;">RANG</th>
                            <th width="30" style="padding:4px; font-size:8px; color:#C9A227;"></th>
                            <th style="padding:4px; font-size:8px; color:#C9A227;">PARTICIPANT</th>
                            <th style="padding:4px; font-size:8px; color:#C9A227;">CLUB</th>
                            <th width="60" align="center" style="padding:4px; font-size:8px; color:#C9A227;">SCORE</th>
                            <th width="70" align="center" style="padding:4px; font-size:8px; color:#C9A227;">MEDAILLE</th>
                        </tr>
                        {rows_cat}
                    </table>
                </div>
                """

            html_content = f"""<!DOCTYPE html><html><head><meta charset="utf-8">
            <style>
                @page {{ size: A4; margin: 1.5cm; }}
                body {{ font-family: Helvetica, sans-serif; font-size: 10px; color: #333; background: white; }}
                table {{ width: 100%; border-collapse: collapse; }}
                th {{ text-align: left; }}
            </style></head><body>
            <table width="100%"><tr>
                <td style="background-color: #1A202C; color: #C9A227; padding: 15px; font-size: 22px; font-weight: bold;" align="center">{competition.title}</td>
            </tr></table>
            <p style="text-align:center; color:#666; font-size:9px; margin:5px 0 15px;">Date: {date_str} | Categories: {categories.count()} | Participants: {len(rows)}</p>
            {categories_html}
            <p style="text-align: center; font-size: 7px; color: #999; margin-top: 20px; border-top: 1px solid #ddd; padding-top: 8px;">Document genere le {timezone.now().strftime('%d/%m/%Y %H:%M')} - MartialComp</p>
            </body></html>"""

            result = BytesIO()
            pdf = pisa.CreatePDF(html_content, dest=result, encoding='utf-8')
            if not pdf.err:
                response = HttpResponse(result.getvalue(), content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="{competition.title}_tous_resultats.pdf"'
                return response
        except Exception as e:
            import traceback
            return HttpResponse(f"<pre>PDF Error:\n{e}\n\n{traceback.format_exc()}</pre>", status=500)

    if export_format == 'excel':
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Resultats'
            total_cols = len(headers)

            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
            ws['A1'] = competition.title
            ws['A1'].font = Font(bold=True, size=16, color='C9A227')
            ws['A1'].fill = PatternFill(start_color='1A202C', end_color='1A202C', fill_type='solid')
            ws['A1'].alignment = Alignment(horizontal='center')
            ws.row_dimensions[1].height = 30

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.font = Font(bold=True, size=10, color='FFFFFF')
                cell.fill = PatternFill(start_color='4A5568', end_color='4A5568', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')

            medal_colors = {'Or': 'FFF9DB', 'Argent': 'F1F3F5', 'Bronze': 'FFF4E6'}
            for row_idx, row_data in enumerate(rows, 4):
                medal_val = row_data[-1]
                bg = medal_colors.get(medal_val)
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal='center')
                    if bg:
                        cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')

            col_widths = [20, 8, 15, 12, 20, 10, 10, 10]
            for i, w in enumerate(col_widths[:total_cols], 1):
                ws.column_dimensions[get_column_letter(i)].width = w

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{competition.title}_tous_resultats.xlsx"'
            wb.save(response)
            return response
        except Exception:
            pass

    # CSV par defaut
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{competition.title}_tous_resultats.csv"'
    response.write('\ufeff')
    writer = csv.writer(response, delimiter=';')
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    return response


@login_required
@competition_management_permission_required
def club_results(request, competition_id):
    """
    Affiche les résultats groupés par club.
    Supporte les deux systèmes de notation : CompetitionRanking (technique) et RankingEntry (unifié).
    """
    # Récupérer la compétition
    competition = get_object_or_404(Competition, pk=competition_id)

    # Récupérer tous les clubs participants via l'organisation
    # La relation est : CompetitionRegistration -> Practitioner -> Organization -> Club
    clubs = Club.objects.filter(
        organization__practitioners__registrations__competition=competition
    ).distinct()

    # Récupérer les résultats pour chaque club
    clubs_results = {}
    for club in clubs:
        # Vérifier que le club a une organisation
        if not club.organization:
            continue

        # Initialiser les compteurs
        gold_medals = 0
        silver_medals = 0
        bronze_medals = 0
        all_rankings = []

        # 1. Essayer d'abord avec CompetitionRanking (système technique)
        tech_rankings = CompetitionRanking.objects.filter(
            competition=competition,
            practitioner__organization=club.organization
        ).select_related('practitioner', 'category').order_by('category__name', 'rank')

        if tech_rankings.exists():
            gold_medals += tech_rankings.filter(rank=1).count()
            silver_medals += tech_rankings.filter(rank=2).count()
            bronze_medals += tech_rankings.filter(rank=3).count()
            all_rankings.extend(list(tech_rankings))

        # 2. Ajouter aussi les RankingEntry (système unifié) si existants
        unified_rankings = RankingEntry.objects.filter(
            ranking__competition=competition,
            practitioner__organization=club.organization
        ).select_related('practitioner', 'ranking__category').order_by('ranking__category__name', 'rank')

        if unified_rankings.exists():
            gold_medals += unified_rankings.filter(rank=1).count()
            silver_medals += unified_rankings.filter(rank=2).count()
            bronze_medals += unified_rankings.filter(rank=3).count()
            all_rankings.extend(list(unified_rankings))

        # Calculer un score global (3 points pour l'or, 2 pour l'argent, 1 pour le bronze)
        total_score = (gold_medals * 3) + (silver_medals * 2) + bronze_medals

        clubs_results[club.id] = {
            'club': club,
            'rankings': all_rankings,
            'gold_medals': gold_medals,
            'silver_medals': silver_medals,
            'bronze_medals': bronze_medals,
            'total_medals': gold_medals + silver_medals + bronze_medals,
            'total_score': total_score
        }

    # Trier les clubs par score total
    sorted_clubs = sorted(clubs_results.values(), key=lambda x: (-x['total_score'], -x['gold_medals'], -x['silver_medals'], -x['bronze_medals']))

    context = {
        'competition': competition,
        'clubs_results': sorted_clubs,
    }

    return render(request, 'competitions/management/club_results.html', context)


@login_required
@competition_management_permission_required
def publish_all_results(request, competition_id):
    """
    Publie tous les résultats de la compétition.
    """
    # Récupérer la compétition
    competition = get_object_or_404(Competition, pk=competition_id)
    
    if request.method == 'POST':
        # Marquer la compétition comme terminée
        competition.status = 'completed'
        competition.save(update_fields=['status'])

        # Fermer les inscriptions pour toutes les catégories (résultats publiés)
        CompetitionCategory.objects.filter(competition=competition).update(registration_status='closed')

        # Publier tous les classements de catégorie
        CategoryRanking.objects.filter(competition=competition).update(
            is_published=True,
            is_final=True,
            published_at=timezone.now()
        )
        
        # Synchroniser avec le modèle CompetitionResult pour la compatibilité
        with transaction.atomic():
            for ranking in CategoryRanking.objects.filter(competition=competition):
                for entry in RankingEntry.objects.filter(ranking=ranking):
                    # Déterminer le type de médaille
                    medal = 'none'
                    if entry.rank == 1:
                        medal = 'gold'
                    elif entry.rank == 2:
                        medal = 'silver'
                    elif entry.rank == 3:
                        medal = 'bronze'
                    
                    # Créer ou mettre Ã  jour le résultat de compétition
                    CompetitionResult.objects.update_or_create(
                        competition=competition,
                        category=ranking.category,
                        practitioner=entry.practitioner,
                        defaults={
                            'rank': entry.rank,
                            'score': entry.score,
                            'medal': medal,
                            'date': timezone.now().date(),
                        }
                    )
        
        messages.success(request, _("Tous les résultats ont été publiés."))

    return redirect('competitions:management:results', competition_id=competition_id)


@login_required
@competition_management_permission_required
def podium_preview(request, competition_id, category_id):
    """
    Prévisualisation du podium pour une catégorie.
    """
    # Récupérer la compétition et la catégorie
    competition = get_object_or_404(Competition, pk=competition_id)
    category = get_object_or_404(CompetitionCategory, pk=category_id, competition=competition)
    
    # Essayer d'abord avec CompetitionRanking (scoring technique)
    try:
        from apps.competitions.models.technical_scoring import CompetitionRanking
        podium = CompetitionRanking.objects.filter(
            competition=competition,
            category=category,
            rank__lte=3
        ).select_related('practitioner', 'practitioner__organization').order_by('rank')
        
        # Enrichir les données du podium avec les informations du club et du pays
        for ranking in podium:
            # Récupérer le club depuis l'organisation
            if ranking.practitioner.organization:
                try:
                    from apps.competitions.models import Club
                    ranking.club = Club.objects.filter(organization=ranking.practitioner.organization).first()
                except Exception:
                    ranking.club = None
            else:
                ranking.club = None
            
            # Récupérer le pays (nationalité du pratiquant ou pays du club)
            ranking.country_code = ranking.practitioner.nationality or (ranking.club.country if ranking.club and ranking.club.country else None)
        
        # Si aucun résultat avec CompetitionRanking, essayer avec CategoryRanking (scoring unifié)
        if not podium.exists():
            raise CompetitionRanking.DoesNotExist
    except (ImportError, AttributeError, Exception):
        # Fallback vers CategoryRanking (système unifié)
        try:
            category_ranking = get_object_or_404(
                CategoryRanking,
                competition=competition,
                category=category
            )
            
            # Récupérer les 3 premiers du classement
            podium = RankingEntry.objects.filter(
                ranking=category_ranking,
                rank__lte=3
            ).select_related('practitioner').order_by('rank')
            
            # Enrichir les données du podium avec les informations du club et du pays
            for ranking in podium:
                # Récupérer le club depuis l'organisation
                if ranking.practitioner.organization:
                    try:
                        from apps.competitions.models import Club
                        ranking.club = Club.objects.filter(organization=ranking.practitioner.organization).first()
                    except Exception:
                        ranking.club = None
                else:
                    ranking.club = None
                
                # Récupérer le pays (nationalité du pratiquant ou pays du club)
                ranking.country_code = ranking.practitioner.nationality or (ranking.club.country if ranking.club and ranking.club.country else None)
        except Exception:
            podium = []
    
    context = {
        'competition': competition,
        'category': category,
        'podium': podium,
    }
    
    return render(request, 'competitions/management/podium_preview.html', context)


@login_required
@competition_management_permission_required
def medals_report(request, competition_id):
    """
    Rapport sur les médailles distribuées.
    Supporte les deux systèmes : CompetitionRanking (technique) et CategoryRanking (unifié).
    """
    # Récupérer la compétition
    competition = get_object_or_404(Competition, pk=competition_id)

    # Compter les médailles par catégorie
    categories = CompetitionCategory.objects.filter(
        competition=competition
    ).select_related('competition_type')

    # Extraire les types de compétition distincts pour le filtre
    competition_types = []
    seen_type_ids = set()
    for cat in categories:
        if cat.competition_type and cat.competition_type.pk not in seen_type_ids:
            seen_type_ids.add(cat.competition_type.pk)
            competition_types.append(cat.competition_type)
    competition_types.sort(key=lambda ct: ct.name)

    medals_by_category = {}
    total_medals = {'gold': 0, 'silver': 0, 'bronze': 0, 'total': 0}

    for category in categories:
        gold = 0
        silver = 0
        bronze = 0

        # 1. Essayer avec CompetitionRanking (système technique)
        tech_rankings = CompetitionRanking.objects.filter(
            competition=competition,
            category=category
        )
        if tech_rankings.exists():
            gold += tech_rankings.filter(rank=1).count()
            silver += tech_rankings.filter(rank=2).count()
            bronze += tech_rankings.filter(rank=3).count()

        # 2. Ajouter CategoryRanking (système unifié) si existant
        try:
            category_ranking = CategoryRanking.objects.get(competition=competition, category=category)
            gold += RankingEntry.objects.filter(ranking=category_ranking, rank=1).count()
            silver += RankingEntry.objects.filter(ranking=category_ranking, rank=2).count()
            bronze += RankingEntry.objects.filter(ranking=category_ranking, rank=3).count()
        except CategoryRanking.DoesNotExist:
            pass

        # Ajouter à la liste si des médailles existent
        if gold > 0 or silver > 0 or bronze > 0:
            medals_by_category[category.id] = {
                'category': category,
                'gold': gold,
                'silver': silver,
                'bronze': bronze,
                'total': gold + silver + bronze
            }

            total_medals['gold'] += gold
            total_medals['silver'] += silver
            total_medals['bronze'] += bronze
            total_medals['total'] += gold + silver + bronze

    # Compter les médailles par club
    clubs_medals = []

    # Récupérer tous les clubs participants
    participating_clubs = Club.objects.filter(
        organization__practitioners__registrations__competition=competition
    ).distinct()

    for club in participating_clubs:
        if not club.organization:
            continue

        gold = 0
        silver = 0
        bronze = 0

        # 1. CompetitionRanking (système technique)
        tech_entries = CompetitionRanking.objects.filter(
            competition=competition,
            practitioner__organization=club.organization,
            rank__lte=3
        )
        gold += tech_entries.filter(rank=1).count()
        silver += tech_entries.filter(rank=2).count()
        bronze += tech_entries.filter(rank=3).count()

        # 2. RankingEntry (système unifié)
        unified_entries = RankingEntry.objects.filter(
            ranking__competition=competition,
            practitioner__organization=club.organization,
            rank__lte=3
        )
        gold += unified_entries.filter(rank=1).count()
        silver += unified_entries.filter(rank=2).count()
        bronze += unified_entries.filter(rank=3).count()

        total = gold + silver + bronze

        if total > 0:
            # Calculer le score (3 points or, 2 argent, 1 bronze)
            total_score = (gold * 3) + (silver * 2) + bronze
            clubs_medals.append({
                'club': club,
                'gold': gold,
                'silver': silver,
                'bronze': bronze,
                'total': total,
                'total_score': total_score
            })
    
    # Trier par or, puis argent, puis bronze
    clubs_medals.sort(key=lambda x: (-x['gold'], -x['silver'], -x['bronze']))
    
    context = {
        'competition': competition,
        'medals_by_category': medals_by_category,
        'competition_types': competition_types,
        'total_medals': total_medals,
        'clubs_medals': clubs_medals,
    }
    
    return render(request, 'competitions/management/medals_report.html', context)


@login_required
@competition_management_permission_required
def public_results_link(request, competition_id):
    """
    Génère un lien pour les résultats publics.
    """
    # Récupérer la compétition
    competition = get_object_or_404(Competition, pk=competition_id)
    
    # Vérifier si la compétition est publiée (accessible publiquement)
    is_public = getattr(competition, 'is_published', False) or getattr(competition, 'public_results', False)
    
    # Construire l'URL publique des résultats
    # Utiliser l'URL de partage public de la compétition si elle existe
    if hasattr(competition, 'share_token') and competition.share_token:
        results_url = request.build_absolute_uri(
            f'/competitions/{competition_id}/?token={competition.share_token}'
        )
    else:
        # URL simple vers la page publique de la compétition
        results_url = request.build_absolute_uri(
            f'/competitions/{competition_id}/'
        )
    
    # Alternative : URL vers la page de résultats de la compétition
    results_url_alt = request.build_absolute_uri(
        f'/en/competitions/management/{competition_id}/results/'
    )
    
    context = {
        'competition': competition,
        'is_public': is_public,
        'results_url': results_url,
        'results_url_alt': results_url_alt,
    }
    
    return render(request, 'competitions/management/public_results_link.html', context)


def public_results(request, competition_id, token=None):
    """
    Affiche les résultats publics de la compétition (accessible sans connexion).
    """
    # Récupérer la compétition
    competition = get_object_or_404(Competition, pk=competition_id)
    
    # Vérifier si la compétition est publiée (accessible publiquement)
    is_public = getattr(competition, 'is_published', False) or getattr(competition, 'public_results', False)
    
    # Si un token est fourni, vérifier sa validité (simplifié)
    if token:
        # Vérifier si le token correspond à un token de partage de la compétition
        if hasattr(competition, 'share_token') and competition.share_token == token:
            is_public = True
        else:
            # Token invalide
            from django.http import Http404
            raise Http404("Token invalide ou expiré")
    
    if not is_public:
        from django.http import Http404
        raise Http404("Les résultats de cette compétition ne sont pas publics")
    
    # Récupérer les catégories avec résultats
    categories = CompetitionCategory.objects.filter(
        competition=competition
    ).select_related('competition_type')
    
    # Ajouter les informations sur les résultats pour chaque catégorie
    for category in categories:
        # Vérifier si des résultats existent (CompetitionRanking ou CategoryRanking)
        try:
            from apps.competitions.models.technical_scoring import CompetitionRanking
            category.has_results = CompetitionRanking.objects.filter(
                competition=competition,
                category=category
            ).exists()
        except (ImportError, AttributeError):
            category.has_results = CategoryRanking.objects.filter(
                competition=competition,
                category=category
            ).exists()
    
    context = {
        'competition': competition,
        'categories': categories,
        'is_public': is_public,
    }
    
    return render(request, 'competitions/management/public_results.html', context)

