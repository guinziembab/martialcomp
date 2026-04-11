from django.core.exceptions import PermissionDenied
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.utils.translation import gettext_lazy as _
from django.contrib.auth.decorators import login_required
from django.db.models import Q, Avg, Sum, Max, Min, Count
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.utils import timezone
from django.db import transaction

from apps.competitions.models import (
    Competition, CompetitionCategory, Match, CompetitionRegistration,
    JudgeAssignment
)
from apps.competitions.models.technical_scoring import (
    ScoringCriterion, JudgeSubmissionStatus, ScoringConfiguration,
    CompetitionRanking, TechnicalPerformance, TechnicalScore, ScoringPreset,
    Performance, Score
)
from apps.competitions.utils.decorators import competition_management_permission_required
from apps.competitions.forms.scoring import (
    ScoringCriterionForm, ScoringConfigurationForm,
    TechnicalPerformanceForm, ScoringForm
)


@login_required
@competition_management_permission_required
def scoring_dashboard(request, competition_id):
    """
    Tableau de bord du système de notation pour une compétition.
    """
    # Récupérer la compétition
    competition = get_object_or_404(Competition, pk=competition_id)
    
    # Récupérer les catégories de la compétition
    categories = CompetitionCategory.objects.filter(
        competition=competition
    ).select_related('competition_type')

    # Extraire les types de compétition distincts pour le filtre
    competition_types = []
    seen_type_ids = set()
    for cat in categories:
        if cat.competition_type and cat.competition_type.pk not in seen_type_ids:
            seen_type_ids.add(cat.competition_type.pk)
            competition_types.append(cat.competition_type)
    competition_types.sort(key=lambda ct: ct.name)

    # Ajouter des statistiques Ã  chaque catégorie
    for category in categories:
        # Nombre de critères de notation
        category.scoring_criteria_count = ScoringCriterion.objects.filter(
            category=category
        ).count()
        
        # Nombre total de performances (modèle Performance utilisé par la feuille de notation)
        total_perfs = Performance.objects.filter(category=category)
        category.performances_count = total_perfs.count()

        # Performances notées : celles qui ont au moins un score enregistré
        scored_count = Score.objects.filter(
            performance__category=category
        ).values('performance').distinct().count()
        category.completed_performances_count = scored_count

        # Progression
        if category.performances_count > 0:
            category.scoring_progress = int((category.completed_performances_count / category.performances_count) * 100)
        else:
            category.scoring_progress = 0
    
    # Récupérer les juges assignés
    judges = JudgeAssignment.objects.filter(
        category__competition=competition,
        assignment_type__in=['technical_judge', 'chief_judge']
    ).select_related('user', 'category')
    
    # Récupérer les performances récentes
    recent_performances = TechnicalPerformance.objects.filter(
        category__competition=competition
    ).select_related('practitioner', 'category').order_by('-end_time')[:5]
    
    context = {
        'competition': competition,
        'categories': categories,
        'competition_types': competition_types,
        'judges': judges,
        'recent_performances': recent_performances,
    }
    
    return render(request, 'competitions/management/scoring_dashboard.html', context)


@login_required
@competition_management_permission_required
def competition_scoring_global_setup(request, competition_id):
    """
    Configuration globale de notation pour toute la compétition.
    Permet de définir les paramètres et critères une seule fois
    et de les appliquer à toutes les catégories.
    """
    import logging
    logger = logging.getLogger(__name__)

    try:
        return _competition_scoring_global_setup_inner(request, competition_id)
    except Exception as e:
        import traceback
        with open('/tmp/scoring_global_error.log', 'a') as f:
            f.write(f"\n{'='*60}\n")
            f.write(f"Error at {timezone.now()}\n")
            f.write(f"Method: {request.method}\n")
            f.write(f"POST data: {dict(request.POST)}\n")
            f.write(traceback.format_exc())
        raise


def _competition_scoring_global_setup_inner(request, competition_id):
    competition = get_object_or_404(Competition, pk=competition_id)
    categories = CompetitionCategory.objects.filter(competition=competition).order_by('name')

    if request.method == 'POST':
        action = request.POST.get('action', '')

        if action == 'apply_config':
            from decimal import Decimal, InvalidOperation

            def parse_decimal(value, default='0.00'):
                """Parse decimal value, handling comma as decimal separator."""
                try:
                    return Decimal(str(value).replace(',', '.').strip())
                except (InvalidOperation, ValueError):
                    return Decimal(default)

            # Apply scoring configuration to all categories
            min_score = parse_decimal(request.POST.get('min_score', '4.00'), '4.00')
            max_score = parse_decimal(request.POST.get('max_score', '7.00'), '7.00')
            score_step = parse_decimal(request.POST.get('score_step', '0.25'), '0.25')
            exclude_extreme = request.POST.get('exclude_extreme_scores') == 'on'
            allow_ties = request.POST.get('allow_ties') == 'on'
            allow_modification = request.POST.get('allow_score_modification') == 'on'
            real_time = request.POST.get('real_time_results') == 'on'
            training_judges = request.POST.get('training_judges_included') == 'on'

            # Configuration Tour 2
            tour2_mode = request.POST.get('tour2_mode', 'all')  # all, top_n, tiebreak
            tour2_top_n = int(request.POST.get('tour2_top_n', '6') or '6')

            # Mode de classement (average ou ccp)
            ranking_mode = request.POST.get('ranking_mode', 'average')
            ccp_coeff = float(request.POST.get('ccp_coeff', '1.5') or '1.5')

            advanced_config = {
                'tour2_mode': tour2_mode,
                'tour2_top_n': tour2_top_n,
                'ranking_mode': ranking_mode,
                'ccp_coeff': ccp_coeff,
            }

            updated = 0
            config_defaults = {
                'min_score': min_score,
                'max_score': max_score,
                'score_step': score_step,
                'exclude_extreme_scores': exclude_extreme,
                'allow_ties': allow_ties,
                'allow_score_modification': allow_modification,
                'real_time_results': real_time,
                'training_judges_included': training_judges,
                'advanced_config': advanced_config,
            }
            for cat in categories:
                config, created = ScoringConfiguration.objects.get_or_create(
                    category=cat,
                    defaults=config_defaults,
                )
                if not created:
                    for key, val in config_defaults.items():
                        setattr(config, key, val)
                    config.save()
                updated += 1

            messages.success(request, _("Configuration appliquée à %(count)d catégories.") % {'count': updated})
            return redirect('competitions:management:competition_scoring_global', competition_id=competition_id)

        elif action == 'apply_criteria':
            # Apply criteria to all categories that don't have criteria yet
            # Get criteria names and details from POST
            criteria_names = request.POST.getlist('criterion_name')
            criteria_weights = request.POST.getlist('criterion_weight')
            criteria_min = request.POST.getlist('criterion_min')
            criteria_max = request.POST.getlist('criterion_max')
            criteria_step = request.POST.getlist('criterion_step')

            if not criteria_names or not criteria_names[0]:
                messages.warning(request, _("Aucun critère défini."))
                return redirect('competitions:management:competition_scoring_global', competition_id=competition_id)

            apply_mode = request.POST.get('apply_mode', 'missing')  # 'missing' or 'all'
            applied = 0

            for cat in categories:
                existing_count = ScoringCriterion.objects.filter(category=cat).count()

                if apply_mode == 'all' or existing_count == 0:
                    if apply_mode == 'all':
                        # Delete existing criteria first
                        ScoringCriterion.objects.filter(category=cat).delete()

                    for i, name in enumerate(criteria_names):
                        if not name.strip():
                            continue
                        def _float(val, default):
                            try:
                                return float(str(val).replace(',', '.').strip())
                            except (ValueError, TypeError):
                                return default
                        ScoringCriterion.objects.create(
                            category=cat,
                            name=name.strip(),
                            weight=_float(criteria_weights[i], 1.0) if i < len(criteria_weights) else 1.0,
                            min_score=_float(criteria_min[i], 4.0) if i < len(criteria_min) else 4.0,
                            max_score=_float(criteria_max[i], 7.0) if i < len(criteria_max) else 7.0,
                            step=_float(criteria_step[i], 0.25) if i < len(criteria_step) else 0.25,
                            order=i + 1,
                            is_active=True,
                        )
                    applied += 1

            messages.success(request, _("Critères appliqués à %(count)d catégories.") % {'count': applied})
            return redirect('competitions:management:competition_scoring_global', competition_id=competition_id)

        elif action == 'save_preset':
            # Sauvegarder la config actuelle comme preset
            preset_name = request.POST.get('preset_name', '').strip()
            if not preset_name:
                messages.warning(request, _("Veuillez donner un nom au preset."))
                return redirect('competitions:management:competition_scoring_global', competition_id=competition_id)

            # Récupérer la discipline de la compétition
            discipline = competition.discipline if hasattr(competition, 'discipline') and competition.discipline else None
            if not discipline:
                messages.error(request, _("La compétition doit avoir une discipline pour sauvegarder un preset."))
                return redirect('competitions:management:competition_scoring_global', competition_id=competition_id)

            # Construire les données du preset depuis la config existante
            sample_cat = categories.first()
            config_data = {}
            if sample_cat:
                try:
                    sc = ScoringConfiguration.objects.get(category=sample_cat)
                    config_data['scoring'] = {
                        'min_score': str(sc.min_score),
                        'max_score': str(sc.max_score),
                        'score_step': str(sc.score_step),
                        'exclude_extreme_scores': sc.exclude_extreme_scores,
                        'allow_ties': sc.allow_ties,
                        'allow_score_modification': sc.allow_score_modification,
                        'real_time_results': sc.real_time_results,
                        'training_judges_included': sc.training_judges_included,
                        'advanced_config': sc.advanced_config or {},
                    }
                except ScoringConfiguration.DoesNotExist:
                    pass

                criteria = ScoringCriterion.objects.filter(category=sample_cat).order_by('order')
                config_data['criteria'] = [
                    {'name': c.name, 'weight': str(c.weight), 'min_score': str(c.min_score),
                     'max_score': str(c.max_score), 'step': str(c.step)}
                    for c in criteria
                ]

            preset, created = ScoringPreset.objects.update_or_create(
                name=preset_name,
                discipline=discipline,
                defaults={
                    'config_data': config_data,
                    'created_by': request.user,
                }
            )
            action_text = _("créé") if created else _("mis à jour")
            messages.success(request, _("Preset \"%(name)s\" %(action)s avec succès.") % {'name': preset_name, 'action': action_text})
            return redirect('competitions:management:competition_scoring_global', competition_id=competition_id)

        elif action == 'load_preset':
            preset_id = request.POST.get('preset_id')
            if not preset_id:
                messages.warning(request, _("Aucun preset sélectionné."))
                return redirect('competitions:management:competition_scoring_global', competition_id=competition_id)

            try:
                preset = ScoringPreset.objects.get(pk=preset_id)
            except ScoringPreset.DoesNotExist:
                messages.error(request, _("Preset introuvable."))
                return redirect('competitions:management:competition_scoring_global', competition_id=competition_id)

            data = preset.config_data
            scoring = data.get('scoring', {})
            criteria_list = data.get('criteria', [])

            from decimal import Decimal
            updated = 0
            for cat in categories:
                if scoring:
                    config, created = ScoringConfiguration.objects.get_or_create(
                        category=cat,
                        defaults={
                            'min_score': Decimal(scoring.get('min_score', '0')),
                            'max_score': Decimal(scoring.get('max_score', '10')),
                            'score_step': Decimal(scoring.get('score_step', '0.25')),
                        }
                    )
                    if not created:
                        config.min_score = Decimal(scoring.get('min_score', '0'))
                        config.max_score = Decimal(scoring.get('max_score', '10'))
                        config.score_step = Decimal(scoring.get('score_step', '0.25'))
                    config.exclude_extreme_scores = scoring.get('exclude_extreme_scores', False)
                    config.allow_ties = scoring.get('allow_ties', True)
                    config.allow_score_modification = scoring.get('allow_score_modification', False)
                    config.real_time_results = scoring.get('real_time_results', False)
                    config.training_judges_included = scoring.get('training_judges_included', False)
                    config.advanced_config = scoring.get('advanced_config', {})
                    config.save()

                if criteria_list:
                    ScoringCriterion.objects.filter(category=cat).delete()
                    for i, crit in enumerate(criteria_list):
                        ScoringCriterion.objects.create(
                            category=cat,
                            name=crit['name'],
                            weight=float(crit.get('weight', 1)),
                            min_score=float(crit.get('min_score', 0)),
                            max_score=float(crit.get('max_score', 10)),
                            step=float(crit.get('step', 0.25)),
                            order=i + 1,
                            is_active=True,
                        )
                updated += 1

            messages.success(request, _("Preset \"%(name)s\" importé sur %(count)d catégories.") % {'name': preset.name, 'count': updated})
            return redirect('competitions:management:competition_scoring_global', competition_id=competition_id)

    # Get existing global config from first category that has one
    sample_config = None
    for cat in categories:
        try:
            sample_config = ScoringConfiguration.objects.get(category=cat)
            break
        except ScoringConfiguration.DoesNotExist:
            continue

    # Get existing criteria from first category that has them
    sample_criteria = []
    for cat in categories:
        criteria = ScoringCriterion.objects.filter(category=cat).order_by('order')
        if criteria.exists():
            sample_criteria = list(criteria)
            break

    # Stats per category
    cat_stats = []
    for cat in categories:
        criteria_count = ScoringCriterion.objects.filter(category=cat).count()
        has_config = ScoringConfiguration.objects.filter(category=cat).exists()
        cat_stats.append({
            'category': cat,
            'criteria_count': criteria_count,
            'has_config': has_config,
        })

    # Presets disponibles pour la discipline de cette compétition
    available_presets = []
    if hasattr(competition, 'discipline') and competition.discipline:
        available_presets = list(ScoringPreset.objects.filter(
            discipline=competition.discipline
        ).order_by('-updated_at'))

    context = {
        'competition': competition,
        'categories': categories,
        'cat_stats': cat_stats,
        'sample_config': sample_config,
        'sample_criteria': sample_criteria,
        'total_categories': categories.count(),
        'available_presets': available_presets,
    }

    return render(request, 'competitions/management/competition_scoring_global.html', context)


@login_required
@competition_management_permission_required
def category_scoring_setup(request, competition_id, category_id):
    """
    Configure le système de notation pour une catégorie.
    """
    # Récupérer la compétition et la catégorie
    competition = get_object_or_404(Competition, pk=competition_id)
    category = get_object_or_404(CompetitionCategory, pk=category_id, competition=competition)
    
    # Récupérer la configuration de notation existante
    
    try:
        scoring_config = ScoringConfiguration.objects.get(category=category)
    except ScoringConfiguration.DoesNotExist:
        scoring_config = None
    
    if request.method == 'POST':
        config_form = ScoringConfigurationForm(request.POST, instance=scoring_config)
        if config_form.is_valid():
            config = config_form.save(commit=False)
            config.category = category
            config.save()
            
            messages.success(request, _("La configuration de notation a été mise Ã  jour."))
            return redirect('competitions:management:category_scoring_setup', 
                          competition_id=competition_id, 
                          category_id=category_id)
    else:
        config_form = ScoringConfigurationForm(instance=scoring_config)
    
    # Récupérer les critères de notation
    criteria = ScoringCriterion.objects.filter(category=category).order_by('order')
    
    context = {
        'competition': competition,
        'category': category,
        'scoring_config': scoring_config,
        'config_form': config_form,
        'criteria': criteria,
    }
    
    return render(request, 'competitions/management/category_scoring_setup.html', context)


@login_required
@competition_management_permission_required
def add_scoring_criterion(request, competition_id, category_id):
    """
    Ajoute un critère de notation pour une catégorie.
    """
    # Récupérer la compétition et la catégorie
    competition = get_object_or_404(Competition, pk=competition_id)
    category = get_object_or_404(CompetitionCategory, pk=category_id, competition=competition)
    
    # Récupérer la config de notation pour synchroniser les défauts
    try:
        scoring_config = ScoringConfiguration.objects.get(category=category)
    except ScoringConfiguration.DoesNotExist:
        scoring_config = None

    if request.method == 'POST':
        form = ScoringCriterionForm(request.POST)
        if form.is_valid():
            criterion = form.save(commit=False)
            criterion.category = category

            # Déterminer l'ordre si non spécifié
            if not criterion.order:
                last_order = ScoringCriterion.objects.filter(
                    category=category
                ).order_by('-order').values_list('order', flat=True).first() or 0
                criterion.order = last_order + 1

            criterion.save()

            messages.success(request, _("Le critère de notation a été ajouté."))
            return redirect('competitions:management:category_scoring_setup',
                          competition_id=competition_id,
                          category_id=category_id)
    else:
        # Pré-remplir avec les valeurs de la ScoringConfiguration
        initial = {}
        if scoring_config:
            initial['min_score'] = float(scoring_config.min_score)
            initial['max_score'] = float(scoring_config.max_score)
            initial['step'] = float(scoring_config.score_step)
        form = ScoringCriterionForm(initial=initial)

    context = {
        'competition': competition,
        'category': category,
        'form': form,
        'scoring_config': scoring_config,
    }

    return render(request, 'competitions/management/add_scoring_criterion.html', context)


@login_required
@competition_management_permission_required
def edit_scoring_criterion(request, competition_id, criterion_id):
    """
    Modifie un critère de notation.
    """
    # Récupérer la compétition et le critère
    competition = get_object_or_404(Competition, pk=competition_id)
    criterion = get_object_or_404(
        ScoringCriterion, 
        pk=criterion_id, 
        category__competition=competition
    )
    
    # Récupérer la config de notation
    try:
        scoring_config = ScoringConfiguration.objects.get(category=criterion.category)
    except ScoringConfiguration.DoesNotExist:
        scoring_config = None

    if request.method == 'POST':
        form = ScoringCriterionForm(request.POST, instance=criterion)
        if form.is_valid():
            form.save()

            messages.success(request, _("Le critère de notation a été mis à jour."))
            return redirect('competitions:management:category_scoring_setup',
                          competition_id=competition_id,
                          category_id=criterion.category.id)
    else:
        form = ScoringCriterionForm(instance=criterion)

    context = {
        'competition': competition,
        'category': criterion.category,
        'criterion': criterion,
        'form': form,
        'scoring_config': scoring_config,
    }

    return render(request, 'competitions/management/edit_scoring_criterion.html', context)


@login_required
@competition_management_permission_required
@require_POST
def delete_scoring_criterion(request, competition_id, criterion_id):
    """
    Supprime un critère de notation.
    """
    # Récupérer la compétition et le critère
    competition = get_object_or_404(Competition, pk=competition_id)
    criterion = get_object_or_404(
        ScoringCriterion, 
        pk=criterion_id, 
        category__competition=competition
    )
    
    category_id = criterion.category.id
    
    # Vérifier si des scores existent pour ce critère
    scores_exist = TechnicalScore.objects.filter(criterion=criterion).exists()
    
    if scores_exist:
        messages.error(request, _("Impossible de supprimer ce critère car des scores existent déjÃ ."))
    else:
        criterion.delete()
        messages.success(request, _("Le critère de notation a été supprimé."))
    
    return redirect('competitions:management:category_scoring_setup', 
                  competition_id=competition_id, 
                  category_id=category_id)


@login_required
@competition_management_permission_required
def reorder_scoring_criteria(request, competition_id, category_id):
    """
    Réorganise l'ordre des critères de notation.
    """
    # Récupérer la compétition et la catégorie
    competition = get_object_or_404(Competition, pk=competition_id)
    category = get_object_or_404(CompetitionCategory, pk=category_id, competition=competition)
    
    if request.method == 'POST':
        # Récupérer les données JSON
        import json
        try:
            data = json.loads(request.body)
            criteria_orders = data.get('criteriaOrders', [])
            
            with transaction.atomic():
                for item in criteria_orders:
                    criterion_id = item.get('id')
                    new_order = item.get('order')
                    
                    if criterion_id and new_order is not None:
                        ScoringCriterion.objects.filter(
                            id=criterion_id,
                            category=category
                        ).update(order=new_order)
                
                return JsonResponse({'status': 'success'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    
    # Si méthode GET, afficher la page de réorganisation
    criteria = ScoringCriterion.objects.filter(
        category=category
    ).order_by('order')
    
    context = {
        'competition': competition,
        'category': category,
        'criteria': criteria,
    }
    
    return render(request, 'competitions/management/reorder_criteria.html', context)


@login_required
@competition_management_permission_required
def manage_performances(request, competition_id, category_id):
    """
    Gère les performances pour une catégorie.
    """
    # Récupérer la compétition et la catégorie
    competition = get_object_or_404(Competition, pk=competition_id)
    category = get_object_or_404(CompetitionCategory, pk=category_id, competition=competition)
    
    # Récupérer les performances existantes
    performances = TechnicalPerformance.objects.filter(
        category=category
    ).select_related('practitioner', 'practitioner__organization').order_by('performance_order')

    # Détecter mode équipe
    is_team_mode = False
    try:
        if category.competition_type and category.competition_type.team_based:
            is_team_mode = True
    except Exception:
        pass

    # Enrichir les performances avec les données d'équipe via MembreEquipe
    performances_list = list(performances)
    if is_team_mode:
        from apps.competitions.models.combat import MembreEquipe
        for perf in performances_list:
            membership = MembreEquipe.objects.filter(
                pratiquant=perf.practitioner,
                equipe__category=category,
                equipe__is_active=True,
            ).select_related('equipe', 'equipe__club').first()
            if membership:
                team = membership.equipe
                perf.is_team = True
                perf.team_name = team.nom
                perf.team_club = team.club.name if team.club else ''
                perf.team_members = list(
                    team.memberships.filter(est_remplacant=False)
                    .select_related('pratiquant')
                    .order_by('ordre')
                )
            else:
                perf.is_team = False
                perf.team_members = []

    # Récupérer les participants inscrits qui n'ont pas encore de performance
    participants = CompetitionRegistration.objects.filter(
        competition=competition,
        categories=category,
        is_competitor=True,
        status='approved'
    ).select_related('practitioner').exclude(
        practitioner__id__in=performances.values_list('practitioner__id', flat=True)
    )

    context = {
        'competition': competition,
        'category': category,
        'performances': performances_list,
        'participants': participants,
        'is_team_mode': is_team_mode,
    }

    return render(request, 'competitions/management/manage_performances.html', context)


@login_required
@competition_management_permission_required
def add_performance(request, competition_id, category_id):
    """
    Ajoute une performance pour un participant.
    """
    # Récupérer la compétition et la catégorie
    competition = get_object_or_404(Competition, pk=competition_id)
    category = get_object_or_404(CompetitionCategory, pk=category_id, competition=competition)
    
    if request.method == 'POST':
        form = TechnicalPerformanceForm(request.POST, competition=competition, category=category)
        if form.is_valid():
            performance = form.save(commit=False)
            performance.competition = competition
            performance.category = category
            
            # Déterminer l'ordre si non spécifié
            if not performance.performance_order:
                # Prendre le dernier ordre + 1
                last_order = TechnicalPerformance.objects.filter(
                    category=category
                ).order_by('-performance_order').values_list('performance_order', flat=True).first() or 0
                performance.performance_order = last_order + 1
            
            performance.save()
            
            messages.success(request, _("La performance a été ajoutée."))
            return redirect('competitions:management:manage_performances', 
                          competition_id=competition_id, 
                          category_id=category_id)
    else:
        form = TechnicalPerformanceForm(competition=competition, category=category)
    
    context = {
        'competition': competition,
        'category': category,
        'form': form,
    }
    
    return render(request, 'competitions/management/add_performance.html', context)


@login_required
@competition_management_permission_required
def edit_performance(request, competition_id, performance_id):
    """
    Modifie une performance.
    """
    # Récupérer la compétition et la performance
    competition = get_object_or_404(Competition, pk=competition_id)
    performance = get_object_or_404(
        TechnicalPerformance, 
        pk=performance_id, 
        competition=competition
    )
    
    if request.method == 'POST':
        form = TechnicalPerformanceForm(
            request.POST, 
            instance=performance, 
            competition=competition, 
            category=performance.category
        )
        if form.is_valid():
            form.save()
            
            messages.success(request, _("La performance a été mise Ã  jour."))
            return redirect('competitions:management:manage_performances', 
                          competition_id=competition_id, 
                          category_id=performance.category.id)
    else:
        form = TechnicalPerformanceForm(
            instance=performance, 
            competition=competition, 
            category=performance.category
        )
    
    context = {
        'competition': competition,
        'category': performance.category,
        'performance': performance,
        'form': form,
    }
    
    return render(request, 'competitions/management/edit_performance.html', context)


@login_required
@competition_management_permission_required
@require_POST
def delete_performance(request, competition_id, performance_id):
    """
    Supprime une performance.
    """
    # Récupérer la compétition et la performance
    competition = get_object_or_404(Competition, pk=competition_id)
    performance = get_object_or_404(
        TechnicalPerformance, 
        pk=performance_id, 
        competition=competition
    )
    
    category_id = performance.category.id
    
    # Vérifier si des scores existent pour cette performance
    scores_exist = TechnicalScore.objects.filter(performance=performance).exists()
    
    if scores_exist:
        messages.error(request, _("Impossible de supprimer cette performance car des scores existent déjÃ ."))
    else:
        performance.delete()
        messages.success(request, _("La performance a été supprimée."))
    
    return redirect('competitions:management:manage_performances', 
                  competition_id=competition_id, 
                  category_id=category_id)


@login_required
@competition_management_permission_required
@require_POST
def start_performance(request, competition_id, performance_id):
    """
    Démarre une performance (pour le chronométrage).
    """
    # Récupérer la compétition et la performance
    competition = get_object_or_404(Competition, pk=competition_id)
    performance = get_object_or_404(
        TechnicalPerformance, 
        pk=performance_id, 
        competition=competition
    )
    
    # Démarrer la performance
    performance.start_performance()
    
    messages.success(request, _("La performance a démarré."))
    
    # Redirection selon le paramètre next
    next_url = request.POST.get('next')
    if next_url:
        return redirect(next_url)
    
    return redirect('competitions:management:manage_performances', 
                  competition_id=competition_id, 
                  category_id=performance.category.id)


@login_required
@competition_management_permission_required
@require_POST
def end_performance(request, competition_id, performance_id):
    """
    Termine une performance (pour le chronométrage).
    """
    # Récupérer la compétition et la performance
    competition = get_object_or_404(Competition, pk=competition_id)
    performance = get_object_or_404(
        TechnicalPerformance, 
        pk=performance_id, 
        competition=competition,
        status='in_progress'
    )
    
    # Terminer la performance
    performance.end_performance()
    
    messages.success(request, _("La performance est terminée."))
    
    # Redirection selon le paramètre next
    next_url = request.POST.get('next')
    if next_url:
        return redirect(next_url)
    
    return redirect('competitions:management:manage_performances', 
                  competition_id=competition_id, 
                  category_id=performance.category.id)


@login_required
@competition_management_permission_required
def reorder_performances(request, competition_id, category_id):
    """
    Réorganise l'ordre des performances.
    """
    # Récupérer la compétition et la catégorie
    competition = get_object_or_404(Competition, pk=competition_id)
    category = get_object_or_404(CompetitionCategory, pk=category_id, competition=competition)
    
    if request.method == 'POST':
        # Récupérer les données JSON
        import json
        try:
            data = json.loads(request.body)
            performance_orders = data.get('performanceOrders', [])
            
            with transaction.atomic():
                for item in performance_orders:
                    performance_id = item.get('id')
                    new_order = item.get('order')
                    
                    if performance_id and new_order is not None:
                        TechnicalPerformance.objects.filter(
                            id=performance_id,
                            category=category
                        ).update(performance_order=new_order)
                
                return JsonResponse({'status': 'success'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    
    # Si méthode GET, afficher la page de réorganisation
    performances = TechnicalPerformance.objects.filter(
        category=category
    ).select_related('practitioner').order_by('performance_order')
    
    context = {
        'competition': competition,
        'category': category,
        'performances': performances,
    }
    
    return render(request, 'competitions/management/reorder_performances.html', context)


@login_required
@competition_management_permission_required
def performance_scores(request, competition_id, performance_id):
    """
    Affiche les scores d'une performance.
    """
    # Récupérer la compétition et la performance
    competition = get_object_or_404(Competition, pk=competition_id)
    performance = get_object_or_404(
        TechnicalPerformance, 
        pk=performance_id, 
        competition=competition
    )
    
    # Récupérer les scores par juge et par critère
    scores = TechnicalScore.objects.filter(
        performance=performance
    ).select_related('judge', 'criterion')
    
    # Organiser les scores par juge et par critère
    judges = {}
    criteria = ScoringCriterion.objects.filter(
        category=performance.category
    ).order_by('order')
    
    for score in scores:
        judge_id = score.judge.id
        criterion_id = score.criterion.id
        
        if judge_id not in judges:
            judges[judge_id] = {
                'name': f"{score.judge.first_name} {score.judge.last_name}",
                'scores': {}
            }
        
        judges[judge_id]['scores'][criterion_id] = score.value
    
    context = {
        'competition': competition,
        'performance': performance,
        'category': performance.category,
        'criteria': criteria,
        'judges': judges,
        'scores': scores,
    }
    
    return render(request, 'competitions/management/performance_scores.html', context)


@login_required
@competition_management_permission_required
def add_score(request, competition_id, performance_id):
    """
    Ajoute un score pour une performance (pour les administrateurs).
    """
    # Récupérer la compétition et la performance
    competition = get_object_or_404(Competition, pk=competition_id)
    performance = get_object_or_404(
        TechnicalPerformance, 
        pk=performance_id, 
        competition=competition
    )
    
    if request.method == 'POST':
        form = ScoringForm(request.POST, performance=performance)
        if form.is_valid():
            judge = form.cleaned_data['judge']
            criterion = form.cleaned_data['criterion']
            value = form.cleaned_data['value']
            
            # Vérifier si un score existe déjÃ 
            try:
                score = TechnicalScore.objects.get(
                    performance=performance,
                    judge=judge,
                    criterion=criterion
                )
                score.value = value
                score.save()
                messages.success(request, _("Le score a été mis Ã  jour."))
            except TechnicalScore.DoesNotExist:
                # Créer un nouveau score
                TechnicalScore.objects.create(
                    performance=performance,
                    judge=judge,
                    criterion=criterion,
                    value=value
                )
                messages.success(request, _("Le score a été ajouté."))
            
            return redirect('competitions:management:performance_scores', 
                          competition_id=competition_id, 
                          performance_id=performance_id)
    else:
        form = ScoringForm(performance=performance)
    
    context = {
        'competition': competition,
        'performance': performance,
        'category': performance.category,
        'form': form,
    }
    
    return render(request, 'competitions/management/add_score.html', context)


@login_required
@competition_management_permission_required
@require_POST
def delete_score(request, competition_id, score_id):
    """
    Supprime un score.
    """
    # Récupérer la compétition et le score
    competition = get_object_or_404(Competition, pk=competition_id)
    score = get_object_or_404(
        TechnicalScore, 
        pk=score_id, 
        performance__competition=competition
    )
    
    performance_id = score.performance.id
    
    # Supprimer le score
    score.delete()
    
    messages.success(request, _("Le score a été supprimé."))
    return redirect('competitions:management:performance_scores', 
                  competition_id=competition_id, 
                  performance_id=performance_id)


@login_required
@competition_management_permission_required
def calculate_results(request, competition_id, category_id):
    """
    Calcule les résultats pour une catégorie.
    """
    # Récupérer la compétition et la catégorie
    competition = get_object_or_404(Competition, pk=competition_id)
    category = get_object_or_404(CompetitionCategory, pk=category_id, competition=competition)

    # Calcul en masse des scores finaux
    # Inclure toutes les performances qui ont des scores (pas seulement status='completed')
    performances = TechnicalPerformance.objects.filter(
        category=category
    ).filter(
        # Au moins un score existe pour cette performance
        scores__isnull=False
    ).distinct().select_related('practitioner')
    
    # Récupérer la configuration de notation
    try:
        config = ScoringConfiguration.objects.get(category=category)
        exclude_extreme_scores = config.exclude_extreme_scores
        allow_ties = config.allow_ties
        advanced = config.advanced_config or {}
        ranking_mode = advanced.get('ranking_mode', 'average')
        ccp_coeff = float(advanced.get('ccp_coeff', 1.5))
    except ScoringConfiguration.DoesNotExist:
        exclude_extreme_scores = False
        allow_ties = True
        ranking_mode = 'average'
        ccp_coeff = 1.5
    
    # Supprimer les classements existants
    CompetitionRanking.objects.filter(
        competition=competition,
        category=category
    ).delete()

    if ranking_mode == 'ccp':
        # ====== MODE CCP (Consensus de Position) ======
        # Chaque juge produit son propre classement independant.
        # Le classement final = agregation des positions ponderees de chaque juge.
        # Le score affiche = valeur CCP (plus bas = meilleur).
        #
        # TOUR 2 (barrage) : Le CCP est calcule sur les scores TOUR 1 uniquement.
        # Les scores Tour 2 servent uniquement a departager les ex-aequo.

        from collections import defaultdict

        # --- Etape 1 : CCP sur les scores Tour 1 ---
        judge_scores_map = defaultdict(dict)
        perf_map = {}

        for performance in performances:
            perf_map[performance.id] = performance
            scores = TechnicalScore.objects.filter(
                performance=performance,
                round_number=1,
            ).values('judge').annotate(avg_score=Avg('value'))
            for s in scores:
                judge_scores_map[s['judge']][performance.id] = float(s['avg_score'])

        # Pour chaque juge, classer et attribuer des positions
        perf_positions = defaultdict(list)

        for judge_id, scores_dict in judge_scores_map.items():
            sorted_perfs = sorted(scores_dict.items(), key=lambda x: x[1], reverse=True)
            current_pos = 1
            prev_score = None
            for idx, (perf_id, score) in enumerate(sorted_perfs):
                if prev_score is not None and score != prev_score:
                    current_pos = idx + 1
                perf_positions[perf_id].append(current_pos)
                prev_score = score

        # Calculer CCP Tour 1 (normalisé par le nombre de juges)
        total_judges = len(judge_scores_map)
        ccp_results = []
        for perf_id, positions in perf_positions.items():
            weighted = [1.0 + (pos - 1) * ccp_coeff for pos in positions]
            # Normaliser : si un pratiquant n'a pas été noté par tous les juges,
            # les juges manquants comptent comme dernière position
            while len(weighted) < total_judges:
                last_pos = len(perf_map) + 1
                weighted.append(1.0 + (last_pos - 1) * ccp_coeff)
            ccp_score = sum(weighted) if weighted else 999
            ccp_results.append({
                'perf_id': perf_id,
                'practitioner': perf_map[perf_id].practitioner,
                'ccp_score': round(ccp_score, 2),
                'tour2_score': None,
            })

        # --- Etape 2 : Scores Tour 2 pour departage ---
        for item in ccp_results:
            t2_scores = TechnicalScore.objects.filter(
                performance_id=item['perf_id'],
                round_number=2,
            ).values('judge').annotate(avg_score=Avg('value'))
            if t2_scores:
                # Moyenne simple des scores Tour 2 pour le departage
                item['tour2_score'] = round(
                    sum(s['avg_score'] for s in t2_scores) / len(t2_scores), 2
                )

        # --- Etape 3 : Tri CCP puis departage par Tour 2 ---
        # Tri principal : CCP croissant (meilleur = plus bas)
        # Tri secondaire : Tour 2 score croissant (meilleur = plus bas, None = dernier)
        def sort_key(item):
            t2 = item['tour2_score'] if item['tour2_score'] is not None else 9999
            return (item['ccp_score'], -t2 if item['tour2_score'] is not None else 9999)

        # Tri : CCP croissant, puis pour les ex-aequo CCP, Tour 2 decroissant
        # (score T2 plus haut = meilleur dans le barrage)
        ccp_results.sort(key=lambda x: (x['ccp_score'], -(x['tour2_score'] or 0)))

        for item in ccp_results:
            CompetitionRanking.objects.create(
                competition=competition,
                category=category,
                practitioner=item['practitioner'],
                rank=0,
                final_score=item['ccp_score'],
                first_places=0,
            )

        # --- Etape 4 : Attribuer les rangs ---
        ccp_rankings = list(CompetitionRanking.objects.filter(
            competition=competition,
            category=category
        ).order_by('id'))

        # Construire la map des resultats
        ccp_data = {}
        for r in ccp_results:
            ccp_data[r['practitioner'].id] = {
                'ccp': r['ccp_score'],
                't2': r['tour2_score'],
            }

        prev_ccp = None
        prev_t2 = None
        current_rank = 1

        for i, ranking in enumerate(ccp_rankings):
            data = ccp_data.get(ranking.practitioner_id, {})
            cur_ccp = data.get('ccp')
            cur_t2 = data.get('t2')

            is_same_ccp = (i > 0 and cur_ccp is not None and cur_ccp == prev_ccp)
            is_same_t2 = (cur_t2 is not None and prev_t2 is not None and cur_t2 == prev_t2)
            # Ex-aequo seulement si meme CCP ET (pas de T2 OU meme T2)
            is_tied = is_same_ccp and (cur_t2 is None or prev_t2 is None or is_same_t2)

            if is_tied:
                ranking.rank = current_rank
                ranking.is_tie = True
                if not ccp_rankings[i - 1].is_tie:
                    ccp_rankings[i - 1].is_tie = True
                    ccp_rankings[i - 1].save()
            else:
                current_rank = i + 1
                ranking.rank = current_rank
                ranking.is_tie = False
            ranking.save()
            prev_ccp = cur_ccp
            prev_t2 = cur_t2

        # En mode équipe, agréger les rankings par équipe
        is_team_mode = False
        try:
            if category.competition_type and category.competition_type.team_based:
                is_team_mode = True
        except Exception:
            pass

        if is_team_mode:
            from apps.competitions.models.combat import MembreEquipe
            # Grouper les rankings par équipe
            team_rankings = {}
            solo_rankings = []
            all_rankings = CompetitionRanking.objects.filter(
                competition=competition, category=category
            ).select_related('practitioner')

            for ranking in all_rankings:
                membership = MembreEquipe.objects.filter(
                    pratiquant=ranking.practitioner,
                    equipe__category=category,
                    equipe__is_active=True,
                ).select_related('equipe').first()
                if membership:
                    team_id = membership.equipe_id
                    if team_id not in team_rankings:
                        team_rankings[team_id] = {
                            'rankings': [],
                            'ccp_scores': [],
                            'keeper': None,
                        }
                    team_rankings[team_id]['rankings'].append(ranking)
                    team_rankings[team_id]['ccp_scores'].append(ranking.final_score)
                    # Garder le premier membre comme représentant
                    if team_rankings[team_id]['keeper'] is None:
                        team_rankings[team_id]['keeper'] = ranking
                else:
                    solo_rankings.append(ranking)

            # Pour chaque équipe : garder 1 ranking avec le CCP moyen, supprimer les doublons
            for team_id, data in team_rankings.items():
                avg_ccp = sum(data['ccp_scores']) / len(data['ccp_scores'])
                keeper = data['keeper']
                keeper.final_score = round(avg_ccp, 2)
                keeper.save()
                # Supprimer les autres rankings du même team
                for r in data['rankings']:
                    if r.id != keeper.id:
                        r.delete()

            # Recalculer les rangs après agrégation
            final_rankings = list(CompetitionRanking.objects.filter(
                competition=competition, category=category
            ).order_by('final_score'))
            prev_score = None
            current_rank = 1
            for i, ranking in enumerate(final_rankings):
                if i > 0 and ranking.final_score == prev_score:
                    ranking.rank = current_rank
                    ranking.is_tie = True
                    final_rankings[i-1].is_tie = True
                    final_rankings[i-1].save()
                else:
                    current_rank = i + 1
                    ranking.rank = current_rank
                    ranking.is_tie = False
                ranking.save()
                prev_score = ranking.final_score

        messages.success(request, _("Les resultats CCP ont ete calcules avec succes."))
        return redirect('competitions:management:category_results',
                       competition_id=competition_id,
                       category_id=category_id)

    else:
        # ====== MODE MOYENNE (classique) ======
        for performance in performances:
            final_score = performance.calculate_final_score()
            CompetitionRanking.objects.create(
                competition=competition,
                category=category,
                practitioner=performance.practitioner,
                rank=0,
                final_score=final_score,
                first_places=0,
            )

    # Attribuer les rangs
    rankings_list = list(CompetitionRanking.objects.filter(
        competition=competition,
        category=category
    ).order_by('-final_score', '-first_places'))

    previous_score = None
    previous_first_places = None
    current_rank = 1
    for i, ranking in enumerate(rankings_list):
        if i > 0 and ranking.final_score == previous_score and ranking.first_places == previous_first_places:
            # Ex-aequo: meme rang que le precedent
            if not allow_ties and current_rank == 3:
                ranking.rank = i + 1
            else:
                ranking.rank = current_rank
            ranking.is_tie = True
            # Marquer aussi le premier du groupe comme ex-aequo
            if not rankings_list[i - 1].is_tie:
                rankings_list[i - 1].is_tie = True
                rankings_list[i - 1].save()
        else:
            # Nouveau rang = position reelle (saute les rangs des ex-aequo)
            current_rank = i + 1
            ranking.rank = current_rank
            ranking.is_tie = False

        ranking.save()
        previous_score = ranking.final_score
        previous_first_places = ranking.first_places
    
    messages.success(request, _("Les résultats ont été calculés avec succès."))
    return redirect('competitions:management:category_results', 
                   competition_id=competition_id, 
                   category_id=category_id)


@login_required
@competition_management_permission_required
def export_results(request, competition_id, category_id):
    """
    Exporte les resultats d'une categorie au format CSV ou Excel.
    """
    try:
        return _export_results_impl(request, competition_id, category_id)
    except Exception as e:
        import traceback
        from django.http import HttpResponse
        return HttpResponse(
            f"<pre>Export Error:\n{e}\n\n{traceback.format_exc()}</pre>",
            status=500
        )


def _export_results_impl(request, competition_id, category_id):
    import csv
    from django.http import HttpResponse

    competition = get_object_or_404(Competition, pk=competition_id)
    category = get_object_or_404(CompetitionCategory, pk=category_id, competition=competition)

    rankings = CompetitionRanking.objects.filter(
        competition=competition,
        category=category
    ).select_related('practitioner').order_by('rank')

    # Verifier si Tour 2 existe
    has_tour2 = TechnicalScore.objects.filter(
        performance__category=category, round_number=2
    ).exists()

    # Preparer les donnees
    headers = ['Rang', 'Nom', 'Prenom', 'Club', 'Score Tour 1']
    if has_tour2:
        headers.append('Score Tour 2')
    headers.extend(['Score CCP', 'Ex-aequo', 'Medaille'])

    rows = []
    for ranking in rankings:
        p = ranking.practitioner
        perf = TechnicalPerformance.objects.filter(category=category, practitioner=p).first()
        score_t1 = ''
        score_t2 = ''
        if perf:
            t1 = TechnicalScore.objects.filter(performance=perf, round_number=1).values('judge').annotate(avg=Avg('value'))
            if t1:
                score_t1 = round(sum(float(s['avg']) for s in t1) / len(t1), 2)
            t2 = TechnicalScore.objects.filter(performance=perf, round_number=2).values('judge').annotate(avg=Avg('value'))
            if t2:
                score_t2 = round(sum(float(s['avg']) for s in t2) / len(t2), 2)

        club_name = ''
        try:
            if p.organization:
                club_name = p.organization.name
        except Exception:
            pass
        if not club_name:
            try:
                if p.club:
                    club_name = p.club.name
            except Exception:
                pass

        medal = ''
        if ranking.rank == 1:
            medal = 'Or'
        elif ranking.rank == 2:
            medal = 'Argent'
        elif ranking.rank == 3:
            medal = 'Bronze'

        row = [ranking.rank, p.last_name, p.first_name, club_name, score_t1]
        if has_tour2:
            row.append(score_t2)
        row.extend([float(ranking.final_score), 'Oui' if ranking.is_tie else 'Non', medal])
        rows.append(row)

    export_format = request.GET.get('format', 'csv')

    if export_format == 'pdf':
        from xhtml2pdf import pisa
        from io import BytesIO

        date_str = ''
        if competition.start_date:
            date_str = competition.start_date.strftime('%d/%m/%Y')

        # Construction du HTML pour le PDF - approche simple compatible xhtml2pdf
        rows_html = ''
        for row in rows:
            medal = row[-1]
            bg = ''
            if 'Or' in str(medal):
                bg = ' bgcolor="#FFF9DB"'
            elif 'Argent' in str(medal):
                bg = ' bgcolor="#F1F3F5"'
            elif 'Bronze' in str(medal):
                bg = ' bgcolor="#FFF4E6"'
            cells = ''
            for i, v in enumerate(row):
                align = 'left' if i in (1, 2, 3) else 'center'
                cells += f'<td align="{align}">{v}</td>'
            rows_html += f'<tr{bg}>{cells}</tr>'

        # En-tetes avec largeurs fixes
        if has_tour2:
            th_list = [
                ('5%', 'Rang'), ('17%', 'Nom'), ('13%', 'Prenom'), ('20%', 'Club'),
                ('9%', 'Tour 1'), ('9%', 'Tour 2'), ('9%', 'Score CCP'),
                ('7%', 'Ex-aeq'), ('11%', 'Medaille'),
            ]
        else:
            th_list = [
                ('6%', 'Rang'), ('18%', 'Nom'), ('15%', 'Prenom'), ('24%', 'Club'),
                ('11%', 'Tour 1'), ('11%', 'Score CCP'),
                ('7%', 'Ex-aeq'), ('11%', 'Medaille'),
            ]
        headers_html = ''.join(f'<th width="{w}">{h}</th>' for w, h in th_list)

        html_content = f"""<!DOCTYPE html><html><head><meta charset="utf-8">
        <style>
            @page {{ size: A4 landscape; margin: 1.5cm; }}
            body {{ font-family: Helvetica, sans-serif; }}
            table {{ width: 100%; border-collapse: collapse; }}
            th {{ background-color: #2D3748; color: white; padding: 5px 8px; font-size: 9px; }}
            td {{ padding: 4px 8px; font-size: 10px; border-bottom: 1px solid #ccc; }}
        </style></head><body>
        <table width="100%"><tr>
            <td style="background-color: #1A202C; color: #C9A227; padding: 12px; font-size: 20px; font-weight: bold;" align="center">{competition.title}</td>
        </tr><tr>
            <td style="background-color: #2D3748; color: white; padding: 8px; font-size: 14px;" align="center">{category.name}</td>
        </tr></table>
        <p style="text-align: center; color: #666; font-size: 9px;">Date: {date_str} | Participants: {len(rows)} | Mode: CCP</p>
        <table>
            <tr>{headers_html}</tr>
            {rows_html}
        </table>
        <p style="text-align: center; font-size: 8px; color: #999; margin-top: 15px; border-top: 1px solid #ddd; padding-top: 5px;">Document genere le {timezone.now().strftime('%d/%m/%Y %H:%M')} - MartialComp</p>
        </body></html>"""

        result = BytesIO()
        pdf = pisa.CreatePDF(html_content, dest=result, encoding='utf-8')
        if not pdf.err:
            response = HttpResponse(result.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{category.name}_resultats.pdf"'
            return response

    elif export_format == 'excel':
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = category.name[:31]
        total_cols = len(headers)

        # --- Banniere competition ---
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        title_cell = ws['A1']
        title_cell.value = competition.title
        title_cell.font = Font(bold=True, size=16, color='C9A227')
        title_cell.fill = PatternFill(start_color='1A202C', end_color='1A202C', fill_type='solid')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 35

        # Sous-titre categorie
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
        sub_cell = ws['A2']
        sub_cell.value = category.name
        sub_cell.font = Font(bold=True, size=13, color='FFFFFF')
        sub_cell.fill = PatternFill(start_color='2D3748', end_color='2D3748', fill_type='solid')
        sub_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 28

        # Infos competition
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=total_cols)
        info_cell = ws['A3']
        date_str = ''
        if competition.start_date:
            date_str = competition.start_date.strftime('%d/%m/%Y')
        info_cell.value = f'Date: {date_str}  |  Participants: {rankings.count()}  |  Mode: CCP'
        info_cell.font = Font(size=10, italic=True, color='A0AEC0')
        info_cell.fill = PatternFill(start_color='2D3748', end_color='2D3748', fill_type='solid')
        info_cell.alignment = Alignment(horizontal='center')
        ws.row_dimensions[3].height = 22

        # Ligne vide
        ws.row_dimensions[4].height = 8

        # En-tetes tableau
        header_row = 5
        thin_border = Border(
            bottom=Side(style='thin', color='4A5568'),
        )
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = Font(bold=True, size=10, color='FFFFFF')
            cell.fill = PatternFill(start_color='4A5568', end_color='4A5568', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        ws.row_dimensions[header_row].height = 25

        # Donnees
        medal_colors = {
            'Or': 'FFF9DB',
            'Argent': 'F1F3F5',
            'Bronze': 'FFF4E6',
        }
        for row_idx, row_data in enumerate(rows, header_row + 1):
            medal_val = row_data[-1]
            bg_color = medal_colors.get(medal_val, None)
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(size=10)
                if bg_color:
                    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')

        # Ajuster largeurs
        from openpyxl.utils import get_column_letter
        col_widths = [8, 15, 12, 18, 12, 12, 12, 10, 10]
        for i, w in enumerate(col_widths[:total_cols], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{category.name}_resultats.xlsx"'
        wb.save(response)
        return response

    # CSV par defaut
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{category.name}_resultats.csv"'
    response.write('\ufeff')  # BOM UTF-8 pour Excel

    writer = csv.writer(response, delimiter=';')
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)

    return response


@login_required
@competition_management_permission_required
def judge_scoring_interface(request, competition_id, category_id, judge_id):
    """
    Interface de notation pour un juge spécifique (vue administrateur).
    """
    try:
        return _judge_scoring_impl(request, competition_id, category_id, judge_id)
    except Exception as e:
        import traceback
        from django.http import HttpResponse
        return HttpResponse(f"<pre>{e}\n\n{traceback.format_exc()}</pre>", status=500)


def _judge_scoring_impl(request, competition_id, category_id, judge_id):
    # Récupérer la compétition, la catégorie et le juge
    competition = get_object_or_404(Competition, pk=competition_id)
    category = get_object_or_404(CompetitionCategory, pk=category_id, competition=competition)
    
    from django.contrib.auth import get_user_model
    User = get_user_model()
    
    judge = get_object_or_404(User, pk=judge_id)
    
    # Vérifier que le juge est bien assigné Ã  cette catégorie
    assignment = get_object_or_404(
        JudgeAssignment,
        category=category,
        user=judge,
        assignment_type__in=['technical_judge', 'chief_judge']
    )
    
    # Récupérer les performances en cours ou Ã  venir
    performances = TechnicalPerformance.objects.filter(
        category=category,
        status__in=['pending', 'in_progress']
    ).select_related('practitioner').order_by('performance_order')
    
    # Récupérer la performance actuelle
    current_performance = performances.filter(status='in_progress').first()
    
    # Sinon, prendre la première performance en attente
    if not current_performance and performances.exists():
        current_performance = performances.first()
    
    # Récupérer les critères de notation
    criteria = ScoringCriterion.objects.filter(
        category=category
    ).order_by('order')
    
    # Si une performance est en cours, récupérer les scores existants
    scores = {}
    if current_performance:
        existing_scores = TechnicalScore.objects.filter(
            performance=current_performance,
            judge=judge
        )
        
        for score in existing_scores:
            scores[score.criterion.id] = score.value
    
    context = {
        'competition': competition,
        'category': category,
        'judge': judge,
        'assignment': assignment,
        'performances': performances,
        'current_performance': current_performance,
        'criteria': criteria,
        'scores': scores,
    }
    
    return render(request, 'competitions/management/judge_scoring_interface.html', context)


@login_required
@competition_management_permission_required
@require_POST
def save_judge_scores(request, competition_id, category_id, judge_id, performance_id):
    """
    Sauvegarde les scores d'un juge pour une performance.
    """
    # Récupérer la compétition, la catégorie, le juge et la performance
    competition = get_object_or_404(Competition, pk=competition_id)
    category = get_object_or_404(CompetitionCategory, pk=category_id, competition=competition)
    
    from django.contrib.auth import get_user_model
    User = get_user_model()
    
    judge = get_object_or_404(User, pk=judge_id)
    performance = get_object_or_404(TechnicalPerformance, pk=performance_id, category=category)
    
    # Vérifier que le juge est bien assigné Ã  cette catégorie
    assignment = get_object_or_404(
        JudgeAssignment,
        category=category,
        user=judge,
        assignment_type__in=['technical_judge', 'chief_judge']
    )
    
    # Récupérer les critères et leurs scores
    criteria = ScoringCriterion.objects.filter(category=category)
    
    with transaction.atomic():
        for criterion in criteria:
            score_key = f'score_{criterion.id}'
            if score_key in request.POST:
                try:
                    score_value = float(request.POST[score_key])
                    
                    # Vérifier les bornes du score
                    if score_value < criterion.min_score or score_value > criterion.max_score:
                        messages.error(request, _("Le score pour {} doit Ãªtre entre {} et {}.").format(
                            criterion.name, criterion.min_score, criterion.max_score))
                        continue
                    
                    # Vérifier si un score existe déjÃ 
                    try:
                        score = TechnicalScore.objects.get(
                            performance=performance,
                            judge=judge,
                            criterion=criterion
                        )
                        score.value = score_value
                        score.save()
                    except TechnicalScore.DoesNotExist:
                        # Créer un nouveau score
                        TechnicalScore.objects.create(
                            performance=performance,
                            judge=judge,
                            criterion=criterion,
                            value=score_value
                        )
                except ValueError:
                    messages.error(request, _("Valeur de score invalide pour {}.").format(criterion.name))
    
    messages.success(request, _("Les scores ont été enregistrés avec succès."))
    
    # Redirection vers l'interface de notation
    return redirect('competitions:management:judge_scoring_interface', 
                   competition_id=competition_id, 
                   category_id=category_id, 
                   judge_id=judge_id)


@login_required
@competition_management_permission_required
def performance_scorecard(request, competition_id, performance_id):
    """
    Affiche une fiche de scores détaillée pour une performance.
    """
    # Récupérer la compétition et la performance
    competition = get_object_or_404(Competition, pk=competition_id)
    performance = get_object_or_404(
        TechnicalPerformance, 
        pk=performance_id, 
        competition=competition
    )
    
    # Récupérer la catégorie
    category = performance.category
    
    # Récupérer les critères de notation
    criteria = ScoringCriterion.objects.filter(
        category=category
    ).order_by('order')
    
    # Récupérer les juges et leurs scores
    scores = TechnicalScore.objects.filter(
        performance=performance
    ).select_related('judge', 'criterion')
    
    # Organiser les scores dans une matrice juge/critère
    from django.contrib.auth import get_user_model
    User = get_user_model()
    
    judge_assignments = JudgeAssignment.objects.filter(
        category=category,
        assignment_type__in=['technical_judge', 'chief_judge']
    ).select_related('user')
    
    judges = {}
    for assignment in judge_assignments:
        judge = assignment.user
        judges[judge.id] = {
            'name': f"{judge.first_name} {judge.last_name}",
            'role': assignment.get_assignment_type_display(),
            'scores': {}
        }
    
    # Ajouter les scores Ã  la matrice
    for score in scores:
        judge_id = score.judge.id
        criterion_id = score.criterion.id
        
        if judge_id in judges:
            judges[judge_id]['scores'][criterion_id] = score.value
    
    # Calculer les statistiques de score
    stats = {}
    for criterion in criteria:
        criterion_scores = [j['scores'].get(criterion.id) for j in judges.values() 
                           if criterion.id in j['scores']]
        if criterion_scores:
            stats[criterion.id] = {
                'min': min(criterion_scores),
                'max': max(criterion_scores),
                'avg': sum(criterion_scores) / len(criterion_scores),
                'count': len(criterion_scores)
            }
    
    context = {
        'competition': competition,
        'performance': performance,
        'category': category,
        'criteria': criteria,
        'judges': judges,
        'stats': stats,
    }
    
    return render(request, 'competitions/management/performance_scorecard.html', context)


@login_required
@competition_management_permission_required
def scoring_statistics(request, competition_id, category_id):
    """
    Affiche des statistiques détaillées sur la notation d'une catégorie.
    """
    # Récupérer la compétition et la catégorie
    competition = get_object_or_404(Competition, pk=competition_id)
    category = get_object_or_404(CompetitionCategory, pk=category_id, competition=competition)
    
    # Récupérer les performances
    performances = TechnicalPerformance.objects.filter(
        category=category
    ).select_related('practitioner')
    
    # Récupérer les juges
    judge_assignments = JudgeAssignment.objects.filter(
        category=category,
        assignment_type__in=['technical_judge', 'chief_judge']
    ).select_related('user')
    
    # Récupérer les critères
    criteria = ScoringCriterion.objects.filter(
        category=category
    ).order_by('order')
    
    # Récupérer tous les scores
    scores = TechnicalScore.objects.filter(
        performance__in=performances
    ).select_related('judge', 'criterion', 'performance')
    
    # Analyser les tendances des juges
    judge_stats = {}
    for assignment in judge_assignments:
        judge = assignment.user
        judge_id = judge.id
        
        judge_scores = [s for s in scores if s.judge.id == judge_id]
        if not judge_scores:
            continue
        
        avg_score = sum(s.value for s in judge_scores) / len(judge_scores) if judge_scores else 0
        score_count = len(judge_scores)
        
        # Calculer l'écart-type
        if score_count > 1:
            import math
            variance = sum((s.value - avg_score) ** 2 for s in judge_scores) / score_count
            std_dev = math.sqrt(variance)
        else:
            std_dev = 0
        
        judge_stats[judge_id] = {
            'name': f"{judge.first_name} {judge.last_name}",
            'role': assignment.get_assignment_type_display(),
            'avg_score': avg_score,
            'score_count': score_count,
            'std_dev': std_dev,
        }
    
    # Analyser les tendances par critère
    criteria_stats = {}
    for criterion in criteria:
        criterion_id = criterion.id
        
        criterion_scores = [s for s in scores if s.criterion.id == criterion_id]
        if not criterion_scores:
            continue
        
        avg_score = sum(s.value for s in criterion_scores) / len(criterion_scores) if criterion_scores else 0
        score_count = len(criterion_scores)
        
        # Calculer l'écart-type
        if score_count > 1:
            import math
            variance = sum((s.value - avg_score) ** 2 for s in criterion_scores) / score_count
            std_dev = math.sqrt(variance)
        else:
            std_dev = 0
        
        criteria_stats[criterion_id] = {
            'name': criterion.name,
            'weight': criterion.weight,
            'avg_score': avg_score,
            'score_count': score_count,
            'std_dev': std_dev,
        }
    
    # Compteurs par statut
    completed_count = performances.filter(status='completed').count()
    in_progress_count = performances.filter(status='in_progress').count()
    pending_count = performances.filter(status__in=['scheduled', 'pending']).count()
    total_count = performances.count()

    # Dernières performances (10 max)
    recent_performances = performances.order_by('-end_time', '-performance_order')[:10]

    context = {
        'competition': competition,
        'category': category,
        'judge_stats': judge_stats,
        'criteria_stats': criteria_stats,
        'performances_count': total_count,
        'completed_count': completed_count,
        'in_progress_count': in_progress_count,
        'pending_count': pending_count,
        'total_scores': scores.count(),
        'recent_performances': recent_performances,
    }

    return render(request, 'competitions/management/scoring_statistics.html', context)


def _calculate_category_results_internal(competition, category):
    """
    Fonction interne pour calculer les résultats d'une catégorie.
    Utilisée par generate_all_results pour éviter les redirects multiples.
    """
    # Récupérer les performances avec des scores
    performances = TechnicalPerformance.objects.filter(
        category=category
    ).filter(
        scores__isnull=False
    ).distinct().select_related('practitioner')

    # Récupérer la configuration de notation
    try:
        config = ScoringConfiguration.objects.get(category=category)
        allow_ties = config.allow_ties
    except ScoringConfiguration.DoesNotExist:
        allow_ties = True

    # Supprimer les classements existants
    CompetitionRanking.objects.filter(
        competition=competition,
        category=category
    ).delete()

    # Corriger les flags is_active_for_ranking pour les barrages existants
    # Si un participant a des scores Tour 2+, desactiver ses scores Tour 1
    for performance in performances:
        max_round = TechnicalScore.objects.filter(
            performance=performance
        ).aggregate(max_round=Max('round_number'))['max_round'] or 1
        if max_round >= 2:
            # Desactiver les scores des tours precedents
            TechnicalScore.objects.filter(
                performance=performance,
                round_number__lt=max_round,
            ).update(is_active_for_ranking=False)
            # Activer les scores du dernier tour
            TechnicalScore.objects.filter(
                performance=performance,
                round_number=max_round,
            ).update(is_active_for_ranking=True)

    # Recalculer les scores finaux pour chaque performance
    for performance in performances:
        final_score = performance.calculate_final_score()
        first_places = 0
        CompetitionRanking.objects.create(
            competition=competition,
            category=category,
            practitioner=performance.practitioner,
            rank=0,
            final_score=final_score,
            first_places=first_places
        )

    # Attribuer les rangs
    rankings = CompetitionRanking.objects.filter(
        competition=competition,
        category=category
    ).order_by('-final_score', '-first_places')

    current_rank = 1
    previous_score = None
    previous_first_places = None
    for i, ranking in enumerate(rankings):
        if (i > 0 and ranking.final_score == previous_score
                and ranking.first_places == previous_first_places):
            if not allow_ties and current_rank == 3:
                ranking.rank = 4
            else:
                ranking.rank = current_rank
            ranking.is_tie = True
        else:
            ranking.rank = current_rank
            ranking.is_tie = False
            current_rank += 1

        ranking.save()
        previous_score = ranking.final_score
        previous_first_places = ranking.first_places


@login_required
@competition_management_permission_required
def generate_all_results(request, competition_id):
    """
    Calcule les résultats pour toutes les catégories de la compétition.
    """
    # Récupérer la compétition
    competition = get_object_or_404(Competition, pk=competition_id)
    
    # Récupérer toutes les catégories
    categories = CompetitionCategory.objects.filter(competition=competition)
    
    # Compteurs pour les statistiques
    total_categories = categories.count()
    processed_categories = 0
    
    for category in categories:
        try:
            # Récupérer les performances terminées
            performances = TechnicalPerformance.objects.filter(
                category=category,
                status='completed'
            ).exists()
            
            if performances:
                # Calculer les résultats pour cette catégorie directement
                _calculate_category_results_internal(competition, category)
                processed_categories += 1
        except Exception as e:
            # Enregistrer l'erreur mais continuer avec les autres catégories
            messages.warning(request, _("Erreur pour la catégorie {}: {}").format(category.name, str(e)))
    
    messages.success(request, _("Résultats calculés pour {}/{} catégories.").format(
        processed_categories, total_categories))
    
    return redirect('competitions:management:scoring_dashboard', competition_id=competition_id)


@login_required
@competition_management_permission_required
def publish_results(request, competition_id, category_id):
    """
    Publie les résultats d'une catégorie.
    """
    # Récupérer la compétition et la catégorie
    competition = get_object_or_404(Competition, pk=competition_id)
    category = get_object_or_404(CompetitionCategory, pk=category_id, competition=competition)
    
    if request.method == 'POST':
        # Marquer la catégorie comme terminée
        category.status = 'completed'
        category.save()
        
        # Récupérer la configuration de notation
        from apps.competitions.models.technical_scoring import ScoringConfiguration
        try:
            config = ScoringConfiguration.objects.get(category=category)
            config.real_time_results = True
            config.save()
        except ScoringConfiguration.DoesNotExist:
            # Créer une configuration de base
            ScoringConfiguration.objects.create(
                category=category,
                real_time_results=True
            )
        
        messages.success(request, _("Les résultats ont été publiés."))
    
    return redirect('competitions:management:category_results', 
                   competition_id=competition_id, 
                   category_id=category_id)


@login_required
@competition_management_permission_required
def podium_view(request, competition_id, category_id):
    """
    Affiche une vue podium pour une catégorie (pour l'affichage public).
    """
    from apps.competitions.models.scoring_results import PodiumEntry

    # Récupérer la compétition et la catégorie
    competition = get_object_or_404(Competition, pk=competition_id)
    category = get_object_or_404(CompetitionCategory, pk=category_id, competition=competition)

    # Détecter mode équipe
    is_team_mode = False
    try:
        if category.competition_type and category.competition_type.team_based:
            is_team_mode = True
    except Exception:
        pass

    # Récupérer les 3 premiers avec organisation
    podium = CompetitionRanking.objects.filter(
        competition=competition,
        category=category,
        rank__lte=3
    ).select_related('practitioner', 'practitioner__organization').order_by('rank')

    # Enrichir chaque entrée avec club, drapeau, logo + données d'équipe
    for rank in podium:
        org = rank.practitioner.organization
        rank.club = org
        rank.country_code = PodiumEntry._get_country_code(org.country) if org else ''
        rank.flag_url = ('https://flagcdn.com/w40/%s.png' % rank.country_code) if rank.country_code else ''
        rank.logo_url = ''
        if org and org.logo:
            try:
                rank.logo_url = org.logo.url
            except Exception:
                pass

        # Enrichir avec données d'équipe si mode équipe
        rank.is_team = False
        rank.team_name = ''
        rank.team_members = []
        if is_team_mode:
            from apps.competitions.models.combat import MembreEquipe
            membership = MembreEquipe.objects.filter(
                pratiquant=rank.practitioner,
                equipe__category=category,
                equipe__is_active=True,
            ).select_related('equipe', 'equipe__club').first()
            if membership:
                team = membership.equipe
                rank.is_team = True
                rank.team_name = team.nom
                rank.team_members = list(
                    team.memberships.filter(est_remplacant=False)
                    .select_related('pratiquant')
                    .order_by('ordre')
                )

    context = {
        'competition': competition,
        'category': category,
        'podium': podium,
        'is_fullscreen': request.GET.get('fullscreen') == '1',
        'is_team_mode': is_team_mode,
    }

    return render(request, 'competitions/management/podium_view.html', context)

@login_required
@competition_management_permission_required
def category_results(request, competition_id, category_id):
    """
    Affiche les résultats finaux d'une catégorie.
    """
    # Récupérer la compétition et la catégorie
    competition = get_object_or_404(Competition, pk=competition_id)
    category = get_object_or_404(CompetitionCategory, pk=category_id, competition=competition)
    
    # Récupérer les performances terminées
    performances = TechnicalPerformance.objects.filter(
        category=category,
        status='completed'
    ).select_related('practitioner')
    
    # Récupérer les résultats
    rankings = CompetitionRanking.objects.filter(
        competition=competition,
        category=category
    ).select_related('practitioner', 'practitioner__organization').order_by('rank')
    
    # Ajouter les inscriptions pour chaque ranking
    for ranking in rankings:
        try:
            ranking.registration = CompetitionRegistration.objects.get(
                competition=competition,
                practitioner=ranking.practitioner
            )
        except CompetitionRegistration.DoesNotExist:
            ranking.registration = None
    
    # Vérifier si les résultats ont été calculés
    results_calculated = rankings.exists()

    # Récupérer la configuration de notation
    scoring_config = None
    are_results_public = False
    tour2_mode = 'all'
    try:
        scoring_config = ScoringConfiguration.objects.get(category=category)
        are_results_public = scoring_config.real_time_results
        advanced = scoring_config.advanced_config or {}
        tour2_mode = advanced.get('tour2_mode', 'all')
    except ScoringConfiguration.DoesNotExist:
        pass

    # Détecter les ex-aequo sur le podium pour proposer le Tour 2
    has_podium_ties = False
    tour2_eligible_ids = set()
    if results_calculated:
        rankings_by_rank = {}
        for r in rankings:
            rankings_by_rank.setdefault(r.rank, []).append(r)
        for rank, group in rankings_by_rank.items():
            if rank <= 3 and len(group) > 1:
                has_podium_ties = True
                for r in group:
                    tour2_eligible_ids.add(r.practitioner_id)

    # Vérifier si un Tour 2 est déjà en cours
    tour2_launched = False
    tour2_closed = False
    if scoring_config and scoring_config.advanced_config:
        tour2_launched = scoring_config.advanced_config.get('tour2_launched', False)
        tour2_closed = scoring_config.advanced_config.get('tour2_closed', False)

    has_tour2_scores = TechnicalScore.objects.filter(
        performance__category=category,
        round_number__gte=2,
    ).exists()

    # Tour 2 en cours = lancé ET pas clôturé
    tour2_in_progress = (tour2_launched or has_tour2_scores) and not tour2_closed
    
    # Calculer le nombre de participants inscrits dans cette catégorie
    registrations_count = CompetitionRegistration.objects.filter(
        competition=competition,
        categories=category
    ).count()
    
    # Ajouter les informations à la catégorie pour le template
    category.registrations_count = registrations_count

    # Détecter mode équipe
    is_team_mode = False
    try:
        if category.competition_type and category.competition_type.team_based:
            is_team_mode = True
    except Exception:
        pass

    # Enrichir les rankings avec les données d'équipe si mode équipe
    rankings_list = list(rankings)
    if is_team_mode:
        from apps.competitions.models.combat import MembreEquipe
        for ranking in rankings_list:
            membership = MembreEquipe.objects.filter(
                pratiquant=ranking.practitioner,
                equipe__category=category,
                equipe__is_active=True,
            ).select_related('equipe', 'equipe__club').first()
            if membership:
                team = membership.equipe
                ranking.is_team = True
                ranking.team_name = team.nom
                ranking.team_club = team.club.name if team.club else ''
                ranking.team_members = list(
                    team.memberships.filter(est_remplacant=False)
                    .select_related('pratiquant')
                    .order_by('ordre')
                )
            else:
                ranking.is_team = False
                ranking.team_members = []

    # En mode équipe, dédupliquer : garder 1 seule ligne par équipe
    if is_team_mode:
        seen_teams = set()
        deduplicated = []
        for ranking in rankings_list:
            team_name = getattr(ranking, 'team_name', None)
            if team_name:
                if team_name not in seen_teams:
                    seen_teams.add(team_name)
                    deduplicated.append(ranking)
            else:
                deduplicated.append(ranking)
        rankings_list = deduplicated

    # Marquer les rankings éligibles au Tour 2 et ajouter les scores par tour
    for ranking in rankings_list:
        ranking.tour2_eligible = ranking.practitioner_id in tour2_eligible_ids

        # Récupérer les scores Tour 1 et Tour 2 séparément
        perf = TechnicalPerformance.objects.filter(
            category=category,
            practitioner=ranking.practitioner,
        ).first()
        if perf:
            t1_scores = TechnicalScore.objects.filter(
                performance=perf, round_number=1
            ).values('judge').annotate(avg=Avg('value'))
            if t1_scores:
                ranking.score_tour1 = round(sum(s['avg'] for s in t1_scores) / len(t1_scores), 2)
            else:
                ranking.score_tour1 = None

            t2_scores = TechnicalScore.objects.filter(
                performance=perf, round_number=2
            ).values('judge').annotate(avg=Avg('value'))
            if t2_scores:
                ranking.score_tour2 = round(sum(s['avg'] for s in t2_scores) / len(t2_scores), 2)
            else:
                ranking.score_tour2 = None
        else:
            ranking.score_tour1 = None
            ranking.score_tour2 = None

    context = {
        'competition': competition,
        'category': category,
        'performances': performances,
        'rankings': rankings_list,
        'results_calculated': results_calculated,
        'has_results': results_calculated,
        'are_results_public': are_results_public,
        'scoring_config': scoring_config,
        'is_team_mode': is_team_mode,
        'has_podium_ties': has_podium_ties,
        'tour2_in_progress': tour2_in_progress,
        'has_tour2_scores': has_tour2_scores,
        'tour2_mode': tour2_mode,
    }

    return render(request, 'competitions/management/category_results.html', context)


@login_required
@require_POST
def launch_tour2(request, competition_id, category_id):
    """
    Lance le Tour 2 pour une catégorie.
    Identifie les ex-aequo du podium et redirige vers la feuille de notation.
    """
    competition = get_object_or_404(Competition, pk=competition_id)
    category = get_object_or_404(CompetitionCategory, pk=category_id, competition=competition)

    # Trouver les ex-aequo sur le podium
    rankings = CompetitionRanking.objects.filter(
        competition=competition,
        category=category,
    ).order_by('rank')

    tour2_practitioner_ids = set()
    rankings_by_rank = {}
    for r in rankings:
        rankings_by_rank.setdefault(r.rank, []).append(r)
    for rank, group in rankings_by_rank.items():
        if rank <= 3 and len(group) > 1:
            for r in group:
                tour2_practitioner_ids.add(r.practitioner_id)

    if not tour2_practitioner_ids:
        messages.warning(request, _("Aucun ex-aequo sur le podium. Le Tour 2 n'est pas necessaire."))
        return redirect('competitions:management:category_results',
                       competition_id=competition_id,
                       category_id=category_id)

    # Sauvegarder les IDs des éligibles dans la config avancée
    try:
        config = ScoringConfiguration.objects.get(category=category)
        advanced = config.advanced_config or {}
        advanced['tour2_eligible_ids'] = list(tour2_practitioner_ids)
        advanced['tour2_launched'] = True
        config.advanced_config = advanced
        config.save()
    except ScoringConfiguration.DoesNotExist:
        pass

    practitioner_names = []
    for r in rankings:
        if r.practitioner_id in tour2_practitioner_ids:
            practitioner_names.append(r.practitioner.full_name)

    messages.success(
        request,
        _("Tour 2 lance pour %(count)d participants : %(names)s") % {
            'count': len(tour2_practitioner_ids),
            'names': ', '.join(practitioner_names),
        }
    )

    return redirect('competitions:management:scoring_dashboard',
                   competition_id=competition_id)


@login_required
def close_tour2(request, competition_id, category_id):
    """
    Cloture le Tour 2 : recalcule les resultats et marque le tour comme termine.
    """
    competition = get_object_or_404(Competition, pk=competition_id)
    category = get_object_or_404(CompetitionCategory, pk=category_id, competition=competition)

    # Marquer le Tour 2 comme termine dans la config
    try:
        config = ScoringConfiguration.objects.get(category=category)
        advanced = config.advanced_config or {}
        advanced['tour2_launched'] = False
        advanced['tour2_closed'] = True
        config.advanced_config = advanced
        config.save()
    except ScoringConfiguration.DoesNotExist:
        pass

    # Forcer la cloture de la session si elle existe
    try:
        from apps.competitions.models.session_workflow import CategorySession
        session = CategorySession.objects.filter(
            category=category
        ).exclude(status='closed').first()
        if session:
            session.status = 'closed'
            session.save()
    except Exception:
        pass

    # Mettre le statut de la categorie a "closed" (termine)
    category.registration_status = 'closed'
    category.save(update_fields=['registration_status'])

    # Recalculer les resultats (appel direct de la logique)
    return redirect('competitions:management:calculate_results',
                   competition_id=competition_id,
                   category_id=category_id)


@login_required
def category_scores_api(request, competition_id, category_id):
    """API: scores en temps réel pour une catégorie."""
    competition = get_object_or_404(Competition, pk=competition_id)
    category = get_object_or_404(CompetitionCategory, pk=category_id, competition=competition)

    performances = TechnicalPerformance.objects.filter(
        category=category
    ).select_related('practitioner').order_by('performance_order')

    data = []
    for perf in performances:
        score = None
        try:
            score = perf.calculate_final_score()
        except Exception:
            pass
        if score is None:
            # Calculer manuellement
            scores = TechnicalScore.objects.filter(performance=perf)
            if scores.exists():
                avg = scores.aggregate(avg=Avg('value'))['avg']
                score = float(avg) if avg else None

        round_num = 1
        try:
            round_num = perf.get_current_round() or 1
        except Exception:
            pass

        data.append({
            'practitioner': perf.practitioner.full_name,
            'club': str(perf.practitioner.organization) if perf.practitioner.organization else '',
            'score': round(score, 2) if score is not None else None,
            'round': round_num,
            'status': perf.status,
            'status_display': perf.get_status_display(),
            'order': perf.performance_order,
        })

    # Trier par score décroissant
    data.sort(key=lambda x: (x['score'] or 0), reverse=True)

    return JsonResponse({'performances': data, 'total': len(data), 'category': str(category)})

