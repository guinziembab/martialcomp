from django.core.exceptions import PermissionDenied
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.utils.translation import gettext_lazy as _
from django.core.paginator import Paginator
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_GET
from django.db import models, transaction
from django.db.models import Q
import csv
from datetime import datetime
from ..models import Federation
from ..models import (
    Competition, 
    CompetitionType, 
    CompetitionCategory,
    Discipline, 
    CompetitionRegistration, 
    ParticipantCategoryRegistration, 
    JudgeQualification,
    Practitioner, 
    Club
)
from ..forms import (
    CompetitionForm, 
    CompetitionCategoryForm, 
    CompetitionRegistrationForm
)
from apps.competitions.utils.decorators import federation_required
from django.utils import timezone
from datetime import datetime, timedelta
from django.db.models import Q
from django.utils.translation import gettext as _
from apps.competitions.models import Competition, CompetitionRegistration, Club
from apps.core.isolation import OrganizationIsolationMixin, get_organization_queryset


@require_GET
def get_competition_types(request):
    """
    Vue API pour récupérer les types de compétition.
    Si discipline_id est fourni en paramètre, filtre les types par discipline.
    Renvoie un JsonResponse avec les ID et noms des types de compétition.
    
    Cette API est utilisée par le formulaire de création/modification des compétitions
    pour charger dynamiquement les types de compétition selon la discipline sélectionnée.
    """
    discipline_id = request.GET.get('discipline_id')
    
    if discipline_id:
        try:
            discipline = Discipline.objects.get(id=discipline_id)
            competition_types = CompetitionType.objects.filter(discipline=discipline)
            
            # Appliquer des filtres additionnels selon le contexte
            team_based = request.GET.get('team_based')
            if team_based is not None:
                team_based_bool = team_based.lower() == 'true'
                competition_types = competition_types.filter(team_based=team_based_bool)
            # Ne pas filtrer par défaut - inclure tous les types de compétition
            
        except Discipline.DoesNotExist:
            return JsonResponse({
                'error': 'Discipline not found',
                'competition_types': []
            }, status=404)
    else:
        # Si aucune discipline spécifiée, renvoyer tous les types sans équipes par défaut
        competition_types = CompetitionType.objects.filter(team_based=False)
    
    # Tri par ordre et nom pour une meilleure présentation
    competition_types = competition_types.order_by('order', 'name')
    
    types_data = [
        {
            'id': comp_type.id, 
            'name': comp_type.name,
            'team_based': comp_type.team_based,
            'description': comp_type.description if hasattr(comp_type, 'description') else None
        }
        for comp_type in competition_types
    ]
    
    return JsonResponse({
        'competition_types': types_data,
        'count': len(types_data)
    })

@login_required
def get_competition_types_by_discipline(request, discipline_id):
    """
    API pour récupérer les types de compétition associés à une discipline.
    Utilisé par le JavaScript dans le formulaire de création/modification.
    """
    try:
        discipline = Discipline.objects.get(id=discipline_id)
        competition_types = CompetitionType.objects.filter(discipline=discipline)
        
        types_data = []
        for comp_type in competition_types:
            types_data.append({
                'id': comp_type.id,
                'name': comp_type.name
            })
        
        return JsonResponse({
            'competition_types': types_data
        })
    except Discipline.DoesNotExist:
        return JsonResponse({
            'error': 'Discipline non trouvée',
            'competition_types': []
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'error': str(e),
            'competition_types': []
        }, status=500)


@login_required
def competition_categories(request, competition_id):
    """
    Vue pour la gestion des catégories d'une compétition.
    Utilisée après la création d'une compétition si des données de catégories ont été fournies.
    """
    # Récupérer la compétition
    competition = get_object_or_404(Competition, pk=competition_id)
    
    # Vérifier les permissions selon le rÃ´le et la propriété de la compétition
    role = getattr(request.user, 'role', None)
    
    has_permission = False
    
    # Administrateur global
    if role == 'admin' or request.user.is_staff:
        has_permission = True
    
    # Responsable de club qui possède cette compétition
    elif role == 'club_manager' and hasattr(competition, 'club') and competition.club:
        # Vérifier si l'utilisateur est propriétaire du club
        if (hasattr(request.user, 'club') and request.user.club == competition.club) or \
           (Club.objects.filter(owner=request.user, id=competition.club.id).exists()):
            has_permission = True
    
    # Admin de fédération qui possède cette compétition
    elif role == 'federation_admin' and hasattr(competition, 'federation') and competition.federation:
        # Vérifier si l'utilisateur est propriétaire de la fédération
        if (hasattr(request.user, 'federation') and request.user.federation == competition.federation) or \
           (Federation.objects.filter(owner=request.user, id=competition.federation.id).exists()):
            has_permission = True
    
    if not has_permission:
        messages.error(request, _("Vous n'avez pas les droits nécessaires pour gérer les catégories de cette compétition."))
        return redirect('competitions:competitions:detail', pk=competition.id)
    
    # Récupérer les catégories existantes
    categories = CompetitionCategory.objects.filter(competition=competition)
    
    # Récupérer les types de compétition disponibles
    competition_types = competition.competition_types.all()
    
    # Traitement du formulaire si une nouvelle catégorie est soumise
    if request.method == 'POST':
        form = CompetitionCategoryForm(request.POST)
        
        if form.is_valid():
            try:
                category = form.save(commit=False)
                category.competition = competition
                category.save()
                
                messages.success(request, _("La catégorie a été ajoutée avec succès."))
                return redirect('competitions:competitions:detail', pk=competition.id)
            except Exception as e:
                messages.error(request, _("Une erreur est survenue lors de l'ajout de la catégorie: {}").format(str(e)))
        else:
            # Afficher les erreurs spécifiques du formulaire
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        form = CompetitionCategoryForm()
    
    # Traiter les données de catégories stockées en session, si présentes
    weight_categories_json = request.session.pop('weight_categories', None)
    age_categories_json = request.session.pop('age_categories', None)
    grade_categories_json = request.session.pop('grade_categories', None)
    
    # Si des données existent et qu'aucune catégorie n'a encore été créée,
    # créer automatiquement les catégories
    if (weight_categories_json or age_categories_json or grade_categories_json) and not categories.exists():
        try:
            import json
            
            # Créer les catégories de poids
            if weight_categories_json:
                weight_data = json.loads(weight_categories_json)
                if 'categories' in weight_data:
                    for category_data in weight_data['categories']:
                        for comp_type in competition_types:
                            CompetitionCategory.objects.create(
                                competition=competition,
                                competition_type=comp_type,
                                name=category_data['name'],
                                min_weight=category_data.get('min_kg'),
                                max_weight=category_data.get('max_kg')
                            )
            
            # Créer les catégories d'Ã¢ge
            if age_categories_json:
                age_data = json.loads(age_categories_json)
                if 'categories' in age_data:
                    for category_data in age_data['categories']:
                        for comp_type in competition_types:
                            CompetitionCategory.objects.create(
                                competition=competition,
                                competition_type=comp_type,
                                name=category_data['name'],
                                min_age=category_data.get('min_age'),
                                max_age=category_data.get('max_age')
                            )
            
            # Créer les catégories de grade
            if grade_categories_json:
                grade_data = json.loads(grade_categories_json)
                if 'categories' in grade_data:
                    for category_data in grade_data['categories']:
                        for comp_type in competition_types:
                            CompetitionCategory.objects.create(
                                competition=competition,
                                competition_type=comp_type,
                                name=category_data['name'],
                                min_grade=category_data.get('grades', [])[0] if category_data.get('grades') else None,
                                max_grade=category_data.get('grades', [])[-1] if category_data.get('grades') else None
                            )
            
            messages.success(request, _("Les catégories ont été créées automatiquement."))
            
            # RafraÃ®chir les catégories
            categories = CompetitionCategory.objects.filter(competition=competition)
        except Exception as e:
            messages.error(request, _("Une erreur est survenue lors de la création automatique des catégories: {}").format(str(e)))
    
    context = {
        'competition': competition,
        'categories': categories,
        'competition_types': competition_types,
        'form': form,
    }
    
    return render(request, 'competitions/competition/categories.html', context)


@login_required
def competition_list(request):
    """Liste des compétitions filtrée selon les droits de l'utilisateur."""
    # Contexte par défaut
    context = {
        'title': _("Liste des compétitions"),
    }
    
    # Si c'est un admin système (is_staff)
    if hasattr(request.user, 'is_staff') and request.user.is_staff:
        competitions = get_organization_queryset(Competition, request.user).order_by('-start_date')
        context['is_admin'] = True
    # Si c'est un admin fédération
    elif hasattr(request, 'federation') and request.federation:
        federation = request.federation
        discipline = federation.discipline
        competitions = Competition.objects.filter(discipline=discipline).order_by('-start_date')
        context['federation'] = federation
        context['discipline'] = discipline
    # Si c'est un club manager, montrer toutes les compétitions publiques
    elif hasattr(request.user, 'userprofile') and request.user.userprofile.role == 'club_manager':
        competitions = Competition.objects.filter(status='published').order_by('-start_date')
        context['is_club_manager'] = True
    else:
        # Cas par défaut: compétitions publiées uniquement
        competitions = Competition.objects.filter(
            status='published',
            start_date__gte=timezone.now().date()
        ).order_by('start_date')
    
    # Pagination
    paginator = Paginator(competitions, 10)  # 10 compétitions par page
    page = request.GET.get('page')
    competitions_page = paginator.get_page(page)
    
    context['competitions'] = competitions_page
    context['total_competitions'] = competitions.count()
    
    return render(request, 'competitions/competition/list.html', context)

@login_required
def competition_create(request, federation_id=None):
    """
    Vue unifiée pour la création d'une compétition.
    Accessible par les profils : fédération, club, manager.
    """
    # Déterminer les permissions et le contexte selon le rôle
    # Le rôle peut être sur l'utilisateur ou dans son profil
    role = getattr(request.user, 'role', None)
    if not role and hasattr(request.user, 'userprofile'):
        role = getattr(request.user.userprofile, 'role', None)
    if not role and hasattr(request.user, 'profile'):
        role = getattr(request.user.profile, 'role', None)

    # Récupérer le club, la fédération ou l'organisation externe associés à l'utilisateur
    club = None
    federation = None
    organization = None

    # Si un federation_id est fourni, vérifier les permissions pour cette fédération
    if federation_id:
        federation = get_object_or_404(Federation, id=federation_id)
        has_federation_permission = False

        # Super admin ou admin global
        if request.user.is_superuser or role == 'admin':
            has_federation_permission = True
        # Owner de la fédération (comparaison par ID)
        elif federation.owner_id == request.user.id:
            has_federation_permission = True
        # Administrateur de la fédération
        elif hasattr(federation, 'administrators'):
            try:
                if federation.administrators.filter(user=request.user).exists():
                    has_federation_permission = True
            except Exception:
                pass

        if not has_federation_permission:
            messages.error(request, _("Vous n'avez pas accès à cette fédération."))
            return redirect('competitions:dashboard:dashboard')

        # Permission accordée via fédération, récupérer l'organisation
        organization = getattr(federation, 'organization', None) or getattr(federation, 'as_organization', None)
    # Sinon, vérifier les permissions selon le rôle
    elif role not in ['federation_admin', 'club_manager', 'admin', 'external_organizer']:
        messages.error(request, _("Vous n'avez pas les droits nécessaires pour créer une compétition."))
        return redirect('competitions:competitions:list')
    elif role == 'club_manager':
        if hasattr(request.user, 'club') and request.user.club:
            club = request.user.club
        else:
            try:
                from ..models import Club
                club = Club.objects.filter(owner=request.user).first()
            except ImportError:
                club = None
        if not club:
            messages.error(request, _( "Vous devez Ãªtre associé Ã  un club pour créer une compétition."))
            return redirect('competitions:dashboard:dashboard')
        organization = getattr(club, 'organization', None) or getattr(club, 'as_organization', None)
    elif role == 'federation_admin':
        try:
            if hasattr(request.user, 'federation') and request.user.federation:
                federation = request.user.federation
            else:
                federation = Federation.objects.filter(owner=request.user).first()
            if not federation:
                messages.error(request, _( "Vous devez Ãªtre associé Ã  une fédération pour créer une compétition."))
                return redirect('competitions:dashboard:dashboard')
            organization = getattr(federation, 'organization', None) or getattr(federation, 'as_organization', None)
        except:
            messages.error(request, _( "La fonctionnalité de fédération n'est pas disponible."))
            return redirect('competitions:dashboard:dashboard')
    elif role == 'external_organizer':
        # Utiliser l'organisation liée au profil, ou en créer une si besoin
        profile = getattr(request.user, 'profile', None)
        if profile and profile.organization:
            organization = profile.organization
        else:
            # Créer une organisation "externe" pour cet utilisateur si elle n'existe pas
            from apps.organizations.models import Organization
            from apps.competitions.models import Club
            organization = Organization.objects.create(
                name=f"Organisation externe {request.user.username}",
                organization_type='other',
                created_by=request.user,
                is_active=True
            )
            # Créer un club fictif lié Ã  cette organisation
            club = Club.objects.create(
                name=f"Club externe {request.user.username}",
                organization=organization,
                owner=request.user,
                is_active=True
            )
            profile.organization = club
            profile.save()
    
    # Traitement du formulaire
    if request.method == 'POST':
        form = CompetitionForm(request.POST, request.FILES, user=request.user)

        if form.is_valid():
            try:
                with transaction.atomic():
                    competition = form.save(commit=False)
                    # Associer l'utilisateur comme organizer
                    competition.organizer = request.user
                    competition.created_by = request.user
                    # Associer Ã  l'organisation trouvée ou créée
                    if organization:
                        competition.organizing_organization = organization
                    if role == 'club_manager' and club:
                        competition.club = club
                    elif (role == 'federation_admin' or federation_id) and federation:
                        competition.federation = federation
                    
                    # Déterminer le statut selon la case "publier immédiatement"
                    is_published = form.cleaned_data.get('is_published', False)
                    competition.status = 'published' if is_published else 'draft'

                    # Sauvegarder les champs extra (non inclus dans Meta.fields)
                    competition.start_time = form.cleaned_data.get('start_time')
                    competition.end_time = form.cleaned_data.get('end_time')
                    if form.cleaned_data.get('max_participants') is not None:
                        competition.max_participants = form.cleaned_data.get('max_participants')
                    if form.cleaned_data.get('min_age') is not None:
                        competition.min_age = form.cleaned_data.get('min_age')
                    # Sauvegarder postal_code depuis le template
                    competition.postal_code = request.POST.get('postal_code', '').strip()

                    # Définir les dates par défaut si non fournies
                    if not competition.registration_deadline:
                        # Par défaut, la date limite est la veille du début
                        if competition.start_date:
                            competition.registration_deadline = competition.start_date - timezone.timedelta(days=1)

                    competition.save()

                    # Enregistrer les types de compétition
                    competition_types = form.cleaned_data.get('competition_types')
                    if competition_types:
                        competition.competition_types.set(competition_types)
                    
                    # Traiter les données JSON des catégories, si fournies
                    # Ces données serviront Ã  créer les catégories après la redirection
                    weight_categories = request.POST.get('weight_categories')
                    age_categories = request.POST.get('age_categories')
                    grade_categories = request.POST.get('grade_categories')
                    
                    # Stocker temporairement ces données dans la session pour les utiliser plus tard
                    if weight_categories:
                        request.session['weight_categories'] = weight_categories
                    if age_categories:
                        request.session['age_categories'] = age_categories
                    if grade_categories:
                        request.session['grade_categories'] = grade_categories
                    
                    messages.success(request, _("La compétition a été créée avec succès."))
                    
                    # CORRECTION ICI - Rediriger vers la page de détail de la compétition
                    # L'URL 'categories' n'existe pas, on redirige vers le détail de la compétition
                    if weight_categories or age_categories or grade_categories:
                        # Les données de catégories sont stockées en session et seront traitées sur la page de détail
                        return redirect('competitions:competitions:detail', pk=competition.id)
                    
                    # Si création depuis une fédération, rediriger vers le tableau de bord de la fédération
                    if federation_id:
                        return redirect('competitions:federations:detail', slug=federation.slug)
                    
                    # Sinon, rediriger vers la page de détail
                    return redirect('competitions:competitions:detail', pk=competition.id)
                    
            except Exception as e:
                messages.error(request, _("Une erreur est survenue lors de la création de la compétition: {}").format(str(e)))
        else:
            # Afficher les erreurs spécifiques du formulaire
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        # Initialiser le formulaire avec des valeurs par défaut
        initial_data = {
            'start_date': timezone.now().date(),
            'end_date': timezone.now().date() + timezone.timedelta(days=1),
            'start_time': '09:00',
            'end_time': '18:00',
            'min_age': 6,
            'max_participants': 150,
        }
        form = CompetitionForm(initial=initial_data, user=request.user)

    # Liste des disciplines pour le formulaire
    disciplines = Discipline.objects.filter(is_active=True)
    
    context = {
        'form': form,
        'disciplines': disciplines,
        'club': club,
        'federation': federation,
        'role': role,
        'user': request.user,
        'page_title': _("Créer une nouvelle compétition"),
    }
    
    return render(request, 'competitions/competition/create.html', context)


@login_required
def competition_detail(request, pk):
    """Affiche les détails d'une compétition avec gestion des droits d'accès."""
    competition = get_object_or_404(Competition, pk=pk)

    # Vérification des droits d'accès pour les compétitions non publiées
    if competition.status != 'published':
        has_access = False

        # Admin système (is_staff ou is_superuser) peut tout voir
        if request.user.is_staff or request.user.is_superuser:
            has_access = True

        # Le créateur de la compétition peut toujours la voir
        elif competition.created_by and competition.created_by == request.user:
            has_access = True

        # Le propriétaire de la fédération associée peut la voir
        elif hasattr(competition, 'federation') and competition.federation:
            if competition.federation.owner == request.user:
                has_access = True
            # Ou si l'utilisateur est admin de cette fédération
            elif hasattr(competition.federation, 'administrators'):
                try:
                    if competition.federation.administrators.filter(user=request.user).exists():
                        has_access = True
                except Exception:
                    pass

        # Le propriétaire du club associé peut la voir
        elif hasattr(competition, 'club') and competition.club:
            if competition.club.owner == request.user:
                has_access = True

        # Vérifier via le rôle de l'utilisateur
        if not has_access:
            role = getattr(request.user, 'role', None)
            if not role and hasattr(request.user, 'userprofile'):
                role = getattr(request.user.userprofile, 'role', None)
            if not role and hasattr(request.user, 'profile'):
                role = getattr(request.user.profile, 'role', None)

            # Admin fédération peut voir les compétitions de sa discipline ou de sa fédération
            if role == 'federation_admin':
                try:
                    user_federation = Federation.objects.filter(owner=request.user).first()
                    if user_federation:
                        # Même fédération
                        if competition.federation and competition.federation.id == user_federation.id:
                            has_access = True
                        # Même discipline (si définie)
                        elif competition.discipline and user_federation.discipline:
                            if competition.discipline == user_federation.discipline:
                                has_access = True
                except Exception:
                    pass

            # Club manager peut voir les compétitions de son club
            elif role == 'club_manager':
                try:
                    from ..models import Club
                    user_club = Club.objects.filter(owner=request.user).first()
                    if user_club and competition.club and competition.club.id == user_club.id:
                        has_access = True
                except Exception:
                    pass

        if not has_access:
            messages.error(request, _("Cette compétition n'est pas encore publiée."))
            return redirect('competitions:competitions:list')

    # Si l'accès est autorisé, afficher les détails

    # Build categories with participant counts
    categories_with_counts = []
    for category in competition.categories.all():
        participant_count = ParticipantCategoryRegistration.objects.filter(
            category=category, registration__competition=competition
        ).count()
        categories_with_counts.append({
            'category': category,
            'participant_count': participant_count,
        })

    # Get registrations
    registrations = CompetitionRegistration.objects.filter(
        competition=competition
    ).exclude(status='rejected').select_related('practitioner')

    # Separate participants and judges
    # Compter les juges (ceux qui ont un rôle de juge)
    judges = registrations.filter(
        Q(is_technical_judge=True) | Q(is_combat_referee=True)
    )
    total_judges = judges.count()

    # Les participants = tous les inscrits sauf les juges purs
    # (un inscrit qui n'est ni juge technique ni arbitre est un compétiteur)
    participants = registrations.exclude(
        Q(is_technical_judge=True) | Q(is_combat_referee=True),
        is_competitor=False,
    )
    total_participants = participants.count()

    # Check permissions
    from apps.organizations.models import OrganizationMember
    can_manage_competition = (
        request.user.is_staff or
        request.user.is_superuser or
        competition.created_by == request.user or
        (competition.organizing_organization and OrganizationMember.objects.filter(
            user=request.user, organization=competition.organizing_organization
        ).exists())
    )

    # Traitement POST : promouvoir des inscrits en juge/arbitre
    if request.method == 'POST' and request.POST.get('action') == 'promote_judge' and can_manage_competition:
        reg_ids = request.POST.getlist('registration_ids')
        judge_role = request.POST.get('judge_role')
        if reg_ids:
            updated = 0
            for reg_id in reg_ids:
                try:
                    reg = CompetitionRegistration.objects.get(id=reg_id, competition=competition)
                    if judge_role == 'technical_judge':
                        reg.is_technical_judge = True
                    elif judge_role == 'combat_referee':
                        reg.is_combat_referee = True
                    reg.save()
                    updated += 1
                except CompetitionRegistration.DoesNotExist:
                    pass
            messages.success(request, _("{} participant(s) promu(s) juge/arbitre.").format(updated))
        else:
            messages.warning(request, _("Aucun participant sélectionné."))
        return redirect('competitions:competitions:detail', pk=competition.id)

    context = {
        'competition': competition,
        'registration_open': (
            competition.registration_deadline and
            competition.registration_deadline >= timezone.now().date()
        ),
        'categories_with_counts': categories_with_counts,
        'registrations': participants,
        'participants': registrations.select_related('practitioner').order_by('practitioner__last_name'),
        'total_participants': total_participants,
        'judges': judges,
        'total_judges': total_judges,
        'can_manage_competition': can_manage_competition,
    }

    return render(request, 'competitions/competition/detail_enhanced.html', context)


@login_required
def competition_update(request, pk):
    """
    Vue unifiée pour la modification d'une compétition.
    Accessible par les profils : fédération, club, manager.
    """
    # Récupérer la compétition
    competition = get_object_or_404(Competition, pk=pk)
    
    # Vérifier les permissions - logique simplifiée
    has_permission = False
    
    # Administrateur global
    if request.user.is_staff:
        has_permission = True
    
    # Créateur de la compétition (si le champ existe)
    elif hasattr(competition, 'created_by') and competition.created_by == request.user:
        has_permission = True
    
    # Pour l'instant, autoriser tous les utilisateurs connectés à modifier
    # TODO: Ajouter le champ created_by au modèle Competition pour une vraie vérification
    else:
        has_permission = True
    
    if not has_permission:
        messages.error(request, _("Vous n'avez pas les droits nécessaires pour modifier cette compétition."))
        return redirect('competitions:competitions:detail', pk=competition.id)
    
    # Traitement du formulaire
    if request.method == 'POST':
        form = CompetitionForm(request.POST, request.FILES, instance=competition, user=request.user)

        # DEBUG: Log POST data for time/postal fields
        import logging as _logging
        _dbg = _logging.getLogger('competitions.debug')
        _dbg.warning(f"[UPDATE] POST start_time={request.POST.get('start_time')!r}, end_time={request.POST.get('end_time')!r}, postal_code={request.POST.get('postal_code')!r}")
        _dbg.warning(f"[UPDATE] form.is_valid()={form.is_valid()}")
        if not form.is_valid():
            _dbg.warning(f"[UPDATE] form.errors={form.errors}")

        if form.is_valid():
            try:
                with transaction.atomic():
                    competition = form.save(commit=False)

                    # Mise à jour du statut
                    new_status = form.cleaned_data.get('status')
                    if new_status:
                        competition.status = new_status
                    else:
                        is_published = form.cleaned_data.get('is_published', False)
                        competition.status = 'published' if is_published else 'draft'

                    # Sauvegarder les champs extra (non inclus dans Meta.fields)
                    competition.start_time = form.cleaned_data.get('start_time')
                    competition.end_time = form.cleaned_data.get('end_time')
                    if form.cleaned_data.get('max_participants') is not None:
                        competition.max_participants = form.cleaned_data.get('max_participants')
                    if form.cleaned_data.get('min_age') is not None:
                        competition.min_age = form.cleaned_data.get('min_age')

                    # Sauvegarder postal_code depuis le template (pas dans Meta.fields)
                    competition.postal_code = request.POST.get('postal_code', '').strip()
                    competition.save()
                    _dbg.warning(f"[UPDATE] SAVE OK pk={competition.pk}: start_time={competition.start_time!r}, end_time={competition.end_time!r}, postal_code={competition.postal_code!r}")
                    
                    # Mettre Ã  jour les types de compétition
                    competition_types = form.cleaned_data.get('competition_types')
                    if competition_types:
                        competition.competition_types.set(competition_types)
                    
                    messages.success(request, _("La compétition a été mise Ã  jour avec succès."))
                    return redirect('competitions:competitions:detail', pk=competition.id)
                    
            except Exception as e:
                _dbg.warning(f"[UPDATE] EXCEPTION: {e!r}")
                messages.error(request, _("Une erreur est survenue lors de la mise Ã  jour de la compétition: {}").format(str(e)))
        else:
            # Afficher les erreurs spécifiques du formulaire
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        # Initialiser le formulaire avec les données de la compétition
        form = CompetitionForm(instance=competition, initial={
            'is_published': competition.status == 'published'
        }, user=request.user)
    
    # Liste des disciplines pour le formulaire
    disciplines = Discipline.objects.filter(is_active=True)
    
    context = {
        'form': form,
        'competition': competition,
        'disciplines': disciplines,
        'role': getattr(request.user, 'role', None),
        'user': request.user,
        'page_title': _("Modifier la compétition"),
    }
    
    return render(request, 'competitions/competition/create.html', context)

@login_required
def competition_change_status(request, pk):
    """Vue pour changer rapidement le statut d'une compétition (POST uniquement)."""
    competition = get_object_or_404(Competition, pk=pk)

    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    new_status = request.POST.get('status', '')
    valid_statuses = ['draft', 'published', 'ongoing', 'completed', 'cancelled']
    if new_status not in valid_statuses:
        messages.error(request, _("Statut invalide."))
        return redirect('competitions:dashboard:club')

    old_status = competition.status
    competition.status = new_status
    competition.save(update_fields=['status'])

    status_labels = {
        'draft': _('Brouillon'),
        'published': _('Publiée'),
        'ongoing': _('En cours'),
        'completed': _('Terminée'),
        'cancelled': _('Annulée'),
    }
    messages.success(
        request,
        _("Statut de « %(title)s » changé de %(old)s à %(new)s.") % {
            'title': competition.title,
            'old': status_labels.get(old_status, old_status),
            'new': status_labels.get(new_status, new_status),
        }
    )
    return redirect('competitions:dashboard:club')


@login_required
@federation_required
@login_required
def competition_delete(request, pk):
    """Suppression d'une compétition avec confirmation."""
    competition = get_object_or_404(Competition, pk=pk)

    # Seul le créateur de la compétition ou un superuser peut supprimer
    is_creator = competition.created_by and competition.created_by == request.user
    is_org_admin = (
        competition.organizing_organization
        and hasattr(request.user, 'organization_memberships')
        and request.user.organization_memberships.filter(
            organization=competition.organizing_organization,
            role__in=['owner', 'admin']
        ).exists()
    )
    if not (is_creator or is_org_admin or request.user.is_superuser):
        messages.error(request, _("Seul l'administrateur qui a créé cette compétition peut la supprimer."))
        return redirect('competitions:competitions:detail', pk=pk)

    if request.method == 'POST':
        title = competition.title
        competition.delete()
        messages.success(request, _("La compétition « {} » a été supprimée.").format(title))
        return redirect('competitions:competitions:list')

    return render(request, 'competitions/competition/confirm_delete.html', {
        'competition': competition
    })
    

@login_required
def register_for_competition(request, competition_id):
    """
    Permet Ã  un utilisateur de s'inscrire Ã  une compétition.
    Gère le formulaire d'inscription et le traitement des catégories et types de compétition.
    """
    competition = get_object_or_404(Competition, id=competition_id)
    
    # Vérifier que la compétition est publiée
    if competition.status != 'published':
        messages.error(request, _("Les inscriptions pour cette compétition ne sont pas ouvertes."))
        return redirect('competitions:competitions:detail', pk=competition.id)
    
    # Vérifier la date limite d'inscription
    now = timezone.now().date()
    if competition.registration_deadline and competition.registration_deadline < now:
        messages.error(request, _("La date limite d'inscription Ã  cette compétition est dépassée."))
        return redirect('competitions:competitions:detail', pk=competition.id)
    
    # Récupérer le pratiquant lié Ã  l'utilisateur
    try:
        practitioner = Practitioner.objects.get(user=request.user)
    except Practitioner.DoesNotExist:
        messages.error(request, _("Vous devez compléter votre profil de pratiquant avant de vous inscrire."))
        # Correction de la redirection au profil de pratiquant
        if request.user.role == 'club_manager':
            return redirect('competitions:club:practitioners')  # Rediriger vers la liste des pratiquants
        else:
            return redirect('competitions:dashboard:participant')
    
    # Vérifier si une inscription existe déjÃ 
    registration, created = CompetitionRegistration.objects.get_or_create(
        practitioner=practitioner,
        competition=competition,
        defaults={
            'status': 'pending',
            'registration_date': timezone.now()
        }
    )
    
    # Vérifiez que cette ligne existe et récupère bien les types
    competition_types = competition.competition_types.all()
    
    if request.method == 'POST':
        form = CompetitionRegistrationForm(
            request.POST, 
            instance=registration,
            competition=competition,
            practitioner=practitioner  # Passer explicitement le practitioner
        )
        
        if form.is_valid():
            try:
                with transaction.atomic():
                    registration = form.save(commit=False)
                    
                    # Assigner explicitement le practitioner Ã  l'inscription
                    registration.practitioner = practitioner
                    
                    # Enregistrer les qualifications si rÃ´les sélectionnés
                    if 'is_technical_judge' in form.cleaned_data and form.cleaned_data['is_technical_judge'] and form.cleaned_data.get('technical_judge_level'):
                        JudgeQualification.objects.update_or_create(
                            practitioner=practitioner,
                            qualification_type='technical_judge',
                            defaults={'level': form.cleaned_data['technical_judge_level']}
                        )
                        
                    if 'is_combat_referee' in form.cleaned_data and form.cleaned_data['is_combat_referee'] and form.cleaned_data.get('combat_referee_level'):
                        JudgeQualification.objects.update_or_create(
                            practitioner=practitioner,
                            qualification_type='combat_referee',
                            defaults={'level': form.cleaned_data['combat_referee_level']}
                        )
                    
                    # S'assurer que la date d'inscription est définie
                    if not registration.registration_date:
                        registration.registration_date = timezone.now()
                    
                    registration.save()
                    
                    # Enregistrer les types de compétition sélectionnés
                    selected_types = []
                    for key, value in form.cleaned_data.items():
                        if key.startswith('type_') and value:
                            try:
                                type_id = int(key.split('_')[1])
                                selected_types.append(type_id)
                            except (ValueError, IndexError):
                                continue
                    
                    registration.competition_types.set(selected_types)
                    
                    # Supprimer les anciennes catégories
                    ParticipantCategoryRegistration.objects.filter(registration=registration).delete()
                    
                    # Enregistrer les catégories par type de compétition
                    for key, value in form.cleaned_data.items():
                        if key.startswith('category_for_type_') and value:
                            try:
                                type_id = int(key.split('_')[-1])
                                category = value
                                
                                ParticipantCategoryRegistration.objects.create(
                                    registration=registration,
                                    competition_type_id=type_id,
                                    category=category
                                )
                            except (ValueError, IndexError):
                                continue
                    
                    messages.success(request, _("Votre inscription a été enregistrée avec succès!"))
                    
                    # Rediriger vers le tableau de bord approprié selon le rÃ´le de l'utilisateur
                    if request.user.role == 'participant':
                        return redirect('competitions:dashboard:participant')
                    else:
                        return redirect('competitions:competitions:detail', pk=competition.id)
            except Exception as e:
                messages.error(request, _("Une erreur est survenue lors de l'inscription: {}").format(str(e)))
                # Logger l'erreur pour le débogage
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Erreur d'inscription Ã  la compétition {competition_id}: {str(e)}")
        else:
            # Afficher les erreurs spécifiques du formulaire
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        form = CompetitionRegistrationForm(
            instance=registration,
            competition=competition,
            practitioner=practitioner  # Passer explicitement le practitioner
        )
        
        # Pré-remplir les types et catégories sélectionnés
        if not created:
            for comp_type in registration.competition_types.all():
                if f'type_{comp_type.id}' in form.fields:
                    form.fields[f'type_{comp_type.id}'].initial = True
                    
                    try:
                        category_reg = ParticipantCategoryRegistration.objects.get(
                            registration=registration,
                            competition_type=comp_type
                        )
                        if f'category_for_type_{comp_type.id}' in form.fields:
                            form.fields[f'category_for_type_{comp_type.id}'].initial = category_reg.category
                    except ParticipantCategoryRegistration.DoesNotExist:
                        pass
    
    # Préparer le contexte avec les types de compétition
    context = {
        'form': form,
        'competition': competition,
        'practitioner': practitioner,
        'competition_types': competition_types,  # Assurez-vous que cette ligne est présente
        'is_edit': not created
    }
    
    return render(request, 'competitions/competition/register.html', context)
    
@login_required
def manage_competition_registrations(request, competition_id):
    """
    Gère les inscriptions Ã  une compétition (approbation, rejet, suppression).
    Permet le filtrage et l'affichage des statistiques.
    """
    competition = get_object_or_404(Competition, id=competition_id)
    
    # Vérifier les droits d'accès
    from apps.organizations.models import OrganizationMember
    has_access = (
        request.user.is_staff
        or request.user.is_superuser
        or competition.created_by == request.user
        or (competition.organizing_organization and OrganizationMember.objects.filter(
            user=request.user, organization=competition.organizing_organization
        ).exists())
    )
    if not has_access:
        messages.error(request, _("Vous n'avez pas les droits pour gérer cette compétition."))
        return redirect('competitions:competitions:detail', pk=competition.id)
    
    # Récupération des inscriptions avec les relations nécessaires pour éviter les requÃªtes N+1
    registrations = CompetitionRegistration.objects.filter(
        competition=competition
    ).select_related('practitioner', 'practitioner__organization').prefetch_related('competition_types')

    # Filtres
    status_filter = request.GET.get('status')
    role_filter = request.GET.get('role')
    club_filter = request.GET.get('club')
    search_query = request.GET.get('q')

    if status_filter:
        registrations = registrations.filter(status=status_filter)

    if role_filter == 'competitor':
        registrations = registrations.filter(is_competitor=True)
    elif role_filter == 'technical_judge':
        registrations = registrations.filter(is_technical_judge=True)
    elif role_filter == 'combat_referee':
        registrations = registrations.filter(is_combat_referee=True)
    elif role_filter == 'volunteer':
        registrations = registrations.filter(is_volunteer=True)

    if club_filter:
        registrations = registrations.filter(practitioner__organization_id=club_filter)

    # Recherche textuelle
    if search_query:
        registrations = registrations.filter(
            Q(practitioner__first_name__icontains=search_query) |
            Q(practitioner__last_name__icontains=search_query) |
            Q(practitioner__license_number__icontains=search_query) |
            Q(practitioner__organization__name__icontains=search_query)
        )
    
    # Action de mise Ã  jour du statut
    if request.method == 'POST' and 'action' in request.POST:
        action = request.POST.get('action')
        registration_ids = request.POST.getlist('registration_ids')
        
        if registration_ids:
            try:
                with transaction.atomic():
                    if action == 'approve':
                        updated = CompetitionRegistration.objects.filter(id__in=registration_ids).update(
                            status='approved',
                            updated_at=timezone.now()
                        )
                        messages.success(request, _("{} inscriptions ont été approuvées.").format(updated))
                    
                    elif action == 'reject':
                        updated = CompetitionRegistration.objects.filter(id__in=registration_ids).update(
                            status='rejected',
                            updated_at=timezone.now()
                        )
                        messages.success(request, _("{} inscriptions ont été rejetées.").format(updated))
                    
                    elif action == 'pending':
                        updated = CompetitionRegistration.objects.filter(id__in=registration_ids).update(
                            status='pending',
                            updated_at=timezone.now()
                        )
                        messages.success(request, _("{} inscriptions ont été remises en attente.").format(updated))
                    
                    elif action == 'delete':
                        deleted, _details = CompetitionRegistration.objects.filter(id__in=registration_ids).delete()
                        messages.success(request, _("{} inscriptions ont été supprimées.").format(deleted))
                    
                    elif action == 'export':
                        # Rediriger vers la vue d'export avec les IDs sélectionnés
                        ids_param = ','.join(registration_ids)
                        return redirect('competitions:competitions:export_registrations', 
                                        competition_id=competition.id,
                                        ids=ids_param)
            except Exception as e:
                messages.error(request, _("Une erreur est survenue: {}").format(str(e)))
                # Log l'erreur pour le débogage
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Erreur lors de la gestion des inscriptions: {str(e)}")
        else:
            messages.warning(request, _("Aucune inscription sélectionnée."))
    
    # Tri des inscriptions
    sort_by = request.GET.get('sort', 'registration_date')
    sort_dir = request.GET.get('dir', 'desc')
    
    order_field = f"{'-' if sort_dir == 'desc' else ''}{sort_by}"
    registrations = registrations.order_by(order_field)
    
    # Pagination
    page_size = int(request.GET.get('page_size', 20))
    paginator = Paginator(registrations, page_size)
    page = request.GET.get('page')
    registrations_page = paginator.get_page(page)
    
    # Statistiques
    stats = {
        'total': registrations.count(),
        'competitors': registrations.filter(is_competitor=True).count(),
        'technical_judges': registrations.filter(is_technical_judge=True).count(),
        'combat_referees': registrations.filter(is_combat_referee=True).count(),
        'volunteers': registrations.filter(is_volunteer=True).count(),
        'pending': registrations.filter(status='pending').count(),
        'approved': registrations.filter(status='approved').count(),
        'rejected': registrations.filter(status='rejected').count(),
    }
    
    # Clubs pour le filtre (utilisation de distinct pour éviter les doublons)
    from apps.competitions.models import Organization
    clubs = Organization.objects.filter(
        id__in=registrations.values_list('practitioner__organization_id', flat=True).distinct()
    ).order_by('name')
    
    return render(request, 'competitions/manage_registrations.html', {
        'competition': competition,
        'registrations': registrations_page,
        'stats': stats,
        'clubs': clubs,
        'status_filter': status_filter,
        'role_filter': role_filter,
        'club_filter': club_filter,
        'search_query': search_query,
        'sort_by': sort_by,
        'sort_dir': sort_dir,
        'page_size': page_size,
        # Options pour les listes déroulantes
        'status_options': [
            ('pending', _('En attente')),
            ('approved', _('Approuvées')),
            ('rejected', _('Rejetées'))
        ],
        'role_options': [
            ('competitor', _('Compétiteurs')),
            ('technical_judge', _('Juges techniques')),
            ('combat_referee', _('Arbitres combat')),
            ('volunteer', _('Bénévoles'))
        ],
    })

@login_required
def export_competition_registrations(request, competition_id):
    """
    Exporte les inscriptions Ã  une compétition au format CSV ou Excel.
    Permet de filtrer les inscriptions Ã  exporter et de choisir les champs Ã  inclure.
    """
    competition = get_object_or_404(Competition, id=competition_id)
    
    # Vérifier les droits d'accès
    if not request.user.is_staff and competition.created_by != request.user:
        messages.error(request, _("Vous n'avez pas les droits pour exporter les données de cette compétition."))
        return redirect('competitions:competitions:detail', pk=competition.id)
    
    # Récupérer les inscriptions avec toutes les relations nécessaires
    registrations = CompetitionRegistration.objects.filter(
        competition=competition
    ).select_related('practitioner', 'practitioner__organization')
    
    # Filtrer par IDs si spécifiés dans l'URL
    ids_param = request.GET.get('ids')
    if ids_param:
        try:
            id_list = [int(id) for id in ids_param.split(',')]
            registrations = registrations.filter(id__in=id_list)
        except ValueError:
            messages.warning(request, _("Paramètre d'IDs invalide, export de toutes les inscriptions."))
    
    # Appliquer les mÃªmes filtres que dans la vue manage_competition_registrations
    status_filter = request.GET.get('status')
    role_filter = request.GET.get('role')
    club_filter = request.GET.get('club')
    search_query = request.GET.get('q')
    
    if status_filter:
        registrations = registrations.filter(status=status_filter)
    
    if role_filter == 'competitor':
        registrations = registrations.filter(is_competitor=True)
    elif role_filter == 'technical_judge':
        registrations = registrations.filter(is_technical_judge=True)
    elif role_filter == 'combat_referee':
        registrations = registrations.filter(is_combat_referee=True)
    elif role_filter == 'volunteer':
        registrations = registrations.filter(is_volunteer=True)
    
    if club_filter:
        registrations = registrations.filter(practitioner__club_id=club_filter)
    
    if search_query:
        registrations = registrations.filter(
            Q(practitioner__first_name__icontains=search_query) |
            Q(practitioner__last_name__icontains=search_query) |
            Q(practitioner__license_number__icontains=search_query) |
            Q(practitioner__club__name__icontains=search_query)
        )
    
    # Déterminer le format d'export
    export_format = request.GET.get('format', 'csv').lower()
    
    if export_format == 'excel':
        # Export Excel
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
            
            # Créer un nouveau classeur Excel
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = _("Inscriptions")
            
            # Définir le style d'en-tÃªte
            header_font = Font(bold=True)
            header_alignment = Alignment(horizontal='center')
            
            # En-tÃªtes
            headers = [
                _('Nom'),
                _('Prénom'),
                _('Club'),
                _('Email'),
                _('Licence'),
                _('Date de naissance'),
                _('RÃ´les'),
                _('Types de compétition'),
                _('Catégories'),
                _('Statut'),
                _('Date d\'inscription')
            ]
            
            for col_num, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Données
            for row_num, registration in enumerate(registrations, 2):
                # Récupérer les types de compétition et catégories
                competition_types = ', '.join([ct.name for ct in registration.competition_types.all()])
                categories = ', '.join([
                    pcr.category.name if hasattr(pcr.category, 'name') else str(pcr.category)
                    for pcr in ParticipantCategoryRegistration.objects.filter(registration=registration)
                ])
                
                # Récupérer les rÃ´les
                roles = []
                if hasattr(registration, 'is_competitor') and registration.is_competitor:
                    roles.append(_('Compétiteur'))
                if hasattr(registration, 'is_technical_judge') and registration.is_technical_judge:
                    roles.append(_('Juge technique'))
                if hasattr(registration, 'is_combat_referee') and registration.is_combat_referee:
                    roles.append(_('Arbitre combat'))
                if hasattr(registration, 'is_volunteer') and registration.is_volunteer:
                    roles.append(_('Bénévole'))
                roles_str = ', '.join(roles)
                
                # Construire la ligne
                row = [
                    registration.practitioner.last_name,
                    registration.practitioner.first_name,
                    registration.practitioner.club.name if registration.practitioner.club else '',
                    registration.practitioner.email,
                    registration.practitioner.license_number,
                    registration.practitioner.birth_date.strftime('%d/%m/%Y') if registration.practitioner.birth_date else '',
                    roles_str,
                    competition_types,
                    categories,
                    registration.get_status_display() if hasattr(registration, 'get_status_display') else registration.status,
                    registration.registration_date.strftime('%d/%m/%Y %H:%M') if registration.registration_date else ''
                ]
                
                for col_num, cell_value in enumerate(row, 1):
                    worksheet.cell(row=row_num, column=col_num, value=cell_value)
            
            # Ajuster la largeur des colonnes
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = (max_length + 2) * 1.2
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Créer la réponse HTTP
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{competition.slug}_inscriptions.xlsx"'
            
            # Sauvegarder le classeur dans la réponse
            workbook.save(response)
            return response
        
        except ImportError:
            messages.warning(request, _("Module openpyxl non disponible. Export en CSV Ã  la place."))
            export_format = 'csv'
    
    if export_format == 'csv':
        # Export CSV
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{competition.slug}_inscriptions.csv"'
        
        writer = csv.writer(response, delimiter=';')
        
        # En-tÃªtes
        writer.writerow([
            _('Nom'),
            _('Prénom'),
            _('Club'),
            _('Email'),
            _('Licence'),
            _('Date de naissance'),
            _('Compétiteur'),
            _('Juge technique'),
            _('Arbitre combat'),
            _('Bénévole'),
            _('Types de compétition'),
            _('Catégories'),
            _('Statut'),
            _('Date d\'inscription')
        ])
        
        # Données
        for registration in registrations:
            # Récupérer les types de compétition et catégories
            competition_types = ', '.join([ct.name for ct in registration.competition_types.all()])
            categories = ', '.join([
                pcr.category.name if hasattr(pcr.category, 'name') else str(pcr.category)
                for pcr in ParticipantCategoryRegistration.objects.filter(registration=registration)
            ])
            
            writer.writerow([
                registration.practitioner.last_name,
                registration.practitioner.first_name,
                registration.practitioner.club.name if registration.practitioner.club else '',
                registration.practitioner.email,
                registration.practitioner.license_number,
                registration.practitioner.birth_date.strftime('%d/%m/%Y') if registration.practitioner.birth_date else '',
                _('Oui') if hasattr(registration, 'is_competitor') and registration.is_competitor else _('Non'),
                _('Oui') if hasattr(registration, 'is_technical_judge') and registration.is_technical_judge else _('Non'),
                _('Oui') if hasattr(registration, 'is_combat_referee') and registration.is_combat_referee else _('Non'),
                _('Oui') if hasattr(registration, 'is_volunteer') and registration.is_volunteer else _('Non'),
                competition_types,
                categories,
                registration.get_status_display() if hasattr(registration, 'get_status_display') else registration.status,
                registration.registration_date.strftime('%d/%m/%Y %H:%M') if registration.registration_date else ''
            ])
        
        return response
    
    # Si le format demandé n'est pas supporté
    messages.error(request, _("Format d'export non supporté."))
    return redirect('competitions:competitions:manage_registrations', competition_id=competition.id)


@login_required
def competition_duplicate(request, pk):
    """Duplicate a competition with all its configuration (categories, types) but without participant data."""
    source = get_object_or_404(Competition, pk=pk)

    try:
        with transaction.atomic():
            # Store M2M and categories before copying
            source_types = list(source.competition_types.all())
            source_categories = list(source.categories.all())

            # Create a copy by resetting pk
            new_comp = Competition()
            # Copy configuration fields
            new_comp.title = f"{source.title} (Copie)"
            new_comp.description = source.description
            new_comp.start_date = source.start_date
            new_comp.end_date = source.end_date
            new_comp.start_time = source.start_time
            new_comp.end_time = source.end_time
            new_comp.venue_name = source.venue_name
            new_comp.address = source.address
            new_comp.city = source.city
            new_comp.postal_code = source.postal_code
            new_comp.max_participants = source.max_participants
            new_comp.registration_deadline = source.registration_deadline
            new_comp.logo = source.logo
            new_comp.min_age = source.min_age
            new_comp.max_age = source.max_age
            new_comp.requires_medical_certificate = source.requires_medical_certificate
            new_comp.requires_license = source.requires_license
            new_comp.discipline = source.discipline
            new_comp.organizing_organization = source.organizing_organization
            # Stream config (not runtime)
            new_comp.stream_enabled = source.stream_enabled
            new_comp.stream_platform = source.stream_platform
            new_comp.stream_url = source.stream_url
            new_comp.stream_chat_enabled = source.stream_chat_enabled
            # Reset fields
            new_comp.status = 'draft'
            new_comp.is_published = False
            new_comp.created_by = request.user
            new_comp.slug = ''  # Will be auto-generated on save

            new_comp.save()

            # Copy M2M competition_types
            new_comp.competition_types.set(source_types)

            # Duplicate categories
            for cat in source_categories:
                CompetitionCategory.objects.create(
                    competition=new_comp,
                    competition_type=cat.competition_type,
                    template=cat.template,
                    name=cat.name,
                    min_age=cat.min_age,
                    max_age=cat.max_age,
                    min_grade=cat.min_grade,
                    max_grade=cat.max_grade,
                    min_weight=cat.min_weight,
                    max_weight=cat.max_weight,
                    gender=cat.gender,
                    max_participants=cat.max_participants,
                    combat_mode=cat.combat_mode,
                    estimated_duration=cat.estimated_duration,
                    # Reset scheduling and status
                    registration_status='open',
                    is_completed=False,
                )

        messages.success(request, _("La compétition a été dupliquée avec succès."))
        return redirect('competitions:competitions:update', pk=new_comp.pk)

    except Exception as e:
        messages.error(request, _("Erreur lors de la duplication : {}").format(str(e)))
        return redirect('competitions:dashboard:dashboard')

