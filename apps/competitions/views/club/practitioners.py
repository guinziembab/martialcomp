"""
Vue améliorée pour la gestion des pratiquants avec double affichage (cartes/tableau)
Support export, actions en lot et filtres avancés
"""
import logging
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib import messages
from django.utils.translation import gettext_lazy as _
from django.utils.translation import gettext as gettext_func
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST, require_http_methods
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.contrib.auth.models import User
from django.db import models, transaction
from django.db.models import Q, Count, Avg
from django.core.paginator import Paginator
from django.core.exceptions import PermissionDenied
from datetime import datetime, timedelta
import json
import csv
import openpyxl
from io import StringIO, BytesIO

# Import des modèles
from apps.competitions.models import (
    Practitioner, 
    Club,
    Discipline,
    Competition,
    CompetitionRegistration
)

# Import des formulaires
from apps.competitions.forms import (
    PractitionerForm,
)

# Import des décorateurs
from apps.competitions.utils.decorators import club_required, permission_required
from apps.core.isolation import OrganizationIsolationMixin, get_organization_queryset
from apps.competitions.utils.club import get_user_club as get_club_from_utils
from apps.competitions.utils.organization_discipline_filtering import (
    filter_practitioners_by_org_disciplines,
    get_organization_disciplines
)

# Import des modèles de grades avec gestion des erreurs
try:
    from apps.grades.models import GradingSystem as GradeSystem, Grade
    GRADES_MODELS_AVAILABLE = True
except (ImportError, ModuleNotFoundError) as e:
    logger = logging.getLogger(__name__)
    logger.warning(f"Les modèles de grades n'ont pas pu être importés: {str(e)}")
    GRADES_MODELS_AVAILABLE = False

# Création du logger
logger = logging.getLogger(__name__)

def get_user_club(user_or_request):
    """Obtenir le club de l'utilisateur avec gestion des erreurs"""
    try:
        # Détermine si c'est un user ou une request
        if hasattr(user_or_request, 'user'):
            request = user_or_request
            user = request.user
        else:
            user = user_or_request
            request = None
        
        # Si le club est déjà dans la requête (via le décorateur)
        if request and hasattr(request, 'club') and request.club:
            return request.club
        
        # PRIORITÉ 1: Utiliser l'organisation du middleware si disponible dans la requête
        if request and hasattr(request, 'user_organization') and request.user_organization:
            return request.user_organization
        
        # PRIORITÉ 2: Utiliser la logique du middleware directement (UserProfile)
        try:
            from apps.competitions.models.users import UserProfile
            profile = UserProfile.objects.get(user=user)
            if profile.organization:
                return profile.organization
        except:
            pass
        
        # PRIORITÉ 3: Memberships actifs (comme le middleware)
        try:
            from apps.membership.models import MembershipSubscription
            subscription = MembershipSubscription.objects.filter(
                practitioner__user=user,
                status='active'
            ).select_related('package__organization').first()
            
            if subscription and subscription.package.organization:
                return subscription.package.organization
        except:
            pass
        
        # PRIORITÉ 4: Si l'utilisateur a un attribut club direct
        if hasattr(user, 'club') and user.club:
            club = user.club
            if hasattr(club, 'organization') and club.organization:
                return club.organization
            else:
                logger.error(f"Le club {club} de l'utilisateur n'a pas d'organisation associée")
                return None
        
        # PRIORITÉ 5: Si l'utilisateur est propriétaire d'un club
        from apps.competitions.models import Club
        owned_club = Club.objects.filter(owner=user).first()
        if owned_club:
            # IMPORTANT: Toujours retourner l'organisation, jamais le club directement
            if hasattr(owned_club, 'organization') and owned_club.organization:
                return owned_club.organization
            else:
                # Si le club n'a pas d'organisation, log l'erreur et retourner None
                logger.error(f"Le club {owned_club.name} n'a pas d'organisation associée")
                return None
        
        # PRIORITÉ 6: Coach profile
        if hasattr(user, 'coach_profile') and user.coach_profile and user.coach_profile.club:
            coach_club = user.coach_profile.club
            # Retourner l'organisation liée au club
            return coach_club.organization if hasattr(coach_club, 'organization') and coach_club.organization else coach_club
        
        # PRIORITÉ 7: Practitioner record
        practitioner = Practitioner.objects.select_related('organization').filter(
            Q(user=user) | Q(email=user.email)
        ).first()
        
        if practitioner and practitioner.organization:
            return practitioner.organization
        
        # PRIORITÉ 8: Admin roles
        if hasattr(user, 'club_admin_roles'):
            club_admin = user.club_admin_roles.first()
            if club_admin:
                # Retourner l'organisation, pas le club
                if hasattr(club_admin, 'organization') and club_admin.organization:
                    return club_admin.organization
                elif hasattr(club_admin, 'club') and club_admin.club:
                    admin_club = club_admin.club
                    if hasattr(admin_club, 'organization') and admin_club.organization:
                        return admin_club.organization
                    else:
                        logger.error(f"Le club {admin_club} n'a pas d'organisation associée")
                        return None
            
        return None
    except Exception as e:
        user_id = user.id if hasattr(user, 'id') else 'unknown'
        logger.error(f"Erreur lors de la récupération du club de l'utilisateur {user_id}: {str(e)}")
        return None

def manual_permission_check(user, organization):
    """Vérification manuelle des permissions"""
    try:
        if user.is_superuser:
            return True
        
        # Vérifier que l'organisation existe
        if organization is None:
            logger.warning(f"Pas d'organisation pour l'utilisateur {user.username}")
            return False
        
        # Si organisation est une chaîne, essayer de trouver l'organisation
        if isinstance(organization, str):
            from apps.organizations.models import Organization
            try:
                organization = Organization.objects.get(name=organization)
            except Organization.DoesNotExist:
                logger.error(f"Organisation '{organization}' non trouvée")
                return False
        
        # Vérifier si l'utilisateur est propriétaire d'un club lié à cette organisation
        from apps.competitions.models import Club
        owned_club = Club.objects.filter(
            owner=user,
            organization=organization
        ).first()
        if owned_club:
            logger.info(f"Utilisateur {user.username} est propriétaire du club {owned_club}")
            return True
            
        if hasattr(user, 'coach_profile') and user.coach_profile:
            coach_club = user.coach_profile.club
            if coach_club and organization and hasattr(organization, 'id'):
                # Vérifier si le club du coach appartient à cette organisation
                if hasattr(coach_club, 'organization') and coach_club.organization == organization:
                    logger.info(f"Utilisateur {user.username} est coach dans l'organisation {organization}")
                    return True
        
        # Vérifier via le Practitioner avec l'objet Organization
        practitioner = Practitioner.objects.filter(
            Q(user=user) | Q(email=user.email),
            organization=organization
        ).first()
        
        if practitioner:
            logger.info(f"Utilisateur {user.username} est pratiquant dans l'organisation {organization}")
            return True
        
        # NOUVEAU: Si l'utilisateur a un UserProfile avec cette organisation
        try:
            from apps.competitions.models.users import UserProfile
            profile = UserProfile.objects.get(user=user)
            if profile.organization == organization:
                logger.info(f"Utilisateur {user.username} a un profil dans l'organisation {organization}")
                return True
        except:
            pass
        
        # Si aucune permission trouvée mais que l'utilisateur a bien une organisation
        # on peut être plus permissif ici pour permettre la création de pratiquants
        logger.info(f"Permission accordée par défaut pour {user.username} dans {organization}")
        return True
        
    except Exception as e:
        logger.error(f"Erreur lors de la vérification des permissions: {str(e)}", exc_info=True)
        return False

@login_required
def practitioners_list(request):
    """
    Vue améliorée pour la gestion des pratiquants avec:
    - Double affichage (cartes/tableau)
    - Filtres avancés
    - Actions en lot
    - Export (CSV, Excel, PDF)
    - Statistiques enrichies
    """
    try:
        # Récupération du club
        user_club = get_user_club(request)
        
        # DEBUG: Log des informations de débogage
        logger.info(f"DEBUG practitioners_list - User: {request.user.username}")
        logger.info(f"DEBUG practitioners_list - Club found: {user_club}")
        logger.info(f"DEBUG practitioners_list - Request has user_organization: {hasattr(request, 'user_organization')}")
        if hasattr(request, 'user_organization'):
            logger.info(f"DEBUG practitioners_list - Request.user_organization: {request.user_organization}")
        
        if not user_club:
            messages.error(request, _("Vous n'êtes associé à aucun club. Contactez un administrateur."))
            return redirect('competitions:dashboard:dashboard')
        
        # DEBUG: Vérifications des permissions détaillées
        permission_result = manual_permission_check(request.user, user_club)
        logger.info(f"DEBUG practitioners_list - Permission check result: {permission_result}")
        
        # Compter TOUS les pratiquants de l'organisation (sans filtre discipline)
        practitioner_count = Practitioner.objects.filter(organization=user_club).count()
        logger.info(f"DEBUG practitioners_list - Practitioners in {user_club}: {practitioner_count}")
        
        if not permission_result:
            # Au lieu de lever une exception, essayons de trouver un club avec des pratiquants
            logger.warning(f"Permission denied for user {request.user.username} in club {user_club}")
            
            # Chercher un practitioner record pour cet utilisateur
            user_practitioner = Practitioner.objects.filter(
                Q(user=request.user) | Q(email=request.user.email)
            ).first()
            
            if user_practitioner and user_practitioner.organization:
                logger.info(f"DEBUG - Found practitioner organization: {user_practitioner.organization}")
                user_club = user_practitioner.organization
                # Réévaluer les permissions avec la nouvelle organisation
                permission_result = manual_permission_check(request.user, user_club)
                logger.info(f"DEBUG - New permission check result: {permission_result}")
                
                if permission_result:
                    logger.info(f"SUCCESS - Using practitioner organization: {user_club}")
                else:
                    raise PermissionDenied(_("Vous n'avez pas l'autorisation de voir cette page."))
            else:
                raise PermissionDenied(_("Vous n'avez pas l'autorisation de voir cette page."))

        # Base queryset: TOUS les pratiquants de l'organisation (sans filtre discipline)
        queryset = Practitioner.objects.filter(
            organization=user_club
        ).select_related('organization', 'primary_discipline', 'user').prefetch_related('disciplines').order_by('last_name', 'first_name')

        # === FILTRES AVANCÉS ===
        search = request.GET.get('search', '').strip()
        if search:
            queryset = queryset.filter(
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(email__icontains=search) |
                Q(phone__icontains=search) |
                Q(license_number__icontains=search)
            )

        # Filtre par grade
        grade_filter = request.GET.get('grade', '').strip()
        if grade_filter:
            if grade_filter == 'none':
                queryset = queryset.filter(Q(grade__isnull=True) | Q(grade=''))
            else:
                queryset = queryset.filter(grade=grade_filter)

        # Filtre par groupe d'âge
        age_group = request.GET.get('age_group', '').strip()
        if age_group:
            today = timezone.now().date()
            if age_group == 'child':  # Moins de 12 ans
                min_birth_date = today - timedelta(days=12*365)
                queryset = queryset.filter(birth_date__gt=min_birth_date)
            elif age_group == 'teen':  # 12-17 ans
                min_birth_date = today - timedelta(days=18*365)
                max_birth_date = today - timedelta(days=12*365)
                queryset = queryset.filter(birth_date__lte=max_birth_date, birth_date__gt=min_birth_date)
            elif age_group == 'adult':  # 18 ans et plus
                max_birth_date = today - timedelta(days=18*365)
                queryset = queryset.filter(birth_date__lte=max_birth_date)

        # Filtre par statut
        status = request.GET.get('status', '').strip()
        if status == 'active':
            queryset = queryset.filter(is_active=True)
        elif status == 'inactive':
            queryset = queryset.filter(is_active=False)

        # Filtre par statut utilisateur
        user_status = request.GET.get('user_status', '').strip()
        if user_status == 'with_account':
            queryset = queryset.filter(user__isnull=False)
        elif user_status == 'without_account':
            queryset = queryset.filter(user__isnull=True)

        # Filtre par discipline
        discipline = request.GET.get('discipline', '').strip()
        if discipline:
            queryset = queryset.filter(primary_discipline_id=discipline)

        # === GESTION DES EXPORTS ===
        export_format = request.GET.get('format', '').strip()
        if export_format in ['csv', 'excel', 'pdf']:
            return handle_export(request, queryset, export_format, user_club)

        # === STATISTIQUES ENRICHIES ===
        total_practitioners = queryset.count()
        active_practitioners = queryset.filter(is_active=True).count()
        inactive_practitioners = total_practitioners - active_practitioners
        practitioners_with_accounts = queryset.filter(user__isnull=False).count()
        coaches = queryset.filter(is_coach=True).count()

        # Distribution par grade
        grade_distribution = {}
        for practitioner in queryset:
            grade_display = get_grade_display(practitioner)
            if grade_display:
                grade_distribution[grade_display] = grade_distribution.get(grade_display, 0) + 1

        # Distribution par âge
        age_distribution = {'child': 0, 'teen': 0, 'adult': 0, 'unknown': 0}
        today = timezone.now().date()
        
        for practitioner in queryset:
            if practitioner.birth_date:
                age = (today - practitioner.birth_date).days // 365
                if age < 12:
                    age_distribution['child'] += 1
                elif age < 18:
                    age_distribution['teen'] += 1
                else:
                    age_distribution['adult'] += 1
            else:
                age_distribution['unknown'] += 1

        # === PAGINATION ===
        paginator = Paginator(queryset, 12)  # 12 par page pour les cartes
        page_number = request.GET.get('page')
        practitioners = paginator.get_page(page_number)

        # Enrichissement des pratiquants avec des données calculées
        for practitioner in practitioners:
            # Use temporary attributes instead of trying to set properties
            practitioner.computed_grade_display = get_grade_display(practitioner)
            practitioner.computed_grade_css_class = get_grade_css_class(practitioner)
            practitioner.computed_age = get_age(practitioner) if practitioner.birth_date else None

            # Ajouter l'objet grade et son image pour le template tag {% grade_image %}
            if practitioner.grade:
                practitioner.computed_grade_object = practitioner.grade
                practitioner.computed_grade_image_url = practitioner.grade.image_url
                practitioner.computed_grade_color = practitioner.grade.get_display_color()
            else:
                practitioner.computed_grade_object = None
                practitioner.computed_grade_image_url = None
                practitioner.computed_grade_color = '#6c757d'

            # Ajouter la discipline pour l'affichage - utiliser disciplines (ManyToMany) en priorité
            disciplines_list = list(practitioner.disciplines.all())
            if disciplines_list:
                practitioner.computed_discipline_display = ', '.join([d.name for d in disciplines_list])
            elif practitioner.primary_discipline:
                practitioner.computed_discipline_display = practitioner.primary_discipline.name
            else:
                practitioner.computed_discipline_display = None

        # === CHOIX POUR LES FILTRES ===
        grade_choices = get_grade_choices(user_club)
        discipline_choices = get_discipline_choices(user_club)

        # Contexte pour le template
        context = {
            'practitioners': practitioners,
            'club': user_club,
            'grade_choices': grade_choices,
            'discipline_choices': discipline_choices,
            'grade_distribution': grade_distribution,
            'age_distribution': age_distribution,
            'stats': {
                'total_practitioners': total_practitioners,
                'active_practitioners': active_practitioners,
                'inactive_practitioners': inactive_practitioners,
                'practitioners_with_accounts': practitioners_with_accounts,
                'coaches': coaches,
                'completion_rate': round((practitioners_with_accounts / total_practitioners * 100) if total_practitioners > 0 else 0, 1)
            },
            'current_section': 'practitioners',
            'page_title': _('Gestion des Pratiquants'),
        }

        return render(request, 'competitions/club/practitioners_enhanced.html', context)

    except PermissionDenied:
        raise
    except Exception as e:
        logger.error(f"Erreur dans practitioners_list pour l'utilisateur {request.user.id}: {str(e)}")
        messages.error(request, _("Une erreur est survenue lors du chargement de la page."))
        return redirect('competitions:dashboard:club')

def handle_export(request, queryset, format_type, club):
    """Gestion des exports en différents formats"""
    try:
        # Récupération des pratiquants sélectionnés si spécifié
        selected_ids = request.GET.get('selected', '').strip()
        if selected_ids:
            selected_ids = [int(id) for id in selected_ids.split(',') if id.isdigit()]
            queryset = queryset.filter(id__in=selected_ids)

        if format_type == 'csv':
            return export_to_csv(queryset, club)
        elif format_type == 'excel':
            return export_to_excel(queryset, club)
        elif format_type == 'pdf':
            return export_to_pdf(queryset, club)
        
    except Exception as e:
        logger.error(f"Erreur lors de l'export {format_type}: {str(e)}")
        messages.error(request, _("Erreur lors de l'export des données."))
        return redirect('competitions:club:practitioners')

def export_to_csv(queryset, club):
    """Export des pratiquants en CSV"""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="pratiquants_{club.name}_{datetime.now().strftime("%Y%m%d")}.csv"'
    response.write('\ufeff')  # BOM pour Excel
    
    writer = csv.writer(response)
    
    # En-têtes
    headers = [
        'Nom', 'Prénom', 'Email', 'Téléphone', 'Date de naissance',
        'Genre', 'Grade', 'Discipline principale', 'Numéro de licence',
        'Statut', 'Date d\'inscription', 'Est instructeur'
    ]
    writer.writerow(headers)
    
    # Données
    for practitioner in queryset:
        row = [
            practitioner.last_name,
            practitioner.first_name,
            practitioner.email or '',
            practitioner.phone or '',
            practitioner.birth_date.strftime('%d/%m/%Y') if practitioner.birth_date else '',
            practitioner.get_gender_display() if practitioner.gender else '',
            get_grade_display(practitioner) or '',
            practitioner.primary_discipline.name if practitioner.primary_discipline else '',
            practitioner.license_number or '',
            'Actif' if practitioner.is_active else 'Inactif',
            practitioner.registration_date.strftime('%d/%m/%Y') if practitioner.registration_date else '',
            'Oui' if practitioner.is_coach else 'Non'
        ]
        writer.writerow(row)
    
    return response

def export_to_excel(queryset, club):
    """Export des pratiquants en Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pratiquants"
    
    # En-têtes avec style
    headers = [
        'Nom', 'Prénom', 'Email', 'Téléphone', 'Date de naissance',
        'Genre', 'Age', 'Grade', 'Discipline principale', 'Numéro de licence',
        'Statut', 'Date d\'inscription', 'Est instructeur', 'Ville'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Données
    for row_num, practitioner in enumerate(queryset, 2):
        data = [
            practitioner.last_name,
            practitioner.first_name,
            practitioner.email or '',
            practitioner.phone or '',
            practitioner.birth_date if practitioner.birth_date else '',
            practitioner.get_gender_display() if practitioner.gender else '',
            get_age(practitioner) if practitioner.birth_date else '',
            get_grade_display(practitioner) or '',
            practitioner.primary_discipline.name if practitioner.primary_discipline else '',
            practitioner.license_number or '',
            'Actif' if practitioner.is_active else 'Inactif',
            practitioner.registration_date if practitioner.registration_date else '',
            'Oui' if practitioner.is_coach else 'Non',
            practitioner.city or ''
        ]
        
        for col, value in enumerate(data, 1):
            ws.cell(row=row_num, column=col, value=value)
    
    # Ajustement de la largeur des colonnes
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Sauvegarde
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="pratiquants_{club.name}_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    
    return response

def export_to_pdf(queryset, club):
    """Export des pratiquants en PDF (version simple)"""
    from django.template.loader import render_to_string
    from django.http import HttpResponse
    
    # Pour une version simple, on retourne un HTML stylé
    # Dans une version complète, on utiliserait reportlab ou weasyprint
    
    context = {
        'practitioners': queryset,
        'club': club,
        'export_date': datetime.now(),
        'total_count': queryset.count()
    }
    
    html_content = render_to_string('competitions/club/practitioners_export_pdf.html', context)
    
    response = HttpResponse(html_content, content_type='text/html')
    response['Content-Disposition'] = f'attachment; filename="pratiquants_{club.name}_{datetime.now().strftime("%Y%m%d")}.html"'
    
    return response

def get_grade_display(practitioner):
    """Obtenir l'affichage du grade d'un pratiquant"""
    try:
        if hasattr(practitioner, 'grade_text') and practitioner.grade_text:
            return practitioner.grade_text
        
        if practitioner.grade:
            # If grade is a Grade model object
            if hasattr(practitioner.grade, 'name'):
                return practitioner.grade.name
            # If grade is a Grade model object with string representation
            elif not isinstance(practitioner.grade, str):
                return str(practitioner.grade)
            # If grade is a string
            else:
                # Mapping basique des grades
                grade_mapping = {
                    'cap_jaune': 'Ceinture Jaune',
                    'cap_rouge': 'Ceinture Rouge',
                    'cap_blanc': 'Ceinture Blanche',
                    'cap_bleu': 'Ceinture Bleue',
                    'cap_marron': 'Ceinture Marron',
                    'cap_noir': 'Ceinture Noire',
                }
                return grade_mapping.get(practitioner.grade, practitioner.grade.replace('_', ' ').title())
        
        return None
        
    except Exception as e:
        logger.error(f"Erreur lors de la récupération du grade pour le pratiquant {practitioner.id}: {str(e)}")
        return None

def get_grade_css_class(practitioner):
    """Obtenir la classe CSS pour le style du grade"""
    try:
        if practitioner.grade:
            # If grade is a string
            if isinstance(practitioner.grade, str):
                return practitioner.grade.replace('_', '-')
            # If grade is a Grade model object, use its identifier or name
            elif hasattr(practitioner.grade, 'identifier'):
                return practitioner.grade.identifier.replace('_', '-')
            elif hasattr(practitioner.grade, 'name'):
                grade_name = practitioner.grade.name.lower().replace(' ', '-').replace('ceinture ', 'cap-')
                return grade_name
            else:
                return str(practitioner.grade).replace('_', '-').replace(' ', '-')
        return 'cap-blanc'
    except Exception as e:
        logger.error(f"Erreur lors de la récupération de la classe CSS du grade pour le pratiquant {practitioner.id}: {str(e)}")
        return 'cap-blanc'

def get_age(practitioner):
    """Calculer l'âge d'un pratiquant"""
    if practitioner.birth_date:
        today = timezone.now().date()
        return (today - practitioner.birth_date).days // 365
    return None

def get_grade_choices(club):
    """Obtenir les choix de grades pour les filtres"""
    try:
        if GRADES_MODELS_AVAILABLE:
            grades = Grade.objects.filter(is_active=True).order_by('order')
            return [(grade.id, grade.name) for grade in grades]
        else:
            # Choix statiques pour Qwan Ki Do
            return [
                ('cap_blanc', 'Ceinture Blanche'),
                ('cap_jaune', 'Ceinture Jaune'),
                ('cap_rouge', 'Ceinture Rouge'),
                ('cap_bleu', 'Ceinture Bleue'),
                ('cap_marron', 'Ceinture Marron'),
                ('cap_noir', 'Ceinture Noire'),
            ]
    except Exception as e:
        logger.error(f"Erreur lors de la récupération des grades: {str(e)}")
        return []

def get_discipline_choices(club):
    """Obtenir les choix de disciplines pour les filtres (filtrées par disciplines de l'organisation)"""
    try:
        # Récupérer les disciplines de l'organisation
        org_disciplines = get_organization_disciplines(club)
        if org_disciplines.exists():
            # Retourner uniquement les disciplines de l'organisation
            return [(d.id, d.name) for d in org_disciplines.order_by('name')]
        else:
            # Fallback : disciplines des pratiquants de l'organisation
            practitioners_qs = filter_practitioners_by_org_disciplines(club)
            disciplines = Discipline.objects.filter(
                practitioners__in=practitioners_qs
            ).distinct().order_by('name')
            return [(d.id, d.name) for d in disciplines]
    except Exception as e:
        logger.error(f"Erreur lors de la récupération des disciplines: {str(e)}")
        return []

@login_required
@require_POST
def practitioners_bulk_action(request):
    """Gestion des actions en lot sur les pratiquants"""
    try:
        action = request.POST.get('action')
        selected_ids = request.POST.getlist('selected_ids')
        
        if not action or not selected_ids:
            return JsonResponse({'success': False, 'error': _('Paramètres manquants')})

        # Récupération du club
        user_club = get_user_club(request)
        if not user_club or not manual_permission_check(request.user, user_club):
            return JsonResponse({'success': False, 'error': _('Permissions insuffisantes')})
        
        # Récupération des pratiquants de l'organisation
        practitioners = Practitioner.objects.filter(organization=user_club, id__in=selected_ids)
        
        if action == 'activate':
            practitioners.update(is_active=True)
            message = f"{practitioners.count()} pratiquants activés"
        elif action == 'deactivate':
            practitioners.update(is_active=False)
            message = f"{practitioners.count()} pratiquants désactivés"
        elif action == 'delete':
            count = practitioners.count()
            practitioners.delete()
            message = f"{count} pratiquants supprimés"
        else:
            return JsonResponse({'success': False, 'error': _('Action non reconnue')})
        
        return JsonResponse({
            'success': True, 
            'message': message,
            'affected_count': len(selected_ids)
        })
        
    except Exception as e:
        logger.error(f"Erreur dans practitioners_bulk_action: {str(e)}")
        return JsonResponse({'success': False, 'error': _("Erreur lors de l'opération")})

# ===== FONCTIONS ORIGINALES CONSERVÉES POUR COMPATIBILITÉ =====

@login_required
def practitioner_detail(request, pk):
    """Afficher le détail d'un pratiquant"""
    try:
        logger.info(f"practitioner_detail appelé pour pk={pk} par {request.user.username}")
        
        # Récupérer le pratiquant sans filtre d'organisation pour éviter les erreurs
        try:
            practitioner = Practitioner.objects.select_related('organization', 'primary_discipline', 'user').prefetch_related(
                'practitioner_qualifications__discipline', 
                'practitioner_qualifications__certified_by'
            ).get(pk=pk)
            logger.info(f"Pratiquant trouvé: {practitioner.full_name}, org={practitioner.organization}")
        except Practitioner.DoesNotExist:
            logger.error(f"Pratiquant pk={pk} introuvable")
            messages.error(request, _("Pratiquant introuvable."))
            return redirect('competitions:club:practitioners')
        
        # Vérification simplifiée des permissions : 
        # L'utilisateur doit être superuser ou avoir un lien avec le club/organisation
        user_club = get_user_club(request)
        logger.info(f"user_club={user_club}, type={type(user_club)}")
        
        if not request.user.is_superuser:
            # Si l'utilisateur n'a pas de club associé, il ne peut pas voir les pratiquants
            if not user_club:
                logger.error(f"Pas de club associé pour {request.user.username}")
                messages.error(request, _("Vous n'avez pas l'autorisation de voir ce pratiquant."))
                return redirect('competitions:dashboard:club')
            
            # Si le pratiquant n'a pas d'organisation, on autorise l'accès
            # Sinon, vérifier que le pratiquant appartient au même club/organisation
            if practitioner.organization:
                if user_club:
                    # Comparer les IDs pour éviter les problèmes de comparaison d'objets
                    practitioner_org_id = practitioner.organization.id if hasattr(practitioner.organization, 'id') else str(practitioner.organization)
                    user_club_id = user_club.id if hasattr(user_club, 'id') else str(user_club)
                    
                    logger.info(f"Comparaison: practitioner_org_id={practitioner_org_id} vs user_club_id={user_club_id}")
                    
                    if str(practitioner_org_id) != str(user_club_id):
                        # Si les organisations sont différentes, vérifier si l'utilisateur est propriétaire d'un club
                        try:
                            from apps.competitions.models import Club
                            is_club_owner = Club.objects.filter(owner=request.user).exists()
                            logger.info(f"is_club_owner={is_club_owner}")
                            if not is_club_owner:
                                logger.error(f"Accès refusé: orgs différentes et pas propriétaire de club")
                                messages.error(request, _("Vous n'avez pas l'autorisation de voir ce pratiquant."))
                                return redirect('competitions:club:practitioners')
                        except Exception as e:
                            logger.error(f"Erreur lors de la vérification des permissions: {str(e)}", exc_info=True)
                            # En cas d'erreur, autoriser l'accès plutôt que de bloquer
                            pass
            else:
                # Si le pratiquant n'a pas d'organisation, autoriser l'accès
                logger.info(f"Pratiquant sans organisation - accès autorisé")
        
        # Enrichissement des données
        try:
            # Charger les grades actuels du pratiquant
            current_grades = {}
            try:
                from apps.grades.models import PractitionerGrade
                practitioner_grades = PractitionerGrade.objects.filter(
                    practitioner=practitioner,
                    is_current=True
                ).select_related('grade', 'discipline', 'grade__discipline')
                
                for pg in practitioner_grades:
                    if pg.discipline:
                        current_grades[pg.discipline.id] = pg
                    elif pg.grade and pg.grade.discipline:
                        current_grades[pg.grade.discipline.id] = pg
                
                # Déterminer le grade principal (le plus élevé ou le premier)
                main_grade = None
                if practitioner_grades.exists():
                    # Prendre le grade avec le niveau le plus élevé
                    main_grade_obj = max(practitioner_grades, key=lambda pg: pg.grade.level if pg.grade and pg.grade.level else 0)
                    main_grade = main_grade_obj.grade.name if main_grade_obj.grade else None
                
                # Si pas de grade via PractitionerGrade, utiliser le champ grade du pratiquant
                if not main_grade and practitioner.grade:
                    main_grade = practitioner.grade if isinstance(practitioner.grade, str) else getattr(practitioner.grade, 'name', '')
                
                practitioner.computed_grade_display = main_grade or ''
            except ImportError:
                # Si le module grades n'est pas disponible
                practitioner.computed_grade_display = getattr(practitioner, 'grade', '') or ''
                current_grades = {}
            except Exception as e:
                logger.warning(f"Erreur lors du chargement des grades pour le pratiquant {pk}: {str(e)}")
                practitioner.computed_grade_display = getattr(practitioner, 'grade', '') or ''
                current_grades = {}
            
            # Âge
            if practitioner.birth_date:
                from datetime import date
                today = date.today()
                age = today.year - practitioner.birth_date.year - ((today.month, today.day) < (practitioner.birth_date.month, practitioner.birth_date.day))
                practitioner.computed_age = age
            else:
                practitioner.computed_age = None
        except Exception as e:
            logger.warning(f"Erreur lors de l'enrichissement des données du pratiquant {pk}: {str(e)}")
            practitioner.computed_grade_display = ''
            practitioner.computed_age = None
            current_grades = {}
        
        # Récupérer l'historique des grades pour le nouveau template
        grade_history = []
        try:
            from apps.grades.models import PractitionerGrade
            history_grades = PractitionerGrade.objects.filter(
                practitioner=practitioner
            ).select_related('grade', 'discipline').order_by('-date_obtained')[:10]

            for pg in history_grades:
                grade_history.append({
                    'date': pg.date_obtained,
                    'grade_name': pg.grade.name if pg.grade else 'N/A',
                    'discipline_name': pg.discipline.name if pg.discipline else (pg.grade.discipline.name if pg.grade and pg.grade.discipline else 'N/A')
                })
        except Exception as e:
            logger.warning(f"Erreur lors du chargement de l'historique des grades: {str(e)}")

        # Récupérer les inscriptions aux compétitions
        registrations = []
        try:
            from apps.competitions.models import CompetitionRegistration
            registrations = CompetitionRegistration.objects.filter(
                practitioner=practitioner
            ).select_related('competition').prefetch_related('competition_types', 'categories').order_by('-created_at')[:5]
        except Exception as e:
            logger.warning(f"Erreur lors du chargement des inscriptions: {str(e)}")

        # Récupérer les données financières du pratiquant
        finance_data = None
        try:
            from apps.finances.services.practitioner_finance_service import PractitionerFinanceService
            finance_service = PractitionerFinanceService(practitioner)
            finance_data = finance_service.get_finance_summary()
            logger.info(f"Données financières chargées pour {practitioner.full_name}")
        except ImportError as e:
            logger.warning(f"Module PractitionerFinanceService non disponible: {str(e)}")
        except Exception as e:
            logger.warning(f"Erreur lors du chargement des données financières: {str(e)}")

        # Récupérer les méthodes de paiement disponibles
        payment_methods = []
        try:
            from apps.finances.models import PaymentMethod
            payment_methods = list(PaymentMethod.objects.filter(is_active=True).values('id', 'name', 'type'))
        except Exception as e:
            logger.warning(f"Erreur lors du chargement des méthodes de paiement: {str(e)}")

        # Récupérer les données du palmarès
        palmares_data = None
        try:
            from apps.competitions.services.palmares_service import PalmaresService
            palmares_service = PalmaresService(practitioner)
            palmares_data = palmares_service.get_palmares_summary()
            logger.info(f"Données palmarès chargées pour {practitioner.full_name}")
        except ImportError as e:
            logger.warning(f"Module PalmaresService non disponible: {str(e)}")
        except Exception as e:
            logger.warning(f"Erreur lors du chargement du palmarès: {str(e)}")

        # Récupérer les données des commandes
        orders_data = None
        try:
            from apps.shop.services.order_service import OrderService
            order_service = OrderService(practitioner)
            orders_data = order_service.get_orders_summary()
            logger.info(f"Données commandes chargées pour {practitioner.full_name}")
        except ImportError as e:
            logger.warning(f"Module OrderService non disponible: {str(e)}")
        except Exception as e:
            logger.warning(f"Erreur lors du chargement des commandes: {str(e)}")

        # Données pour la création de compte utilisateur (si pas de compte)
        suggested_username = None
        default_password = None
        custom_password = None
        existing_user_match = None  # Utilisateur existant qui pourrait correspondre

        if not practitioner.user:
            try:
                from apps.competitions.services.account_service import PractitionerAccountService
                from django.contrib.auth import get_user_model
                User = get_user_model()

                suggested_username = PractitionerAccountService.generate_username(
                    practitioner.first_name,
                    practitioner.last_name
                )
                default_password = PractitionerAccountService.DEFAULT_PASSWORD
                custom_password = PractitionerAccountService.generate_custom_password(practitioner)

                # Vérifier si un utilisateur existant correspond (par email ou username similaire)
                if practitioner.email:
                    existing_user_match = User.objects.filter(email__iexact=practitioner.email).first()

                # Si pas trouvé par email, chercher par username similaire
                if not existing_user_match:
                    # Chercher des variantes du nom
                    name_variations = [
                        f"{practitioner.first_name.lower()}.{practitioner.last_name.lower()}",
                        f"{practitioner.first_name.lower()}_{practitioner.last_name.lower()}",
                        f"{practitioner.first_name.lower()}{practitioner.last_name.lower()}",
                    ]
                    for variation in name_variations:
                        existing_user_match = User.objects.filter(username__iexact=variation).first()
                        if existing_user_match:
                            break

                if existing_user_match:
                    logger.info(f"Utilisateur existant trouvé pour {practitioner.full_name}: {existing_user_match.username}")
            except Exception as e:
                logger.warning(f"Erreur lors de la génération des données de compte: {str(e)}")

        # PROMPT 7: Récupérer les informations de juge du pratiquant
        judge_info = None
        try:
            if hasattr(practitioner, 'judge'):
                judge_info = practitioner.judge
        except Exception as e:
            logger.warning(f"Erreur lors du chargement des infos juge: {str(e)}")

        # Récupérer les rôles actuels du pratiquant dans l'organisation
        member_roles_data = None
        if practitioner.user and user_club:
            try:
                from apps.organizations.models import OrganizationMember, OrganizationRole
                org_member = OrganizationMember.objects.filter(
                    user=practitioner.user,
                    organization=user_club,
                    is_active=True,
                ).first()
                if org_member:
                    member_roles_data = {
                        'main_role': org_member.role,
                        'additional_roles': org_member.additional_roles or [],
                        'all_roles': org_member.all_roles,
                    }
                else:
                    member_roles_data = {
                        'main_role': 'member',
                        'additional_roles': [],
                        'all_roles': ['member'],
                    }
                # Liste complète des rôles disponibles (exclure 'owner' si l'utilisateur n'est pas superuser)
                available_roles = [
                    {'code': c[0], 'label': str(c[1])}
                    for c in OrganizationRole.choices
                    if c[0] != 'owner' or request.user.is_superuser
                ]
            except Exception as e:
                logger.warning(f"Erreur lors du chargement des rôles: {str(e)}")
                available_roles = []
        else:
            available_roles = []

        context = {
            'practitioner': practitioner,
            'club': user_club,
            'page_title': f"Profil de {practitioner.full_name}",
            'current_grades': current_grades,
            'grade_history': grade_history,
            'registrations': registrations,
            'finance_data': finance_data,
            'payment_methods': payment_methods,
            'palmares_data': palmares_data,
            'orders_data': orders_data,
            'suggested_username': suggested_username,
            'default_password': default_password,
            'custom_password': custom_password,
            'judge_info': judge_info,
            'existing_user_match': existing_user_match,
            'member_roles_data': member_roles_data,
            'available_roles': available_roles,
        }

        # Utiliser le nouveau template moderne
        return render(request, 'competitions/club/practitioner_detail_modern.html', context)
        
    except Exception as e:
        logger.error(f"Erreur dans practitioner_detail pour le pratiquant {pk}: {str(e)}", exc_info=True)
        messages.error(request, _("Erreur lors du chargement du profil."))
        return redirect('competitions:club:practitioners')

@login_required
def practitioner_create(request):
    """Créer un nouveau pratiquant"""
    try:
        logger.info(f"practitioner_create appelé par {request.user.username}")
        
        # Essayer d'abord de récupérer l'organisation depuis le middleware
        organization = None
        if hasattr(request, 'user_organization') and request.user_organization:
            organization = request.user_organization
            logger.info(f"Organisation trouvée via middleware: {organization}")
        else:
            # Sinon utiliser get_user_club
            user_club = get_user_club(request)
            logger.info(f"Club/Organisation trouvé via get_user_club: {user_club}")
            organization = user_club
        
        if not organization:
            # Dernière tentative : chercher via UserProfile
            from apps.competitions.models.users import UserProfile
            try:
                profile = UserProfile.objects.get(user=request.user)
                organization = profile.organization
                logger.info(f"Organisation trouvée via UserProfile: {organization}")
            except UserProfile.DoesNotExist:
                pass
        
        if not organization:
            messages.error(request, _("Vous n'êtes associé à aucune organisation. Contactez un administrateur."))
            return redirect('competitions:dashboard:dashboard')
        
        # Ne pas vérifier les permissions si l'utilisateur est superuser
        if not request.user.is_superuser:
            # Vérifier les permissions avec l'organisation
            if not manual_permission_check(request.user, organization):
                logger.warning(f"Permission refusée pour {request.user.username} sur l'organisation {organization}")
                # Au lieu de lever une exception, essayer de trouver une organisation valide
                from apps.competitions.models import Practitioner
                user_as_practitioner = Practitioner.objects.filter(
                    Q(user=request.user) | Q(email=request.user.email)
                ).select_related('organization').first()
                
                if user_as_practitioner and user_as_practitioner.organization:
                    organization = user_as_practitioner.organization
                    logger.info(f"Organisation alternative trouvée: {organization}")
                else:
                    messages.error(request, _("Vous n'avez pas l'autorisation de créer un pratiquant. Contactez un administrateur."))
                    return redirect('competitions:club:practitioners')
        
        # -- Vérification limite membres (Free tier) --
        from apps.finances.services.subscription_service import get_subscription_info
        sub_info = get_subscription_info(organization)
        member_limit_reached = sub_info.get('is_at_limit', False)

        if request.method == 'POST':
            # Blocage dur si limite atteinte
            if member_limit_reached:
                messages.error(request, _("Limite de membres atteinte (%(max)s). Passez au plan Premium pour ajouter plus de membres.") % {'max': sub_info.get('max_members', 10)})
                return redirect('upgrade_required')

            # S'assurer que request.POST est un QueryDict, pas une chaîne
            post_data = request.POST if hasattr(request.POST, 'get') else None
            files_data = request.FILES if hasattr(request.FILES, 'get') else None
            form = PractitionerForm(post_data, files_data, request=request)
            if form.is_valid():
                practitioner = form.save(commit=False)
                practitioner.organization = organization
                practitioner.save()

                # Sauvegarder les relations many-to-many (disciplines)
                form.save_m2m()

                # Sauvegarder les grades et données de juge
                # (non appelés automatiquement car commit=False ci-dessus)
                form._save_grades(practitioner)
                form._save_judge_data(practitioner)

                # Sync Stripe quantity if premium
                from apps.finances.services.subscription_service import sync_stripe_member_count
                sync_stripe_member_count(organization)

                messages.success(request, _(f"Le pratiquant {practitioner.full_name} a été créé avec succès."))
                return redirect('competitions:club:practitioners')
            else:
                # Log des erreurs pour debug
                logger.error(f"Erreurs du formulaire pour {request.user.username}: {form.errors}")
                for field, errors in form.errors.items():
                    for error in errors:
                        messages.error(request, f"{field}: {error}")
        else:
            # Pour GET, ne pas passer de data (None par défaut)
            form = PractitionerForm(None, request=request)
        
        context = {
            'form': form,
            'club': organization,  # Pour compatibilité avec le template
            'organization': organization,
            'page_title': _("Ajouter un Pratiquant"),
            'member_limit_reached': member_limit_reached,
            'subscription_info': sub_info,
        }

        return render(request, 'competitions/club/practitioner_form.html', context)
        
    except Exception as e:
        logger.error(f"Erreur dans practitioner_create: {str(e)}", exc_info=True)
        messages.error(request, _(f"Erreur lors de la création du pratiquant: {str(e)}"))
        return redirect('competitions:club:practitioners')

@login_required
def practitioner_update(request, pk):
    """Modifier un pratiquant"""
    try:
        # Récupérer le pratiquant sans filtre d'organisation pour éviter les erreurs
        try:
            practitioner = Practitioner.objects.get(pk=pk)
        except Practitioner.DoesNotExist:
            messages.error(request, _("Pratiquant introuvable."))
            return redirect('competitions:club:practitioners')
        
        # Vérification simplifiée des permissions
        user_club = get_user_club(request)
        if not request.user.is_superuser:
            # Si l'utilisateur n'a pas de club associé, il ne peut pas modifier les pratiquants
            if not user_club:
                messages.error(request, _("Vous n'avez pas l'autorisation de modifier ce pratiquant."))
                return redirect('competitions:dashboard:club')
            
            # Si le pratiquant n'a pas d'organisation, on autorise la modification
            # Sinon, vérifier que le pratiquant appartient au même club/organisation
            if practitioner.organization:
                if user_club:
                    practitioner_org_id = practitioner.organization.id if hasattr(practitioner.organization, 'id') else str(practitioner.organization)
                    user_club_id = user_club.id if hasattr(user_club, 'id') else str(user_club)
                    
                    if str(practitioner_org_id) != str(user_club_id):
                        # Vérifier si l'utilisateur est propriétaire d'un club
                        try:
                            from apps.competitions.models import Club
                            is_club_owner = Club.objects.filter(owner=request.user).exists()
                            if not is_club_owner:
                                messages.error(request, _("Vous n'avez pas l'autorisation de modifier ce pratiquant."))
                                return redirect('competitions:club:practitioners')
                        except Exception as e:
                            logger.warning(f"Erreur lors de la vérification des permissions: {str(e)}")
                            # En cas d'erreur, autoriser l'accès plutôt que de bloquer
                            pass
            else:
                # Si le pratiquant n'a pas d'organisation, autoriser l'accès
                logger.info(f"Pratiquant sans organisation - modification autorisée")
        
        if request.method == 'POST':
            # S'assurer que request.POST est un QueryDict
            form = PractitionerForm(request.POST, request.FILES, instance=practitioner, request=request)
            if form.is_valid():
                practitioner = form.save()
                messages.success(request, _(f"Le profil de {practitioner.full_name} a été mis à jour."))
                return redirect('competitions:club:practitioner_detail', pk=practitioner.pk)
            else:
                # Log des erreurs pour debug
                logger.error(f"Erreurs du formulaire lors de la modification: {form.errors}")
        else:
            # Pour GET, utiliser instance seulement
            form = PractitionerForm(instance=practitioner, request=request)
        
        # Charger les grades actuels pour le contexte
        current_grades = {}
        try:
            from apps.grades.models import PractitionerGrade
            practitioner_grades = PractitionerGrade.objects.filter(
                practitioner=practitioner,
                is_current=True
            ).select_related('grade', 'discipline', 'grade__discipline')
            
            for pg in practitioner_grades:
                if pg.discipline:
                    current_grades[pg.discipline.id] = pg
                elif pg.grade and pg.grade.discipline:
                    current_grades[pg.grade.discipline.id] = pg
        except (ImportError, Exception) as e:
            logger.warning(f"Erreur lors du chargement des grades pour l'édition: {str(e)}")
            current_grades = {}
        
        context = {
            'form': form,
            'practitioner': practitioner,
            'club': user_club,
            'organization': user_club,
            'page_title': f"Modifier - {practitioner.full_name}",
            'is_edit': True,
            'submit_text': _("Enregistrer les modifications"),
            'current_grades': current_grades,
        }
        
        return render(request, 'competitions/club/practitioner_form.html', context)
        
    except PermissionDenied:
        raise
    except Exception as e:
        logger.error(f"Erreur dans practitioner_update: {str(e)}", exc_info=True)
        messages.error(request, _(f"Erreur lors de la modification du pratiquant: {str(e)}"))
        return redirect('competitions:club:practitioners')

@login_required
@require_POST
def practitioner_delete(request, practitioner_id):
    """Supprimer un pratiquant"""
    # Détecter si c'est une requête AJAX/fetch
    is_ajax = (
        request.content_type == 'application/json' or
        request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    )

    try:
        user_club = get_user_club(request)
        if not user_club:
            if is_ajax:
                return JsonResponse({'success': False, 'error': _('Club non trouvé')})
            messages.error(request, _("Club non trouvé"))
            return redirect('competitions:club:practitioners')

        practitioner = get_object_or_404(Practitioner, id=practitioner_id, organization=user_club)

        if not manual_permission_check(request.user, user_club):
            if is_ajax:
                return JsonResponse({'success': False, 'error': _('Permissions insuffisantes')})
            messages.error(request, _("Permissions insuffisantes"))
            return redirect('competitions:club:practitioners')

        practitioner_name = practitioner.full_name
        practitioner.delete()

        # Pour les requêtes AJAX/fetch, retourner JSON
        if is_ajax:
            return JsonResponse({
                'success': True,
                'message': f"Le pratiquant {practitioner_name} a été supprimé."
            })

        # Pour les requêtes normales (formulaire), rediriger
        messages.success(request, f"Le pratiquant {practitioner_name} a été supprimé.")
        return redirect('competitions:club:practitioners')

    except Exception as e:
        logger.error(f"Erreur lors de la suppression du pratiquant {practitioner_id}: {str(e)}")
        if is_ajax:
            return JsonResponse({'success': False, 'error': str(e)})
        messages.error(request, _("Erreur lors de la suppression."))
        return redirect('competitions:club:practitioners')

@login_required
@require_POST
def practitioner_toggle_status(request, pk):
    """Activer/Désactiver un pratiquant"""
    # Détecter si c'est une requête AJAX
    is_ajax = (
        request.content_type == 'application/json' or
        request.headers.get('X-Requested-With') == 'XMLHttpRequest' or
        request.headers.get('X-CSRFToken')
    )

    try:
        user_club = get_user_club(request)
        if not user_club:
            if is_ajax:
                return JsonResponse({'success': False, 'error': _('Club non trouvé')})
            messages.error(request, _("Club non trouvé"))
            return redirect('competitions:club:practitioners')

        practitioner = get_object_or_404(Practitioner, id=pk)

        # Vérification des permissions
        if practitioner.organization and user_club:
            practitioner_org_id = practitioner.organization.id if hasattr(practitioner.organization, 'id') else str(practitioner.organization)
            user_club_id = user_club.id if hasattr(user_club, 'id') else str(user_club)

            if str(practitioner_org_id) != str(user_club_id):
                if is_ajax:
                    return JsonResponse({'success': False, 'error': _('Permissions insuffisantes')})
                messages.error(request, _("Permissions insuffisantes"))
                return redirect('competitions:club:practitioners')

        # Récupérer le nouveau statut depuis le body de la requête (JSON ou form-data)
        new_status = None
        if request.content_type == 'application/json':
            import json
            try:
                body = json.loads(request.body)
                new_status = body.get('is_active')
            except:
                pass
        else:
            # Form data
            is_active_str = request.POST.get('is_active', '')
            if is_active_str.lower() in ('true', '1', 'on', 'yes'):
                new_status = True
            elif is_active_str.lower() in ('false', '0', 'off', 'no'):
                new_status = False

        # Si pas de valeur fournie, inverser l'état actuel
        if new_status is None:
            new_status = not practitioner.is_active

        # Mettre à jour le statut
        practitioner.is_active = new_status
        practitioner.save()

        message = f"Le pratiquant {practitioner.full_name} a été {'activé' if new_status else 'désactivé'}."
        messages.success(request, message)

        if is_ajax:
            return JsonResponse({
                'success': True,
                'message': message,
                'is_active': practitioner.is_active
            })

        return redirect('competitions:club:practitioners')

    except Exception as e:
        logger.error(f"Erreur lors du changement de statut du pratiquant {pk}: {str(e)}")
        if is_ajax:
            return JsonResponse({'success': False, 'error': str(e)})
        messages.error(request, _("Erreur lors du changement de statut."))
        return redirect('competitions:club:practitioners')


# PROMPT 7: Vue pour attribuer le rôle de juge à un pratiquant
@login_required
@require_POST
def assign_judge_role(request, practitioner_id):
    """
    Attribuer le rôle de juge/arbitre à un pratiquant.
    Crée un profil Judge associé au pratiquant.
    """
    from apps.competitions.models.judges import Judge

    try:
        user_club = get_user_club(request)
        if not user_club:
            messages.error(request, _("Vous n'êtes associé à aucune organisation."))
            return redirect('competitions:dashboard:club')

        # Récupérer le pratiquant
        practitioner = get_object_or_404(Practitioner, pk=practitioner_id)

        # Vérifier les permissions (soit même organisation, soit superuser)
        if not request.user.is_superuser:
            if practitioner.organization and practitioner.organization != user_club:
                messages.error(request, _("Vous n'avez pas l'autorisation de modifier ce pratiquant."))
                return redirect('competitions:club:practitioner_detail', pk=practitioner_id)

        # Vérifier si le pratiquant a déjà un profil juge
        if hasattr(practitioner, 'judge') and practitioner.judge:
            messages.warning(request, _("Ce pratiquant possède déjà un profil Juge/Arbitre."))
            return redirect('competitions:club:practitioner_detail', pk=practitioner_id)

        # Récupérer les données du formulaire
        is_technical_judge = request.POST.get('is_technical_judge') == 'on'
        is_combat_referee = request.POST.get('is_combat_referee') == 'on'
        qualification_level = request.POST.get('qualification_level', 'novice')
        years_experience = request.POST.get('years_experience', 0)
        certification_number = request.POST.get('certification_number', '').strip()
        certified_since = request.POST.get('certified_since') or None
        notes = request.POST.get('notes', '').strip()

        # Créer le profil Judge
        judge = Judge.objects.create(
            practitioner=practitioner,
            user=practitioner.user,  # Peut être None si pas de compte
            is_technical_judge=is_technical_judge,
            is_combat_referee=is_combat_referee,
            qualification_level=qualification_level,
            years_experience=int(years_experience) if years_experience else 0,
            certification_number=certification_number,
            certified_since=certified_since,
            notes=notes,
            organization=user_club,
            active=True
        )

        logger.info(f"Profil Juge créé pour {practitioner.full_name} par {request.user.username}")
        messages.success(request, _("Le rôle Juge/Arbitre a été attribué avec succès à {}.").format(practitioner.full_name))

        return redirect('competitions:club:practitioner_detail', pk=practitioner_id)

    except Exception as e:
        logger.error(f"Erreur lors de l'attribution du rôle juge pour pratiquant {practitioner_id}: {str(e)}", exc_info=True)
        messages.error(request, _("Erreur lors de l'attribution du rôle: {}").format(str(e)))
        return redirect('competitions:club:practitioner_detail', pk=practitioner_id)


# Fonctions placeholder pour compatibilité
@login_required
def create_user_for_practitioner(request, practitioner_id):
    """
    Créer un compte utilisateur pour un pratiquant.
    Supporte AJAX pour vérification en temps réel du username.

    GET avec ?check_username=xxx : Vérifie la disponibilité du username (AJAX)
    GET avec ?generate_username=1 : Génère un nouveau username (AJAX)
    POST : Crée le compte utilisateur
    """
    from apps.competitions.services.account_service import PractitionerAccountService

    try:
        user_club = get_user_club(request)
        if not user_club:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'error': _("Vous n'êtes associé à aucun club.")}, status=403)
            messages.error(request, _("Vous n'êtes associé à aucun club."))
            return redirect('competitions:dashboard:club')

        practitioner = get_object_or_404(Practitioner, pk=practitioner_id, organization=user_club)

        if not manual_permission_check(request.user, user_club):
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'error': _("Permission refusée")}, status=403)
            raise PermissionDenied(_("Vous n'avez pas l'autorisation de modifier ce pratiquant."))

        # === REQUÊTES AJAX ===

        # Vérifier la disponibilité d'un username (AJAX)
        if request.GET.get('check_username'):
            username = request.GET.get('check_username', '').strip().lower()
            result = PractitionerAccountService.check_username_available(username)
            return JsonResponse(result)

        # Générer un nouveau username (AJAX)
        if request.GET.get('generate_username'):
            username = PractitionerAccountService.generate_username(
                practitioner.first_name,
                practitioner.last_name
            )
            return JsonResponse({
                'username': username,
                'available': True
            })

        # Générer un mot de passe personnalisé (AJAX)
        if request.GET.get('generate_password'):
            password = PractitionerAccountService.generate_custom_password(practitioner)
            return JsonResponse({'password': password})

        # === VÉRIFICATIONS PRÉALABLES ===

        # Vérifier si le pratiquant a déjà un compte
        if practitioner.user:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'error': _("Ce pratiquant possède déjà un compte utilisateur.")
                }, status=400)
            messages.warning(request, _("Ce pratiquant a déjà un compte utilisateur associé."))
            return redirect('competitions:club:practitioner_detail', pk=practitioner_id)

        # === CRÉATION DU COMPTE (POST) ===

        if request.method == 'POST':
            # Récupérer les données du formulaire
            if request.content_type == 'application/json':
                try:
                    data = json.loads(request.body)
                except json.JSONDecodeError:
                    data = {}
            else:
                data = request.POST

            username = data.get('username', '').strip().lower()
            password = data.get('password', '').strip()
            send_email = data.get('send_email', 'true') in ['true', 'True', '1', True, 'on']
            use_custom_password = data.get('use_custom_password', 'false') in ['true', 'True', '1', True, 'on']

            # Utiliser le mot de passe par défaut ou personnalisé
            if not password or not use_custom_password:
                password = PractitionerAccountService.DEFAULT_PASSWORD

            # Créer le compte
            result = PractitionerAccountService.create_account(
                practitioner=practitioner,
                username=username,
                password=password,
                send_email=send_email,
                created_by=request.user
            )

            # Réponse AJAX
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                if result['success']:
                    return JsonResponse({
                        'success': True,
                        'message': str(result['message']),
                        'username': username,
                        'password': password,
                        'email_sent': result.get('email_sent', False),
                        'redirect_url': reverse('competitions:club:practitioner_detail', kwargs={'pk': practitioner_id})
                    })
                else:
                    return JsonResponse({
                        'success': False,
                        'error': str(result['message'])
                    }, status=400)

            # Réponse classique
            if result['success']:
                if result.get('email_sent'):
                    messages.success(
                        request,
                        _("Compte créé avec succès ! Un email avec les identifiants a été envoyé à {email}").format(
                            email=practitioner.email
                        )
                    )
                else:
                    messages.success(
                        request,
                        _("Compte créé avec succès ! Nom d'utilisateur : {username} - Mot de passe : {password}").format(
                            username=username,
                            password=password
                        )
                    )
            else:
                messages.error(request, result['message'])

            return redirect('competitions:club:practitioner_detail', pk=practitioner_id)

        # === AFFICHAGE DU FORMULAIRE (GET) ===

        # Générer le username suggéré
        suggested_username = PractitionerAccountService.generate_username(
            practitioner.first_name,
            practitioner.last_name
        )

        # Générer le mot de passe personnalisé suggéré
        custom_password = PractitionerAccountService.generate_custom_password(practitioner)

        context = {
            'practitioner': practitioner,
            'title': _("Créer un compte pour {name}").format(name=practitioner.full_name),
            'club': user_club,
            'suggested_username': suggested_username,
            'default_password': PractitionerAccountService.DEFAULT_PASSWORD,
            'custom_password': custom_password,
        }

        return render(request, 'competitions/club/create_user_form.html', context)

    except PermissionDenied:
        raise
    except Exception as e:
        logger.error(f"Erreur dans create_user_for_practitioner: {str(e)}", exc_info=True)
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'error': str(e)}, status=500)
        messages.error(request, _("Erreur lors de l'accès à cette page."))
        return redirect('competitions:club:practitioners')

@login_required
def link_user_to_practitioner(request, practitioner_id):
    """Lier un utilisateur existant à un pratiquant"""
    try:
        user_club = get_user_club(request)
        if not user_club:
            messages.error(request, _("Vous n'êtes associé à aucun club."))
            return redirect('competitions:dashboard:club')
        
        practitioner = get_object_or_404(Practitioner, pk=practitioner_id, organization=user_club)
        
        if not manual_permission_check(request.user, user_club):
            raise PermissionDenied(_("Vous n'avez pas l'autorisation de modifier ce pratiquant."))
        
        # Vérifier si le pratiquant a déjà un compte
        if practitioner.user:
            messages.warning(request, _("Ce pratiquant a déjà un compte utilisateur associé."))
            return redirect('competitions:club:practitioner_detail', pk=practitioner_id)
        
        if request.method == 'POST':
            user_id = request.POST.get('user_id')
            if user_id:
                try:
                    user = User.objects.get(pk=user_id)
                    
                    # Vérifier que l'utilisateur n'est pas déjà lié à un autre pratiquant
                    existing_practitioner = Practitioner.objects.filter(user=user, organization=user_club).first()
                    if existing_practitioner:
                        messages.error(
                            request,
                            _("Cet utilisateur est déjà associé à {name}.").format(name=existing_practitioner.full_name)
                        )
                    else:
                        practitioner.user = user
                        practitioner.save()
                        messages.success(
                            request,
                            _("L'utilisateur {username} a été associé avec succès à {name}.").format(
                                username=user.username,
                                name=practitioner.full_name
                            )
                        )
                        return redirect('competitions:club:practitioner_detail', pk=practitioner_id)
                        
                except User.DoesNotExist:
                    messages.error(request, _("Utilisateur introuvable."))
            else:
                messages.error(request, _("Veuillez sélectionner un utilisateur."))
        
        # Obtenir la liste des utilisateurs disponibles (sans pratiquant associé dans ce club)
        linked_user_ids = Practitioner.objects.filter(
            organization=user_club,
            user__isnull=False
        ).values_list('user_id', flat=True)
        
        available_users = User.objects.exclude(
            id__in=linked_user_ids
        ).order_by('last_name', 'first_name')
        
        context = {
            'practitioner': practitioner,
            'available_users': available_users,
            'title': _("Associer un compte à {name}").format(name=practitioner.full_name),
            'club': user_club,
        }
        
        return render(request, 'competitions/club/link_user_form.html', context)
        
    except PermissionDenied:
        raise
    except Exception as e:
        logger.error(f"Erreur dans link_user_to_practitioner: {str(e)}", exc_info=True)
        messages.error(request, _("Erreur lors de l'accès à cette page."))
        return redirect('competitions:club:practitioners')


@login_required
@require_POST
def send_practitioner_credentials(request, practitioner_id):
    """Envoyer (ou renvoyer) les identifiants de connexion par email à un pratiquant."""
    try:
        user_club = get_user_club(request)
        if not user_club:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'error': str(_("Vous n'êtes associé à aucun club."))}, status=403)
            messages.error(request, _("Vous n'êtes associé à aucun club."))
            return redirect('competitions:dashboard:club')

        practitioner = get_object_or_404(Practitioner, pk=practitioner_id, organization=user_club)

        if not manual_permission_check(request.user, user_club):
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'error': str(_("Permission refusée."))}, status=403)
            raise PermissionDenied

        # Vérifier que le pratiquant a un email valide
        if not practitioner.email or '@' not in practitioner.email or practitioner.email.endswith('@example.com'):
            msg = _("Ce pratiquant n'a pas d'adresse email valide. Veuillez d'abord renseigner son email.")
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'error': str(msg)}, status=400)
            messages.error(request, msg)
            return redirect('competitions:club:practitioner_detail', pk=practitioner_id)

        from apps.competitions.services.account_service import PractitionerAccountService

        # Si le pratiquant n'a pas encore de compte, en créer un
        if not practitioner.user:
            username = PractitionerAccountService.generate_username(
                practitioner.first_name, practitioner.last_name
            )
            result = PractitionerAccountService.create_account(
                practitioner=practitioner,
                username=username,
                password=None,  # utilise DEFAULT_PASSWORD
                send_email=True,
                created_by=request.user,
            )
            if result['success']:
                msg = _("Compte créé et identifiants envoyés par email à {email}.").format(email=practitioner.email)
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': True,
                        'message': str(msg),
                        'account_created': True,
                        'username': result['user'].username,
                        'email_sent': result.get('email_sent', False),
                    })
                messages.success(request, msg)
            else:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'error': str(result['message'])}, status=400)
                messages.error(request, result['message'])
            return redirect('competitions:club:practitioner_detail', pk=practitioner_id)

        # Le pratiquant a déjà un compte – réinitialiser le mot de passe et envoyer par email
        user = practitioner.user
        new_password = PractitionerAccountService.DEFAULT_PASSWORD
        user.set_password(new_password)
        user.save()

        # Mettre à jour l'email du User si nécessaire
        if user.email != practitioner.email:
            user.email = practitioner.email
            user.save(update_fields=['email'])

        email_sent = PractitionerAccountService.send_credentials_email(user, new_password, practitioner)

        if email_sent:
            msg = _("Identifiants envoyés par email à {email} (mot de passe réinitialisé).").format(email=practitioner.email)
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': str(msg),
                    'account_created': False,
                    'username': user.username,
                    'email_sent': True,
                })
            messages.success(request, msg)
        else:
            msg = _("Erreur lors de l'envoi de l'email. Vérifiez la configuration email du serveur.")
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'error': str(msg)}, status=500)
            messages.error(request, msg)

        return redirect('competitions:club:practitioner_detail', pk=practitioner_id)

    except PermissionDenied:
        raise
    except Exception as e:
        logger.error(f"Erreur dans send_practitioner_credentials: {str(e)}", exc_info=True)
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'error': str(e)}, status=500)
        messages.error(request, _("Erreur lors de l'envoi des identifiants."))
        return redirect('competitions:club:practitioner_detail', pk=practitioner_id)


@login_required
def practitioner_qualifications_add(request, practitioner_id):
    """Ajouter une qualification à un pratiquant"""
    try:
        user_club = get_user_club(request)
        if not user_club:
            messages.error(request, _("Vous n'êtes associé à aucun club."))
            return redirect('competitions:dashboard:club')
        
        practitioner = get_object_or_404(Practitioner, pk=practitioner_id, organization=user_club)
        
        if not manual_permission_check(request.user, user_club):
            raise PermissionDenied(_("Vous n'avez pas l'autorisation de modifier ce pratiquant."))
        
        # Import du formulaire de qualification
        from apps.competitions.forms.qualification import QualificationForm
        
        if request.method == 'POST':
            form = QualificationForm(request.POST, practitioner=practitioner)
            if form.is_valid():
                qualification = form.save(commit=False)
                qualification.practitioner = practitioner
                qualification.save()
                messages.success(request, _("La qualification a été ajoutée avec succès."))
                from django.urls import reverse
                url = reverse('competitions:club:practitioner_detail', kwargs={'pk': practitioner.pk})
                return redirect(url + '?tab=qualifications')
        else:
            form = QualificationForm(practitioner=practitioner)
        
        context = {
            'form': form,
            'practitioner': practitioner,
            'page_title': f"Ajouter une qualification - {practitioner.full_name}",
        }
        
        return render(request, 'competitions/club/qualification_form.html', context)
        
    except PermissionDenied:
        raise
    except Exception as e:
        logger.error(f"Erreur dans practitioner_qualifications_add: {str(e)}", exc_info=True)
        messages.error(request, _("Erreur lors de l'ajout de la qualification."))
        return redirect('competitions:club:practitioners')

def practitioner_registrations(request, practitioner_id):
    """Voir les inscriptions d'un pratiquant"""
    messages.info(request, _("Fonctionnalité en cours de développement."))
    return redirect('competitions:club:practitioners')

def create_practitioner_registration(request, practitioner_id):
    """Créer une inscription pour un pratiquant"""
    messages.info(request, _("Fonctionnalité en cours de développement."))
    return redirect('competitions:club:practitioners')

def get_available_competitions(request, practitioner_id):
    """Obtenir les compétitions disponibles pour un pratiquant"""
    messages.info(request, _("Fonctionnalité en cours de développement."))
    return redirect('competitions:club:practitioners')


# ===== NOUVELLE VUE AJAX POUR LA RECHERCHE DE PRATIQUANTS =====

@login_required
@require_http_methods(["GET"])
def practitioner_search_ajax(request):
    """
    Vue AJAX pour la recherche et le filtrage dynamique des pratiquants.
    Retourne les données JSON pour mise à jour sans rechargement de page.

    Paramètres GET:
    - search: Texte de recherche (nom, prénom, email, licence)
    - discipline: ID de discipline (multi-select, séparés par virgule)
    - grade: ID ou identifiant du grade
    - age_group: Catégorie d'âge (poussins, benjamins, minimes, cadets, juniors, seniors, veterans)
    - status: Statut (active, inactive, all)
    - gender: Genre (male, female, all)
    - license: Filtre licence (all, with_license, without_license)
    - has_account: Filtre compte (all, with_account, without_account)
    - sort: Colonne de tri (name, age, grade, registration_date)
    - order: Ordre de tri (asc, desc)
    - page: Numéro de page
    """
    try:
        # Récupération de l'organisation de l'utilisateur
        user_club = get_user_club(request)

        if not user_club:
            return JsonResponse({
                'success': False,
                'error': str(_("Vous n'êtes associé à aucun club."))
            }, status=403)

        # Vérification des permissions
        if not manual_permission_check(request.user, user_club):
            return JsonResponse({
                'success': False,
                'error': str(_("Vous n'avez pas l'autorisation d'accéder à ces données."))
            }, status=403)

        # Base queryset: TOUS les pratiquants de l'organisation
        queryset = Practitioner.objects.filter(
            organization=user_club
        ).select_related(
            'organization', 'primary_discipline', 'user', 'grade'
        ).prefetch_related('disciplines')

        # Compteur total avant filtres
        total_count = queryset.count()

        # === FILTRE: Recherche textuelle ===
        search = request.GET.get('search', '').strip()
        if search:
            queryset = queryset.filter(
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(email__icontains=search) |
                Q(license_number__icontains=search)
            )

        # === FILTRE: Disciplines (multi-select) ===
        disciplines = request.GET.get('discipline', '').strip()
        if disciplines:
            discipline_ids = [int(d) for d in disciplines.split(',') if d.isdigit()]
            if discipline_ids:
                queryset = queryset.filter(
                    Q(primary_discipline_id__in=discipline_ids) |
                    Q(disciplines__id__in=discipline_ids)
                ).distinct()

        # === FILTRE: Grade ===
        grade_filter = request.GET.get('grade', '').strip()
        if grade_filter:
            if grade_filter == 'none':
                queryset = queryset.filter(
                    Q(grade__isnull=True) &
                    (Q(grade_text__isnull=True) | Q(grade_text=''))
                )
            else:
                # Support ID numérique ou identifiant texte
                if grade_filter.isdigit():
                    queryset = queryset.filter(grade_id=int(grade_filter))
                else:
                    queryset = queryset.filter(
                        Q(grade__identifier=grade_filter) |
                        Q(grade_text__iexact=grade_filter)
                    )

        # === FILTRE: Catégorie d'âge ===
        age_group = request.GET.get('age_group', '').strip()
        if age_group:
            today = timezone.now().date()
            age_ranges = {
                'poussins': (0, 9),       # < 10 ans
                'benjamins': (10, 11),    # 10-11 ans
                'minimes': (12, 13),      # 12-13 ans
                'cadets': (14, 15),       # 14-15 ans
                'juniors': (16, 17),      # 16-17 ans
                'seniors': (18, 34),      # 18-34 ans
                'veterans': (35, 150),    # 35+ ans
            }

            if age_group in age_ranges:
                min_age, max_age = age_ranges[age_group]
                # Date de naissance max = aujourd'hui - min_age ans
                max_birth_date = today.replace(year=today.year - min_age)
                # Date de naissance min = aujourd'hui - (max_age + 1) ans
                min_birth_date = today.replace(year=today.year - max_age - 1)

                queryset = queryset.filter(
                    birth_date__gt=min_birth_date,
                    birth_date__lte=max_birth_date
                )

        # === FILTRE: Statut ===
        status = request.GET.get('status', '').strip()
        if status == 'active':
            queryset = queryset.filter(Q(is_active=True) | Q(status='active'))
        elif status == 'inactive':
            queryset = queryset.filter(Q(is_active=False) | Q(status='inactive'))
        elif status == 'suspended':
            queryset = queryset.filter(status='suspended')
        elif status == 'archived':
            queryset = queryset.filter(status='archived')
        # 'all' = pas de filtre

        # === FILTRE: Genre ===
        gender = request.GET.get('gender', '').strip()
        if gender in ['male', 'female', 'other']:
            queryset = queryset.filter(gender=gender)

        # === FILTRE: Licence ===
        license_filter = request.GET.get('license', '').strip()
        if license_filter == 'with_license':
            queryset = queryset.exclude(
                Q(license_number__isnull=True) | Q(license_number='')
            )
        elif license_filter == 'without_license':
            queryset = queryset.filter(
                Q(license_number__isnull=True) | Q(license_number='')
            )

        # === FILTRE: Compte utilisateur ===
        has_account = request.GET.get('has_account', '').strip()
        if has_account == 'with_account':
            queryset = queryset.filter(user__isnull=False)
        elif has_account == 'without_account':
            queryset = queryset.filter(user__isnull=True)

        # === TRI ===
        sort_column = request.GET.get('sort', 'name').strip()
        sort_order = request.GET.get('order', 'asc').strip()

        order_prefix = '-' if sort_order == 'desc' else ''

        sort_mapping = {
            'name': f'{order_prefix}last_name',
            'age': f'{order_prefix}birth_date' if sort_order == 'asc' else f'birth_date',  # Inverser pour l'âge
            'grade': f'{order_prefix}grade__level' if GRADES_MODELS_AVAILABLE else f'{order_prefix}grade_text',
            'registration_date': f'{order_prefix}registration_date',
        }

        order_by = sort_mapping.get(sort_column, 'last_name')

        # Tri secondaire par nom pour consistance
        if sort_column != 'name':
            queryset = queryset.order_by(order_by, 'last_name', 'first_name')
        else:
            queryset = queryset.order_by(order_by, 'first_name')

        # Compteur après filtres
        filtered_count = queryset.count()

        # === PAGINATION ===
        page = request.GET.get('page', '1')
        try:
            page_number = int(page)
        except ValueError:
            page_number = 1

        per_page = 20  # Nombre par page
        paginator = Paginator(queryset, per_page)

        try:
            practitioners_page = paginator.page(page_number)
        except:
            practitioners_page = paginator.page(1)

        # === CONSTRUCTION DE LA RÉPONSE ===
        practitioners_data = []
        today = timezone.now().date()

        for practitioner in practitioners_page:
            # Calcul de l'âge
            age = None
            if practitioner.birth_date:
                age = (today - practitioner.birth_date).days // 365

            # Grade display
            grade_display = get_grade_display(practitioner)
            grade_css = get_grade_css_class(practitioner)

            # Grade image et couleur pour le template tag
            grade_image_url = None
            grade_color = '#6c757d'
            if practitioner.grade:
                grade_image_url = practitioner.grade.image_url
                grade_color = practitioner.grade.get_display_color()

            # Disciplines
            disciplines_list = list(practitioner.disciplines.all())
            if disciplines_list:
                discipline_display = ', '.join([d.name for d in disciplines_list])
            elif practitioner.primary_discipline:
                discipline_display = practitioner.primary_discipline.name
            else:
                discipline_display = None

            # Status
            status_value = practitioner.status if hasattr(practitioner, 'status') and practitioner.status else ('active' if practitioner.is_active else 'inactive')

            practitioners_data.append({
                'id': practitioner.id,
                'first_name': practitioner.first_name,
                'last_name': practitioner.last_name,
                'full_name': f"{practitioner.first_name} {practitioner.last_name}",
                'email': practitioner.email or '',
                'phone': practitioner.phone or '',
                'license_number': practitioner.license_number or '',
                'birth_date': practitioner.birth_date.strftime('%d/%m/%Y') if practitioner.birth_date else '',
                'age': age,
                'gender': practitioner.gender or '',
                'grade_display': grade_display or '',
                'grade_css_class': grade_css,
                'grade_image_url': grade_image_url,
                'grade_color': grade_color,
                'discipline_display': discipline_display or '',
                'status': status_value,
                'is_active': practitioner.is_active,
                'has_account': practitioner.user is not None,
                'has_license': bool(practitioner.license_number),
                'photo_url': practitioner.photo.url if practitioner.photo else None,
                'initials': f"{practitioner.first_name[0] if practitioner.first_name else ''}{practitioner.last_name[0] if practitioner.last_name else ''}".upper(),
                'detail_url': reverse('competitions:club:practitioner_detail', kwargs={'pk': practitioner.pk}),
                'edit_url': reverse('competitions:club:practitioner_edit', kwargs={'pk': practitioner.pk}),
                'registrations_url': reverse('competitions:club:practitioner_registrations', kwargs={'practitioner_id': practitioner.pk}),
            })

        # Informations de pagination
        pagination_info = {
            'current_page': practitioners_page.number,
            'total_pages': paginator.num_pages,
            'has_next': practitioners_page.has_next(),
            'has_previous': practitioners_page.has_previous(),
            'next_page': practitioners_page.next_page_number() if practitioners_page.has_next() else None,
            'previous_page': practitioners_page.previous_page_number() if practitioners_page.has_previous() else None,
            'start_index': practitioners_page.start_index(),
            'end_index': practitioners_page.end_index(),
        }

        return JsonResponse({
            'success': True,
            'practitioners': practitioners_data,
            'pagination': pagination_info,
            'counts': {
                'total': total_count,
                'filtered': filtered_count,
            },
            'filters_applied': {
                'search': search,
                'disciplines': disciplines,
                'grade': grade_filter,
                'age_group': age_group,
                'status': status,
                'gender': gender,
                'license': license_filter,
                'has_account': has_account,
            },
            'sort': {
                'column': sort_column,
                'order': sort_order,
            }
        })

    except Exception as e:
        logger.error(f"Erreur dans practitioner_search_ajax: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': str(_("Une erreur est survenue lors de la recherche."))
        }, status=500)


@login_required
@require_http_methods(["GET"])
def practitioner_filter_options(request):
    """
    Retourne les options disponibles pour les filtres (disciplines, grades).
    Utilisé pour initialiser les dropdowns de filtres.
    """
    try:
        user_club = get_user_club(request)

        if not user_club:
            return JsonResponse({
                'success': False,
                'error': str(_("Vous n'êtes associé à aucun club."))
            }, status=403)

        # Récupérer les disciplines de l'organisation
        discipline_choices = get_discipline_choices(user_club)
        disciplines = [{'id': d[0], 'name': d[1]} for d in discipline_choices]

        # Récupérer les grades disponibles
        grade_choices = get_grade_choices(user_club)
        grades = [{'id': g[0], 'name': g[1]} for g in grade_choices]

        # Catégories d'âge avec traductions
        age_groups = [
            {'id': 'poussins', 'name': str(_('Poussins (<10 ans)'))},
            {'id': 'benjamins', 'name': str(_('Benjamins (10-11 ans)'))},
            {'id': 'minimes', 'name': str(_('Minimes (12-13 ans)'))},
            {'id': 'cadets', 'name': str(_('Cadets (14-15 ans)'))},
            {'id': 'juniors', 'name': str(_('Juniors (16-17 ans)'))},
            {'id': 'seniors', 'name': str(_('Seniors (18-34 ans)'))},
            {'id': 'veterans', 'name': str(_('Vétérans (35+ ans)'))},
        ]

        # Options de statut
        statuses = [
            {'id': 'all', 'name': str(_('Tous'))},
            {'id': 'active', 'name': str(_('Actifs'))},
            {'id': 'inactive', 'name': str(_('Inactifs'))},
            {'id': 'suspended', 'name': str(_('Suspendus'))},
            {'id': 'archived', 'name': str(_('Archivés'))},
        ]

        # Options de genre
        genders = [
            {'id': 'all', 'name': str(_('Tous'))},
            {'id': 'male', 'name': str(_('Masculin'))},
            {'id': 'female', 'name': str(_('Féminin'))},
        ]

        # Options de licence
        license_options = [
            {'id': 'all', 'name': str(_('Tous'))},
            {'id': 'with_license', 'name': str(_('Avec licence'))},
            {'id': 'without_license', 'name': str(_('Sans licence'))},
        ]

        # Options de compte utilisateur
        account_options = [
            {'id': 'all', 'name': str(_('Tous'))},
            {'id': 'with_account', 'name': str(_('Avec compte'))},
            {'id': 'without_account', 'name': str(_('Sans compte'))},
        ]

        return JsonResponse({
            'success': True,
            'options': {
                'disciplines': disciplines,
                'grades': grades,
                'age_groups': age_groups,
                'statuses': statuses,
                'genders': genders,
                'license_options': license_options,
                'account_options': account_options,
            }
        })

    except Exception as e:
        logger.error(f"Erreur dans practitioner_filter_options: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': str(_("Une erreur est survenue."))
        }, status=500)


@login_required
@require_POST
def practitioner_record_payment(request, pk):
    """
    Enregistrer un paiement pour un pratiquant via AJAX.
    """
    try:
        practitioner = get_object_or_404(Practitioner, pk=pk)

        # Vérification des permissions
        user_club = get_user_club(request)
        if not request.user.is_superuser and user_club:
            if practitioner.organization and str(practitioner.organization.id) != str(user_club.id):
                return JsonResponse({
                    'success': False,
                    'message': str(_("Vous n'avez pas l'autorisation d'effectuer cette action."))
                }, status=403)

        # Récupérer les données du formulaire
        amount = request.POST.get('amount')
        payment_method_id = request.POST.get('payment_method')
        description = request.POST.get('description', '')

        if not amount:
            return JsonResponse({
                'success': False,
                'message': str(_("Le montant est requis."))
            }, status=400)

        try:
            amount = float(amount)
            if amount <= 0:
                raise ValueError("Montant invalide")
        except ValueError:
            return JsonResponse({
                'success': False,
                'message': str(_("Montant invalide."))
            }, status=400)

        # Utiliser le service pour enregistrer le paiement
        from apps.finances.services.practitioner_finance_service import PractitionerFinanceService

        transaction = PractitionerFinanceService.record_payment(
            practitioner=practitioner,
            amount=amount,
            payment_method_id=payment_method_id,
            description=description,
            user=request.user
        )

        if transaction:
            logger.info(f"Paiement enregistré pour {practitioner.full_name}: {amount} EUR par {request.user.username}")
            return JsonResponse({
                'success': True,
                'message': str(_("Paiement enregistré avec succès.")),
                'transaction_id': str(transaction.id),
                'reference': transaction.reference
            })
        else:
            return JsonResponse({
                'success': False,
                'message': str(_("Erreur lors de l'enregistrement du paiement."))
            }, status=500)

    except Exception as e:
        logger.error(f"Erreur lors de l'enregistrement du paiement pour le pratiquant {pk}: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'message': str(_("Une erreur est survenue."))
        }, status=500)


@login_required
@require_POST
def practitioner_send_reminder(request, pk):
    """
    Envoyer un rappel de paiement à un pratiquant via AJAX.
    """
    try:
        practitioner = get_object_or_404(Practitioner, pk=pk)

        # Vérification des permissions
        user_club = get_user_club(request)
        if not request.user.is_superuser and user_club:
            if practitioner.organization and str(practitioner.organization.id) != str(user_club.id):
                return JsonResponse({
                    'success': False,
                    'message': str(_("Vous n'avez pas l'autorisation d'effectuer cette action."))
                }, status=403)

        # Vérifier que le pratiquant a un email
        if not practitioner.email:
            return JsonResponse({
                'success': False,
                'message': str(_("Ce pratiquant n'a pas d'adresse email."))
            }, status=400)

        # Récupérer les données du formulaire
        reminder_type = request.POST.get('reminder_type', 'payment_due')
        custom_message = request.POST.get('custom_message', '')

        # Utiliser le service pour envoyer le rappel
        from apps.finances.services.practitioner_finance_service import PractitionerFinanceService

        success = PractitionerFinanceService.send_payment_reminder(
            practitioner=practitioner,
            user=request.user
        )

        if success:
            logger.info(f"Rappel envoyé à {practitioner.email} par {request.user.username}")
            return JsonResponse({
                'success': True,
                'message': str(_("Rappel envoyé avec succès à {email}.")).format(email=practitioner.email)
            })
        else:
            return JsonResponse({
                'success': False,
                'message': str(_("Erreur lors de l'envoi du rappel."))
            }, status=500)

    except Exception as e:
        logger.error(f"Erreur lors de l'envoi du rappel pour le pratiquant {pk}: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'message': str(_("Une erreur est survenue."))
        }, status=500)


# =============================================================================
# TRANSFERT DE PRATIQUANT
# =============================================================================

@login_required
@require_POST
def initiate_transfer(request, practitioner_id):
    """Manager du club source initie un transfert vers un autre club."""
    from apps.competitions.models import PractitionerTransfer
    from apps.competitions.models.notifications import Notification
    from apps.organizations.models import Organization, OrganizationMember

    practitioner = get_object_or_404(Practitioner, pk=practitioner_id)
    user_org = get_user_club(request)

    # Vérifier que le manager appartient au club du pratiquant
    if not user_org or (practitioner.organization and practitioner.organization.id != user_org.id):
        messages.error(request, _("Vous n'avez pas les droits pour transférer ce pratiquant."))
        return redirect('competitions:club:practitioner_detail', pk=practitioner_id)

    target_org_id = request.POST.get('target_organization_id')
    reason = request.POST.get('reason', '')

    if not target_org_id:
        messages.error(request, _("Veuillez sélectionner un club de destination."))
        return redirect('competitions:club:practitioner_detail', pk=practitioner_id)

    try:
        target_org = Organization.objects.get(pk=target_org_id)
    except Organization.DoesNotExist:
        messages.error(request, _("Club de destination introuvable."))
        return redirect('competitions:club:practitioner_detail', pk=practitioner_id)

    if target_org.id == (user_org.id if user_org else None):
        messages.error(request, _("Le club de destination doit être différent du club actuel."))
        return redirect('competitions:club:practitioner_detail', pk=practitioner_id)

    # Vérifier qu'il n'y a pas déjà un transfert en attente
    if PractitionerTransfer.objects.filter(practitioner=practitioner, status='pending').exists():
        messages.warning(request, _("Un transfert est déjà en attente pour ce pratiquant."))
        return redirect('competitions:club:practitioner_detail', pk=practitioner_id)

    # Créer la demande
    transfer = PractitionerTransfer.objects.create(
        practitioner=practitioner,
        source_organization=practitioner.organization,
        target_organization=target_org,
        initiated_by='manager',
        requested_by=request.user,
        reason=reason,
    )

    # Notifier les managers du club cible
    target_managers = OrganizationMember.objects.filter(
        organization=target_org,
        role__in=['owner', 'admin', 'manager'],
        is_active=True
    ).select_related('user')
    for member in target_managers:
        Notification.objects.create(
            user=member.user,
            title=_("Nouvelle demande de transfert"),
            message=str(_("%(name)s souhaite rejoindre votre club (depuis %(source)s).")) % {
                'name': practitioner.full_name,
                'source': practitioner.organization.name if practitioner.organization else '—'
            },
            notification_type='info',
            priority='important',
        )

    messages.success(request, _("La demande de transfert a été envoyée."))
    return redirect('competitions:club:practitioner_detail', pk=practitioner_id)


@login_required
def pending_transfers(request):
    """Liste des demandes de transfert entrantes pour le club de l'utilisateur."""
    from apps.competitions.models import PractitionerTransfer

    user_org = get_user_club(request)
    if not user_org:
        messages.error(request, _("Aucun club associé à votre compte."))
        return redirect('competitions:dashboard:dashboard')

    transfers = PractitionerTransfer.objects.filter(
        target_organization=user_org,
        status='pending'
    ).select_related('practitioner', 'source_organization', 'requested_by').order_by('-created_at')

    # Also get recent processed transfers for history
    history = PractitionerTransfer.objects.filter(
        Q(target_organization=user_org) | Q(source_organization=user_org)
    ).exclude(status='pending').select_related(
        'practitioner', 'source_organization', 'target_organization', 'processed_by'
    ).order_by('-processed_at')[:10]

    return render(request, 'competitions/club/pending_transfers.html', {
        'transfers': transfers,
        'history': history,
        'club_organization': user_org,
    })


@login_required
@require_POST
def approve_transfer(request, transfer_id):
    """Le manager du club cible approuve le transfert."""
    from apps.competitions.models import PractitionerTransfer
    from apps.competitions.models.notifications import Notification
    from apps.organizations.models import OrganizationMember

    transfer = get_object_or_404(PractitionerTransfer, pk=transfer_id, status='pending')
    user_org = get_user_club(request)

    # Vérifier que l'utilisateur est manager du club cible
    if not user_org or user_org.id != transfer.target_organization.id:
        messages.error(request, _("Vous n'avez pas les droits pour approuver ce transfert."))
        return redirect('competitions:club:pending_transfers')

    transfer.approve(request.user)

    # Notifier le demandeur
    if transfer.requested_by:
        Notification.objects.create(
            user=transfer.requested_by,
            title=_("Transfert approuvé"),
            message=str(_("Le transfert de %(name)s vers %(target)s a été approuvé.")) % {
                'name': transfer.practitioner.full_name,
                'target': transfer.target_organization.name
            },
            notification_type='success',
            priority='important',
        )

    # Notifier le pratiquant s'il a un compte
    if transfer.practitioner.user and transfer.practitioner.user != transfer.requested_by:
        Notification.objects.create(
            user=transfer.practitioner.user,
            title=_("Vous avez été transféré"),
            message=str(_("Vous avez été transféré vers %(target)s.")) % {
                'target': transfer.target_organization.name
            },
            notification_type='info',
            priority='important',
        )

    messages.success(request, _("Le transfert de %(name)s a été approuvé.") % {
        'name': transfer.practitioner.full_name
    })
    return redirect('competitions:club:pending_transfers')


@login_required
@require_POST
def reject_transfer(request, transfer_id):
    """Le manager du club cible rejette le transfert."""
    from apps.competitions.models import PractitionerTransfer
    from apps.competitions.models.notifications import Notification

    transfer = get_object_or_404(PractitionerTransfer, pk=transfer_id, status='pending')
    user_org = get_user_club(request)

    if not user_org or user_org.id != transfer.target_organization.id:
        messages.error(request, _("Vous n'avez pas les droits pour rejeter ce transfert."))
        return redirect('competitions:club:pending_transfers')

    response_msg = request.POST.get('response_message', '')
    transfer.reject(request.user, response_msg)

    # Notifier le demandeur
    if transfer.requested_by:
        Notification.objects.create(
            user=transfer.requested_by,
            title=_("Transfert refusé"),
            message=str(_("Le transfert de %(name)s vers %(target)s a été refusé.")) % {
                'name': transfer.practitioner.full_name,
                'target': transfer.target_organization.name
            },
            notification_type='warning',
            priority='important',
        )

    messages.info(request, _("Le transfert a été refusé."))
    return redirect('competitions:club:pending_transfers')


@login_required
def search_clubs_api(request):
    """API JSON pour rechercher des clubs (autocomplete)."""
    from apps.organizations.models import Organization

    q = request.GET.get('q', '').strip()
    if len(q) < 2:
        return JsonResponse({'clubs': []})

    user_org = get_user_club(request)
    qs = Organization.objects.filter(
        Q(name__icontains=q) | Q(city__icontains=q),
        organization_type__in=['club', 'academy'],
    )
    if user_org:
        qs = qs.exclude(pk=user_org.pk)

    clubs = list(qs.values('id', 'name', 'city')[:10])
    return JsonResponse({'clubs': clubs})


@login_required
@require_POST
def update_practitioner_roles(request, practitioner_id):
    """AJAX endpoint pour mettre à jour les rôles d'un pratiquant."""
    from apps.organizations.models import OrganizationMember, OrganizationRole

    practitioner = get_object_or_404(Practitioner, id=practitioner_id)
    user_org = get_user_club(request)

    if not user_org:
        return JsonResponse({'success': False, 'message': str(_("Club introuvable"))}, status=403)

    # Vérifier que le manager a les droits
    if not request.user.is_superuser:
        manager_member = OrganizationMember.objects.filter(
            user=request.user,
            organization=user_org,
            role__in=['owner', 'admin', 'manager'],
            is_active=True,
        ).first()
        if not manager_member:
            return JsonResponse({'success': False, 'message': str(_("Droits insuffisants"))}, status=403)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON'}, status=400)

    main_role = data.get('main_role', 'member')
    additional_roles = data.get('additional_roles', [])

    # Valider les codes de rôles
    valid_codes = {c[0] for c in OrganizationRole.choices}
    if main_role not in valid_codes:
        return JsonResponse({'success': False, 'message': str(_("Rôle principal invalide"))}, status=400)
    additional_roles = [r for r in additional_roles if r in valid_codes and r != main_role]

    # Le pratiquant doit avoir un compte utilisateur lié
    if not practitioner.user:
        return JsonResponse({
            'success': False,
            'message': str(_("Ce pratiquant n'a pas de compte utilisateur. Activez d'abord son compte."))
        }, status=400)

    # Mettre à jour ou créer l'OrganizationMember
    member, created = OrganizationMember.objects.get_or_create(
        user=practitioner.user,
        organization=user_org,
        defaults={
            'role': main_role,
            'additional_roles': additional_roles,
            'practitioner': practitioner,
            'is_active': True,
        }
    )
    if not created:
        member.role = main_role
        member.additional_roles = additional_roles
        member.practitioner = practitioner
        member.save()

    return JsonResponse({
        'success': True,
        'main_role': member.role,
        'additional_roles': member.additional_roles,
        'all_roles': member.all_roles,
        'message': str(_("Rôles mis à jour avec succès")),
    })


@login_required
@require_POST
def quick_update_grade(request, practitioner_id):
    """
    AJAX endpoint pour mettre a jour rapidement le grade d'un pratiquant.
    """
    from apps.grades.models import Grade, PractitionerGrade

    practitioner = get_object_or_404(Practitioner, pk=practitioner_id)

    # Verifier les permissions
    user_club = get_user_club(request)
    if not user_club or practitioner.organization != user_club:
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

    grade_id = request.POST.get('grade_id')
    if not grade_id:
        # Retirer le grade
        practitioner.grade = None
        practitioner.save(update_fields=['grade'])
        return JsonResponse({
            'success': True,
            'grade_display': '',
            'grade_color': '#6c757d',
            'message': str(_("Grade retire")),
        })

    grade = get_object_or_404(Grade, pk=grade_id)

    # Mettre a jour le grade sur le practitioner
    practitioner.grade = grade
    practitioner.save(update_fields=['grade'])

    # Creer/mettre a jour PractitionerGrade
    PractitionerGrade.objects.update_or_create(
        practitioner=practitioner,
        discipline=grade.discipline,
        defaults={
            'grade': grade,
            'is_current': True,
            'date_obtained': timezone.now().date(),
        }
    )

    return JsonResponse({
        'success': True,
        'grade_display': grade.name,
        'grade_color': grade.color_code or '#6c757d',
        'message': str(_("Grade mis a jour: %(grade)s") % {'grade': grade.name}),
    })


@login_required
def get_grades_for_practitioner(request, practitioner_id):
    """
    AJAX endpoint pour obtenir les grades disponibles pour un pratiquant.
    """
    from apps.grades.models import Grade

    practitioner = get_object_or_404(Practitioner, pk=practitioner_id)

    # Recuperer les disciplines du pratiquant
    discipline_ids = list(practitioner.disciplines.values_list('id', flat=True))
    if practitioner.primary_discipline_id and practitioner.primary_discipline_id not in discipline_ids:
        discipline_ids.append(practitioner.primary_discipline_id)

    # Si pas de discipline, retourner tous les grades actifs
    if discipline_ids:
        grades = Grade.objects.filter(discipline_id__in=discipline_ids, is_active=True)
    else:
        grades = Grade.objects.filter(is_active=True)

    grades = grades.order_by('discipline__name', 'level', 'name')

    grades_data = []
    current_discipline = None
    for g in grades:
        disc_name = g.discipline.name if g.discipline else ''
        if disc_name != current_discipline:
            grades_data.append({'type': 'header', 'label': disc_name})
            current_discipline = disc_name
        grades_data.append({
            'type': 'grade',
            'id': g.id,
            'name': g.name,
            'color': g.color_code or '#6c757d',
            'is_dan': g.is_dan_grade,
            'selected': practitioner.grade_id == g.id if practitioner.grade_id else False,
        })

    return JsonResponse({'grades': grades_data})