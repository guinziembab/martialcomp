import logging
from django.core.exceptions import PermissionDenied, ValidationError
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils.translation import gettext as _
from django.http import JsonResponse, HttpResponse
from django.db import transaction
from django.utils import timezone
from django.views.decorators.csrf import ensure_csrf_cookie
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO

from apps.competitions.utils.permission_helpers import get_user_club
from apps.competitions.models import Practitioner, Discipline
from apps.core.isolation import OrganizationIsolationMixin, get_organization_queryset

logger = logging.getLogger(__name__)

def get_organization_from_club(club):
    """
    Extrait l'organisation depuis un club.
    Si le club n'a pas d'organisation, en crée une automatiquement.

    Args:
        club: Objet Club ou Organization

    Returns:
        Organization: L'organisation associée ou créée
    """
    from apps.organizations.models import Organization

    # Si c'est déjà une organisation
    if isinstance(club, Organization):
        return club

    # Si le club a une organisation associée
    if hasattr(club, 'organization') and club.organization:
        return club.organization

    # Si le club n'a pas d'organisation, en créer une automatiquement
    if club:
        logger.info(f"Club {club.name} (id={club.id}) n'a pas d'organisation. Création automatique...")
        try:
            # Créer une organisation pour ce club
            org = Organization.objects.create(
                name=club.name,
                organization_type='club',
                owner=club.owner,
                email=club.contact_email or '',
                phone=club.contact_phone or '',
                address=club.address or '',
                city=club.city or '',
                postal_code=club.postal_code or '',
                is_active=True
            )
            # Copier les disciplines du club vers l'organisation
            if club.disciplines.exists():
                org.disciplines.set(club.disciplines.all())

            # Lier l'organisation au club
            club.organization = org
            club.save(update_fields=['organization'])

            logger.info(f"Organisation '{org.name}' (id={org.id}) créée et liée au club {club.id}")
            return org
        except Exception as e:
            logger.error(f"Erreur lors de la création de l'organisation pour le club {club.id}: {e}")
            return None

    return None

def import_practitioners_from_excel(file, organization):
    """
    Importe des pratiquants depuis un fichier Excel.
    
    Args:
        file: Fichier Excel uploadé
        organization: Organisation à laquelle associer les pratiquants
    
    Returns:
        dict: Résultat de l'importation avec statistiques
    """
    results = {
        'success': 0,
        'errors': [],
        'skipped': 0,
        'total': 0
    }

    # -- Vérification limite membres (Free tier) --
    try:
        from apps.finances.services.subscription_service import get_subscription_info
        sub_info = get_subscription_info(organization)
        if sub_info.get('is_at_limit', False):
            results['errors'].append({
                'row': 0,
                'error': _(
                    "Import bloqué : limite de %(max)s membres atteinte "
                    "(plan Free). Passez au plan Premium pour importer des membres."
                ) % {'max': sub_info.get('max_members', 10)}
            })
            return results
    except Exception as e:
        logger.warning(f"Erreur vérification limite membres pour import: {e}")

    try:
        # Lire le contenu du fichier
        file_content = file.read()
        # Réinitialiser le pointeur pour permettre une relecture si nécessaire
        file.seek(0)
        
        # Charger le workbook
        wb = openpyxl.load_workbook(BytesIO(file_content))
        ws = wb.active
        
        # Lire les en-têtes (première ligne)
        headers = [cell.value for cell in ws[1]]
        
        # Normaliser les noms de colonnes (enlever les espaces, mettre en minuscules)
        header_map = {}
        for idx, header in enumerate(headers, start=1):
            if header:
                normalized = str(header).strip().lower()
                header_map[normalized] = idx
        
        logger.info(f"En-têtes détectés dans le fichier Excel: {header_map}")
        logger.info(f"En-têtes bruts: {headers}")

        # Mapper les colonnes attendues avec plusieurs variantes possibles
        col_mapping = {}
        
        # Mapping flexible pour chaque champ
        field_mappings = {
            'last_name': ['nom', 'lastname', 'last name', 'surname', 'family name', 'nom de famille'],
            'first_name': ['prénom', 'firstname', 'first name', 'prenom', 'given name', 'prénom '],
            'birth_date': ['date de naissance', 'date_de_naissance', 'date naissance', 'datenaissance',
                          'birthdate', 'birth date', 'birth_date', 'date_naissance', 'dob',
                          'né le', 'né(e) le', 'naissance', 'anniversaire'],
            'grade': ['grade', 'niveau', 'level', 'belt', 'ceinture', 'rang'],
            'email': ['email', 'e-mail', 'mail', 'courriel', 'adresse email', 'adresse e-mail'],
            'phone': ['téléphone', 'telephone', 'phone', 'tél', 'tel', 'mobile', 'portable', 'numéro'],
            'license_number': ['licence', 'license', 'numéro de licence', 'numero de licence', 'license number', 'numéro licence', 'n° licence'],
        }
        
        for field_name, possible_keys in field_mappings.items():
            for key in possible_keys:
                if key in header_map:
                    col_mapping[field_name] = header_map[key]
                    break  # Prendre la première correspondance trouvée
        
        logger.info(f"Mapping des colonnes utilisé: {col_mapping}")

        # Vérifier si les colonnes requises sont présentes
        if 'birth_date' not in col_mapping:
            logger.warning(f"Colonne 'date de naissance' non trouvée. En-têtes disponibles: {list(header_map.keys())}")
            results['errors'].append({
                'row': 0,
                'error': _("Colonne 'Date de naissance' non trouvée. En-têtes détectés: {}").format(
                    ', '.join(str(h) for h in headers if h) if headers else 'aucun'
                )
            })
            return results

        if 'first_name' not in col_mapping or 'last_name' not in col_mapping:
            logger.warning(
                f"Colonnes nom/prénom manquantes. "
                f"En-têtes disponibles: {list(header_map.keys())}"
            )
            results['errors'].append({
                'row': 0,
                'error': _("Colonnes 'Nom' et/ou 'Prénom' non trouvées. "
                          "En-têtes détectés: {}").format(
                    ', '.join(str(h) for h in headers if h) if headers else 'aucun'
                )
            })
            return results

        # Parcourir les lignes de données (à partir de la ligne 2)
        logger.debug(f"Nombre total de lignes dans le fichier: {ws.max_row}")
        for row_num in range(2, ws.max_row + 1):
            row = ws[row_num]

            # Vérifier si la ligne est vide (toutes les cellules sont None ou vides)
            row_values = [cell.value for cell in row if cell.value is not None and str(cell.value).strip()]
            if not row_values:
                continue  # Ignorer les lignes vides

            results['total'] += 1

            try:
                # Extraire les données de la ligne
                data = {}
                for field, col_idx in col_mapping.items():
                    if col_idx and col_idx <= len(row):
                        cell_value = row[col_idx - 1].value
                        if cell_value is not None:
                            # Convertir en string et nettoyer
                            if isinstance(cell_value, str):
                                cell_value = cell_value.strip()
                                if cell_value:  # Ignorer les chaînes vides
                                    data[field] = cell_value
                            else:
                                data[field] = cell_value

                # Vérifier les champs requis
                if 'first_name' not in data or 'last_name' not in data:
                    error_msg = _("Nom ou prénom manquant")
                    logger.error(f"Ligne {row_num}: {error_msg} - data={data}")
                    results['errors'].append({
                        'row': row_num,
                        'error': error_msg
                    })
                    results['skipped'] += 1
                    continue
                
                # Traiter la date de naissance avec support de multiples formats
                birth_date = None
                if 'birth_date' in data:
                    birth_date_value = data['birth_date']
                    
                    # Si c'est déjà un objet datetime, extraire la date
                    if isinstance(birth_date_value, datetime):
                        birth_date = birth_date_value.date()
                    # Si c'est une date (openpyxl peut retourner date directement)
                    elif hasattr(birth_date_value, 'year'):
                        birth_date = birth_date_value
                    # Si c'est une chaîne, essayer plusieurs formats
                    elif isinstance(birth_date_value, str):
                        birth_date_str = birth_date_value.strip()
                        
                        # Liste des formats de date à essayer (ordre important)
                        date_formats = [
                            '%Y-%m-%d',      # AAAA-MM-JJ (ISO standard)
                            '%d/%m/%Y',      # JJ/MM/AAAA (format français)
                            '%m/%d/%Y',      # MM/JJ/AAAA (format américain)
                            '%d-%m-%Y',      # JJ-MM-AAAA
                            '%m-%d-%Y',      # MM-JJ-AAAA
                            '%Y/%m/%d',      # AAAA/MM/JJ
                            '%d.%m.%Y',      # JJ.MM.AAAA
                            '%m.%d.%Y',      # MM.JJ.AAAA
                            '%Y.%m.%d',      # AAAA.MM.JJ
                            '%d %m %Y',      # JJ MM AAAA
                            '%Y %m %d',      # AAAA MM JJ
                        ]
                        
                        birth_date = None
                        for date_format in date_formats:
                            try:
                                birth_date = datetime.strptime(birth_date_str, date_format).date()
                                logger.debug(f"Date parsée avec le format {date_format}: {birth_date}")
                                break
                            except ValueError:
                                continue
                        
                        # Si aucun format n'a fonctionné, essayer de détecter automatiquement
                        if birth_date is None:
                            # Essayer de détecter le format en analysant la structure
                            parts = birth_date_str.replace('/', '-').replace('.', '-').replace(' ', '-').split('-')
                            if len(parts) == 3:
                                try:
                                    # Essayer de déterminer l'ordre (AAAA, MM, JJ)
                                    # Si une partie fait 4 chiffres, c'est probablement l'année
                                    year_idx = None
                                    for i, part in enumerate(parts):
                                        if len(part) == 4 and part.isdigit():
                                            year_idx = i
                                            break
                                    
                                    if year_idx is not None:
                                        year = int(parts[year_idx])
                                        # Les deux autres parties sont le mois et le jour
                                        other_parts = [int(p) for i, p in enumerate(parts) if i != year_idx]
                                        
                                        # Essayer les deux ordres possibles
                                        for month, day in [(other_parts[0], other_parts[1]), (other_parts[1], other_parts[0])]:
                                            if 1 <= month <= 12 and 1 <= day <= 31:
                                                try:
                                                    birth_date = datetime(year, month, day).date()
                                                    logger.debug(f"Date détectée automatiquement: {birth_date}")
                                                    break
                                                except ValueError:
                                                    continue
                                        
                                        if birth_date:
                                            break
                                    
                                    # Si pas d'année à 4 chiffres, essayer JJ/MM/AAAA ou MM/JJ/AAAA
                                    if birth_date is None and all(p.isdigit() for p in parts):
                                        # Essayer JJ/MM/AAAA (format français le plus courant)
                                        try:
                                            part1, part2, part3 = int(parts[0]), int(parts[1]), int(parts[2])
                                            if part3 < 100:
                                                # Si l'année est à 2 chiffres, ajouter 2000 ou 1900
                                                part3 = 2000 + part3 if part3 < 50 else 1900 + part3
                                            
                                            # Détecter l'ordre : si part1 > 12, c'est probablement le jour (JJ/MM/AAAA)
                                            # Si part1 <= 12 et part2 > 12, c'est probablement MM/JJ/AAAA
                                            if part1 > 12:
                                                # Format JJ/MM/AAAA
                                                day, month, year = part1, part2, part3
                                                if 1 <= month <= 12 and 1 <= day <= 31:
                                                    birth_date = datetime(year, month, day).date()
                                                    logger.debug(f"Date détectée (JJ/MM/AAAA): {birth_date}")
                                                    break
                                            elif part2 > 12:
                                                # Format MM/JJ/AAAA
                                                month, day, year = part1, part2, part3
                                                if 1 <= month <= 12 and 1 <= day <= 31:
                                                    birth_date = datetime(year, month, day).date()
                                                    logger.debug(f"Date détectée (MM/JJ/AAAA): {birth_date}")
                                                    break
                                            else:
                                                # Ambiguïté : les deux peuvent être valides
                                                # Essayer d'abord JJ/MM/AAAA (plus courant en français)
                                                day, month, year = part1, part2, part3
                                                if 1 <= month <= 12 and 1 <= day <= 31:
                                                    try:
                                                        birth_date = datetime(year, month, day).date()
                                                        logger.debug(f"Date détectée (JJ/MM/AAAA, ambigu): {birth_date}")
                                                        break
                                                    except ValueError:
                                                        pass
                                                
                                                # Si ça n'a pas marché, essayer MM/JJ/AAAA
                                                if birth_date is None:
                                                    month, day, year = part1, part2, part3
                                                    if 1 <= month <= 12 and 1 <= day <= 31:
                                                        try:
                                                            birth_date = datetime(year, month, day).date()
                                                            logger.debug(f"Date détectée (MM/JJ/AAAA, ambigu): {birth_date}")
                                                            break
                                                        except ValueError:
                                                            pass
                                        except (ValueError, IndexError) as e:
                                            logger.warning(f"Erreur lors du parsing de date: {e}")
                                            pass
                                except (ValueError, IndexError, AttributeError) as e:
                                    logger.warning(f"Erreur lors de la détection automatique de date: {e}")
                                    pass
                        
                        # Si toujours aucune date trouvée, signaler une erreur
                        if birth_date is None:
                            error_msg = _("Format de date invalide ou non reconnu: {}").format(birth_date_value)
                            logger.error(f"Ligne {row_num}: {error_msg} - data={data}")
                            results['errors'].append({
                                'row': row_num,
                                'error': error_msg
                            })
                            results['skipped'] += 1
                            continue
                    else:
                        error_msg = _("Date de naissance invalide (type: {})").format(type(birth_date_value).__name__)
                        logger.error(f"Ligne {row_num}: {error_msg} - data={data}")
                        results['errors'].append({
                            'row': row_num,
                            'error': error_msg
                        })
                        results['skipped'] += 1
                        continue
                else:
                    error_msg = _("Date de naissance manquante")
                    logger.error(f"Ligne {row_num}: {error_msg} - data={data}")
                    results['errors'].append({
                        'row': row_num,
                        'error': error_msg
                    })
                    results['skipped'] += 1
                    continue
                
                # Vérifier si le pratiquant existe déjà (par nom, prénom et date de naissance)
                existing = Practitioner.objects.filter(
                    first_name__iexact=data['first_name'],
                    last_name__iexact=data['last_name'],
                    birth_date=birth_date,
                    organization=organization
                ).first()

                if existing:
                    error_msg = _("Pratiquant déjà existant: {} {}").format(
                        data['first_name'], data['last_name']
                    )
                    logger.warning(f"Ligne {row_num}: {error_msg}")
                    results['errors'].append({
                        'row': row_num,
                        'error': error_msg
                    })
                    results['skipped'] += 1
                    continue
                
                # Traiter le grade (optionnel)
                grade = None
                if 'grade' in data and data['grade']:
                    grade_text = str(data['grade']).strip()
                    if grade_text:
                        try:
                            from apps.grades.models import Grade
                            # Chercher le grade par nom
                            grade = Grade.objects.filter(
                                name__iexact=grade_text
                            ).first()
                            if not grade:
                                # Si le grade n'existe pas, on continue sans grade
                                logger.warning(f"Grade non trouvé: {grade_text}")
                        except ImportError:
                            pass
                
                # Créer le pratiquant
                # Gérer les champs optionnels (None si vide)
                email = data.get('email')
                if email:
                    email = str(email).strip()
                    email = email if email else None
                else:
                    email = None
                
                phone = data.get('phone')
                if phone:
                    phone = str(phone).strip()
                    phone = phone if phone else None
                else:
                    phone = None
                
                license_number = data.get('license_number', '').strip() if data.get('license_number') else ''

                # Vérifier la limite de membres à chaque ajout (Free tier)
                try:
                    _sub = get_subscription_info(organization)
                    if _sub.get('is_at_limit', False):
                        remaining_rows = ws.max_row - row_num
                        results['errors'].append({
                            'row': row_num,
                            'error': _(
                                "Limite de %(max)s membres atteinte. "
                                "%(remaining)s ligne(s) restante(s) ignorée(s). "
                                "Passez au plan Premium."
                            ) % {'max': _sub.get('max_members', 10), 'remaining': remaining_rows}
                        })
                        break
                except Exception:
                    pass

                # Log des données avant création
                logger.info(f"Ligne {row_num}: Tentative de création - first_name='{data['first_name']}', last_name='{data['last_name']}', birth_date={birth_date}, email={email}, grade={grade}")

                practitioner = Practitioner.objects.create(
                    first_name=data['first_name'],
                    last_name=data['last_name'],
                    birth_date=birth_date,
                    organization=organization,
                    email=email,
                    phone=phone,
                    license_number=license_number,
                    grade=grade,
                    status='active',
                    is_active=True
                )

                # Assigner automatiquement les disciplines de l'organisation au pratiquant
                org_disciplines = organization.disciplines.all()
                if org_disciplines.exists():
                    practitioner.disciplines.set(org_disciplines)

                results['success'] += 1
                logger.info(f"Pratiquant créé: {practitioner.first_name} {practitioner.last_name}")
                
            except Exception as e:
                import traceback
                logger.error(
                    f"Erreur lors de l'importation de la ligne {row_num}: {str(e)}\n"
                    f"Traceback: {traceback.format_exc()}"
                )
                results['errors'].append({
                    'row': row_num,
                    'error': str(e)
                })
                results['skipped'] += 1
                continue
        
    except Exception as e:
        import traceback
        logger.error(f"Erreur lors de la lecture du fichier Excel: {str(e)}\nTraceback: {traceback.format_exc()}")
        results['errors'].append({
            'row': 0,
            'error': _("Erreur lors de la lecture du fichier: {}").format(str(e))
        })

    # Log final des résultats
    logger.info(f"=== RÉSULTAT IMPORT: success={results['success']}, errors={len(results['errors'])}, skipped={results['skipped']}, total={results['total']} ===")
    if results['errors']:
        for err in results['errors']:
            logger.info(f"  Erreur ligne {err.get('row', '?')}: {err.get('error', 'N/A')}")

    # Sync Stripe quantity if any members were imported
    if results['success'] > 0:
        from apps.finances.services.subscription_service import sync_stripe_member_count
        sync_stripe_member_count(organization)

    return results

@login_required
@ensure_csrf_cookie
def import_export_data(request):
    """Vue pour l'import/export de données du club"""
    
    # Vérifier si c'est une requête AJAX pour l'import CSV
    if request.path.endswith('/ajax/') and request.method == 'POST':
        club = get_user_club(request)
        if not club:
            return JsonResponse({
                'success': False,
                'error': _("Veuillez sélectionner un club valide.")
            }, status=400)
        
        # Vérifier si un fichier a été uploadé
        if 'file' not in request.FILES:
            return JsonResponse({
                'success': False,
                'error': _("Aucun fichier n'a été sélectionné.")
            }, status=400)
        
        uploaded_file = request.FILES['file']
        
        # Vérifier l'extension du fichier
        file_name = uploaded_file.name.lower()
        if not (file_name.endswith('.csv') or file_name.endswith('.xlsx') or file_name.endswith('.xls')):
            return JsonResponse({
                'success': False,
                'error': _("Format de fichier non supporté. Utilisez CSV ou Excel.")
            }, status=400)
        
        # Importer les pratiquants
        organization = get_organization_from_club(club)
        if not organization:
            return JsonResponse({
                'success': False,
                'error': _("Aucune organisation associée au club.")
            }, status=400)
        results = import_practitioners_from_excel(uploaded_file, organization)
        
        if results['success'] > 0:
            return JsonResponse({
                'success': True,
                'message': _("{} pratiquants importés avec succès. {} erreurs.").format(
                    results['success'], len(results['errors'])
                ),
                'results': results
            })
        else:
            return JsonResponse({
                'success': False,
                'error': _("Aucun pratiquant n'a pu être importé. {} erreurs.").format(len(results['errors'])),
                'results': results
            }, status=400)
    
    club = get_user_club(request)
    organization = get_organization_from_club(club) if club else None
    
    if not club:
        messages.error(request, _("Aucun club associé à votre compte."))
        return render(request, 'competitions/club/import_export.html', {
            'page_title': _('Import/Export de données'),
            'section': 'import_export',
            'club': None
        })
    
    context = {
        'page_title': _('Import/Export de données'),
        'section': 'import_export',
        'club': club
    }
    
    if request.method == 'POST':
        action = request.POST.get('action', '')

        # Gérer l'import via le formulaire normal (pas AJAX)
        # Vérifier soit le bouton "import" soit l'action "import"
        import_requested = ('import' in request.POST) or (action == 'import')
        has_file = 'excel_file' in request.FILES

        if import_requested and has_file:
            uploaded_file = request.FILES['excel_file']
            
            # Vérifier l'extension du fichier
            file_name = uploaded_file.name.lower()
            if not (file_name.endswith('.csv') or file_name.endswith('.xlsx') or file_name.endswith('.xls')):
                messages.error(request, _("Format de fichier non supporté. Utilisez CSV ou Excel."))
            elif not organization:
                messages.error(request, _("Aucune organisation associée au club."))
            else:
                # Importer les pratiquants
                try:
                    with transaction.atomic():
                        results = import_practitioners_from_excel(uploaded_file, organization)

                        if results['success'] > 0:
                            messages.success(
                                request,
                                _("{} pratiquants importés avec succès.").format(results['success'])
                            )
                            if results['errors']:
                                error_count = len(results['errors'])
                                messages.warning(
                                    request,
                                    _("{} lignes n'ont pas pu être importées.").format(error_count)
                                )
                        else:
                            messages.error(
                                request,
                                _("Aucun pratiquant n'a pu être importé. {} erreurs.").format(len(results['errors']))
                            )
                            
                        # Stocker les résultats dans le contexte pour affichage détaillé
                        context['import_results'] = results
                        
                except Exception as e:
                    logger.exception(f"Erreur lors de l'importation: {str(e)}")
                    messages.error(request, _("Erreur lors de l'importation: {}").format(str(e)))
        
        elif action == 'export':
            messages.info(request, _('Fonctionnalité d\'export en cours de développement'))
    
    # Toujours afficher le template, jamais de redirection
    return render(request, 'competitions/club/import_export.html', context)


@login_required
def download_import_template(request):
    """
    Génère et télécharge un modèle Excel pour l'import des pratiquants.
    """
    # Créer un nouveau workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pratiquants"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    example_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    # En-têtes
    headers = [
        ('Nom', 20),
        ('Prénom', 20),
        ('Date de naissance', 20),
        ('Grade', 20),
        ('Email', 30),
        ('Numéro de licence', 20)
    ]

    for col_num, (header, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width

    # Exemples de données
    examples = [
        ['DUPONT', 'Jean', '15/03/1990', 'Ceinture noire', 'jean.dupont@email.com', 'LIC001'],
        ['MARTIN', 'Marie', '22/07/1985', 'Ceinture marron', 'marie.martin@email.com', 'LIC002'],
        ['BERNARD', 'Pierre', '10/01/2005', 'Ceinture orange', 'pierre.bernard@email.com', 'LIC003'],
    ]

    for row_num, example in enumerate(examples, 2):
        for col_num, value in enumerate(example, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.fill = example_fill
            cell.border = thin_border

    # Ajouter une feuille d'instructions
    ws_instructions = wb.create_sheet(title="Instructions")
    instructions = [
        ("INSTRUCTIONS D'IMPORT", ""),
        ("", ""),
        ("Colonnes obligatoires:", ""),
        ("- Nom", "Le nom de famille du pratiquant"),
        ("- Prénom", "Le prénom du pratiquant"),
        ("- Date de naissance", "Format JJ/MM/AAAA ou AAAA-MM-JJ"),
        ("", ""),
        ("Colonnes optionnelles:", ""),
        ("- Grade", "Ex: Ceinture noire, Ceinture marron, 1er dan..."),
        ("- Email", "L'adresse email du pratiquant"),
        ("- Numéro de licence", "Le numéro de licence fédérale"),
        ("", ""),
        ("Notes importantes:", ""),
        ("1.", "Supprimez les exemples avant d'ajouter vos données"),
        ("2.", "Ne modifiez pas les noms des colonnes"),
        ("3.", "Les pratiquants existants (même nom, prénom, date) seront ignorés"),
    ]

    for row_num, (col1, col2) in enumerate(instructions, 1):
        ws_instructions.cell(row=row_num, column=1, value=col1)
        ws_instructions.cell(row=row_num, column=2, value=col2)
        if row_num == 1:
            ws_instructions.cell(row=row_num, column=1).font = Font(bold=True, size=14)

    ws_instructions.column_dimensions['A'].width = 25
    ws_instructions.column_dimensions['B'].width = 50

    # Sauvegarder en mémoire
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Créer la réponse HTTP
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Modele_Import_Pratiquants.xlsx"'

    return response


@login_required
def export_practitioners(request):
    """
    Exporte tous les pratiquants du club en fichier Excel.
    """
    club = get_user_club(request)
    organization = get_organization_from_club(club) if club else None

    if not organization:
        messages.error(request, _("Aucune organisation associée à votre compte."))
        return render(request, 'competitions/club/import_export.html', {
            'page_title': _('Import/Export de données'),
            'section': 'import_export',
            'club': club
        })

    # Récupérer tous les pratiquants de l'organisation
    practitioners = Practitioner.objects.filter(
        organization=organization
    ).select_related('grade').order_by('last_name', 'first_name')

    # Créer le workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pratiquants"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # En-têtes
    headers = [
        ('Nom', 20),
        ('Prénom', 20),
        ('Date de naissance', 18),
        ('Grade', 20),
        ('Email', 30),
        ('Téléphone', 18),
        ('Numéro de licence', 20),
        ('Statut', 12)
    ]

    for col_num, (header, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width

    # Données des pratiquants
    for row_num, p in enumerate(practitioners, 2):
        data = [
            p.last_name or '',
            p.first_name or '',
            p.birth_date.strftime('%d/%m/%Y') if p.birth_date else '',
            p.grade.name if p.grade else '',
            p.email or '',
            p.phone or '',
            p.license_number or '',
            'Actif' if p.is_active else 'Inactif'
        ]

        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border

    # Sauvegarder en mémoire
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Nom du fichier avec date
    filename = f"Pratiquants_{organization.name}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    # Nettoyer le nom de fichier
    filename = "".join(c for c in filename if c.isalnum() or c in ('_', '-', '.'))

    # Créer la réponse HTTP
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    logger.info(f"Export de {practitioners.count()} pratiquants pour {organization.name}")

    return response

