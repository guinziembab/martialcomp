"""
Vue améliorée pour la gestion des pratiquants avec double affichage (cartes/tableau)
Support export, actions en lot et filtres avancés
"""
import logging
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.utils.translation import gettext_lazy as _
from django.utils.translation import gettext as gettext_func
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST, require_http_methods
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.contrib.auth.models import User
from django.db import models, transaction
from django.db.models import Q, Count, Avg
from django.core.paginator import Paginator
from django.core.exceptions import PermissionDenied
from datetime import datetime, timedelta
import json
import csv
import openpyxl
from io import StringIO, BytesIO

# Import des modèles
from apps.competitions.models import (
    Practitioner, 
    Club,
    Discipline,
    Competition,
    CompetitionRegistration
)

# Import des formulaires
from apps.competitions.forms import (
    PractitionerForm,
)

# Import des décorateurs
from apps.competitions.utils.decorators import club_required, permission_required
from apps.core.isolation import OrganizationIsolationMixin, get_organization_queryset
from apps.competitions.utils.club import get_user_club as get_club_from_utils

# Import des modèles de grades avec gestion des erreurs
try:
    from apps.grades.models import GradingSystem as GradeSystem, Grade
    GRADES_MODELS_AVAILABLE = True
except (ImportError, ModuleNotFoundError) as e:
    logger = logging.getLogger(__name__)
    logger.warning(f"Les modèles de grades n'ont pas pu être importés: {str(e)}")
    GRADES_MODELS_AVAILABLE = False

# Création du logger
logger = logging.getLogger(__name__)

def get_user_club(user_or_request):
    """Obtenir le club de l'utilisateur avec gestion des erreurs"""
    try:
        # Détermine si c'est un user ou une request
        if hasattr(user_or_request, 'user'):
            request = user_or_request
            user = request.user
        else:
            user = user_or_request
            request = None
        
        # Si le club est déjà dans la requête (via le décorateur)
        if request and hasattr(request, 'club') and request.club:
            return request.club
        
        # PRIORITÉ 1: Utiliser l'organisation du middleware si disponible dans la requête
        if request and hasattr(request, 'user_organization') and request.user_organization:
            return request.user_organization
        
        # PRIORITÉ 2: Utiliser la logique du middleware directement (UserProfile)
        try:
            from apps.competitions.models.users import UserProfile
            profile = UserProfile.objects.get(user=user)
            if profile.organization:
                return profile.organization
        except:
            pass
        
        # PRIORITÉ 3: Memberships actifs (comme le middleware)
        try:
            from apps.membership.models import MembershipSubscription
            subscription = MembershipSubscription.objects.filter(
                practitioner__user=user,
                status='active'
            ).select_related('package__organization').first()
            
            if subscription and subscription.package.organization:
                return subscription.package.organization
        except:
            pass
        
        # PRIORITÉ 4: Si l'utilisateur a un attribut club direct
        if hasattr(user, 'club') and user.club:
            return user.club
        
        # PRIORITÉ 5: Si l'utilisateur est propriétaire d'un club
        from apps.competitions.models import Club
        owned_club = Club.objects.filter(owner=user).first()
        if owned_club:
            # Retourner l'organisation liée au club, pas le club lui-même
            return owned_club.organization if hasattr(owned_club, 'organization') and owned_club.organization else owned_club
        
        # PRIORITÉ 6: Coach profile
        if hasattr(user, 'coach_profile') and user.coach_profile and user.coach_profile.club:
            coach_club = user.coach_profile.club
            # Retourner l'organisation liée au club
            return coach_club.organization if hasattr(coach_club, 'organization') and coach_club.organization else coach_club
        
        # PRIORITÉ 7: Practitioner record
        practitioner = Practitioner.objects.select_related('organization').filter(
            Q(user=user) | Q(email=user.email)
        ).first()
        
        if practitioner and practitioner.organization:
            return practitioner.organization
        
        # PRIORITÉ 8: Admin roles
        if hasattr(user, 'club_admin_roles'):
            club_admin = user.club_admin_roles.first()
            if club_admin:
                # Retourner l'organisation, pas le club
                if hasattr(club_admin, 'organization') and club_admin.organization:
                    return club_admin.organization
                elif hasattr(club_admin, 'club') and club_admin.club:
                    admin_club = club_admin.club
                    return admin_club.organization if hasattr(admin_club, 'organization') and admin_club.organization else admin_club
                return club_admin.club
            
        return None
    except Exception as e:
        user_id = user.id if hasattr(user, 'id') else 'unknown'
        logger.error(f"Erreur lors de la récupération du club de l'utilisateur {user_id}: {str(e)}")
        return None

def manual_permission_check(user, club):
    """Vérification manuelle des permissions"""
    try:
        if user.is_superuser:
            return True
        
        # Vérifier que club est bien un objet Organization
        if club is None:
            return False
        
        # Si club est une chaîne, essayer de trouver l'organisation
        if isinstance(club, str):
            from apps.organizations.models import Organization
            try:
                club = Organization.objects.get(name=club)
            except Organization.DoesNotExist:
                logger.error(f"Organisation '{club}' non trouvée")
                return False
        
        # Vérifier si l'utilisateur est propriétaire du club lié à cette organisation
        from apps.competitions.models import Club
        owned_club = Club.objects.filter(
            owner=user,
            organization=club
        ).first()
        if owned_club:
            return True
            
        if hasattr(user, 'coach_profile') and user.coach_profile:
            coach_club = user.coach_profile.club
            if coach_club and club and hasattr(club, 'id') and coach_club.id == club.id:
                return True
        
        # Vérifier via le Practitioner avec l'objet Organization
        practitioner = Practitioner.objects.filter(
            Q(user=user) | Q(email=user.email),
            organization=club
        ).first()
        
        return practitioner is not None
        
    except Exception as e:
        logger.error(f"Erreur lors de la vérification des permissions: {str(e)}")
        return False

@login_required
def practitioners_list(request):
    """
    Vue améliorée pour la gestion des pratiquants avec:
    - Double affichage (cartes/tableau)
    - Filtres avancés
    - Actions en lot
    - Export (CSV, Excel, PDF)
    - Statistiques enrichies
    """
    try:
        # Récupération du club
        user_club = get_user_club(request)
        
        # DEBUG: Log des informations de débogage
        logger.info(f"DEBUG practitioners_list - User: {request.user.username}")
        logger.info(f"DEBUG practitioners_list - Club found: {user_club}")
        logger.info(f"DEBUG practitioners_list - Request has user_organization: {hasattr(request, 'user_organization')}")
        if hasattr(request, 'user_organization'):
            logger.info(f"DEBUG practitioners_list - Request.user_organization: {request.user_organization}")
        
        if not user_club:
            messages.error(request, _("Vous n'êtes associé à aucun club. Contactez un administrateur."))
            return redirect('competitions:dashboard:dashboard')
        
        # DEBUG: Vérifications des permissions détaillées
        permission_result = manual_permission_check(request.user, user_club)
        logger.info(f"DEBUG practitioners_list - Permission check result: {permission_result}")
        
        # Compter les pratiquants dans ce club pour diagnostic
        practitioner_count = Practitioner.objects.filter(organization=user_club).count()
        logger.info(f"DEBUG practitioners_list - Practitioners in {user_club}: {practitioner_count}")
        
        if not permission_result:
            # Au lieu de lever une exception, essayons de trouver un club avec des pratiquants
            logger.warning(f"Permission denied for user {request.user.username} in club {user_club}")
            
            # Chercher un practitioner record pour cet utilisateur
            user_practitioner = Practitioner.objects.filter(
                Q(user=request.user) | Q(email=request.user.email)
            ).first()
            
            if user_practitioner and user_practitioner.organization:
                logger.info(f"DEBUG - Found practitioner organization: {user_practitioner.organization}")
                user_club = user_practitioner.organization
                # Réévaluer les permissions avec la nouvelle organisation
                permission_result = manual_permission_check(request.user, user_club)
                logger.info(f"DEBUG - New permission check result: {permission_result}")
                
                if permission_result:
                    logger.info(f"SUCCESS - Using practitioner organization: {user_club}")
                else:
                    raise PermissionDenied(_("Vous n'avez pas l'autorisation de voir cette page."))
            else:
                raise PermissionDenied(_("Vous n'avez pas l'autorisation de voir cette page."))

        # Base queryset avec optimisation
        queryset = Practitioner.objects.filter(
            organization=user_club
        ).select_related('organization', 'primary_discipline', 'user').order_by('last_name', 'first_name')

        # === FILTRES AVANCÉS ===
        search = request.GET.get('search', '').strip()
        if search:
            queryset = queryset.filter(
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(email__icontains=search) |
                Q(phone__icontains=search) |
                Q(license_number__icontains=search)
            )

        # Filtre par grade
        grade_filter = request.GET.get('grade', '').strip()
        if grade_filter:
            if grade_filter == 'none':
                queryset = queryset.filter(Q(grade__isnull=True) | Q(grade=''))
            else:
                queryset = queryset.filter(grade=grade_filter)

        # Filtre par groupe d'âge
        age_group = request.GET.get('age_group', '').strip()
        if age_group:
            today = timezone.now().date()
            if age_group == 'child':  # Moins de 12 ans
                min_birth_date = today - timedelta(days=12*365)
                queryset = queryset.filter(birth_date__gt=min_birth_date)
            elif age_group == 'teen':  # 12-17 ans
                min_birth_date = today - timedelta(days=18*365)
                max_birth_date = today - timedelta(days=12*365)
                queryset = queryset.filter(birth_date__lte=max_birth_date, birth_date__gt=min_birth_date)
            elif age_group == 'adult':  # 18 ans et plus
                max_birth_date = today - timedelta(days=18*365)
                queryset = queryset.filter(birth_date__lte=max_birth_date)

        # Filtre par statut
        status = request.GET.get('status', '').strip()
        if status == 'active':
            queryset = queryset.filter(is_active=True)
        elif status == 'inactive':
            queryset = queryset.filter(is_active=False)

        # Filtre par statut utilisateur
        user_status = request.GET.get('user_status', '').strip()
        if user_status == 'with_account':
            queryset = queryset.filter(user__isnull=False)
        elif user_status == 'without_account':
            queryset = queryset.filter(user__isnull=True)

        # Filtre par discipline
        discipline = request.GET.get('discipline', '').strip()
        if discipline:
            queryset = queryset.filter(primary_discipline_id=discipline)

        # === GESTION DES EXPORTS ===
        export_format = request.GET.get('format', '').strip()
        if export_format in ['csv', 'excel', 'pdf']:
            return handle_export(request, queryset, export_format, user_club)

        # === STATISTIQUES ENRICHIES ===
        total_practitioners = queryset.count()
        active_practitioners = queryset.filter(is_active=True).count()
        inactive_practitioners = total_practitioners - active_practitioners
        practitioners_with_accounts = queryset.filter(user__isnull=False).count()
        coaches = queryset.filter(is_coach=True).count()

        # Distribution par grade
        grade_distribution = {}
        for practitioner in queryset:
            grade_display = get_grade_display(practitioner)
            if grade_display:
                grade_distribution[grade_display] = grade_distribution.get(grade_display, 0) + 1

        # Distribution par âge
        age_distribution = {'child': 0, 'teen': 0, 'adult': 0, 'unknown': 0}
        today = timezone.now().date()
        
        for practitioner in queryset:
            if practitioner.birth_date:
                age = (today - practitioner.birth_date).days // 365
                if age < 12:
                    age_distribution['child'] += 1
                elif age < 18:
                    age_distribution['teen'] += 1
                else:
                    age_distribution['adult'] += 1
            else:
                age_distribution['unknown'] += 1

        # === PAGINATION ===
        paginator = Paginator(queryset, 12)  # 12 par page pour les cartes
        page_number = request.GET.get('page')
        practitioners = paginator.get_page(page_number)

        # Enrichissement des pratiquants avec des données calculées
        for practitioner in practitioners:
            # Use temporary attributes instead of trying to set properties
            practitioner.computed_grade_display = get_grade_display(practitioner)
            practitioner.computed_grade_css_class = get_grade_css_class(practitioner)
            practitioner.computed_age = get_age(practitioner) if practitioner.birth_date else None

        # === CHOIX POUR LES FILTRES ===
        grade_choices = get_grade_choices(user_club)
        discipline_choices = get_discipline_choices(user_club)

        # Contexte pour le template
        context = {
            'practitioners': practitioners,
            'club': user_club,
            'grade_choices': grade_choices,
            'discipline_choices': discipline_choices,
            'grade_distribution': grade_distribution,
            'age_distribution': age_distribution,
            'stats': {
                'total_practitioners': total_practitioners,
                'active_practitioners': active_practitioners,
                'inactive_practitioners': inactive_practitioners,
                'practitioners_with_accounts': practitioners_with_accounts,
                'coaches': coaches,
                'completion_rate': round((practitioners_with_accounts / total_practitioners * 100) if total_practitioners > 0 else 0, 1)
            },
            'current_section': 'practitioners',
            'page_title': _('Gestion des Pratiquants'),
        }

        return render(request, 'competitions/club/practitioners_enhanced.html', context)

    except PermissionDenied:
        raise
    except Exception as e:
        logger.error(f"Erreur dans practitioners_list pour l'utilisateur {request.user.id}: {str(e)}")
        messages.error(request, _("Une erreur est survenue lors du chargement de la page."))
        return redirect('competitions:dashboard:club')

def handle_export(request, queryset, format_type, club):
    """Gestion des exports en différents formats"""
    try:
        # Récupération des pratiquants sélectionnés si spécifié
        selected_ids = request.GET.get('selected', '').strip()
        if selected_ids:
            selected_ids = [int(id) for id in selected_ids.split(',') if id.isdigit()]
            queryset = queryset.filter(id__in=selected_ids)

        if format_type == 'csv':
            return export_to_csv(queryset, club)
        elif format_type == 'excel':
            return export_to_excel(queryset, club)
        elif format_type == 'pdf':
            return export_to_pdf(queryset, club)
        
    except Exception as e:
        logger.error(f"Erreur lors de l'export {format_type}: {str(e)}")
        messages.error(request, _("Erreur lors de l'export des données."))
        return redirect('competitions:club:practitioners')

def export_to_csv(queryset, club):
    """Export des pratiquants en CSV"""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="pratiquants_{club.name}_{datetime.now().strftime("%Y%m%d")}.csv"'
    response.write('\ufeff')  # BOM pour Excel
    
    writer = csv.writer(response)
    
    # En-têtes
    headers = [
        'Nom', 'Prénom', 'Email', 'Téléphone', 'Date de naissance',
        'Genre', 'Grade', 'Discipline principale', 'Numéro de licence',
        'Statut', 'Date d\'inscription', 'Est instructeur'
    ]
    writer.writerow(headers)
    
    # Données
    for practitioner in queryset:
        row = [
            practitioner.last_name,
            practitioner.first_name,
            practitioner.email or '',
            practitioner.phone or '',
            practitioner.birth_date.strftime('%d/%m/%Y') if practitioner.birth_date else '',
            practitioner.get_gender_display() if practitioner.gender else '',
            get_grade_display(practitioner) or '',
            practitioner.primary_discipline.name if practitioner.primary_discipline else '',
            practitioner.license_number or '',
            'Actif' if practitioner.is_active else 'Inactif',
            practitioner.registration_date.strftime('%d/%m/%Y') if practitioner.registration_date else '',
            'Oui' if practitioner.is_coach else 'Non'
        ]
        writer.writerow(row)
    
    return response

def export_to_excel(queryset, club):
    """Export des pratiquants en Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pratiquants"
    
    # En-têtes avec style
    headers = [
        'Nom', 'Prénom', 'Email', 'Téléphone', 'Date de naissance',
        'Genre', 'Age', 'Grade', 'Discipline principale', 'Numéro de licence',
        'Statut', 'Date d\'inscription', 'Est instructeur', 'Ville'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Données
    for row_num, practitioner in enumerate(queryset, 2):
        data = [
            practitioner.last_name,
            practitioner.first_name,
            practitioner.email or '',
            practitioner.phone or '',
            practitioner.birth_date if practitioner.birth_date else '',
            practitioner.get_gender_display() if practitioner.gender else '',
            get_age(practitioner) if practitioner.birth_date else '',
            get_grade_display(practitioner) or '',
            practitioner.primary_discipline.name if practitioner.primary_discipline else '',
            practitioner.license_number or '',
            'Actif' if practitioner.is_active else 'Inactif',
            practitioner.registration_date if practitioner.registration_date else '',
            'Oui' if practitioner.is_coach else 'Non',
            practitioner.city or ''
        ]
        
        for col, value in enumerate(data, 1):
            ws.cell(row=row_num, column=col, value=value)
    
    # Ajustement de la largeur des colonnes
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Sauvegarde
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="pratiquants_{club.name}_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    
    return response

def export_to_pdf(queryset, club):
    """Export des pratiquants en PDF (version simple)"""
    from django.template.loader import render_to_string
    from django.http import HttpResponse
    
    # Pour une version simple, on retourne un HTML stylé
    # Dans une version complète, on utiliserait reportlab ou weasyprint
    
    context = {
        'practitioners': queryset,
        'club': club,
        'export_date': datetime.now(),
        'total_count': queryset.count()
    }
    
    html_content = render_to_string('competitions/club/practitioners_export_pdf.html', context)
    
    response = HttpResponse(html_content, content_type='text/html')
    response['Content-Disposition'] = f'attachment; filename="pratiquants_{club.name}_{datetime.now().strftime("%Y%m%d")}.html"'
    
    return response

def get_grade_display(practitioner):
    """Obtenir l'affichage du grade d'un pratiquant"""
    try:
        if hasattr(practitioner, 'grade_text') and practitioner.grade_text:
            return practitioner.grade_text
        
        if practitioner.grade:
            # If grade is a Grade model object
            if hasattr(practitioner.grade, 'name'):
                return practitioner.grade.name
            # If grade is a Grade model object with string representation
            elif not isinstance(practitioner.grade, str):
                return str(practitioner.grade)
            # If grade is a string
            else:
                # Mapping basique des grades
                grade_mapping = {
                    'cap_jaune': 'Ceinture Jaune',
                    'cap_rouge': 'Ceinture Rouge',
                    'cap_blanc': 'Ceinture Blanche',
                    'cap_bleu': 'Ceinture Bleue',
                    'cap_marron': 'Ceinture Marron',
                    'cap_noir': 'Ceinture Noire',
                }
                return grade_mapping.get(practitioner.grade, practitioner.grade.replace('_', ' ').title())
        
        return None
        
    except Exception as e:
        logger.error(f"Erreur lors de la récupération du grade pour le pratiquant {practitioner.id}: {str(e)}")
        return None

def get_grade_css_class(practitioner):
    """Obtenir la classe CSS pour le style du grade"""
    try:
        if practitioner.grade:
            # If grade is a string
            if isinstance(practitioner.grade, str):
                return practitioner.grade.replace('_', '-')
            # If grade is a Grade model object, use its identifier or name
            elif hasattr(practitioner.grade, 'identifier'):
                return practitioner.grade.identifier.replace('_', '-')
            elif hasattr(practitioner.grade, 'name'):
                grade_name = practitioner.grade.name.lower().replace(' ', '-').replace('ceinture ', 'cap-')
                return grade_name
            else:
                return str(practitioner.grade).replace('_', '-').replace(' ', '-')
        return 'cap-blanc'
    except Exception as e:
        logger.error(f"Erreur lors de la récupération de la classe CSS du grade pour le pratiquant {practitioner.id}: {str(e)}")
        return 'cap-blanc'

def get_age(practitioner):
    """Calculer l'âge d'un pratiquant"""
    if practitioner.birth_date:
        today = timezone.now().date()
        return (today - practitioner.birth_date).days // 365
    return None

def get_grade_choices(club):
    """Obtenir les choix de grades pour les filtres"""
    try:
        if GRADES_MODELS_AVAILABLE:
            grades = Grade.objects.filter(system__is_active=True).order_by('order')
            return [(grade.id, grade.name) for grade in grades]
        else:
            # Choix statiques pour Qwan Ki Do
            return [
                ('cap_blanc', 'Ceinture Blanche'),
                ('cap_jaune', 'Ceinture Jaune'),
                ('cap_rouge', 'Ceinture Rouge'),
                ('cap_bleu', 'Ceinture Bleue'),
                ('cap_marron', 'Ceinture Marron'),
                ('cap_noir', 'Ceinture Noire'),
            ]
    except Exception as e:
        logger.error(f"Erreur lors de la récupération des grades: {str(e)}")
        return []

def get_discipline_choices(club):
    """Obtenir les choix de disciplines pour les filtres"""
    try:
        disciplines = Discipline.objects.filter(
            practitioners__organization=club
        ).distinct().order_by('name')
        return [(d.id, d.name) for d in disciplines]
    except Exception as e:
        logger.error(f"Erreur lors de la récupération des disciplines: {str(e)}")
        return []

@login_required
@require_POST
def practitioners_bulk_action(request):
    """Gestion des actions en lot sur les pratiquants"""
    try:
        action = request.POST.get('action')
        selected_ids = request.POST.getlist('selected_ids')
        
        if not action or not selected_ids:
            return JsonResponse({'success': False, 'error': 'Paramètres manquants'})
        
        # Récupération du club
        user_club = get_user_club(request)
        if not user_club or not manual_permission_check(request.user, user_club):
            return JsonResponse({'success': False, 'error': 'Permissions insuffisantes'})
        
        # Récupération des pratiquants
        practitioners = Practitioner.objects.filter(
            id__in=selected_ids,
            organization=user_club
        )
        
        if action == 'activate':
            practitioners.update(is_active=True)
            message = f"{practitioners.count()} pratiquants activés"
        elif action == 'deactivate':
            practitioners.update(is_active=False)
            message = f"{practitioners.count()} pratiquants désactivés"
        elif action == 'delete':
            count = practitioners.count()
            practitioners.delete()
            message = f"{count} pratiquants supprimés"
        else:
            return JsonResponse({'success': False, 'error': 'Action non reconnue'})
        
        return JsonResponse({
            'success': True, 
            'message': message,
            'affected_count': len(selected_ids)
        })
        
    except Exception as e:
        logger.error(f"Erreur dans practitioners_bulk_action: {str(e)}")
        return JsonResponse({'success': False, 'error': 'Erreur lors de l\'opération'})

# ===== FONCTIONS ORIGINALES CONSERVÉES POUR COMPATIBILITÉ =====

@login_required
def practitioner_detail(request, pk):
    """Afficher le détail d'un pratiquant"""
    try:
        user_club = get_user_club(request)
        if not user_club:
            messages.error(request, _("Vous n'êtes associé à aucun club."))
            return redirect('competitions:dashboard:club')
        
        practitioner = get_object_or_404(
            Practitioner.objects.select_related('organization', 'primary_discipline', 'user'),
            pk=pk,
            organization=user_club
        )
        
        # Vérifications des permissions
        if not manual_permission_check(request.user, user_club):
            raise PermissionDenied(_("Vous n'avez pas l'autorisation de voir ce pratiquant."))
        
        # Enrichissement des données
        practitioner.computed_grade_display = get_grade_display(practitioner)
        practitioner.computed_age = get_age(practitioner)
        
        context = {
            'practitioner': practitioner,
            'club': user_club,
            'page_title': f"Profil de {practitioner.full_name}",
        }
        
        return render(request, 'competitions/club/practitioner_detail.html', context)
        
    except PermissionDenied:
        raise
    except Exception as e:
        logger.error(f"Erreur dans practitioner_detail: {str(e)}")
        messages.error(request, _("Erreur lors du chargement du profil."))
        return redirect('competitions:club:practitioners')

@login_required
def practitioner_create(request):
    """Créer un nouveau pratiquant"""
    try:
        user_club = get_user_club(request)
        if not user_club:
            messages.error(request, _("Vous n'êtes associé à aucun club."))
            return redirect('competitions:dashboard:club')
        
        if not manual_permission_check(request.user, user_club):
            raise PermissionDenied(_("Vous n'avez pas l'autorisation de créer un pratiquant."))
        
        if request.method == 'POST':
            form = PractitionerForm(request.POST, request.FILES)
            if form.is_valid():
                practitioner = form.save(commit=False)
                practitioner.organization = user_club
                practitioner.save()
                
                messages.success(request, _(f"Le pratiquant {practitioner.full_name} a été créé avec succès."))
                return redirect('competitions:club:practitioners')
        else:
            form = PractitionerForm()
        
        context = {
            'form': form,
            'club': user_club,
            'page_title': _("Ajouter un Pratiquant"),
        }
        
        return render(request, 'competitions/club/practitioner_form.html', context)
        
    except PermissionDenied:
        raise
    except Exception as e:
        logger.error(f"Erreur dans practitioner_create: {str(e)}")
        messages.error(request, _("Erreur lors de la création du pratiquant."))
        return redirect('competitions:club:practitioners')

@login_required
def practitioner_update(request, pk):
    """Modifier un pratiquant"""
    try:
        user_club = get_user_club(request)
        if not user_club:
            messages.error(request, _("Vous n'êtes associé à aucun club."))
            return redirect('competitions:dashboard:club')
        
        practitioner = get_object_or_404(Practitioner, pk=pk, organization=user_club)
        
        if not manual_permission_check(request.user, user_club):
            raise PermissionDenied(_("Vous n'avez pas l'autorisation de modifier ce pratiquant."))
        
        if request.method == 'POST':
            form = PractitionerForm(request.POST, request.FILES, instance=practitioner)
            if form.is_valid():
                practitioner = form.save()
                messages.success(request, _(f"Le profil de {practitioner.full_name} a été mis à jour."))
                return redirect('competitions:club:practitioner_detail', pk=practitioner.pk)
        else:
            form = PractitionerForm(instance=practitioner)
        
        context = {
            'form': form,
            'practitioner': practitioner,
            'club': user_club,
            'page_title': f"Modifier - {practitioner.full_name}",
        }
        
        return render(request, 'competitions/club/practitioner_form.html', context)
        
    except PermissionDenied:
        raise
    except Exception as e:
        logger.error(f"Erreur dans practitioner_update: {str(e)}")
        messages.error(request, _("Erreur lors de la modification du pratiquant."))
        return redirect('competitions:club:practitioners')

@login_required
@require_POST
def practitioner_delete(request, practitioner_id):
    """Supprimer un pratiquant"""
    try:
        user_club = get_user_club(request)
        if not user_club:
            return JsonResponse({'success': False, 'error': 'Club non trouvé'})
        
        practitioner = get_object_or_404(Practitioner, id=practitioner_id, organization=user_club)
        
        if not manual_permission_check(request.user, user_club):
            return JsonResponse({'success': False, 'error': 'Permissions insuffisantes'})
        
        practitioner_name = practitioner.full_name
        practitioner.delete()
        
        messages.success(request, f"Le pratiquant {practitioner_name} a été supprimé.")
        return redirect('competitions:club:practitioners')
        
    except Exception as e:
        logger.error(f"Erreur lors de la suppression du pratiquant {practitioner_id}: {str(e)}")
        messages.error(request, _("Erreur lors de la suppression."))
        return redirect('competitions:club:practitioners')

# Fonctions placeholder pour compatibilité
def create_user_for_practitioner(request, practitioner_id):
    """Créer un compte utilisateur pour un pratiquant"""
    messages.info(request, _("Fonctionnalité en cours de développement."))
    return redirect('competitions:club:practitioners')

def link_user_to_practitioner(request, practitioner_id):
    """Lier un utilisateur existant à un pratiquant"""
    messages.info(request, _("Fonctionnalité en cours de développement."))
    return redirect('competitions:club:practitioners')

def practitioner_qualifications_add(request, practitioner_id):
    """Ajouter une qualification à un pratiquant"""
    messages.info(request, _("Fonctionnalité en cours de développement."))
    return redirect('competitions:club:practitioners')

def practitioner_registrations(request, practitioner_id):
    """Voir les inscriptions d'un pratiquant"""
    messages.info(request, _("Fonctionnalité en cours de développement."))
    return redirect('competitions:club:practitioners')

def create_practitioner_registration(request, practitioner_id):
    """Créer une inscription pour un pratiquant"""
    messages.info(request, _("Fonctionnalité en cours de développement."))
    return redirect('competitions:club:practitioners')

def get_available_competitions(request, practitioner_id):
    """Obtenir les compétitions disponibles pour un pratiquant"""
    messages.info(request, _("Fonctionnalité en cours de développement."))
    return redirect('competitions:club:practitioners')