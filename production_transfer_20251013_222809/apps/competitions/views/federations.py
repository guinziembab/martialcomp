from django.core.exceptions import PermissionDenied
# Importations standard Django
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required

# Importation de notre helper de permission personnalisé
from apps.competitions.utils.permission_helpers import manual_permission_check, get_user_club, manual_check_federation_permission
from django.utils.translation import gettext_lazy as _
from django.db.models import ProtectedError
from django.db import transaction
from django.core.paginator import Paginator
from django.views.decorators.http import require_POST
from django.http import JsonResponse, HttpResponse, Http404, FileResponse
from django.urls import reverse
from django.utils import timezone
from django.template.exceptions import TemplateDoesNotExist
from django.contrib.auth.models import User
from django.db import models

# Configuration du logger
import logging
logger = logging.getLogger(__name__)

# Importations pour l'export de données
import csv
import io
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import xlwt
from io import BytesIO
import json
from datetime import timedelta

# Importations des formulaires
from ..forms.federations import FederationForm
from ..forms.club_forms import ClubAffiliationForm
from ..forms.club import ClubForm
from ..forms.competitions import CompetitionForm
from ..forms.categories import CompetitionCategoryForm, CategoryTemplateForm
from apps.grades.forms import GradeCategoryForm

# Importations des modèles
from ..models import (
    Federation,
    Club,
    Discipline,
    CompetitionCategory,
    CompetitionType,
    Competition,
    Practitioner,
    JudgeQualification,
    FederationAdministrator,
    Judge,
    JudgeTraining,
    JudgeTrainingRegistration,
)
# Importation correcte des modèles de l'application grades

# Importation directe des modèles de certification
from ..models.certifications import JudgeCertification, CertificationRegistration

# Importations des utilitaires
from ..utils.upload import handle_uploaded_image
from ..utils.decorators import federation_admin_required, federation_permission_required

# Importations des vues complémentaires
from .dashboard.federations import federation_dashboard


from ..forms.certifications import (
    JudgeCertificationForm, 
    CertificationRegistrationForm,
    ReviewCertificationRegistrationForm
)
from ..models.certifications import JudgeCertification, CertificationRegistration
from ..models.administrators import ClubAdministrator


from apps.grades.models import GradeCategory, Grade, GradingSystem
from apps.core.isolation import OrganizationIsolationMixin, get_organization_queryset
def federation_list(request):
    """Liste toutes les fédérations actives."""
    federations = Federation.objects.filter(is_active=True).order_by('name')
    
    # Filtrage par pays si spécifié
    country_filter = request.GET.get('country')
    if country_filter:
        federations = federations.filter(country__icontains=country_filter)
    
    # Recherche par nom si spécifiée
    search_query = request.GET.get('q')
    if search_query:
        federations = federations.filter(name__icontains=search_query)
    
    # Pagination
    paginator = Paginator(federations, 12)  # 12 fédérations par page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Liste des pays distincts pour le filtrage
    countries = Federation.objects.filter(is_active=True).values_list('country', flat=True).distinct().order_by('country')
    
    context = {
        'page_obj': page_obj,
        'countries': countries,
        'current_country': country_filter,
        'search_query': search_query,
        'total_federations': federations.count()
    }
    
    return render(request, 'competitions/federations/list.html', context)

def federation_detail(request, slug):
    """Affiche les détails d'une fédération spécifique."""
    federation = get_object_or_404(Federation, slug=slug, is_active=True)
    
    # Récupérer les clubs associés Ã  cette fédération via l'organisation
    clubs = Club.objects.filter(
        organization=federation.organization, 
        is_active=True
    ).order_by('name') if federation.organization else Club.objects.none()
    
    # Pagination des clubs
    paginator = Paginator(clubs, 10)  # 10 clubs par page
    page_number = request.GET.get('page')
    clubs_page = paginator.get_page(page_number)
    
    context = {
        'federation': federation,
        'clubs_page': clubs_page,
        'total_clubs': clubs.count()
    }
    
    return render(request, 'competitions/federations/detail.html', context)

@login_required
def federation_create(request):
    """Crée une nouvelle fédération et associe l'utilisateur comme administrateur principal."""
    # Vérifier si l'utilisateur a le droit de créer une fédération
    if not request.user.has_perm('competitions.add_federation'):
        if not hasattr(request.user, 'userprofile') or request.user.userprofile.role != 'federation_admin':
            messages.error(request, _("Vous n'avez pas la permission de créer une fédération."))
            return redirect('competitions:federations:list')
    
    if request.method == 'POST':
        form = FederationForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                with transaction.atomic():
                    # Créer la fédération avec l'utilisateur comme propriétaire
                    federation = form.save(commit=False)
                    federation.owner = request.user
                    
                    # Traitement du logo si fourni
                    if 'logo' in request.FILES:
                        federation.logo = handle_uploaded_image(request.FILES['logo'], 'federations/logos/')
                    
                    # Enregistrer la fédération
                    federation.save()
                    
                    # S'assurer que l'association administrateur est créée
                    # (ne pas compter uniquement sur le signal)
                    from apps.competitions.models import FederationAdministrator
                    FederationAdministrator.objects.get_or_create(
                        user=request.user,
                        federation=federation,
                        defaults={
                            'role': 'owner',
                            'is_primary': True
                        }
                    )
                    
                    # Créer manuellement l'Organisation associée
                    try:
                        from apps.organizations.models import Organization, OrganizationMember, OrganizationRole
                        
                        # Vérifier si l'organisation existe déjÃ 
                        if not Organization.objects.filter(old_federation_id=federation.id).exists():
                            org = Organization.objects.create(
                                name=federation.name,
                                organization_type='national_federation',
                                description=federation.description,
                                email=federation.contact_email,
                                phone=federation.contact_phone,
                                website=federation.website,
                                address=federation.address,
                                city=federation.city,
                                country=federation.country,
                                postal_code=federation.postal_code,
                                is_active=federation.is_active,
                                created_by=federation.owner,
                                old_federation_id=federation.id
                            )
                            
                            # Associer l'organisation Ã  la fédération
                            federation.organization = org
                            federation.save(update_fields=['organization'])
                            
                            # Associer l'utilisateur comme membre
                            OrganizationMember.objects.get_or_create(
                                organization=org,
                                user=request.user,
                                defaults={
                                    'role': OrganizationRole.OWNER,
                                    'can_manage_members': True,
                                    'can_edit_organization': True,
                                    'can_manage_competitions': True
                                }
                            )
                    except Exception as org_error:
                        # Log l'erreur mais continuer sans bloquer la création de fédération
                        logger.warning(f"Impossible de créer Organization pour Federation {federation.id}: {org_error}")
                    
                    # Mettre Ã  jour le profil de l'utilisateur si nécessaire
                    if hasattr(request.user, 'profile') and request.user.profile.role != 'federation_admin':
                        request.user.profile.role = 'federation_admin'
                        request.user.profile.save(update_fields=['role'])
                    
                    messages.success(request, _("Votre fédération a été créée avec succès ! Redirection vers le tableau de bord."))
                    
                    # Ajouter un message de log pour le débogage
                    import logging
                    logger = logging.getLogger('django')
                    logger.info(f"Fédération '{federation.name}' créée par l'utilisateur '{request.user.username}' (ID: {request.user.id})")
                    
                    return redirect('competitions:federations:federation_dashboard', federation_id=federation.id)
            except Exception as e:
                # Log détaillé de l'erreur pour faciliter le débogage
                import logging
                logger = logging.getLogger('django')
                logger.error(f"Erreur lors de la création de fédération: {str(e)}", exc_info=True)
                
                messages.error(request, _("Une erreur est survenue lors de la création de la fédération : {}").format(str(e)))
        else:
            # Amélioration de l'affichage des erreurs
            for field, errors in form.errors.items():
                for error in errors:
                    field_label = form[field].label or field
                    messages.error(request, f"{field_label}: {error}")
    else:
        form = FederationForm()
    
    context = {
        'form': form,
        'title': _('Créer une nouvelle fédération'),
        'submit_text': _('Créer la fédération')
    }
    
    return render(request, 'competitions/federations/form.html', context)

@login_required
def federation_update(request, slug):
    """Met Ã  jour une fédération existante."""
    federation = get_object_or_404(Federation, slug=slug)
    
    # Vérifier si l'utilisateur a le droit de modifier cette fédération
    if federation.owner != request.user and not request.user.has_perm('competitions.change_federation'):
        messages.error(request, _("Vous n'avez pas la permission de modifier cette fédération."))
        return redirect('competitions:federations:federation_dashboard', federation_id=federation.id)
    
    if request.method == 'POST':
        form = FederationForm(request.POST, request.FILES, instance=federation)
        if form.is_valid():
            try:
                with transaction.atomic():
                    federation = form.save(commit=False)
                    
                    # Traitement du logo si fourni
                    if 'logo' in request.FILES:
                        federation.logo = handle_uploaded_image(request.FILES['logo'], 'federations/logos/')
                    
                    federation.save()
                    messages.success(request, _("La fédération a été mise Ã  jour avec succès !"))
                    return redirect('competitions:federations:federation_dashboard', federation_id=federation.id)
            except Exception as e:
                messages.error(request, _("Une erreur est survenue lors de la mise Ã  jour de la fédération : {}").format(str(e)))
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        form = FederationForm(instance=federation)
    
    context = {
        'form': form,
        'federation': federation,
        'title': _('Modifier la fédération'),
        'submit_text': _('Enregistrer les modifications')
    }
    
    return render(request, 'competitions/federations/form.html', context)

@login_required
def federation_delete(request, pk):
    """Supprime une fédération ou la marque comme inactive."""
    federation = get_object_or_404(Federation, pk=pk)
    
    # Vérifier si l'utilisateur a le droit de supprimer cette fédération
    if federation.owner != request.user and not request.user.has_perm('competitions.delete_federation'):
        messages.error(request, _("Vous n'avez pas la permission de supprimer cette fédération."))
        return redirect('competitions:federations:federation_dashboard', federation_id=federation.id)
    
    if request.method != 'POST':
        # Si ce n'est pas une requÃªte POST, rediriger vers la page de détail
        return redirect('competitions:federations:federation_dashboard', federation_id=federation.id)
    
    try:
        # Option 1: Marquer comme inactive au lieu de supprimer
        federation.is_active = False
        federation.save()
        
        # Option 2: Supprimer réellement (décommenter si nécessaire)
        # federation.delete()
        
        messages.success(request, _("La fédération a été supprimée avec succès."))
        
        # Si appel AJAX, renvoyer une réponse JSON
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': True})
            
        return redirect('competitions:federations:list')
    except Exception as e:
        messages.error(request, _("Une erreur est survenue lors de la suppression : {}").format(str(e)))
        
        # Si appel AJAX, renvoyer une réponse JSON
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': str(e)})
            
        return redirect('competitions:federations:federation_dashboard', federation_id=federation.id)
    
@login_required
def federation_create_competition(request, federation_id):
    """Vue pour créer une compétition liée Ã  une fédération spécifique."""
    
    # Récupérer la fédération et vérifier les permissions
    federation = get_object_or_404(Federation, id=federation_id)
    
    # Vérifier que l'utilisateur est bien le propriétaire de la fédération
    if federation.owner != request.user:
        messages.error(request, _("Vous n'avez pas les droits pour créer une compétition pour cette fédération."))
        return redirect('competitions:dashboard:federations')
    
    if request.method == 'POST':
        form = CompetitionForm(request.POST, request.FILES)
        
        # Pour le débogage - afficher les données POST
        logger.debug(f"POST data: {request.POST}")
        
        # Vérifier si des types de compétition sont sélectionnés
        competition_types = request.POST.getlist('competition_types')
        logger.debug(f"Competition types selected: {competition_types}")
        
        if form.is_valid():
            try:
                with transaction.atomic():
                    competition = form.save(commit=False)
                    competition.federation = federation
                    
                    # Traitement du code postal si fourni
                    postal_code = form.cleaned_data.get('postal_code')
                    if postal_code and competition.address:
                        if not any(c.isdigit() for c in competition.address.split()[-1]):
                            competition.address = f"{competition.address}, {postal_code}"
                    
                    # Traitement du logo si fourni
                    if 'logo' in request.FILES:
                        competition.logo = handle_uploaded_image(
                            request.FILES['logo'], 
                            'competitions/logos/',
                            max_size=(800, 800)  # Redimensionner si nécessaire
                        )
                    
                    # Ajouter d'autres valeurs par défaut si nécessaire
                    if not competition.status:
                        competition.status = 'published'
                    
                    competition.save()
                    
                    # Sauvegarde des relations ManyToMany
                    if hasattr(form, 'save_m2m'):
                        form.save_m2m()
                    
                    messages.success(request, _("La compétition a été créée avec succès !"))
                    
                    # Rediriger vers la page détaillée de la compétition
                    # Correction du namespace
                    return redirect('competitions:detail', pk=competition.id)
                        
            except Exception as e:
                messages.error(request, _("Une erreur est survenue lors de la création de la compétition : {}").format(str(e)))
                logger.error(f"Erreur lors de la création de compétition: {str(e)}")
        else:
            logger.error(f"Form errors: {form.errors}")
            # Afficher les erreurs de validation
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        # Préinitialiser le formulaire avec la fédération
        form = CompetitionForm(initial={'federation': federation})
        
        # Si une discipline est spécifiée en paramètre, la présélectionner
        discipline_id = request.GET.get('discipline')
        if discipline_id:
            try:
                form.initial['discipline'] = int(discipline_id)
            except ValueError:
                pass
    
    # Préparer le contexte avec toutes les données nécessaires
    context = {
        'form': form,
        'federation': federation,
        'disciplines': Discipline.objects.filter(is_active=True).order_by('name'),
        'title': _('Créer une compétition'),
        'submit_text': _('Créer la compétition'),
        'form_id': 'competitionForm',  # ID pour le formulaire, important pour le JavaScript
        'is_edit': False
    }
    
    return render(request, 'competitions/federations/competition_form.html', context)


@login_required
def federation_add_club(request, federation_id):
    """Vue pour ajouter/créer un club dans une fédération."""
    federation = get_object_or_404(Federation, id=federation_id)
    
    # Vérifier que l'utilisateur est bien le propriétaire ou administrateur
    if federation.owner != request.user:
        messages.error(request, _("Vous n'avez pas les droits pour modifier cette fédération."))
        return redirect('competitions:federations:federation_dashboard', federation_id=federation.id)
    
    # Récupérer l'organisation de la fédération
    from apps.organizations.models import Organization, Affiliation
    if hasattr(federation, 'organization') and federation.organization:
        federation_org = federation.organization
    else:
        federation_org = Organization.objects.filter(old_federation_id=federation.id).first()
    
    # Gérer l'affiliation d'un club existant
    if request.method == 'POST' and 'existing_club' in request.POST:
        existing_club_id = request.POST.get('existing_club')
        if existing_club_id:
            try:
                club = Club.objects.get(id=existing_club_id, is_active=True)
                
                # Créer une organisation pour le club s'il n'en a pas déjà une
                if not club.organization:
                    from apps.organizations.models import Organization
                    club_org = Organization.objects.create(
                        name=club.name,
                        organization_type='club',
                        is_active=True
                    )
                    club.organization = club_org
                    club.save()
                
                # Créer l'affiliation
                if federation_org and club.organization:
                    from django.utils import timezone
                    affiliation, created = Affiliation.objects.get_or_create(
                        parent_organization=federation_org,
                        child_organization=club.organization,
                        defaults={
                            'is_active': True,
                            'start_date': timezone.now().date()
                        }
                    )
                    
                    if created:
                        messages.success(request, _("Le club {} a été affilié à votre fédération avec succès.").format(club.name))
                    else:
                        messages.warning(request, _("Ce club est déjà affilié à votre fédération."))
                else:
                    messages.error(request, _("Erreur lors de l'affiliation du club."))
                
                return redirect('competitions:federations:clubs', federation_id=federation.id)
                
            except Club.DoesNotExist:
                messages.error(request, _("Le club sélectionné n'existe pas."))
    
    # Gérer la création d'un nouveau club  
    elif request.method == 'POST':
        form = ClubAffiliationForm(request.POST, request.FILES)
        
        if form.is_valid():
            try:
                # Créer le club
                club = form.save(commit=False)
                club.owner = request.user
                club.save()
                
                # Sauvegarder les relations ManyToMany (disciplines)
                form.save_m2m()
                
                # Créer une organisation pour le club
                from apps.organizations.models import Organization
                club_org = Organization.objects.create(
                    name=club.name,
                    organization_type='club',
                    is_active=True
                )
                club.organization = club_org
                club.save()
                
                # Créer l'affiliation à la fédération
                if federation_org:
                    from django.utils import timezone
                    Affiliation.objects.create(
                        parent_organization=federation_org,
                        child_organization=club_org,
                        is_active=True,
                        start_date=timezone.now().date()
                    )
                
                messages.success(request, _("Le club {} a été créé et affilié à votre fédération avec succès.").format(club.name))
                return redirect('competitions:federations:clubs', federation_id=federation.id)
                
            except Exception as e:
                messages.error(request, _("Erreur lors de la création du club: {}").format(str(e)))
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, _("Erreur dans le champ {}: {}").format(field, error))
    else:
        form = ClubAffiliationForm()
    
    # Récupérer les clubs non affiliés pour l'onglet "Affilier un club existant"
    if federation_org:
        affiliated_org_ids = Affiliation.objects.filter(
            parent_organization=federation_org,
            is_active=True
        ).values_list('child_organization_id', flat=True)
        unaffiliated_clubs = Club.objects.filter(
            is_active=True,
            organization__isnull=True
        ) | Club.objects.filter(
            is_active=True,
            organization__isnull=False
        ).exclude(organization_id__in=affiliated_org_ids)
    else:
        unaffiliated_clubs = Club.objects.filter(is_active=True)
    
    context = {
        'form': form,
        'federation': federation,
        'unaffiliated_clubs': unaffiliated_clubs[:50],  # Limiter pour les performances
    }
    
    return render(request, 'competitions/federations/add_club.html', context)

@login_required
def federation_manage_clubs(request, federation_id):
    """Vue pour gérer les clubs d'une fédération."""
    federation = get_object_or_404(Federation, id=federation_id, owner=request.user)
    
    # Récupérer l'organisation de la fédération
    from apps.organizations.models import Organization, Affiliation
    if hasattr(federation, 'organization') and federation.organization:
        federation_org = federation.organization
    else:
        federation_org = Organization.objects.filter(old_federation_id=federation.id).first()
    
    if not federation_org:
        messages.error(request, _("La fédération n'a pas d'organisation associée."))
        return redirect('competitions:federations:federation_dashboard', federation_id=federation.id)
    
    # Récupérer les organisations affiliées de type club
    affiliated_org_ids = Affiliation.objects.filter(
        parent_organization=federation_org,
        is_active=True
    ).values_list('child_organization_id', flat=True)
    
    # Récupérer les clubs associés Ã  ces organisations
    clubs = Club.objects.filter(organization_id__in=affiliated_org_ids).order_by('name')
    
    # Debug: afficher le nombre de clubs trouvés
    print(f"DEBUG: Federation {federation.name} - Found {clubs.count()} affiliated clubs")
    print(f"DEBUG: Affiliated org IDs: {list(affiliated_org_ids)}")
    
    # Récupérer les disciplines disponibles pour les filtres
    disciplines = clubs.values_list('disciplines__name', flat=True).distinct()
    
    # Calculer des statistiques
    total_practitioners = sum(club.practitioners.count() for club in clubs)
    cities_count = clubs.values_list('city', flat=True).distinct().count()
    
    context = {
        'federation': federation,
        'clubs': clubs,
        'affiliated_clubs': clubs,  # Alias pour le template
        'disciplines': disciplines,
        'total_practitioners': total_practitioners,
        'cities_count': cities_count,
        'stats': {
            'clubs_count': clubs.count(),
            'practitioners_count': total_practitioners,
            'cities_count': cities_count,
        }
    }
    
    return render(request, 'competitions/federations/manage_clubs.html', context)

@login_required
@require_POST
def federation_remove_club(request, federation_id, club_id):
    """Retirer un club d'une fédération."""
    federation = get_object_or_404(Federation, id=federation_id, owner=request.user)
    club = get_object_or_404(Club, id=club_id)
    
    # Récupérer l'organisation de la fédération
    from apps.organizations.models import Organization, Affiliation
    if hasattr(federation, 'organization') and federation.organization:
        federation_org = federation.organization
    else:
        federation_org = Organization.objects.filter(old_federation_id=federation.id).first()
    
    if not federation_org:
        messages.error(request, _("La fédération n'a pas d'organisation associée."))
        return redirect('competitions:federations:manage_clubs', federation_id=federation.id)
    
    # Vérifier que le club est bien affilié Ã  cette fédération
    affiliation = Affiliation.objects.filter(
        parent_organization=federation_org,
        child_organization=club.organization,
        is_active=True
    ).first()
    
    if not affiliation:
        messages.error(request, _("Ce club n'est pas affilié Ã  votre fédération."))
        return redirect('competitions:federations:manage_clubs', federation_id=federation.id)
    
    # Désactiver l'affiliation
    affiliation.is_active = False
    affiliation.save()
    
    messages.success(request, _("Le club {} a été retiré de votre fédération.").format(club.name))
    return redirect('competitions:federations:manage_clubs', federation_id=federation.id)

@login_required
def federation_club_detail(request, federation_id, club_id):
    """Vue détaillée pour gérer un club spécifique depuis la fédération."""
    federation = get_object_or_404(Federation, id=federation_id, owner=request.user)
    club = get_object_or_404(Club, id=club_id)
    
    # Récupérer l'organisation de la fédération
    from apps.organizations.models import Organization, Affiliation
    if hasattr(federation, 'organization') and federation.organization:
        federation_org = federation.organization
    else:
        federation_org = Organization.objects.filter(old_federation_id=federation.id).first()
    
    # Vérifier que le club est bien affilié à cette fédération
    affiliation = None
    if federation_org and club.organization:
        affiliation = Affiliation.objects.filter(
            parent_organization=federation_org,
            child_organization=club.organization,
            is_active=True
        ).first()
    
    if not affiliation:
        messages.error(request, _("Ce club n'est pas affilié à votre fédération."))
        return redirect('competitions:federations:clubs', federation_id=federation.id)
    
    # Gérer la modification du club si POST
    if request.method == 'POST':
        from ..forms.club_forms import ClubAffiliationForm
        form = ClubAffiliationForm(request.POST, request.FILES, instance=club)
        
        if form.is_valid():
            try:
                club = form.save()
                messages.success(request, _("Les informations du club {} ont été mises à jour.").format(club.name))
                return redirect('competitions:federations:club_detail', federation_id=federation.id, club_id=club.id)
            except Exception as e:
                messages.error(request, _("Erreur lors de la mise à jour: {}").format(str(e)))
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, _("Erreur dans le champ {}: {}").format(field, error))
    else:
        from ..forms.club_forms import ClubAffiliationForm
        form = ClubAffiliationForm(instance=club)
    
    # Récupérer les statistiques du club
    practitioners_count = getattr(club, 'practitioners', club).count() if hasattr(club, 'practitioners') else 0
    active_competitions = 0  # À implémenter selon votre modèle
    
    # Récupérer les disciplines disponibles pour le formulaire d'édition
    all_disciplines = get_organization_queryset(Discipline, self.request.user).order_by('name')
    
    context = {
        'federation': federation,
        'club': club,
        'affiliation': affiliation,
        'form': form,
        'practitioners_count': practitioners_count,
        'active_competitions': active_competitions,
        'all_disciplines': all_disciplines,
        'can_edit': True,  # La fédération peut éditer les clubs qu'elle a créés
    }
    
    return render(request, 'competitions/federations/club_detail.html', context)

@login_required
def federation_settings(request, federation_id):
    """Vue pour modifier les paramètres d'une fédération."""
    federation = get_object_or_404(Federation, id=federation_id, owner=request.user)
    
    if request.method == 'POST':
        form = FederationForm(request.POST, request.FILES, instance=federation)
        if form.is_valid():
            try:
                federation = form.save(commit=False)
                
                # Traitement du logo si présent
                if 'logo' in request.FILES:
                    federation.logo = handle_uploaded_image(request.FILES['logo'], 'federations/logos/')
                
                federation.save()
                
                messages.success(request, _("Les paramètres de la fédération ont été mis Ã  jour avec succès !"))
                return redirect('competitions:federations:federation_dashboard', federation_id=federation.id)
            except Exception as e:
                messages.error(request, _("Une erreur est survenue lors de la mise Ã  jour : {}").format(str(e)))
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        form = FederationForm(instance=federation)
    
    context = {
        'form': form,
        'federation': federation,
        'title': _('Paramètres de la fédération'),
        'submit_text': _('Enregistrer les modifications')
    }
    
    return render(request, 'competitions/federations/settings.html', context)

@login_required
def federation_export_data(request, federation_id):
    """Vue pour exporter les données d'une fédération."""
    federation = get_object_or_404(Federation, id=federation_id, owner=request.user)
    
    export_type = request.GET.get('type', 'clubs')
    
    if export_type == 'clubs':
        # Exporter les clubs
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="federation_{federation.id}_clubs.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Nom', 'Ville', 'Adresse', 'Email', 'Téléphone', 'Responsable'])
        
        clubs = Club.objects.filter(federation=federation)
        for club in clubs:
            writer.writerow([
                club.name,
                club.city,
                club.address,
                club.contact_email,
                club.contact_phone,
                club.owner.get_full_name() if club.owner else ''
            ])
        
        return response
    
    elif export_type == 'competitions':
        # Exporter les compétitions
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="federation_{federation.id}_competitions.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Titre', 'Date début', 'Date fin', 'Lieu', 'Statut', 'Participants'])
        
        competitions = Competition.objects.filter(organizing_organization__old_federation_id=federation.id)
        for competition in competitions:
            writer.writerow([
                competition.title,
                competition.start_date,
                competition.end_date,
                competition.address,
                competition.status,
                competition.participants.count()
            ])
        
        return response
    
    # Si on arrive ici, on affiche la page de choix d'export
    return render(request, 'competitions/federations/export.html', {
        'federation': federation
    })
    
@login_required
def select_club(request, competition_id):
    """Permet Ã  un admin de fédération de sélectionner un club pour inscrire un participant."""
    # Récupérer les fédérations dont l'utilisateur est administrateur
    administered_federations = request.user.get_administered_federations()
    
    if not administered_federations.exists():
        messages.error(request, _("Vous n'Ãªtes administrateur d'aucune fédération."))
        return redirect('competitions:dashboard:index')
    
    # Si une fédération est spécifiée dans l'URL
    federation_id = request.GET.get('federation_id')
    if federation_id:
        try:
            federation = Federation.objects.get(pk=federation_id)
            if not request.user.is_federation_admin(federation):
                messages.error(request, _("Vous n'Ãªtes pas administrateur de cette fédération."))
                federation = administered_federations.first()
        except Federation.DoesNotExist:
            messages.error(request, _("Fédération non trouvée."))
            federation = administered_federations.first()
    else:
        # Utiliser la première fédération administrée par défaut
        federation = administered_federations.first()
    
    # Récupérer la compétition
    try:
        competition = get_object_or_404(Competition, pk=competition_id)
    except (ValueError, Competition.DoesNotExist):
        messages.error(request, _("Compétition non trouvée."))
        return redirect('competitions:list')
    
    # Récupérer les clubs affiliés Ã  cette fédération
    clubs = Club.objects.filter(federation=federation).order_by('name')
    
    # Si l'utilisateur administre plusieurs fédérations, les afficher pour permettre de changer
    other_federations = None
    if administered_federations.count() > 1:
        other_federations = administered_federations.exclude(pk=federation.pk)
    
    return render(request, 'competitions/federations/select_club.html', {
        'competition': competition,
        'clubs': clubs,
        'federation': federation,
        'other_federations': other_federations
    })
    
def federation_calendar(request, slug):
    """
    Affiche le calendrier des événements pour une fédération spécifique.
    """
    federation = get_object_or_404(Federation, slug=slug)
    
    # Récupérer les clubs appartenant Ã  cette fédération
    clubs_ids = Club.objects.filter(federation=federation).values_list('id', flat=True)
    
    # Récupérer les compétitions organisées par ces clubs
    events = Competition.objects.filter(organizing_club_id__in=clubs_ids)
    
    # Alternative: Si vous avez une relation indirecte via discipline
    # Vérifiez si les fédérations sont liées aux disciplines
    # federation_disciplines = federation.disciplines.all()
    # events = Competition.objects.filter(discipline__in=federation_disciplines)
    
    context = {
        'federation': federation,
        'events': events,
    }
    
    return render(request, 'competitions/federations/calendar.html', context)

def federation_competitions(request, federation_id):
    """
    Affiche toutes les compétitions pour une fédération spécifique.
    """
    federation = get_object_or_404(Federation, pk=federation_id)
    
    # Récupérer les disciplines gérées par cette fédération
    disciplines = federation.disciplines.all()
    
    # Récupérer l'organisation de la fédération
    from apps.organizations.models import Organization, Affiliation
    if hasattr(federation, 'organization') and federation.organization:
        federation_org = federation.organization
    else:
        federation_org = Organization.objects.filter(old_federation_id=federation.id).first()
    
    # Récupérer les organisations affiliées
    affiliated_org_ids = []
    if federation_org:
        affiliated_org_ids = Affiliation.objects.filter(
            parent_organization=federation_org,
            is_active=True
        ).values_list('child_organization_id', flat=True)
    
    # Récupérer les compétitions liées Ã  la fédération via les disciplines ou les organisations
    competitions = Competition.objects.filter(
        models.Q(discipline__in=disciplines) | 
        models.Q(organizing_organization_id__in=affiliated_org_ids)
    ).distinct().order_by('-start_date')
    
    # Récupérer les clubs affiliés pour les statistiques
    affiliated_clubs = Club.objects.filter(organization_id__in=affiliated_org_ids) if affiliated_org_ids else Club.objects.none()
    
    # Compter les praticiens liés aux organisations affiliées
    practitioner_count = Practitioner.objects.filter(
        organization_id__in=affiliated_org_ids
    ).count() if affiliated_org_ids else 0
    
    # Compter les juges (utiliser l'organisation si le modèle Judge n'a pas de lien direct avec federation)
    judge_count = Judge.objects.filter(
        organization_id__in=affiliated_org_ids
    ).count() if affiliated_org_ids else 0
    
    context = {
        'federation': federation,
        'competitions': competitions,
        # Ajout des compteurs pour le sidebar
        'competition_count': competitions.count(),
        'club_count': affiliated_clubs.count(),
        'practitioner_count': practitioner_count,
        'judge_count': judge_count
    }
    
    return render(request, 'competitions/federations/competitions.html', context)

@login_required
def federation_categories(request, federation_id):
    """Affiche et gère les catégories de compétition pour une fédération."""
    federation = get_object_or_404(Federation, id=federation_id)
    
    print(f"DEBUG: Federation categories - Federation: {federation.name}")
    
    # Récupérer toutes les catégories disponibles pour débogage
    all_categories = CompetitionCategory.objects.all()
    print(f"DEBUG: Total categories in database: {all_categories.count()}")
    
    # Récupérer les disciplines de la fédération
    disciplines = federation.disciplines.all()
    print(f"DEBUG: Federation disciplines: {[d.name for d in disciplines]}")
    
    # Pour l'instant, récupérer toutes les catégories pour voir si ça fonctionne
    categories = all_categories[:20]  # Limiter à 20 pour le test
    
    context = {
        'federation': federation,
        'categories': categories,
        'disciplines': disciplines,
        'debug_info': {
            'total_categories': all_categories.count(),
            'federation_disciplines': [d.name for d in disciplines],
        }
    }
    
    return render(request, 'competitions/federations/categories.html', context)

def federation_roles(request, federation_id):
    """Gestion des rÃ´les et permissions de la fédération."""
    federation = get_object_or_404(Federation, id=federation_id)
    # Logique de la vue...
    context = {
        'federation': federation,
    }
    return render(request, 'competitions/federations/roles.html', context)

def federation_judges(request, federation_id):
    """Gestion des juges et arbitres de la fédération."""
    federation = get_object_or_404(Federation, id=federation_id)
    
    # Détecter si nous sommes en mode "ajout"
    action = request.GET.get('action')
    
    if action == 'add':
        # Logique pour l'ajout d'un juge
        # Si c'est un POST, traiter le formulaire
        # Sinon, afficher le formulaire vide
        if request.method == 'POST':
            # Traiter le formulaire
            # Rediriger vers la liste des juges après succès
            pass
        
        # Vous pourriez mÃªme utiliser un template différent pour l'ajout
        context = {
            'federation': federation,
            'is_add_mode': True,
            # Ajouter un formulaire si nécessaire
        }
        return render(request, 'competitions/federations/judge_form.html', context)
    
    # Logique habituelle pour la vue de liste des juges
    context = {
        'federation': federation,
        # Autres données de contexte
    }
    return render(request, 'competitions/federations/judges.html', context)

@login_required
def federation_manage_users(request, federation_id):
    # Vérifier manuellement les permissions
    if not manual_check_federation_permission(request.user, 'competitions.change_federation'):
        messages.error(request, _("Vous n'avez pas les permissions nécessaires pour cette action."))
        return redirect('competitions:dashboard:index')
    """
    Vue permettant de gérer les utilisateurs associés Ã  une fédération.
    Les administrateurs peuvent ajouter/supprimer des utilisateurs et modifier leurs rÃ´les.
    """
    federation = get_object_or_404(Federation, id=federation_id)
    
    # Vérifier si l'utilisateur a les droits d'administration sur cette fédération
    if not request.user.is_federation_admin(federation) and not request.user.is_staff:
        raise PermissionDenied(_("Vous n'avez pas les droits pour gérer les utilisateurs de cette fédération."))
    
    # Obtenir les administrateurs actuels
    administrators = FederationAdministrator.objects.filter(federation=federation).select_related('user')
    
    # Traitement du formulaire d'ajout
    if request.method == 'POST':
        if 'add_admin' in request.POST:
            # Ajouter un nouvel administrateur
            user_id = request.POST.get('user_id')
            role = request.POST.get('role', 'admin')
            
            if user_id:
                try:
                    user = User.objects.get(id=user_id)
                    
                    # Vérifier si l'utilisateur est déjÃ  administrateur
                    if not FederationAdministrator.objects.filter(user=user, federation=federation).exists():
                        is_primary = request.POST.get('is_primary') == 'on'
                        
                        # Si on définit un nouvel admin principal, il faut retirer ce statut aux autres
                        if is_primary:
                            FederationAdministrator.objects.filter(
                                federation=federation, 
                                is_primary=True
                            ).update(is_primary=False)
                        
                        admin = FederationAdministrator.objects.create(
                            user=user,
                            federation=federation,
                            role=role,
                            is_primary=is_primary
                        )
                        
                        messages.success(
                            request, 
                            _("L'utilisateur {} a été ajouté comme {} de la fédération.").format(
                                user.get_full_name() or user.username, 
                                admin.get_role_display().lower()
                            )
                        )
                        
                        # Mettre Ã  jour le rÃ´le utilisateur si nécessaire
                        if hasattr(user, 'profile') and user.profile.role != 'federation_admin':
                            user.profile.role = 'federation_admin'
                            user.profile.save()
                    else:
                        messages.warning(request, _("Cet utilisateur est déjÃ  administrateur de cette fédération."))
                
                except User.DoesNotExist:
                    messages.error(request, _("L'utilisateur sélectionné n'existe pas."))
        
        elif 'remove_admin' in request.POST:
            # Supprimer un administrateur
            admin_id = request.POST.get('admin_id')
            
            if admin_id:
                try:
                    admin = FederationAdministrator.objects.get(id=admin_id, federation=federation)
                    
                    # EmpÃªcher la suppression du dernier administrateur principal
                    if admin.is_primary and FederationAdministrator.objects.filter(federation=federation, is_primary=True).count() <= 1:
                        messages.error(request, _("Impossible de supprimer le dernier administrateur principal."))
                    else:
                        user = admin.user
                        admin.delete()
                        
                        messages.success(
                            request, 
                            _("{} n'est plus {} de la fédération.").format(
                                user.get_full_name() or user.username, 
                                admin.get_role_display().lower()
                            )
                        )
                except FederationAdministrator.DoesNotExist:
                    messages.error(request, _("L'administrateur sélectionné n'existe pas."))
        
        elif 'update_admin' in request.POST:
            # Mettre Ã  jour le rÃ´le d'un administrateur
            admin_id = request.POST.get('admin_id')
            new_role = request.POST.get('role')
            is_primary = request.POST.get('is_primary') == 'on'
            
            if admin_id and new_role:
                try:
                    admin = FederationAdministrator.objects.get(id=admin_id, federation=federation)
                    
                    # Si on fait de cet admin un admin principal
                    if is_primary and not admin.is_primary:
                        # Retirer ce statut aux autres
                        FederationAdministrator.objects.filter(
                            federation=federation, 
                            is_primary=True
                        ).update(is_primary=False)
                    
                    admin.role = new_role
                    admin.is_primary = is_primary
                    admin.save()
                    
                    messages.success(
                        request, 
                        _("Le rÃ´le de {} a été mis Ã  jour Ã  {}.").format(
                            admin.user.get_full_name() or admin.user.username, 
                            admin.get_role_display().lower()
                        )
                    )
                except FederationAdministrator.DoesNotExist:
                    messages.error(request, _("L'administrateur sélectionné n'existe pas."))
    
    # Recherche d'utilisateurs pour l'ajout (exclusion des admins existants)
    search_query = request.GET.get('q', '')
    if search_query:
        # Exclure les utilisateurs qui sont déjÃ  administrateurs
        existing_admin_ids = administrators.values_list('user_id', flat=True)
        
        users = User.objects.filter(
            Q(username__icontains=search_query) | 
            Q(email__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query)
        ).exclude(id__in=existing_admin_ids)[:10]  # Limiter Ã  10 résultats pour la performance
    else:
        users = User.objects.none()
    
    # Contexte pour le template
    context = {
        'federation': federation,
        'administrators': administrators,
        'users': users,
        'search_query': search_query,
        'roles': FederationAdministrator._meta.get_field('role').choices
    }
    
    return render(request, 'competitions/federations/manage_users.html', context)

@login_required
def import_export(request, federation_id):
    """Vue principale pour l'import/export des données de la fédération."""
    federation = get_object_or_404(Federation, id=federation_id)
    
    # Ajouter les statistiques pour le menu latéral
    context = {
        'federation': federation,
        'title': _('Import/Export - {0}').format(federation.name),
        'competition_count': Competition.objects.filter(
            discipline__in=federation.disciplines.all()
        ).count(),
        'club_count': Club.objects.filter(federation=federation).count(),
        'practitioner_count': Practitioner.objects.filter(
            organization__old_federation_id=federation.id
        ).count(),
        'judge_count': Judge.objects.filter(
            practitioner__organization__old_federation_id=federation.id
        ).count(),
    }
    
    return render(request, 'competitions/federations/import_export.html', context)


@login_required
def export_clubs(request, federation_id):
    """Exporte les clubs de la fédération au format CSV/Excel."""
    federation = get_object_or_404(Federation, id=federation_id)
    clubs = Club.objects.filter(federation=federation)
    export_format = request.GET.get('format', 'csv')
    
    if export_format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="clubs_{federation.slug}.csv"'
        
        writer = csv.writer(response)
        # En-tÃªtes
        writer.writerow([
            'Nom', 'Adresse', 'Ville', 'Code Postal', 'Email', 'Téléphone', 
            'Site Web', 'Description', 'Nombre de Pratiquants'
        ])
        
        # Données
        for club in clubs:
            practitioner_count = club.practitioners.count()
            writer.writerow([
                club.name,
                club.address,
                club.city,
                club.postal_code,
                club.contact_email,
                club.contact_phone,
                club.website,
                club.description,
                practitioner_count
            ])
    
    elif export_format == 'excel':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="clubs_{federation.slug}.xlsx"'
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Clubs"
        
        # Style d'en-tÃªte
        header_fill = PatternFill(start_color="00CCCCCC", end_color="00CCCCCC", fill_type="solid")
        header_font = Font(bold=True)
        
        # En-tÃªtes
        headers = [
            'Nom', 'Adresse', 'Ville', 'Code Postal', 'Email', 'Téléphone', 
            'Site Web', 'Description', 'Nombre de Pratiquants'
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Données
        for row, club in enumerate(clubs, 2):
            practitioner_count = club.practitioners.count()
            ws.cell(row=row, column=1, value=club.name)
            ws.cell(row=row, column=2, value=club.address)
            ws.cell(row=row, column=3, value=club.city)
            ws.cell(row=row, column=4, value=club.postal_code)
            ws.cell(row=row, column=5, value=club.contact_email)
            ws.cell(row=row, column=6, value=club.contact_phone)
            ws.cell(row=row, column=7, value=club.website)
            ws.cell(row=row, column=8, value=club.description)
            ws.cell(row=row, column=9, value=practitioner_count)
        
        # Ajuster la largeur des colonnes
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        
        wb.save(response)
    
    return response


@login_required
def export_practitioners(request, federation_id):
    """Exporte les pratiquants de la fédération au format CSV/Excel."""
    federation = get_object_or_404(Federation, id=federation_id)
    practitioners = Practitioner.objects.filter(organization__old_federation_id=federation.id)
    export_format = request.GET.get('format', 'csv')
    
    if export_format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="pratiquants_{federation.slug}.csv"'
        
        writer = csv.writer(response)
        # En-tÃªtes
        writer.writerow([
            'Prénom', 'Nom', 'Club', 'Email', 'Téléphone', 'Date de naissance', 
            'Grade', 'Numéro de licence', 'Date du certificat médical'
        ])
        
        # Données
        for practitioner in practitioners:
            writer.writerow([
                practitioner.first_name,
                practitioner.last_name,
                practitioner.organization.name if practitioner.organization else '',
                practitioner.email,
                practitioner.phone,
                practitioner.birth_date.strftime('%d/%m/%Y') if practitioner.birth_date else '',
                practitioner.grade,
                practitioner.license_number,
                practitioner.medical_certificate_date.strftime('%d/%m/%Y') if practitioner.medical_certificate_date else ''
            ])
    
    elif export_format == 'excel':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="pratiquants_{federation.slug}.xlsx"'
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pratiquants"
        
        # Style d'en-tÃªte
        header_fill = PatternFill(start_color="00CCCCCC", end_color="00CCCCCC", fill_type="solid")
        header_font = Font(bold=True)
        
        # En-tÃªtes
        headers = [
            'Prénom', 'Nom', 'Club', 'Email', 'Téléphone', 'Date de naissance', 
            'Grade', 'Numéro de licence', 'Date du certificat médical'
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Données
        for row, practitioner in enumerate(practitioners, 2):
            ws.cell(row=row, column=1, value=practitioner.first_name)
            ws.cell(row=row, column=2, value=practitioner.last_name)
            ws.cell(row=row, column=3, value=practitioner.organization.name if practitioner.organization else '')
            ws.cell(row=row, column=4, value=practitioner.email)
            ws.cell(row=row, column=5, value=practitioner.phone)
            ws.cell(row=row, column=6, value=practitioner.birth_date.strftime('%d/%m/%Y') if practitioner.birth_date else '')
            ws.cell(row=row, column=7, value=practitioner.grade)
            ws.cell(row=row, column=8, value=practitioner.license_number)
            ws.cell(row=row, column=9, value=practitioner.medical_certificate_date.strftime('%d/%m/%Y') if practitioner.medical_certificate_date else '')
        
        # Ajuster la largeur des colonnes
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        
        wb.save(response)
    
    return response


@login_required
def import_clubs(request, federation_id):
    """Importe des clubs Ã  partir d'un fichier CSV/Excel."""
    federation = get_object_or_404(Federation, id=federation_id)
    
    if request.method == 'POST' and 'import_file' in request.FILES:
        file = request.FILES['import_file']
        
        if file.name.endswith('.csv'):
            # Importation CSV
            decoded_file = file.read().decode('utf-8')
            io_string = io.StringIO(decoded_file)
            reader = csv.DictReader(io_string)
            
            success_count = 0
            error_count = 0
            
            with transaction.atomic():
                for row in reader:
                    try:
                        club = Club.objects.create(
                            federation=federation,
                            name=row.get('Nom', ''),
                            address=row.get('Adresse', ''),
                            city=row.get('Ville', ''),
                            postal_code=row.get('Code Postal', ''),
                            contact_email=row.get('Email', ''),
                            contact_phone=row.get('Téléphone', ''),
                            website=row.get('Site Web', ''),
                            description=row.get('Description', '')
                        )
                        success_count += 1
                    except Exception as e:
                        error_count += 1
                        continue
            
            messages.success(request, _('{} clubs importés avec succès.').format(success_count))
            if error_count > 0:
                messages.warning(request, _('{} erreurs lors de l\'importation.').format(error_count))
        
        elif file.name.endswith('.xlsx'):
            # Importation Excel
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            
            success_count = 0
            error_count = 0
            
            with transaction.atomic():
                for row in ws.iter_rows(min_row=2, values_only=True):
                    try:
                        if all(cell is None for cell in row):
                            continue
                            
                        club = Club.objects.create(
                            federation=federation,
                            name=row[0] if row[0] else '',
                            address=row[1] if row[1] else '',
                            city=row[2] if row[2] else '',
                            postal_code=row[3] if row[3] else '',
                            contact_email=row[4] if row[4] else '',
                            contact_phone=row[5] if row[5] else '',
                            website=row[6] if row[6] else '',
                            description=row[7] if row[7] else ''
                        )
                        success_count += 1
                    except Exception as e:
                        error_count += 1
                        continue
            
            messages.success(request, _('{} clubs importés avec succès.').format(success_count))
            if error_count > 0:
                messages.warning(request, _('{} erreurs lors de l\'importation.').format(error_count))
    
    return redirect('competitions:federations:import_export', federation_id=federation.id)


@login_required
@federation_admin_required
def download_sample_clubs(request, federation_id):
    """Télécharge un exemple de fichier pour l'import de clubs."""
    export_format = request.GET.get('format', 'csv')
    
    if export_format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="exemple_clubs.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            'Nom', 'Adresse', 'Ville', 'Code Postal', 'Email', 'Téléphone', 
            'Site Web', 'Description'
        ])
        writer.writerow([
            'Club Martial Paris', '123 Rue des Arts Martiaux', 'Paris', '75001', 
            'contact@club-martial-paris.fr', '0123456789', 'https://club-martial-paris.fr', 
            'Club d\'arts martiaux basé Ã  Paris'
        ])
    
    elif export_format == 'excel':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="exemple_clubs.xlsx"'
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Exemple Clubs"
        
        headers = [
            'Nom', 'Adresse', 'Ville', 'Code Postal', 'Email', 'Téléphone', 
            'Site Web', 'Description'
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        example_data = [
            'Club Martial Paris', '123 Rue des Arts Martiaux', 'Paris', '75001', 
            'contact@club-martial-paris.fr', '0123456789', 'https://club-martial-paris.fr', 
            'Club d\'arts martiaux basé Ã  Paris'
        ]
        for col, value in enumerate(example_data, 1):
            ws.cell(row=2, column=col, value=value)
        
        wb.save(response)
    
    return response

@login_required
@federation_admin_required
def federation_grades(request, federation_id):
    """
    Vue pour la gestion des grades et certifications d'une fédération.
    Permet de visualiser, ajouter, modifier et supprimer des grades.
    """
    federation = get_object_or_404(Federation, id=federation_id)
    
    # Récupération des catégories de grades existantes
    # Correction: Utiliser GradeCategory au lieu de CategoryGrade
    grade_categories = GradeCategory.objects.filter(discipline__in=federation.disciplines.all())
    
    # Récupérer la discipline sélectionnée si présente
    selected_discipline_id = request.GET.get('discipline')
    selected_discipline = None
    selected_discipline_name = None
    discipline_grades = []
    
    if selected_discipline_id:
        # Appliquer le filtre
        grade_categories = grade_categories.filter(discipline_id=selected_discipline_id)
        selected_discipline = get_object_or_404(Discipline, id=selected_discipline_id)
        selected_discipline_name = selected_discipline.name
        
        # Récupérer les grades spécifiques Ã  cette discipline pour les listes déroulantes
        try:
            # Vérifier si GradeSystem existe, sinon utiliser une approche alternative
            if 'GradeSystem' in globals():
                # Chercher le système de grades pour cette discipline
                grade_system = GradeSystem.objects.filter(discipline=selected_discipline).first()
                if grade_system:
                    # Récupérer les catégories puis les grades
                    categories = GradeCategory.objects.filter(system=grade_system)
                    discipline_grades = Grade.objects.filter(category__in=categories).order_by('level')
            else:
                # Approche alternative si GradeSystem n'existe pas
                categories = GradeCategory.objects.filter(discipline=selected_discipline)
                discipline_grades = Grade.objects.filter(category__in=categories).order_by('level')
        except Exception as e:
            messages.warning(request, _("Erreur lors de la récupération des grades: {}").format(str(e)))
    
    # Importer le formulaire au moment oÃ¹ il est nécessaire
    from apps.grades.forms import GradeCategoryForm as CategoryGradeForm
    
    # Formulaire d'ajout de grade
    if request.method == 'POST':
        if 'add_grade' in request.POST:
            form = CategoryGradeForm(request.POST)
            if form.is_valid():
                grade = form.save(commit=False)
                # Vérifier que la discipline appartient Ã  la fédération
                if grade.discipline in federation.disciplines.all():
                    # Vérifier si les champs grade_min et grade_max existent
                    if hasattr(grade, 'grade_min') and hasattr(grade, 'grade_max'):
                        # Récupérer les valeurs pour grade_min et grade_max
                        grade.grade_min = request.POST.get('grade_min', '')
                        grade.grade_max = request.POST.get('grade_max', '')
                        
                        # Si grade_min et grade_max sont spécifiés, vérifier que grade_min < grade_max
                        if grade.grade_min and grade.grade_max:
                            try:
                                # Pour les vérifications d'ordre, on pourrait utiliser les systèmes de grades
                                # et vérifier les niveaux, mais pour simplifier, on accepte la validation du formulaire
                                pass
                            except Exception as e:
                                form.add_error('grade_max', _("Erreur de validation des grades: {}").format(str(e)))
                                form.is_valid = lambda: False
                    
                    # Si tout est valide, sauvegarder
                    if form.is_valid():
                        grade.save()
                        messages.success(request, _("Catégorie de grade ajoutée avec succès"))
                        
                        # Rediriger vers la mÃªme page avec le filtre de discipline maintenu
                        redirect_url = reverse('competitions:federations:grades', kwargs={'federation_id': federation_id})
                        if selected_discipline_id:
                            redirect_url += f"?discipline={selected_discipline_id}"
                        return redirect(redirect_url)
                else:
                    messages.error(request, _("La discipline sélectionnée n'appartient pas Ã  cette fédération"))
        
        elif 'delete_grade' in request.POST:
            grade_id = request.POST.get('grade_id')
            # Correction: Utiliser GradeCategory au lieu de CategoryGrade
            grade = get_object_or_404(GradeCategory, id=grade_id)
            # Vérifier que le grade appartient Ã  une discipline de la fédération
            if grade.discipline in federation.disciplines.all():
                try:
                    grade.delete()
                    messages.success(request, _("Catégorie de grade supprimée avec succès"))
                except ProtectedError:
                    messages.error(request, _("Impossible de supprimer cette catégorie car elle est utilisée par des pratiquants"))
            else:
                messages.error(request, _("Vous n'avez pas les permissions pour supprimer cette catégorie de grade"))
            
            # Rediriger vers la mÃªme page avec le filtre de discipline maintenu
            redirect_url = reverse('competitions:federations:grades', kwargs={'federation_id': federation_id})
            if selected_discipline_id:
                redirect_url += f"?discipline={selected_discipline_id}"
            return redirect(redirect_url)
    else:
        form = CategoryGradeForm(initial={'discipline': selected_discipline_id})
    
    # Toujours limiter les choix aux disciplines de la fédération
    form.fields['discipline'].queryset = federation.disciplines.all()
    
    # Disciplines de la fédération pour le filtre
    disciplines = federation.disciplines.all()
    
    # Ajouter des classes CSS aux champs du formulaire
    form.fields['name'].widget.attrs.update({'class': 'form-control'})
    form.fields['discipline'].widget.attrs.update({'class': 'form-select'})
    form.fields['order'].widget.attrs.update({'class': 'form-control', 'min': 0})
    if hasattr(form.fields, 'description'):
        form.fields['description'].widget.attrs.update({'class': 'form-control', 'rows': 3})
    
    context = {
        'federation': federation,
        'grade_categories': grade_categories,
        'form': form,
        'disciplines': disciplines,
        'selected_discipline': selected_discipline_id,
        'selected_discipline_name': selected_discipline_name,
        'discipline_grades': discipline_grades,
        'tab': 'grades',  # Ajout pour la navigation par onglets
    }
    
    return render(request, 'competitions/federations/grades.html', context)

@login_required
@federation_admin_required
def federation_edit_grade(request, federation_id, grade_id):
    """Vue pour modifier une catégorie de grade existante."""
    federation = get_object_or_404(Federation, id=federation_id)
    grade = get_object_or_404(CategoryGrade, id=grade_id)
    
    # Vérifier que le grade appartient Ã  une discipline de la fédération
    if grade.discipline not in federation.disciplines.all():
        messages.error(request, _("Vous n'avez pas les permissions pour modifier ce grade"))
        return redirect('competitions:federations:grades', federation_id=federation_id)
    
    if request.method == 'POST':
        form = CategoryGradeForm(request.POST, instance=grade)
        if form.is_valid():
            updated_grade = form.save(commit=False)
            # Vérifier que la discipline appartient Ã  la fédération
            if updated_grade.discipline in federation.disciplines.all():
                updated_grade.save()
                messages.success(request, _("Grade mis Ã  jour avec succès"))
                return redirect('competitions:federations:grades', federation_id=federation_id)
            else:
                messages.error(request, _("La discipline sélectionnée n'appartient pas Ã  cette fédération"))
    else:
        form = CategoryGradeForm(instance=grade)
        # Limiter les choix aux disciplines de la fédération
        form.fields['discipline'].queryset = federation.disciplines.all()
    
    context = {
        'federation': federation,
        'grade': grade,
        'form': form,
        'edit_mode': True
    }
    
    return render(request, 'competitions/federations/edit_grade.html', context)

@login_required
def federation_trainings(request, federation_id):
    """Liste des formations de juges de la fédération."""
    # Assurez-vous que les modèles nécessaires sont importés
    from ..models import (
        Federation, Competition, Club, Judge, 
        JudgeTraining, JudgeTrainingRegistration
    )
    
    federation = get_object_or_404(Federation, pk=federation_id)
    
    # Vérifier les permissions
    if not request.user.is_federation_admin(federation) and not request.user.is_superuser:
        messages.error(request, _("Vous n'avez pas les droits pour accéder Ã  cette page."))
        return redirect('competitions:federations:federation_dashboard', federation_id=federation.id)
    
    # Vérifier que JudgeTraining est correctement défini
    if not JudgeTraining:
        messages.error(request, _("Le modèle JudgeTraining n'est pas disponible."))
        return redirect('competitions:federations:federation_dashboard', federation_id=federation.id)
    
    try:
        # Récupérer les formations
        trainings = JudgeTraining.objects.filter(federation=federation).order_by('-start_date')
        
        # Filtrer par statut si nécessaire
        status_filter = request.GET.get('status')
        if status_filter:
            trainings = trainings.filter(status=status_filter)
        
        # Filtrer par type si nécessaire
        type_filter = request.GET.get('type')
        if type_filter:
            trainings = trainings.filter(training_type=type_filter)
        
        # Statistiques avec gestion des erreurs
        stats = {
            'total_trainings': trainings.count(),
            'planned_trainings': trainings.filter(status='planned').count(),
            'ongoing_trainings': trainings.filter(status='ongoing').count(),
            'completed_trainings': trainings.filter(status='completed').count(),
        }
        
        # Vérifier que JudgeTrainingRegistration est correctement défini
        if JudgeTrainingRegistration:
            stats['total_participants'] = JudgeTrainingRegistration.objects.filter(
                training__federation=federation
            ).values('user').distinct().count()
        else:
            stats['total_participants'] = 0
            
        # Vérifier les autres modèles avant d'y accéder
        competition_count = 0
        club_count = 0
        judge_count = 0
        
        if hasattr(federation, 'disciplines'):
            competition_count = Competition.objects.filter(
                discipline__in=federation.disciplines.all()
            ).count() if Competition else 0
            
        if Club and hasattr(Club, 'federation'):
            club_count = Club.objects.filter(federation=federation).count()
            
        if Judge and hasattr(Judge, 'federation'):
            judge_count = Judge.objects.filter(federation=federation).count()
        
    except Exception as e:
        # Gérer les exceptions et afficher un message d'erreur
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Erreur dans federation_trainings: {str(e)}")
        
        messages.error(request, _("Une erreur est survenue lors du chargement des formations."))
        return redirect('competitions:federations:federation_dashboard', federation_id=federation.id)
    
    context = {
        'federation': federation,
        'trainings': trainings,
        'stats': stats,
        'title': _("Formations de juges"),
        'competition_count': competition_count,
        'club_count': club_count,
        'practitioner_count': 0,  # Cette valeur est toujours Ã  0 dans le code original
        'judge_count': judge_count,
        'tab': 'trainings'
    }
    
    return render(request, 'competitions/federations/trainings.html', context)


@login_required
def create_training(request, federation_id):
    """Crée une nouvelle formation pour les juges."""
    federation = get_object_or_404(Federation, pk=federation_id)
    
    # Vérifier les permissions
    if not request.user.is_federation_admin(federation) and not request.user.is_superuser:
        messages.error(request, _("Vous n'avez pas les droits pour accéder Ã  cette page."))
        return redirect('competitions:federations:federation_dashboard', federation_id=federation.id)
    
    if request.method == 'POST':
        form = JudgeTrainingForm(request.POST, federation=federation)
        if form.is_valid():
            training = form.save(commit=False)
            training.federation = federation
            training.save()
            
            # Enregistrer les relations ManyToMany
            form.save_m2m()
            
            messages.success(request, _("Formation créée avec succès."))
            return redirect('competitions:federations:trainings', federation_id=federation.id)
    else:
        form = JudgeTrainingForm(
            federation=federation,
            initial={'instructor': request.user}
        )
    
    context = {
        'federation': federation,
        'form': form,
        'title': _("Créer une formation"),
        'submit_text': _("Créer la formation"),
        'competition_count': Competition.objects.filter(discipline__in=federation.disciplines.all()).count(),
        'club_count': Club.objects.filter(federation=federation).count() if hasattr(Club, 'federation') else 0,
        'practitioner_count': 0,
        'judge_count': Judge.objects.filter(federation=federation).count() if hasattr(Judge, 'federation') else 0,
        'tab': 'trainings'
    }
    
    return render(request, 'competitions/federations/training_form.html', context)


@login_required
def edit_training(request, federation_id, training_id):
    """Modifie une formation existante."""
    federation = get_object_or_404(Federation, pk=federation_id)
    training = get_object_or_404(JudgeTraining, pk=training_id, federation=federation)
    
    # Vérifier les permissions
    if not request.user.is_federation_admin(federation) and not request.user.is_superuser:
        messages.error(request, _("Vous n'avez pas les droits pour accéder Ã  cette page."))
        return redirect('competitions:federations:federation_dashboard', federation_id=federation.id)
    
    if request.method == 'POST':
        form = JudgeTrainingForm(request.POST, instance=training, federation=federation)
        if form.is_valid():
            form.save()
            messages.success(request, _("Formation mise Ã  jour avec succès."))
            return redirect('competitions:federations:trainings', federation_id=federation.id)
    else:
        form = JudgeTrainingForm(instance=training, federation=federation)
    
    context = {
        'federation': federation,
        'form': form,
        'training': training,
        'title': _("Modifier la formation"),
        'submit_text': _("Enregistrer les modifications"),
        'competition_count': Competition.objects.filter(discipline__in=federation.disciplines.all()).count(),
        'club_count': Club.objects.filter(federation=federation).count() if hasattr(Club, 'federation') else 0,
        'practitioner_count': 0,
        'judge_count': Judge.objects.filter(federation=federation).count() if hasattr(Judge, 'federation') else 0,
        'tab': 'trainings'
    }
    
    return render(request, 'competitions/federations/training_form.html', context)


@login_required
def delete_training(request, federation_id, training_id):
    """Supprime une formation."""
    federation = get_object_or_404(Federation, pk=federation_id)
    training = get_object_or_404(JudgeTraining, pk=training_id, federation=federation)
    
    # Vérifier les permissions
    if not request.user.is_federation_admin(federation) and not request.user.is_superuser:
        messages.error(request, _("Vous n'avez pas les droits pour accéder Ã  cette page."))
        return redirect('competitions:federations:federation_dashboard', federation_id=federation.id)
    
    if request.method == 'POST':
        training.delete()
        messages.success(request, _("Formation supprimée avec succès."))
        return redirect('competitions:federations:trainings', federation_id=federation.id)
    
    context = {
        'federation': federation,
        'training': training,
        'title': _("Supprimer la formation"),
        'competition_count': Competition.objects.filter(discipline__in=federation.disciplines.all()).count(),
        'club_count': Club.objects.filter(federation=federation).count() if hasattr(Club, 'federation') else 0,
        'practitioner_count': 0,
        'judge_count': Judge.objects.filter(federation=federation).count() if hasattr(Judge, 'federation') else 0,
        'tab': 'trainings'
    }
    
    return render(request, 'competitions/federations/training_confirm_delete.html', context)


@login_required
def training_detail(request, federation_id, training_id):
    """Affiche les détails d'une formation."""
    federation = get_object_or_404(Federation, pk=federation_id)
    training = get_object_or_404(JudgeTraining, pk=training_id, federation=federation)
    
    # Récupérer les inscriptions
    registrations = training.registrations.select_related('user').order_by('status', '-registration_date')
    
    context = {
        'federation': federation,
        'training': training,
        'registrations': registrations,
        'title': training.title,
        'competition_count': Competition.objects.filter(discipline__in=federation.disciplines.all()).count(),
        'club_count': Club.objects.filter(federation=federation).count() if hasattr(Club, 'federation') else 0,
        'practitioner_count': 0,
        'judge_count': Judge.objects.filter(federation=federation).count() if hasattr(Judge, 'federation') else 0,
        'tab': 'trainings'
    }
    
    return render(request, 'competitions/federations/training_detail.html', context)


@login_required
def start_training(request, federation_id, training_id):
    """Démarre une formation (change le statut de 'planned' à 'ongoing')."""
    federation = get_object_or_404(Federation, pk=federation_id)
    training = get_object_or_404(JudgeTraining, pk=training_id, federation=federation)
    
    # Vérifier les permissions
    if not request.user.is_federation_admin(federation) and not request.user.is_superuser:
        messages.error(request, _("Vous n'avez pas les droits pour effectuer cette action."))
        return redirect('competitions:federations:training_detail', federation_id=federation.id, training_id=training.id)
    
    # Vérifier que la formation peut être démarrée
    if training.status != 'planned':
        messages.error(request, _("Cette formation ne peut pas être démarrée (statut: {}).").format(training.get_status_display()))
        return redirect('competitions:federations:training_detail', federation_id=federation.id, training_id=training.id)
    
    # Démarrer la formation
    training.status = 'ongoing'
    training.save()
    
    messages.success(request, _("La formation '{}' a été démarrée avec succès.").format(training.title))
    return redirect('competitions:federations:training_detail', federation_id=federation.id, training_id=training.id)


@login_required
def complete_training(request, federation_id, training_id):
    """Termine une formation (change le statut de 'ongoing' à 'completed')."""
    federation = get_object_or_404(Federation, pk=federation_id)
    training = get_object_or_404(JudgeTraining, pk=training_id, federation=federation)
    
    # Vérifier les permissions
    if not request.user.is_federation_admin(federation) and not request.user.is_superuser:
        messages.error(request, _("Vous n'avez pas les droits pour effectuer cette action."))
        return redirect('competitions:federations:training_detail', federation_id=federation.id, training_id=training.id)
    
    # Vérifier que la formation peut être terminée
    if training.status != 'ongoing':
        messages.error(request, _("Cette formation ne peut pas être terminée (statut: {}).").format(training.get_status_display()))
        return redirect('competitions:federations:training_detail', federation_id=federation.id, training_id=training.id)
    
    # Terminer la formation
    training.status = 'completed'
    training.completion_date = timezone.now().date()
    training.save()
    
    messages.success(request, _("La formation '{}' a été terminée avec succès.").format(training.title))
    return redirect('competitions:federations:training_detail', federation_id=federation.id, training_id=training.id)


@login_required
def federation_certifications(request, federation_id):
    """Liste des certifications de juges de la fédération."""
    federation = get_object_or_404(Federation, pk=federation_id)
    
    # Vérifier les permissions
    if not request.user.is_federation_admin(federation) and not request.user.is_superuser:
        messages.error(request, _("Vous n'avez pas les droits pour accéder Ã  cette page."))
        return redirect('competitions:federations:federation_dashboard', federation_id=federation.id)
    
    # Récupérer ou créer l'organisation de la fédération
    federation_org = federation.organization
    if not federation_org:
        # Créer automatiquement une organisation pour cette fédération
        try:
            from apps.organizations.models import Organization, OrganizationType
            federation_org = Organization.objects.create(
                name=federation.name,
                short_name=federation.name[:50] if federation.name else "Federation",
                organization_type=OrganizationType.NATIONAL_FEDERATION,
                description=federation.description or "",
                email=federation.contact_email or "",
                phone=federation.contact_phone or "",
                website=federation.website or "",
                country="",  # Laisser vide pour éviter l'erreur de longueur
                address=federation.address or "",
                city=federation.city or "",
                postal_code=(federation.postal_code or "")[:10],  # Limiter la longueur
                created_by=federation.owner,
                is_active=federation.is_active
            )
            
            # Copier les disciplines si elles existent
            if hasattr(federation, 'disciplines'):
                for discipline in federation.disciplines.all():
                    federation_org.disciplines.add(discipline)
            
            # Lier l'organisation à la fédération
            federation.organization = federation_org
            federation.save()
            
            messages.success(request, _("Organisation créée automatiquement pour cette fédération."))
            
        except Exception as e:
            messages.error(request, _("Impossible de créer l'organisation pour cette fédération: {}").format(str(e)))
            return redirect('competitions:federations:federation_dashboard', federation_id=federation.id)
    
    # Récupérer les certifications
    certifications = JudgeCertification.objects.filter(organization=federation_org).order_by('-created_at')
    
    # Filtrer par statut si nécessaire
    status_filter = request.GET.get('status')
    if status_filter:
        certifications = certifications.filter(is_active=(status_filter == 'active'))
    
    # Filtrer par discipline si nécessaire
    discipline_filter = request.GET.get('discipline')
    if discipline_filter:
        certifications = certifications.filter(discipline__id=discipline_filter)
    
    # Statistiques
    stats = {
        'total_certifications': certifications.count(),
        'active_certifications': certifications.filter(is_active=True).count(),
        'inactive_certifications': certifications.filter(is_active=False).count(),
    }
    
    context = {
        'federation': federation,
        'certifications': certifications,
        'stats': stats,
        'title': _("Certifications de juges"),
        'disciplines': federation.disciplines.all() if hasattr(federation, 'disciplines') else [],
        'tab': 'certifications'
    }
    
    return render(request, 'competitions/federations/certifications.html', context)

@login_required
def create_certification(request, federation_id):
    """Création d'une certification pour une fédération."""
    federation = get_object_or_404(Federation, pk=federation_id)
    
    # Vérifier les permissions
    if not request.user.is_federation_admin(federation) and not request.user.is_superuser:
        messages.error(request, _("Vous n'avez pas les droits pour accéder Ã  cette page."))
        return redirect('competitions:federations:federation_dashboard', federation_id=federation.id)
    
    if request.method == 'POST':
        form = JudgeCertificationForm(request.POST)
        if form.is_valid():
            certification = form.save(commit=False)
            certification.organization = federation.organization
            certification.save()
            
            messages.success(request, _("La certification a été créée avec succès."))
            return redirect('competitions:federations:certifications', federation_id=federation.id)
    else:
        form = JudgeCertificationForm()
    
    # Si des disciplines sont associées Ã  la fédération, limiter la liste
    if hasattr(federation, 'disciplines') and federation.disciplines.exists():
        form.fields['discipline'].queryset = federation.disciplines.all()
    
    context = {
        'federation': federation,
        'form': form,
        'title': _("Créer une certification"),
        'tab': 'certifications'
    }
    
    return render(request, 'competitions/federations/certification_form.html', context)

@login_required
def edit_certification(request, federation_id, certification_id):
    """Ã‰dition d'une certification."""
    federation = get_object_or_404(Federation, pk=federation_id)
    certification = get_object_or_404(JudgeCertification, pk=certification_id, organization=federation.organization)
    
    # Vérifier les permissions
    if not request.user.is_federation_admin(federation) and not request.user.is_superuser:
        messages.error(request, _("Vous n'avez pas les droits pour accéder Ã  cette page."))
        return redirect('competitions:federations:federation_dashboard', federation_id=federation.id)
    
    if request.method == 'POST':
        form = JudgeCertificationForm(request.POST, instance=certification)
        if form.is_valid():
            form.save()
            
            messages.success(request, _("La certification a été mise Ã  jour avec succès."))
            return redirect('competitions:federations:certifications', federation_id=federation.id)
    else:
        form = JudgeCertificationForm(instance=certification)
    
    # Si des disciplines sont associées Ã  la fédération, limiter la liste
    if hasattr(federation, 'disciplines') and federation.disciplines.exists():
        form.fields['discipline'].queryset = federation.disciplines.all()
    
    context = {
        'federation': federation,
        'certification': certification,
        'form': form,
        'title': _("Modifier la certification"),
        'tab': 'certifications'
    }
    
    return render(request, 'competitions/federations/certification_form.html', context)

@login_required
def delete_certification(request, federation_id, certification_id):
    """Suppression d'une certification."""
    federation = get_object_or_404(Federation, pk=federation_id)
    certification = get_object_or_404(JudgeCertification, pk=certification_id, organization=federation.organization)
    
    # Vérifier les permissions
    if not request.user.is_federation_admin(federation) and not request.user.is_superuser:
        messages.error(request, _("Vous n'avez pas les droits pour accéder Ã  cette page."))
        return redirect('competitions:federations:federation_dashboard', federation_id=federation.id)
    
    if request.method == 'POST':
        # Vérifier si la certification a des inscriptions
        if certification.registrations.exists():
            messages.error(request, _(
                "Cette certification ne peut pas être supprimée car elle a des inscriptions associées. "
                "Vous pouvez la désactiver Ã  la place."
            ))
            return redirect('competitions:federations:certifications', federation_id=federation.id)
        
        certification.delete()
        messages.success(request, _("La certification a été supprimée avec succès."))
        return redirect('competitions:federations:certifications', federation_id=federation.id)
    
    context = {
        'federation': federation,
        'certification': certification,
        'title': _("Supprimer la certification"),
        'tab': 'certifications'
    }
    
    return render(request, 'competitions/federations/confirm_delete.html', context)

@login_required
def certification_registrations(request, federation_id, certification_id):
    """Liste des inscriptions Ã  une certification."""
    federation = get_object_or_404(Federation, pk=federation_id)
    certification = get_object_or_404(JudgeCertification, pk=certification_id, organization=federation.organization)
    
    # Vérifier les permissions
    if not request.user.is_federation_admin(federation) and not request.user.is_superuser:
        messages.error(request, _("Vous n'avez pas les droits pour accéder Ã  cette page."))
        return redirect('competitions:federations:federation_dashboard', federation_id=federation.id)
    
    # Récupérer les inscriptions
    registrations = certification.registrations.all().order_by('-registration_date')
    
    # Filtrer par statut si nécessaire
    status_filter = request.GET.get('status')
    if status_filter:
        registrations = registrations.filter(status=status_filter)
    
    context = {
        'federation': federation,
        'certification': certification,
        'registrations': registrations,
        'title': _("Inscriptions Ã  la certification"),
        'tab': 'certifications'
    }
    
    return render(request, 'competitions/federations/certification_registrations.html', context)

@login_required
def review_certification_registration(request, federation_id, registration_id):
    """Examen d'une demande de certification."""
    federation = get_object_or_404(Federation, pk=federation_id)
    registration = get_object_or_404(
        CertificationRegistration, 
        pk=registration_id, 
        certification__organization=federation.organization
    )
    
    # Vérifier les permissions
    if not request.user.is_federation_admin(federation) and not request.user.is_superuser:
        messages.error(request, _("Vous n'avez pas les droits pour accéder Ã  cette page."))
        return redirect('competitions:federations:federation_dashboard', federation_id=federation.id)
    
    if request.method == 'POST':
        form = ReviewCertificationRegistrationForm(request.POST, instance=registration)
        if form.is_valid():
            decision = form.cleaned_data['decision']
            
            if decision == 'approve':
                registration.approve(reviewer=request.user)
                messages.success(request, _("L'inscription a été approuvée."))
            else:
                registration.reject(reviewer=request.user, notes=form.cleaned_data['notes'])
                messages.success(request, _("L'inscription a été rejetée."))
            
            return redirect('competitions:federations:certification_registrations', 
                           federation_id=federation.id, 
                           certification_id=registration.certification.id)
    else:
        form = ReviewCertificationRegistrationForm(instance=registration)
    
    context = {
        'federation': federation,
        'registration': registration,
        'form': form,
        'title': _("Examiner la demande de certification"),
        'tab': 'certifications'
    }
    
    return render(request, 'competitions/federations/review_registration.html', context)

@login_required
def certification_detail(request, federation_id, certification_id):
    """Affichage détaillé d'une certification."""
    federation = get_object_or_404(Federation, pk=federation_id)
    certification = get_object_or_404(JudgeCertification, pk=certification_id, organization=federation.organization)
    
    # Vérifier les permissions
    if not request.user.is_federation_admin(federation) and not request.user.is_superuser:
        messages.error(request, _("Vous n'avez pas les droits pour accéder Ã  cette page."))
        return redirect('competitions:federations:federation_dashboard', federation_id=federation.id)
    
    # Récupérer les inscriptions Ã  cette certification
    registrations = certification.registrations.all().order_by('-registration_date')
    
    # Statistiques
    stats = {
        'total_registrations': registrations.count(),
        'pending_registrations': registrations.filter(status='pending').count(),
        'approved_registrations': registrations.filter(status='approved').count(),
        'rejected_registrations': registrations.filter(status='rejected').count(),
        'completed_registrations': registrations.filter(status='completed').count(),
    }
    
    context = {
        'federation': federation,
        'certification': certification,
        'registrations': registrations,
        'stats': stats,
        'title': certification.title,
        'tab': 'certifications'
    }
    
    return render(request, 'competitions/federations/certification_detail.html', context)

@login_required
def federation_calendar(request, federation_id=None):
    """
    Affiche le calendrier des compétitions pour une fédération.
    Les compétitions sont organisées par mois.
    """
    # Récupérer la fédération spécifiée par l'ID de l'URL
    federation = None
    if federation_id:
        try:
            federation = Federation.objects.get(id=federation_id)
        except Federation.DoesNotExist:
            # Gérer le cas oÃ¹ la fédération n'existe pas
            messages.error(request, _("La fédération demandée n'existe pas."))
            return redirect('competitions:federations:list')
    elif hasattr(request.user, 'profile') and hasattr(request.user.profile, 'federation'):
        federation = request.user.profile.federation
    else:
        # Rechercher dans les fédérations administrées
        federations = request.user.get_administered_federations() if hasattr(request.user, 'get_administered_federations') else []
        federation = federations.first() if federations.exists() else None
    
    # Définir la période Ã  afficher (par défaut: 3 mois passés et 12 mois Ã  venir)
    today = timezone.now().date()
    start_date = today - timedelta(days=90)  # 3 mois en arrière
    end_date = today + timedelta(days=365)   # 12 mois en avant
    
    # Filtrer les compétitions
    competitions = Competition.objects.filter(
        start_date__gte=start_date,
        start_date__lte=end_date
    ).order_by('start_date')
    
    # Si une fédération est spécifiée, filtrer les compétitions de cette fédération
    if federation:
        # Adapter ce filtre selon votre modèle de données
        if hasattr(Competition, 'federation'):
            competitions = competitions.filter(federation=federation)
        # Si les compétitions sont liées via les clubs
        elif hasattr(Competition, 'club') and hasattr(Club, 'federation'):
            competitions = competitions.filter(organizing_organization__old_federation_id=federation.id)
    
    # Organiser les compétitions par mois pour l'affichage calendrier
    calendar_data = {}
    for competition in competitions:
        month_key = competition.start_date.strftime('%Y-%m')
        month_name = competition.start_date.strftime('%B %Y')
        
        if month_key not in calendar_data:
            calendar_data[month_key] = {
                'name': month_name,
                'competitions': []
            }
        
        calendar_data[month_key]['competitions'].append(competition)
    
    # Trier les mois chronologiquement
    sorted_calendar = {}
    for key in sorted(calendar_data.keys()):
        sorted_calendar[key] = calendar_data[key]
    
    # Préparer les données pour l'affichage du calendrier
    # (Cette partie peut Ãªtre adaptée selon votre logique d'affichage)
    current_month = today.strftime('%B %Y')
    
    context = {
        'title': _('Calendrier des compétitions'),
        'federation': federation,
        'calendar': sorted_calendar,
        'competitions': competitions,
        'today': today,
        'current_month': current_month,
        'can_add_event': request.user.has_perm('competitions.add_event'),
        'can_edit_event': request.user.has_perm('competitions.change_event'),
    }
    
    return render(request, 'competitions/federations/calendar.html', context)

# Les vues manquantes pour gérer les rÃ´les (create_federation_role, etc.)
@login_required
@federation_admin_required
def create_federation_role(request, federation_id):
    """
    Crée un nouveau rÃ´le dans la fédération.
    """
    federation = get_object_or_404(Federation, id=federation_id)
    
    # Vérifier que l'utilisateur a les droits sur cette fédération
    if not (request.user.is_federation_admin(federation) or request.user == federation.owner):
        messages.error(request, _("Vous n'avez pas les droits pour créer des rÃ´les dans cette fédération."))
        return redirect('competitions:federations:roles', federation_id=federation_id)
    
    # Logique pour créer un rÃ´le
    if request.method == 'POST':
        # Récupérer les données du formulaire
        user_id = request.POST.get('user_id')
        role_name = request.POST.get('role')
        is_primary = request.POST.get('is_primary') == 'on'
        
        # Validation des données
        if not user_id or not role_name:
            messages.error(request, _("Veuillez remplir tous les champs obligatoires."))
            return redirect('competitions:federations:create_role', federation_id=federation_id)
        
        try:
            # Récupérer l'utilisateur
            from django.contrib.auth import get_user_model
            User = get_user_model()
            user = User.objects.get(id=user_id)
            
            # Créer le rÃ´le d'administrateur
            admin = FederationAdministrator.objects.create(
                user=user,
                federation=federation,
                role=role_name,
                is_primary=is_primary
            )
            
            # Si l'utilisateur a un profil, mettre Ã  jour son rÃ´le
            if hasattr(user, 'profile'):
                user.profile.role = 'federation_admin'
                user.profile.save()
            
            messages.success(request, _("Le rÃ´le a été créé avec succès."))
            return redirect('competitions:federations:roles', federation_id=federation_id)
        except Exception as e:
            logger.error(f"Erreur lors de la création du rÃ´le: {str(e)}")
            messages.error(request, _("Une erreur est survenue lors de la création du rÃ´le: {}").format(str(e)))
    
    # Récupérer les utilisateurs qui ne sont pas déjÃ  administrateurs
    from django.contrib.auth import get_user_model
    User = get_user_model()
    existing_admin_ids = FederationAdministrator.objects.filter(
        federation=federation
    ).values_list('user_id', flat=True)
    
    available_users = User.objects.exclude(id__in=existing_admin_ids)
    
    context = {
        'federation': federation,
        'available_users': available_users,
        'role_choices': FederationAdministrator._meta.get_field('role').choices,
    }
    
    return render(request, 'competitions/federations/create_role.html', context)

@login_required
@federation_admin_required
def update_federation_role(request, federation_id, role_id):
    """
    Met Ã  jour un rÃ´le existant dans la fédération.
    """
    federation = get_object_or_404(Federation, id=federation_id)
    role = get_object_or_404(FederationAdministrator, id=role_id, federation=federation)
    
    # Vérifier que l'utilisateur a les droits sur cette fédération
    if not (request.user.is_federation_admin(federation) or request.user == federation.owner):
        messages.error(request, _("Vous n'avez pas les droits pour modifier des rÃ´les dans cette fédération."))
        return redirect('competitions:federations:roles', federation_id=federation_id)
    
    if request.method == 'POST':
        # Récupérer les données du formulaire
        role_name = request.POST.get('role')
        is_primary = request.POST.get('is_primary') == 'on'
        
        # Validation des données
        if not role_name:
            messages.error(request, _("Veuillez remplir tous les champs obligatoires."))
            return redirect('competitions:federations:update_role', federation_id=federation_id, role_id=role_id)
        
        try:
            # Si on fait de cet admin un admin principal et qu'il ne l'était pas déjÃ 
            if is_primary and not role.is_primary:
                # Retirer ce statut aux autres
                FederationAdministrator.objects.filter(
                    federation=federation, 
                    is_primary=True
                ).exclude(id=role.id).update(is_primary=False)
            
            # Mettre Ã  jour le rÃ´le
            role.role = role_name
            role.is_primary = is_primary
            role.save()
            
            messages.success(request, _("Le rÃ´le a été mis Ã  jour avec succès."))
            return redirect('competitions:federations:roles', federation_id=federation_id)
        except Exception as e:
            logger.error(f"Erreur lors de la mise Ã  jour du rÃ´le: {str(e)}")
            messages.error(request, _("Une erreur est survenue lors de la mise Ã  jour du rÃ´le: {}").format(str(e)))
    
    context = {
        'federation': federation,
        'role': role,
        'role_choices': FederationAdministrator._meta.get_field('role').choices,
    }
    
    return render(request, 'competitions/federations/update_role.html', context)

@login_required
@federation_admin_required
def delete_federation_role(request, federation_id, role_id):
    """
    Supprime un rÃ´le de la fédération.
    """
    federation = get_object_or_404(Federation, id=federation_id)
    role = get_object_or_404(FederationAdministrator, id=role_id, federation=federation)
    
    # Vérifier que l'utilisateur a les droits sur cette fédération
    if not (request.user.is_federation_admin(federation) or request.user == federation.owner):
        messages.error(request, _("Vous n'avez pas les droits pour supprimer des rÃ´les dans cette fédération."))
        return redirect('competitions:federations:roles', federation_id=federation_id)
    
    # EmpÃªcher la suppression du dernier administrateur principal
    if role.is_primary and FederationAdministrator.objects.filter(federation=federation, is_primary=True).count() <= 1:
        messages.error(request, _("Impossible de supprimer le dernier administrateur principal."))
        return redirect('competitions:federations:roles', federation_id=federation_id)
    
    if request.method == 'POST':
        try:
            user = role.user
            role.delete()
            
            messages.success(request, _("Le rÃ´le a été supprimé avec succès."))
            return redirect('competitions:federations:roles', federation_id=federation_id)
        except Exception as e:
            logger.error(f"Erreur lors de la suppression du rÃ´le: {str(e)}")
            messages.error(request, _("Une erreur est survenue lors de la suppression du rÃ´le: {}").format(str(e)))
    
    context = {
        'federation': federation,
        'role': role,
    }
    
    return render(request, 'competitions/federations/delete_role.html', context)


@login_required
@federation_admin_required
def club_detail(request, federation_id, club_id):
    """
    Affiche les détails d'un club spécifique lié Ã  une fédération.
    """
    federation = get_object_or_404(Federation, id=federation_id)
    club = get_object_or_404(Club, id=club_id, federation=federation)
    
    # Récupérer les données supplémentaires nécessaires
    practitioners = club.practitioners.all().order_by('last_name', 'first_name')
    administrators = ClubAdministrator.objects.filter(club=club)
    
    context = {
        'federation': federation,
        'club': club,
        'practitioners': practitioners,
        'administrators': administrators,
        # Autres données de contexte selon vos besoins
    }
    
    return render(request, 'competitions/federations/club_detail.html', context)

@login_required
@federation_admin_required
def federation_users(request, federation_id):
    """
    Liste et gestion des utilisateurs associés Ã  une fédération.
    Inclut les administrateurs de la fédération et les responsables des clubs affiliés.
    """
    federation = get_object_or_404(Federation, id=federation_id)
    
    # Récupérer les administrateurs de la fédération
    federation_administrators = FederationAdministrator.objects.filter(
        federation=federation
    ).select_related('user').order_by('-is_primary', 'role', 'user__last_name')
    
    # Récupérer les clubs affiliés Ã  la fédération via organization
    try:
        from apps.organizations.models import Organization
        federation_org = federation.as_organization
        if federation_org:
            # Clubs affiliés via le système d'organizations
            affiliated_orgs = Organization.objects.filter(
                parent_affiliations__parent_organization=federation_org,
                organization_type='club'
            )
            # Récupérer les clubs correspondants
            clubs = Club.objects.filter(organization__in=affiliated_orgs)
        else:
            # Fallback vers l'ancienne relation si elle existe
            clubs = Club.objects.filter(old_federation_id=federation.id) if hasattr(Club, 'old_federation_id') else Club.objects.none()
    except ImportError:
        # Si le modèle Organization n'existe pas, utiliser l'ancienne relation
        clubs = Club.objects.filter(old_federation_id=federation.id) if hasattr(Club, 'old_federation_id') else Club.objects.none()
    
    # Récupérer les administrateurs des clubs affiliés
    club_administrators = ClubAdministrator.objects.filter(
        club__in=clubs
    ).select_related('user', 'club').order_by('club__name', '-is_primary', 'user__last_name')
    
    # Récupérer les juges associés Ã  la fédération via organization
    try:
        # Utiliser les clubs affiliés pour récupérer les juges
        judges = Judge.objects.filter(
            organization__in=affiliated_orgs if 'affiliated_orgs' in locals() else clubs.values_list('organization', flat=True)
        ).select_related('practitioner', 'user').order_by('practitioner__last_name')
    except:
        # Fallback si la relation directe existe encore
        if hasattr(Judge, 'old_federation_id'):
            judges = Judge.objects.filter(
                old_federation_id=federation.id
            ).select_related('practitioner', 'user').order_by('practitioner__last_name')
        else:
            judges = Judge.objects.none()
    
    # Traitement du formulaire d'ajout d'administrateur de fédération
    if request.method == 'POST' and 'action' in request.POST:
        action = request.POST.get('action')
        
        if action == 'add_admin':
            # Traitement du formulaire d'ajout d'administrateur
            user_id = request.POST.get('user_id')
            role = request.POST.get('role')
            is_primary = request.POST.get('is_primary') == 'on'
            
            if user_id and role:
                try:
                    user = User.objects.get(id=user_id)
                    
                    # Vérifier si l'association existe déjÃ 
                    admin_exists = FederationAdministrator.objects.filter(
                        user=user, 
                        federation=federation
                    ).exists()
                    
                    if admin_exists:
                        messages.warning(request, _("Cet utilisateur est déjÃ  administrateur de cette fédération."))
                    else:
                        # Créer l'association administrateur de fédération
                        FederationAdministrator.objects.create(
                            user=user,
                            federation=federation,
                            role=role,
                            is_primary=is_primary
                        )
                        
                        # Mettre Ã  jour le profil utilisateur si nécessaire
                        if hasattr(user, 'profile') and user.profile:
                            profile = user.profile
                            if profile.role != 'federation_admin':
                                profile.role = 'federation_admin'
                                profile.save(update_fields=['role'])
                        
                        messages.success(request, _("Administrateur ajouté avec succès !"))
                    
                    return redirect('competitions:federations:users', federation_id=federation.id)
                except User.DoesNotExist:
                    messages.error(request, _("Utilisateur non trouvé."))
        
        elif action == 'remove_admin':
            # Traitement de la suppression d'un administrateur
            admin_id = request.POST.get('admin_id')
            
            if admin_id:
                try:
                    admin = FederationAdministrator.objects.get(id=admin_id, federation=federation)
                    
                    # Ne pas supprimer l'administrateur principal si c'est le seul
                    if admin.is_primary and federation_administrators.filter(is_primary=True).count() <= 1:
                        messages.error(request, _("Impossible de supprimer le seul administrateur principal."))
                    else:
                        admin.delete()
                        messages.success(request, _("Administrateur supprimé avec succès !"))
                    
                    return redirect('competitions:federations:users', federation_id=federation.id)
                except FederationAdministrator.DoesNotExist:
                    messages.error(request, _("Administrateur non trouvé."))
    
    # Récupérer les utilisateurs disponibles pour l'ajout comme administrateurs
    # (ceux qui ne sont pas déjÃ  administrateurs de cette fédération)
    existing_admin_ids = federation_administrators.values_list('user_id', flat=True)
    available_users = User.objects.exclude(id__in=existing_admin_ids).order_by('last_name', 'first_name')
    
    # Préparation des rÃ´les d'administrateur pour le formulaire
    admin_roles = dict(FederationAdministrator._meta.get_field('role').choices)
    
    context = {
        'federation': federation,
        'federation_administrators': federation_administrators,
        'club_administrators': club_administrators,
        'judges': judges,
        'available_users': available_users,
        'admin_roles': admin_roles,
        'clubs': clubs,
    }
    
    return render(request, 'competitions/federations/users.html', context)

@login_required
@federation_admin_required
def role_users(request, federation_id, role_name):
    """
    Affiche les utilisateurs ayant un rÃ´le spécifique dans une fédération.
    """
    federation = get_object_or_404(Federation, id=federation_id)
    
    # Récupérer les utilisateurs selon le rÃ´le
    if role_name == 'admin':
        # Administrateurs de la fédération
        users = User.objects.filter(
            federationadministrator__federation=federation,
            federationadministrator__role__in=['owner', 'admin']
        ).distinct()
        role_display_name = _("Administrateurs")
    elif role_name == 'manager':
        # Gestionnaires
        users = User.objects.filter(
            federationadministrator__federation=federation,
            federationadministrator__role='manager'
        ).distinct()
        role_display_name = _("Gestionnaires")
    elif role_name == 'secretary':
        # Secrétaires
        users = User.objects.filter(
            federationadministrator__federation=federation,
            federationadministrator__role='secretary'
        ).distinct()
        role_display_name = _("Secrétaires")
    elif role_name == 'treasurer':
        # Trésoriers
        users = User.objects.filter(
            federationadministrator__federation=federation,
            federationadministrator__role='treasurer'
        ).distinct()
        role_display_name = _("Trésoriers")
    elif role_name == 'tech_director':
        # Directeurs techniques
        users = User.objects.filter(
            federationadministrator__federation=federation,
            federationadministrator__role='tech_director'
        ).distinct()
        role_display_name = _("Directeurs techniques")
    else:
        # RÃ´le personnalisé ou inconnu
        users = User.objects.filter(
            federationadministrator__federation=federation,
            federationadministrator__role=role_name
        ).distinct()
        role_display_name = role_name.replace('_', ' ').title()
    
    # Pagination
    paginator = Paginator(users, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'federation': federation,
        'role_name': role_name,
        'role_display_name': role_display_name,
        'page_obj': page_obj,
        'total_users': users.count(),
    }
    
    return render(request, 'competitions/federations/role_users.html', context)


@login_required
def download_sample_clubs(request, federation_id):
    """Télécharge un fichier exemple pour l'import de clubs."""
    federation = get_object_or_404(Federation, id=federation_id)
    export_format = request.GET.get('format', 'csv')
    
    if export_format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="exemple_clubs.csv"'
        
        writer = csv.writer(response)
        # En-têtes
        writer.writerow([
            'Nom', 'Adresse', 'Ville', 'Code Postal', 'Email', 'Téléphone', 
            'Site Web', 'Description'
        ])
        
        # Exemple de données
        writer.writerow([
            'Club Exemple',
            '123 Rue de la Paix',
            'Paris',
            '75001',
            'contact@clubexemple.fr',
            '01 23 45 67 89',
            'https://www.clubexemple.fr',
            'Description du club exemple'
        ])
        
    elif export_format == 'excel':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="exemple_clubs.xlsx"'
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Exemple Clubs"
        
        # Style d'en-tête
        header_fill = PatternFill(start_color="00CCCCCC", end_color="00CCCCCC", fill_type="solid")
        header_font = Font(bold=True)
        
        # En-têtes
        headers = [
            'Nom', 'Adresse', 'Ville', 'Code Postal', 'Email', 'Téléphone', 
            'Site Web', 'Description'
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Exemple de données
        ws.cell(row=2, column=1, value='Club Exemple')
        ws.cell(row=2, column=2, value='123 Rue de la Paix')
        ws.cell(row=2, column=3, value='Paris')
        ws.cell(row=2, column=4, value='75001')
        ws.cell(row=2, column=5, value='contact@clubexemple.fr')
        ws.cell(row=2, column=6, value='01 23 45 67 89')
        ws.cell(row=2, column=7, value='https://www.clubexemple.fr')
        ws.cell(row=2, column=8, value='Description du club exemple')
        
        # Ajuster la largeur des colonnes
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        
        wb.save(response)
    
    return response





# ========================================
# VUES POUR LES EXAMENS
# ========================================

@login_required
def federation_examens(request, federation_id):
    """Liste des examens de la fédération."""
    federation = get_object_or_404(Federation, id=federation_id)
    
    # Récupérer ou créer l'organisation de la fédération
    federation_org = federation.organization
    if not federation_org:
        # Créer automatiquement une organisation pour cette fédération
        try:
            from apps.organizations.models import Organization, OrganizationType
            federation_org = Organization.objects.create(
                name=federation.name,
                short_name=federation.name[:50] if federation.name else "Federation",
                organization_type=OrganizationType.NATIONAL_FEDERATION,
                description=federation.description or "",
                email=federation.contact_email or "",
                phone=federation.contact_phone or "",
                website=federation.website or "",
                country="",  # Laisser vide pour éviter l'erreur de longueur
                address=federation.address or "",
                city=federation.city or "",
                postal_code=(federation.postal_code or "")[:10],  # Limiter la longueur
                created_by=federation.owner,
                is_active=federation.is_active
            )
            
            # Copier les disciplines si elles existent
            if hasattr(federation, 'disciplines'):
                for discipline in federation.disciplines.all():
                    federation_org.disciplines.add(discipline)
            
            # Lier l'organisation à la fédération
            federation.organization = federation_org
            federation.save()
            
            messages.success(request, _("Organisation créée automatiquement pour cette fédération."))
            
        except Exception as e:
            messages.error(request, _("Impossible de créer l'organisation pour cette fédération: {}").format(str(e)))
            return redirect('competitions:federations:federation_dashboard', federation_id=federation.id)
    
    # Récupérer les examens - Gestion temporaire si la table n'existe pas
    try:
        from ..models.certifications import Exam
        from django.utils import timezone
        examens = Exam.objects.filter(organization=federation_org).order_by('-start_date')
        
        # Statistiques
        total_examens = examens.count()
        upcoming_examens = examens.filter(status='published', start_date__gte=timezone.now().date()).count()
        ongoing_examens = examens.filter(status='ongoing').count()
    except ImportError as e:
        # Problème d'import du modèle
        messages.error(request, _("Erreur d'import du modèle Exam: {}").format(str(e)))
        examens = []
        total_examens = 0
        upcoming_examens = 0
        ongoing_examens = 0
    except Exception as e:
        # La table n'existe pas encore - migrations manquantes ou autre erreur
        messages.warning(request, _("Les examens ne sont pas encore disponibles. Erreur: {}").format(str(e)))
        examens = []
        total_examens = 0
        upcoming_examens = 0
        ongoing_examens = 0
    
    context = {
        'federation': federation,
        'examens': examens,
        'stats': {
            'total': total_examens,
            'upcoming': upcoming_examens,
            'ongoing': ongoing_examens,
        },
        'title': _('Examens - {0}').format(federation.name),
    }
    
    return render(request, 'competitions/federations/examens/list.html', context)


@login_required
def create_examen(request, federation_id):
    """Crée un nouvel examen."""
    federation = get_object_or_404(Federation, id=federation_id)
    
    # Récupérer l'organisation de la fédération
    federation_org = federation.organization
    if not federation_org:
        messages.error(request, _("Cette fédération n'a pas d'organisation associée."))
        return redirect('competitions:federations:federation_dashboard', federation_id=federation.id)
    
    from ..models.certifications import Exam
    from ..forms.certifications import ExamForm
    
    if request.method == 'POST':
        form = ExamForm(request.POST, federation=federation)
        if form.is_valid():
            examen = form.save(commit=False)
            examen.organization = federation_org
            examen.created_by = request.user
            examen.save()
            form.save_m2m()  # Pour les relations ManyToMany
            messages.success(request, _("L'examen a été créé avec succès."))
            return redirect('competitions:federations:examens', federation_id=federation.id)
    else:
        form = ExamForm(federation=federation)
    
    context = {
        'federation': federation,
        'form': form,
        'title': _('Créer un examen'),
    }
    
    return render(request, 'competitions/federations/examens/form.html', context)


@login_required  
def edit_examen(request, federation_id, examen_id):
    """Modifie un examen existant."""
    federation = get_object_or_404(Federation, id=federation_id)
    
    from ..models.certifications import Exam
    examen = get_object_or_404(Exam, id=examen_id, organization=federation.organization)
    
    from ..forms.certifications import ExamForm
    
    if request.method == 'POST':
        form = ExamForm(request.POST, instance=examen, federation=federation)
        if form.is_valid():
            form.save()
            messages.success(request, _("L'examen a été mis à jour avec succès."))
            return redirect('competitions:federations:examens', federation_id=federation.id)
    else:
        form = ExamForm(instance=examen, federation=federation)
    
    context = {
        'federation': federation,
        'examen': examen,
        'form': form,
        'title': _('Modifier l\'examen'),
    }
    
    return render(request, 'competitions/federations/examens/form.html', context)


@login_required
def delete_examen(request, federation_id, examen_id):
    """Supprime un examen."""
    federation = get_object_or_404(Federation, id=federation_id)
    
    from ..models.certifications import Exam
    examen = get_object_or_404(Exam, id=examen_id, organization=federation.organization)
    
    if request.method == 'POST':
        examen.delete()
        messages.success(request, _("L'examen a été supprimé avec succès."))
        return redirect('competitions:federations:examens', federation_id=federation.id)
    
    context = {
        'federation': federation,
        'examen': examen,
        'title': _('Supprimer l\'examen'),
    }
    
    return render(request, 'competitions/federations/examens/delete.html', context)


@login_required
def examen_detail(request, federation_id, examen_id):
    """Affiche les détails d'un examen."""
    federation = get_object_or_404(Federation, id=federation_id)
    
    from ..models.certifications import Exam, ExamRegistration
    examen = get_object_or_404(Exam, id=examen_id, organization=federation.organization)
    
    # Récupérer les inscriptions
    registrations = ExamRegistration.objects.filter(exam=examen).select_related('user').order_by('status', '-registration_date')
    
    # Statistiques
    total_registrations = registrations.count()
    approved_registrations = registrations.filter(status='approved').count()
    pending_registrations = registrations.filter(status='pending').count()
    
    context = {
        'federation': federation,
        'examen': examen,
        'registrations': registrations,
        'stats': {
            'total': total_registrations,
            'approved': approved_registrations,
            'pending': pending_registrations,
            'available_spots': examen.available_spots,
        },
        'title': examen.title,
    }
    
    return render(request, 'competitions/federations/examens/detail.html', context)


@login_required
def federation_certification_detail(request, federation_id, certification_id):
    """Affiche les détails d'une certification."""
    federation = get_object_or_404(Federation, id=federation_id)
    
    from ..models.certifications import JudgeCertification, CertificationRegistration
    certification = get_object_or_404(JudgeCertification, id=certification_id, organization=federation.organization)
    
    # Récupérer les inscriptions
    registrations = CertificationRegistration.objects.filter(certification=certification).select_related('user').order_by('status', '-registration_date')
    
    context = {
        'federation': federation,
        'certification': certification,
        'registrations': registrations,
        'title': certification.title,
    }
    
    return render(request, 'competitions/federations/certifications/detail.html', context)

